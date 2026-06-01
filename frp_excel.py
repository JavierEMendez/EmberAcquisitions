"""FRP Excel builder — renders the BS / IS / SCF as the same .xlsx
workbook the accountant's frp_builder.py script generates.

Ported from frp_builder.py (lifted verbatim from the script in
docs/. Originating-team-built; we use it here so the dashboard's
"Download FRP" output matches what the accounting team is already
distributing.

Account classification predicates are imported from tb_parser so the
dashboard's on-screen render and this Excel export use the same
rules — change once, both stay consistent.
"""
from __future__ import annotations
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from tb_parser import (
    RELATED_PARTY_ACCTS,
    is_cash, is_restricted_bonds, is_receivables, is_prepaids,
    is_land, is_development, is_contra_sales, is_contra_mud,
    is_promissory, is_ppe,
    is_other_current_liab, is_other_op_liab,
    is_dev_loan, is_bond_lt, is_members_eq, is_income_stmt,
    is_re_invest,
    sum_accounts, get_prefix,
    _normalize_accounts,
)


# ─── Styling constants ───────────────────────────────────────────
NUM_FMT = '#,##0;(#,##0);"-"'
FONT_HEADER    = Font(name="Arial", size=11, bold=True)
FONT_DATA      = Font(name="Arial", size=10)
FONT_DATA_BOLD = Font(name="Arial", size=10, bold=True)
FONT_TITLE     = Font(name="Arial", size=12, bold=True)
THIN_BOTTOM    = Border(bottom=Side(style="thin"))
DOUBLE_BOTTOM  = Border(bottom=Side(style="double"))


# ─── Sheet/line helpers ──────────────────────────────────────────
def _setup_sheet(ws, title_lines, col_headers, col_widths):
    """Write title rows, column headers, set column widths.
    Returns the row number where the first data line should go."""
    row = 1
    for line in title_lines:
        c = ws.cell(row=row, column=1, value=line)
        c.font = FONT_TITLE
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=len(col_widths))
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        row += 1
    row += 1  # blank line
    if col_headers:
        for i, hdr in enumerate(col_headers):
            c = ws.cell(row=row, column=i + 1, value=hdr)
            c.font = FONT_HEADER
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            c.border = THIN_BOTTOM
        row += 1
    for i, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    return row


def _write_line(ws, row, label, values, indent=0, bold=False,
                underline=None, number_format=NUM_FMT):
    """Write a labeled line with one or more numeric values. Returns
    the next row number. `values` can contain floats (rounded to int)
    or formula strings starting with '='."""
    prefix = "  " * indent
    c = ws.cell(row=row, column=1, value=prefix + label)
    c.font = FONT_DATA_BOLD if bold else FONT_DATA
    c.alignment = Alignment(horizontal="left")
    for i, val in enumerate(values):
        cell = ws.cell(row=row, column=2 + i)
        if isinstance(val, str) and val.startswith("="):
            cell.value = val
        elif isinstance(val, (int, float)):
            cell.value = round(float(val))
        else:
            cell.value = val
        cell.font = FONT_DATA_BOLD if bold else FONT_DATA
        cell.number_format = number_format
        cell.alignment = Alignment(horizontal="right")
        if underline == "thin":
            cell.border = THIN_BOTTOM
        elif underline == "double":
            cell.border = DOUBLE_BOTTOM
    return row + 1


# ─── Balance Sheet ───────────────────────────────────────────────
def build_balance_sheet(ws, monthly_accounts, full_name, date_label):
    """Render the BS into the given worksheet, using the cached
    monthly TB accounts. Returns dict of row-marker keys → row
    numbers (used internally; not currently consumed externally)."""
    accounts = _normalize_accounts(monthly_accounts)
    title_lines = [full_name, "Balance Sheet", date_label]
    col_widths = [50, 18]
    row = _setup_sheet(ws, title_lines, ["", "Balance"], col_widths)
    row += 1

    def gs(predicate): return sum_accounts(accounts, predicate, "closing")
    rm: dict[str, int] = {}

    # ASSETS
    row = _write_line(ws, row, "Assets", [], bold=True); row += 1
    row = _write_line(ws, row, "Current Assets", [], indent=1, bold=True)

    v = gs(is_cash)
    if v != 0:
        rm["cash"] = row
        row = _write_line(ws, row, "Cash and Cash Equivalents", [v], indent=2)
    v = gs(is_restricted_bonds)
    if v != 0:
        rm["restricted"] = row
        row = _write_line(ws, row, "Restricted Funds - Bonds", [v], indent=2)
    v = gs(is_receivables)
    if v != 0:
        rm["receivables"] = row
        row = _write_line(ws, row, "Receivables, Net", [v], indent=2)
    v = gs(is_prepaids)
    if v != 0:
        rm["prepaids"] = row
        row = _write_line(ws, row, "Prepaids and Other Current Assets", [v], indent=2)

    parts = [rm[k] for k in ("cash", "restricted", "receivables", "prepaids") if k in rm]
    formula = "=" + "+".join(f"B{r}" for r in parts) if parts else 0
    rm["total_ca"] = row
    row = _write_line(ws, row, "Total Current Assets", [formula],
                     indent=1, bold=True, underline="thin")
    row += 1

    # Non-Current Assets / Real Estate
    row = _write_line(ws, row, "Non-Current Assets", [], indent=1, bold=True)
    row = _write_line(ws, row, "Real Estate, Net", [], indent=2, bold=True)

    re_parts: list[int] = []
    for key, label, pred in [
        ("land",         "Land",                                is_land),
        ("development",  "Development",                         is_development),
        ("contra_sales", "Contra Real Estate - Sales",          is_contra_sales),
        ("contra_mud",   "Contra Real Estate - MUD Receivable", is_contra_mud),
    ]:
        v = gs(pred)
        if v != 0:
            rm[key] = row; re_parts.append(row)
            row = _write_line(ws, row, label, [v], indent=3)
    formula = "=" + "+".join(f"B{r}" for r in re_parts) if re_parts else 0
    rm["re_net"] = row
    row = _write_line(ws, row, "Real Estate, Net", [formula],
                     indent=2, bold=True, underline="thin")

    v = gs(is_promissory)
    if v != 0:
        rm["promissory"] = row
        row = _write_line(ws, row, "Promissory Note - MUD Board fee", [v], indent=2)
    v = gs(is_ppe)
    if v != 0:
        rm["ppe"] = row
        row = _write_line(ws, row, "Property, Plant and Equipment, Net", [v], indent=2)

    parts = [rm[k] for k in ("re_net", "promissory", "ppe") if k in rm]
    formula = "=" + "+".join(f"B{r}" for r in parts) if parts else 0
    rm["total_nca"] = row
    row = _write_line(ws, row, "Total Non-Current Assets", [formula],
                     indent=1, bold=True, underline="thin")
    row += 1

    parts = [rm[k] for k in ("total_ca", "total_nca") if k in rm]
    formula = "=" + "+".join(f"B{r}" for r in parts) if parts else 0
    rm["total_assets"] = row
    row = _write_line(ws, row, "Total Assets", [formula],
                     bold=True, underline="double")
    row += 1

    # LIABILITIES + EQUITY
    row = _write_line(ws, row, "Liabilities and Equity", [], bold=True); row += 1
    row = _write_line(ws, row, "Current Liabilities", [], indent=1, bold=True)
    row = _write_line(ws, row, "Accounts Payable, Net", [], indent=2, bold=True)

    ap_parts: list[int] = []
    for key, label, pred in [
        ("trade_pay", "Trade Payables",         lambda p: p == 20030),
        ("retention", "Retention",              lambda p: p == 20060),
        ("rp_pay",    "Related Party Payables", lambda p: p in RELATED_PARTY_ACCTS),
        ("tax_pay",   "Taxes Payable",          lambda p: p == 20108),
    ]:
        v = -gs(pred)
        if v != 0:
            rm[key] = row; ap_parts.append(row)
            row = _write_line(ws, row, label, [v], indent=3)

    other_cl_val = sum(
        -a["closing"] for a in accounts
        if is_other_current_liab(get_prefix(a["num"]))
    )
    if other_cl_val != 0:
        rm["other_cl"] = row; ap_parts.append(row)
        row = _write_line(ws, row, "Other Current Liabilities", [other_cl_val], indent=3)

    formula = "=" + "+".join(f"B{r}" for r in ap_parts) if ap_parts else 0
    rm["ap_net"] = row
    row = _write_line(ws, row, "Accounts Payable, Net", [formula],
                     indent=2, bold=True, underline="thin")

    v = -gs(lambda p: p == 21060)
    if v != 0:
        rm["earnest"] = row
        row = _write_line(ws, row, "Builder Earnest Money", [v], indent=2)
    v = -gs(lambda p: p == 21015)
    if v != 0:
        rm["bond_st"] = row
        row = _write_line(ws, row, "Bond Payable - Short Term", [v], indent=2)

    parts = [rm[k] for k in ("ap_net", "earnest", "bond_st") if k in rm]
    formula = "=" + "+".join(f"B{r}" for r in parts) if parts else 0
    rm["total_cl"] = row
    row = _write_line(ws, row, "Total Current Liabilities", [formula],
                     indent=1, bold=True, underline="thin")
    row += 1

    row = _write_line(ws, row, "Non-Current Liabilities", [], indent=1, bold=True)
    ncl_parts: list[int] = []
    for key, label, pred in [
        ("dev_loan", "Development Loan Payable", is_dev_loan),
        ("bond_lt",  "Bond Payable - Long Term", is_bond_lt),
        ("deferred", "Deferred Income",          lambda p: p == 20020),
    ]:
        v = -gs(pred)
        if v != 0:
            rm[key] = row; ncl_parts.append(row)
            row = _write_line(ws, row, label, [v], indent=2)
    formula = "=" + "+".join(f"B{r}" for r in ncl_parts) if ncl_parts else 0
    rm["total_ncl"] = row
    row = _write_line(ws, row, "Total Non-Current Liabilities", [formula],
                     indent=1, bold=True, underline="thin")
    row += 1

    parts = [rm[k] for k in ("total_cl", "total_ncl") if k in rm]
    formula = "=" + "+".join(f"B{r}" for r in parts) if parts else 0
    rm["total_liab"] = row
    row = _write_line(ws, row, "Total Liabilities", [formula],
                     indent=1, bold=True, underline="thin")
    row += 1

    # Equity
    row = _write_line(ws, row, "Equity", [], indent=1, bold=True)
    eq_parts: list[int] = []
    meq = -sum(
        a["closing"] for a in accounts
        if is_members_eq(get_prefix(a["num"]))
    )
    if meq != 0:
        rm["members_eq"] = row; eq_parts.append(row)
        row = _write_line(ws, row, "Members' Equity", [meq], indent=2)
    re_30500 = gs(lambda p: p == 30500)
    re_30550 = gs(lambda p: p == 30550)
    is_sum   = gs(is_income_stmt)
    retained = -(re_30500 + re_30550 + is_sum)
    rm["retained"] = row; eq_parts.append(row)
    row = _write_line(ws, row, "Retained Earnings", [retained], indent=2)
    formula = "=" + "+".join(f"B{r}" for r in eq_parts)
    rm["total_eq"] = row
    row = _write_line(ws, row, "Total Equity", [formula],
                     indent=1, bold=True, underline="thin")
    row += 1

    parts = [rm[k] for k in ("total_liab", "total_eq") if k in rm]
    formula = "=" + "+".join(f"B{r}" for r in parts) if parts else 0
    rm["total_liab_eq"] = row
    row = _write_line(ws, row, "Total Liabilities and Equity", [formula],
                     bold=True, underline="double")
    return rm


# ─── Income Statement ────────────────────────────────────────────
def build_income_statement(ws, monthly_accounts, ytd_accounts, full_name, date_label):
    monthly_accounts = _normalize_accounts(monthly_accounts)
    ytd_accounts     = _normalize_accounts(ytd_accounts)
    title_lines = [full_name, "Income Statement", date_label]
    col_widths = [50, 18, 18]
    col_headers = ["", f"Current Month\n{date_label}", f"Year To Date\n{date_label}"]
    row = _setup_sheet(ws, title_lines, col_headers, col_widths)
    row += 1

    def both(predicate, flip=False):
        m = y = 0.0
        for a in monthly_accounts:
            if predicate(get_prefix(a["num"])):
                chg = a["closing"] - a["opening"]
                m += -chg if flip else chg
        for a in ytd_accounts:
            if predicate(get_prefix(a["num"])):
                chg = a["closing"] - a["opening"]
                y += -chg if flip else chg
        return m, y

    rm: dict[str, int] = {}

    # REVENUE
    row = _write_line(ws, row, "Revenue", [], bold=True)
    row = _write_line(ws, row, "Operating Revenue", [], indent=1, bold=True)
    row = _write_line(ws, row, "Real Estate Revenue", [], indent=2, bold=True)

    re_rev_rows: list[int] = []
    for key, label, pred in [
        ("lot_sales", "Lot Sales Revenue",      lambda p: 40000 <= p <= 40999),
        ("mkt_fee",   "Marketing Fee Income",   lambda p: p == 41020),
        ("fence",     "Fence Credits",          lambda p: p == 41001),
    ]:
        m, y = both(pred, flip=True)
        if m != 0 or y != 0:
            rm[key] = row; re_rev_rows.append(row)
            row = _write_line(ws, row, label, [m, y], indent=3)

    if re_rev_rows:
        fb = "=" + "+".join(f"B{r}" for r in re_rev_rows)
        fc = "=" + "+".join(f"C{r}" for r in re_rev_rows)
    else:
        fb, fc = 0, 0
    rm["total_re_rev"] = row
    row = _write_line(ws, row, "Total Real Estate Revenue", [fb, fc],
                     indent=2, bold=True, underline="thin")

    dev_rev_rows: list[int] = []
    m, y = both(lambda p: 42000 <= p <= 46999, flip=True)
    if m != 0 or y != 0:
        rm["dev_income"] = row; dev_rev_rows.append(row)
        row = _write_line(ws, row, "Development/Management Income", [m, y], indent=2)

    op_rev_parts = [rm[k] for k in ("total_re_rev",) if k in rm] \
                 + [rm[k] for k in ("dev_income",) if k in rm]
    fb = "=" + "+".join(f"B{r}" for r in op_rev_parts) if op_rev_parts else 0
    fc = "=" + "+".join(f"C{r}" for r in op_rev_parts) if op_rev_parts else 0
    rm["total_op_rev"] = row
    row = _write_line(ws, row, "Total Operating Revenue", [fb, fc],
                     indent=1, bold=True, underline="thin")

    rm["total_rev"] = row
    row = _write_line(ws, row, "Total Revenue",
                     [f"=B{rm['total_op_rev']}", f"=C{rm['total_op_rev']}"],
                     bold=True, underline="thin")
    row += 1

    # EXPENSES
    row = _write_line(ws, row, "Expenses", [], bold=True)
    row = _write_line(ws, row, "Operating Expenses", [], indent=1, bold=True)
    exp_parts: list[int] = []
    for key, label, pred in [
        ("cos",     "Cost of Sales - Real Estate", lambda p: 50000 <= p <= 59999),
        ("mkt_exp", "Marketing and Advertising",   lambda p: 70001 <= p <= 70039),
        ("ga",      "General & Administrative",
         lambda p: (60000 <= p <= 69999) or (70040 <= p <= 79999)),
    ]:
        m, y = both(pred)
        if m != 0 or y != 0:
            rm[key] = row; exp_parts.append(row)
            row = _write_line(ws, row, label, [m, y], indent=2)
    fb = "=" + "+".join(f"B{r}" for r in exp_parts) if exp_parts else 0
    fc = "=" + "+".join(f"C{r}" for r in exp_parts) if exp_parts else 0
    rm["total_op_exp"] = row
    row = _write_line(ws, row, "Total Operating Expenses", [fb, fc],
                     indent=1, bold=True, underline="thin")
    rm["total_exp"] = row
    row = _write_line(ws, row, "Total Expenses",
                     [f"=B{rm['total_op_exp']}", f"=C{rm['total_op_exp']}"],
                     bold=True, underline="thin")
    row += 1

    # Other Income / Expense
    other_parts: list[int] = []
    int_m, int_y = both(lambda p: p == 90031, flip=True)
    ft_m,  ft_y  = both(lambda p: p == 80015)
    has_other = (int_m != 0 or int_y != 0) or (ft_m != 0 or ft_y != 0)
    if has_other:
        row = _write_line(ws, row, "Other Income (Expense), Net", [], bold=True)
        if int_m != 0 or int_y != 0:
            rm["int_inc"] = row; other_parts.append(row)
            row = _write_line(ws, row, "Interest Income", [int_m, int_y], indent=1)
        if ft_m != 0 or ft_y != 0:
            rm["franchise_tax"] = row; other_parts.append(row)
            row = _write_line(ws, row, "Franchise Tax", [-ft_m, -ft_y], indent=1)
        fb = "=" + "+".join(f"B{r}" for r in other_parts)
        fc = "=" + "+".join(f"C{r}" for r in other_parts)
        rm["total_other"] = row
        row = _write_line(ws, row, "Total Other Income (Expense), Net", [fb, fc],
                         bold=True, underline="thin")
        row += 1

    # Net Income
    ni_b = f"=B{rm['total_rev']}-B{rm['total_exp']}"
    ni_c = f"=C{rm['total_rev']}-C{rm['total_exp']}"
    if "total_other" in rm:
        ni_b += f"+B{rm['total_other']}"
        ni_c += f"+C{rm['total_other']}"
    rm["net_income"] = row
    row = _write_line(ws, row, "Net Income", [ni_b, ni_c],
                     bold=True, underline="double")
    return rm


# ─── Statement of Cash Flows (indirect method) ───────────────────
def build_scf(ws, monthly_accounts, ytd_accounts, full_name, date_label):
    monthly_accounts = _normalize_accounts(monthly_accounts)
    ytd_accounts     = _normalize_accounts(ytd_accounts)
    title_lines = [full_name, "Statement of Cash Flows", date_label]
    col_widths = [50, 18, 18]
    col_headers = ["", f"Current Month\n{date_label}", f"Year To Date\n{date_label}"]
    row = _setup_sheet(ws, title_lines, col_headers, col_widths)
    row += 1

    def chg_m(predicate):
        return sum(a["opening"] - a["closing"] for a in monthly_accounts
                   if predicate(get_prefix(a["num"])))
    def chg_y(predicate):
        return sum(a["opening"] - a["closing"] for a in ytd_accounts
                   if predicate(get_prefix(a["num"])))

    rm: dict[str, int] = {}

    # Operating
    row = _write_line(ws, row, "Cash Flows from Operating Activities:", [], bold=True)
    ni_m, ni_y = chg_m(is_income_stmt), chg_y(is_income_stmt)
    rm["ni"] = row
    row = _write_line(ws, row, "Net Income (Loss)", [ni_m, ni_y], indent=1)
    row += 1
    op_parts = [rm["ni"]]

    wip_m, wip_y = chg_m(is_contra_sales), chg_y(is_contra_sales)
    has_noncash = abs(wip_m) > 0.005 or abs(wip_y) > 0.005
    mud_m, mud_y = chg_m(lambda p: p == 16902), chg_y(lambda p: p == 16902)
    has_mud = abs(mud_m) > 0.005 or abs(mud_y) > 0.005
    if has_noncash:
        row = _write_line(ws, row, "  Adjustments for Non-Cash Items:", [], bold=True)
        rm["wip_reclass"] = row; op_parts.append(row)
        row = _write_line(ws, row, "Cost of Sales - WIP Reclassification (non-cash)",
                         [wip_m, wip_y], indent=2)
        if has_mud:
            rm["mud_bond"] = row; op_parts.append(row)
            row = _write_line(ws, row, "MUD Receivable - Bond Proceeds Adj (non-cash)",
                             [mud_m, mud_y], indent=2)
        row += 1

    row = _write_line(ws, row, "Changes in Operating Assets and Liabilities:",
                     [], indent=1, bold=True)
    for key, label, pred in [
        ("recv_chg",     "Changes in Receivables",                       is_receivables),
        ("prep_chg",     "Changes in Prepaid Expenses and Other Assets", is_prepaids),
        ("ap_chg",       "Changes in Accounts Payable",                  lambda p: p == 20030),
        ("ret_chg",      "Changes in Retainage Payable",                 lambda p: p == 20060),
        ("tax_chg",      "Changes in Taxes Payable",                     lambda p: p == 20108),
        ("rp_chg",       "Changes in Related Party Payables",            lambda p: p in RELATED_PARTY_ACCTS),
        ("earn_chg",     "Changes in Earnest Money Deposits",            lambda p: p == 21060),
        ("def_chg",      "Changes in Deferred Revenue",                  lambda p: p == 20020),
    ]:
        m, y = chg_m(pred), chg_y(pred)
        if abs(m) > 0.005 or abs(y) > 0.005:
            rm[key] = row; op_parts.append(row)
            row = _write_line(ws, row, label, [m, y], indent=2)
    oom = sum(a["opening"] - a["closing"] for a in monthly_accounts
              if is_other_op_liab(get_prefix(a["num"])))
    ooy = sum(a["opening"] - a["closing"] for a in ytd_accounts
              if is_other_op_liab(get_prefix(a["num"])))
    if abs(oom) > 0.005 or abs(ooy) > 0.005:
        rm["other_op_chg"] = row; op_parts.append(row)
        row = _write_line(ws, row, "Changes in Other Operating Liabilities",
                         [oom, ooy], indent=2)

    fb = "=" + "+".join(f"B{r}" for r in op_parts)
    fc = "=" + "+".join(f"C{r}" for r in op_parts)
    rm["net_op"] = row
    row = _write_line(ws, row, "Net Cash from Operating Activities",
                     [fb, fc], bold=True, underline="thin")
    row += 1

    # Investing
    row = _write_line(ws, row, "Cash Flows from Investing Activities:", [], bold=True)
    inv_parts: list[int] = []
    for key, label, pred in [
        ("re_dev",     "Real Estate Development Expenditures",   is_re_invest),
        ("land_purch", "Purchases of Land",                      is_land),
        ("bond_cash",  "Changes in Restricted Cash (Bond Funds)", is_restricted_bonds),
        ("ppe_chg",    "Purchases of PP&E",                      is_ppe),
        ("prom_chg",   "Changes in Promissory Note",             is_promissory),
    ]:
        m, y = chg_m(pred), chg_y(pred)
        if abs(m) > 0.005 or abs(y) > 0.005:
            rm[key] = row; inv_parts.append(row)
            row = _write_line(ws, row, label, [m, y], indent=1)
    fb = "=" + "+".join(f"B{r}" for r in inv_parts) if inv_parts else "=0"
    fc = "=" + "+".join(f"C{r}" for r in inv_parts) if inv_parts else "=0"
    rm["net_inv"] = row
    row = _write_line(ws, row, "Net Cash from Investing Activities",
                     [fb, fc], bold=True, underline="thin")
    row += 1

    # Financing
    row = _write_line(ws, row, "Cash Flows from Financing Activities:", [], bold=True)
    fin_parts: list[int] = []
    for key, label, pred in [
        ("loan_chg",    "Net Borrowings/(Repayments) - Development Loan", is_dev_loan),
        ("bond_st_chg", "Changes in Bond Payable - Short Term",           lambda p: p == 21015),
        ("bond_lt_chg", "Changes in Bond Payable - Long Term, Net",       is_bond_lt),
        ("mem_chg",     "Member Contributions/Distributions",             is_members_eq),
    ]:
        m, y = chg_m(pred), chg_y(pred)
        if abs(m) > 0.005 or abs(y) > 0.005:
            rm[key] = row; fin_parts.append(row)
            row = _write_line(ws, row, label, [m, y], indent=1)
    fb = "=" + "+".join(f"B{r}" for r in fin_parts) if fin_parts else "=0"
    fc = "=" + "+".join(f"C{r}" for r in fin_parts) if fin_parts else "=0"
    rm["net_fin"] = row
    row = _write_line(ws, row, "Net Cash from Financing Activities",
                     [fb, fc], bold=True, underline="thin")
    row += 1

    # Summary
    rm["net_change"] = row
    row = _write_line(ws, row, "Net Change in Cash",
                     [f"=B{rm['net_op']}+B{rm['net_inv']}+B{rm['net_fin']}",
                      f"=C{rm['net_op']}+C{rm['net_inv']}+C{rm['net_fin']}"],
                     bold=True, underline="thin")
    row += 1

    cash_open_m = sum_accounts(monthly_accounts, is_cash, "opening")
    cash_open_y = sum_accounts(ytd_accounts,     is_cash, "opening")
    rm["cash_beg"] = row
    row = _write_line(ws, row, "Cash, Beginning of Period",
                     [cash_open_m, cash_open_y])

    cash_close = sum_accounts(monthly_accounts, is_cash, "closing")
    rm["cash_end"] = row
    row = _write_line(ws, row, "Cash, End of Period",
                     [cash_close, cash_close], underline="double")
    row += 1

    rm["footing"] = row
    row = _write_line(ws, row, "Footing Check (should be 0)",
                     [f"=B{rm['cash_beg']}+B{rm['net_change']}-B{rm['cash_end']}",
                      f"=C{rm['cash_beg']}+C{rm['net_change']}-C{rm['cash_end']}"],
                     bold=True)
    return rm


# ─── Public API ──────────────────────────────────────────────────
def build_workbook(full_name: str, monthly_accounts: list[dict],
                   ytd_accounts: list[dict], date_label: str) -> BytesIO:
    """Build a 3-sheet FRP workbook (BS / IS / SCF) and return it as
    a BytesIO ready to send to the user. Caller is responsible for
    seek(0) before sending if needed."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"
    build_balance_sheet(ws, monthly_accounts, full_name, date_label)

    if ytd_accounts:
        ws = wb.create_sheet("Income Statement")
        build_income_statement(ws, monthly_accounts, ytd_accounts,
                               full_name, date_label)
        ws = wb.create_sheet("Statement of Cash Flows")
        build_scf(ws, monthly_accounts, ytd_accounts, full_name, date_label)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
