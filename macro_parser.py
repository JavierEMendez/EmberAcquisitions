"""
macro_parser.py — Parses the Houston County macroeconomic Excel file and
returns structured JSON for the Macro dashboard.
"""

import io
from typing import Any
import openpyxl

FIPS = {
    "Austin": "48015", "Brazoria": "48039", "Chambers": "48071",
    "Fort Bend": "48157", "Galveston": "48167", "Harris": "48201",
    "Liberty": "48291", "Montgomery": "48339", "Waller": "48473",
}

_GDP_SECTORS = {
    "All industry total",
    "Private industries",
    "Agriculture, forestry, fishing and hunting",
    "Mining, quarrying, and oil and gas extraction",
    "Construction",
    "Manufacturing",
    "Information",
    "Finance, insurance, real estate, rental, and leasing",
    "Educational services, health care, and social assistance",
    "Arts, entertainment, recreation, accommodation, and food services",
    "Other services (except government and government enterprises)",
    "Government and government enterprises",
}


def _n(val: Any, precision: int | None = 2) -> float | int:
    if val is None:
        return 0
    try:
        n = float(val)
    except (TypeError, ValueError):
        return 0
    if precision is not None:
        n = round(n, precision)
    if precision is not None and n == int(n) and abs(n) < 1e15:
        return int(n)
    return n


def _s(val: Any) -> str:
    return "" if val is None else str(val).strip()


def _build_series(data: dict) -> dict:
    """Convert {county: {year: val}} to {years: [...], county: [...], ...}."""
    all_years = sorted({yr for cdata in data.values() for yr in cdata})
    result = {"years": all_years}
    for county, cdata in data.items():
        result[county] = [cdata.get(yr, None) for yr in all_years]
    return result


def _parse_census_pep(ws) -> dict:
    raw = {m: {} for m in ["population", "births", "deaths",
                            "net_domestic_migration", "net_international_migration",
                            "net_pop_change"]}
    col_map = {
        "births": 4, "deaths": 5, "net_domestic_migration": 6,
        "net_international_migration": 7, "net_pop_change": 10,
        "population": 11,
    }
    for r in range(2, ws.max_row + 1):
        county = _s(ws.cell(row=r, column=2).value)
        year = ws.cell(row=r, column=3).value
        if not county or year is None:
            continue
        year = int(year)
        for metric, col in col_map.items():
            raw[metric].setdefault(county, {})[year] = _n(ws.cell(row=r, column=col).value, 0)

    all_years = sorted({yr for cdata in raw["population"].values() for yr in cdata})
    result = {"years": all_years}
    for metric, cdata in raw.items():
        result[metric] = {county: [cdata.get(county, {}).get(yr, None) for yr in all_years]
                          for county in FIPS}
    return result


def _parse_irs_soi(ws) -> dict:
    raw_agi = {}
    raw_returns = {}
    for r in range(2, ws.max_row + 1):
        county = _s(ws.cell(row=r, column=2).value)
        pair = _s(ws.cell(row=r, column=3).value)
        if not county or not pair:
            continue
        # Use the second year of the pair as the label year
        try:
            year = int(pair.split("-")[1])
        except (IndexError, ValueError):
            continue
        raw_agi.setdefault(county, {})[year] = _n(ws.cell(row=r, column=12).value, 0)
        raw_returns.setdefault(county, {})[year] = _n(ws.cell(row=r, column=10).value, 0)

    all_years = sorted({yr for cdata in raw_agi.values() for yr in cdata})
    return {
        "years": all_years,
        "net_agi": {c: [raw_agi.get(c, {}).get(y, None) for y in all_years] for c in FIPS},
        "net_returns": {c: [raw_returns.get(c, {}).get(y, None) for y in all_years] for c in FIPS},
    }


def _parse_bea_income(ws) -> dict:
    raw_pi = {}
    raw_pci = {}
    for r in range(2, ws.max_row + 1):
        county = _s(ws.cell(row=r, column=2).value)
        year = ws.cell(row=r, column=3).value
        measure = _s(ws.cell(row=r, column=4).value)
        val = ws.cell(row=r, column=5).value
        if not county or year is None:
            continue
        year = int(year)
        if year < 2000:
            continue
        m = measure.rstrip()
        if "Per capita" in m:
            raw_pci.setdefault(county, {})[year] = _n(val, 0)
        elif "Personal income" in m:
            raw_pi.setdefault(county, {})[year] = _n(val, 0)

    all_years = sorted({yr for cdata in raw_pi.values() for yr in cdata})
    return {
        "years": all_years,
        "personal_income": {c: [raw_pi.get(c, {}).get(y, None) for y in all_years] for c in FIPS},
        "per_capita_income": {c: [raw_pci.get(c, {}).get(y, None) for y in all_years] for c in FIPS},
    }


def _parse_bea_gdp(ws) -> dict:
    raw_total = {}   # {county: {year: val}}
    raw_sector = {}  # {county: {industry: {year: val}}}

    for r in range(2, ws.max_row + 1):
        county = _s(ws.cell(row=r, column=2).value)
        year = ws.cell(row=r, column=3).value
        industry_raw = ws.cell(row=r, column=4).value
        val = ws.cell(row=r, column=5).value
        if not county or year is None or industry_raw is None:
            continue
        year = int(year)
        industry = _s(industry_raw)

        if industry == "All industry total":
            raw_total.setdefault(county, {})[year] = _n(val, 0)

        if industry in _GDP_SECTORS:
            raw_sector.setdefault(county, {}).setdefault(industry, {})[year] = _n(val, 0)

    all_years = sorted({yr for cdata in raw_total.values() for yr in cdata})
    sector_names = sorted(_GDP_SECTORS - {"All industry total", "Private industries"})

    return {
        "years": all_years,
        "total": {c: [raw_total.get(c, {}).get(y, None) for y in all_years] for c in FIPS},
        "sector_names": sector_names,
        "by_sector": {
            c: {
                s: [raw_sector.get(c, {}).get(s, {}).get(y, None) for y in all_years]
                for s in sector_names
            }
            for c in FIPS
        },
    }


def _parse_bls_qcew(ws) -> dict:
    raw_emp = {}
    raw_wage = {}
    raw_est = {}
    raw_total_wages = {}

    for r in range(2, ws.max_row + 1):
        county = _s(ws.cell(row=r, column=2).value)
        year = ws.cell(row=r, column=3).value
        if not county or year is None:
            continue
        year = int(year)
        weekly = _n(ws.cell(row=r, column=7).value, 2)
        annual = round(weekly / 7 * 365) if weekly else 0
        raw_emp.setdefault(county, {})[year] = _n(ws.cell(row=r, column=5).value, 0)
        raw_wage.setdefault(county, {})[year] = annual
        raw_est.setdefault(county, {})[year] = _n(ws.cell(row=r, column=4).value, 0)
        raw_total_wages.setdefault(county, {})[year] = _n(ws.cell(row=r, column=6).value, 0)

    all_years = sorted({yr for cdata in raw_emp.values() for yr in cdata})
    return {
        "years": all_years,
        "avg_employment": {c: [raw_emp.get(c, {}).get(y, None) for y in all_years] for c in FIPS},
        "avg_annual_wage": {c: [raw_wage.get(c, {}).get(y, None) for y in all_years] for c in FIPS},
        "establishments": {c: [raw_est.get(c, {}).get(y, None) for y in all_years] for c in FIPS},
        "total_wages": {c: [raw_total_wages.get(c, {}).get(y, None) for y in all_years] for c in FIPS},
    }


def _parse_fdic(ws) -> dict:
    raw_dep = {}
    raw_inst = {}

    for r in range(2, ws.max_row + 1):
        county = _s(ws.cell(row=r, column=2).value)
        year = ws.cell(row=r, column=3).value
        if not county or year is None:
            continue
        year = int(year)
        raw_dep.setdefault(county, {})[year] = _n(ws.cell(row=r, column=4).value, 0)
        raw_inst.setdefault(county, {})[year] = _n(ws.cell(row=r, column=5).value, 0)

    all_years = sorted({yr for cdata in raw_dep.values() for yr in cdata})
    return {
        "years": all_years,
        "total_deposits": {c: [raw_dep.get(c, {}).get(y, None) for y in all_years] for c in FIPS},
        "num_institutions": {c: [raw_inst.get(c, {}).get(y, None) for y in all_years] for c in FIPS},
    }


def _parse_zillow(ws) -> dict:
    raw = {}
    for r in range(2, ws.max_row + 1):
        county = _s(ws.cell(row=r, column=2).value)
        year = ws.cell(row=r, column=3).value
        val = ws.cell(row=r, column=4).value
        if not county or year is None or val is None:
            continue
        raw.setdefault(county, {})[int(year)] = _n(val, 0)

    all_years = sorted({yr for cdata in raw.values() for yr in cdata})
    return {
        "years": all_years,
        "zhvi": {c: [raw.get(c, {}).get(y, None) for y in all_years] for c in FIPS},
    }


def _parse_hud_fmr(ws) -> dict:
    raw = {br: {} for br in range(5)}
    for r in range(2, ws.max_row + 1):
        county = _s(ws.cell(row=r, column=2).value)
        year = ws.cell(row=r, column=3).value
        if not county or year is None:
            continue
        year = int(year)
        for br in range(5):
            # columns 4-8 are fmr0-fmr4, columns 9-13 are fmr_0 - fmr_4
            v = ws.cell(row=r, column=4 + br).value or ws.cell(row=r, column=9 + br).value
            if v is not None:
                raw[br].setdefault(county, {})[year] = _n(v, 0)

    all_years = sorted({yr for br_data in raw.values()
                        for cdata in br_data.values() for yr in cdata})
    br_labels = ["0BR", "1BR", "2BR", "3BR", "4BR"]
    return {
        "years": all_years,
        **{
            f"fmr_{br_labels[br]}": {
                c: [raw[br].get(c, {}).get(y, None) for y in all_years]
                for c in FIPS
            }
            for br in range(5)
        },
    }


def parse_macro(file_bytes: bytes) -> dict:
    wb = openpyxl.load_workbook(
        filename=io.BytesIO(file_bytes), data_only=True, read_only=False
    )

    def _ws(keywords):
        for name in wb.sheetnames:
            nl = name.lower()
            if all(k in nl for k in keywords):
                return wb[name]
        return None

    result = {
        "counties": sorted(FIPS.keys()),
        "fips": FIPS,
        "population": _parse_census_pep(_ws(["census", "pep"]) or _ws(["pep"])),
        "irs": _parse_irs_soi(_ws(["irs"]) or _ws(["soi"])),
        "income": _parse_bea_income(_ws(["bea", "income"]) or _ws(["personal", "income"])),
        "gdp": _parse_bea_gdp(_ws(["bea", "gdp"]) or _ws(["gdp"])),
        "employment": _parse_bls_qcew(_ws(["bls"]) or _ws(["qcew"])),
        "deposits": _parse_fdic(_ws(["fdic"])),
        "home_values": _parse_zillow(_ws(["zillow"])),
        "rents": _parse_hud_fmr(_ws(["hud"]) or _ws(["fmr"]) or _ws(["rent"])),
    }
    wb.close()
    return result
