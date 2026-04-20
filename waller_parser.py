"""
waller_parser.py — Parses the raw Waller ISD monthly-permits workbook
(e.g. "Ember - Waller ISD Comps by Month.xlsx") and returns structured
JSON for the Waller Monthly dashboard tab.

Sheet "by Community" layout:
  Row 0: header
    col 0 COMMUNITY, col 1 LOT SIZE, col 2 BUILDER
    cols 3-14  JAN 23..DEC 23  | col 15 = 2023 Total
    cols 16-27 JAN 24..DEC 24  | col 28 = 2024 Total
    cols 29-40 JAN 25..DEC 25  | col 41 = 2025 Total
  Row 1: "WALLER ISD Total" grand-total row (all monthlies + yearly totals).
  Rows 2+: builder detail rows + per-community aggregator rows
           ("{COMMUNITY} Total" with blank lot/builder) + blank separators.
"""

from __future__ import annotations

import io
from typing import Any

import openpyxl


# 36 monthly columns, skipping yearly-total columns (15, 28, 41).
MONTH_COLS = list(range(3, 15)) + list(range(16, 28)) + list(range(29, 41))
YEAR_COLS = {2023: 15, 2024: 28, 2025: 41}


def _s(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _num(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _int0(v: Any) -> int:
    n = _num(v)
    return int(n) if n else 0


def _monthly(r: tuple) -> list[int]:
    return [_int0(r[c]) if len(r) > c else 0 for c in MONTH_COLS]


def _yearly(r: tuple) -> dict[int, int]:
    return {y: (_int0(r[c]) if len(r) > c else 0) for y, c in YEAR_COLS.items()}


def _month_labels() -> list[str]:
    months = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN",
              "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]
    return [f"{m} {yy}" for yy in (23, 24, 25) for m in months]


def parse_waller_monthly(file_bytes: bytes) -> dict:
    """
    Parse a Waller Monthly xlsx and return:
      {
        "months":        ["JAN 23", ..., "DEC 25"],          # 36 labels
        "waller_total":  {"monthly": [36 ints], "yearly": {2023,2024,2025}},
        "communities": [
          {
            "community":  "THE GRAND PRAIRIE",
            "monthly":    [36 ints],
            "yearly":     {2023, 2024, 2025},
            "builders":   [
              {"builder": "LENNAR HOMES", "lot_size": "40",
               "monthly": [36 ints], "yearly": {...}},
              ...
            ],
          },
          ...
        ],
      }
    Communities are sorted by 2025 total, descending.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    if "by Community" not in wb.sheetnames:
        wb.close()
        return {"months": _month_labels(), "waller_total": None, "communities": []}

    ws = wb["by Community"]
    rows = list(ws.iter_rows(values_only=True))

    data: dict[str, Any] = {
        "months":       _month_labels(),
        "waller_total": None,
        "communities":  [],
    }

    # ── Row 1: WALLER ISD Total grand-total row
    if len(rows) > 1 and _s(rows[1][0]).upper().startswith("WALLER ISD"):
        gr = rows[1]
        data["waller_total"] = {"monthly": _monthly(gr), "yearly": _yearly(gr)}

    # ── Per-community accumulation
    # community name -> {monthly: [..], yearly: {..}, builders: [..]}
    communities: dict[str, dict] = {}

    for i in range(2, len(rows)):
        r = rows[i]
        if not r:
            continue
        name = _s(r[0])
        lot = _s(r[1]) if len(r) > 1 else ""
        builder = _s(r[2]) if len(r) > 2 else ""

        if not name:
            continue

        # Aggregator row: "<Community> Total" (blank lot & builder, yearly totals present).
        # Use these as the authoritative community totals.
        if name.upper().endswith(" TOTAL") and not lot and not builder:
            comm = name[:-6].strip()  # strip trailing " Total"
            if comm.upper() == "WALLER ISD":
                continue
            entry = communities.setdefault(
                comm,
                {"community": comm, "monthly": [0] * 36,
                 "yearly": {2023: 0, 2024: 0, 2025: 0}, "builders": []},
            )
            entry["monthly"] = _monthly(r)
            entry["yearly"] = _yearly(r)
            continue

        # Detail row: needs a builder to be meaningful
        if not builder:
            continue

        entry = communities.setdefault(
            name,
            {"community": name, "monthly": [0] * 36,
             "yearly": {2023: 0, 2024: 0, 2025: 0}, "builders": []},
        )
        entry["builders"].append({
            "builder":  builder,
            "lot_size": lot or "N/A",
            "monthly":  _monthly(r),
            "yearly":   _yearly(r),
        })

    # Sort builders within each community by 2025 total desc
    for e in communities.values():
        e["builders"].sort(key=lambda b: b["yearly"].get(2025, 0), reverse=True)

    # Sort communities by 2025 total desc
    data["communities"] = sorted(
        communities.values(),
        key=lambda c: c["yearly"].get(2025, 0),
        reverse=True,
    )

    wb.close()
    return data


if __name__ == "__main__":
    # Smoke-test against the user's raw file
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else \
        r"C:\Users\Javier\Downloads\Ember - Waller ISD Comps by Month.xlsx"
    with open(path, "rb") as f:
        out = parse_waller_monthly(f.read())
    print(f"months: {len(out['months'])}  first={out['months'][0]}  last={out['months'][-1]}")
    wt = out["waller_total"]
    if wt:
        print(f"WALLER ISD  2023={wt['yearly'][2023]}  2024={wt['yearly'][2024]}  2025={wt['yearly'][2025]}")
        print(f"  monthly[0]={wt['monthly'][0]}  monthly[-1]={wt['monthly'][-1]}  sum={sum(wt['monthly'])}")
    print(f"communities: {len(out['communities'])}")
    print("Top 10 by 2025:")
    for c in out["communities"][:10]:
        print(f"  {c['community']:<30} 2023={c['yearly'][2023]:>4}  2024={c['yearly'][2024]:>4}  "
              f"2025={c['yearly'][2025]:>4}  builders={len(c['builders'])}")
    # Find THE GRAND PRAIRIE
    tgp = next((c for c in out["communities"] if c["community"] == "THE GRAND PRAIRIE"), None)
    if tgp:
        print(f"\nTHE GRAND PRAIRIE 2025 total: {tgp['yearly'][2025]}  (expected 678)")
        print(f"  builders: {len(tgp['builders'])}")
        for b in tgp["builders"][:5]:
            print(f"    lot={b['lot_size']:<4}  {b['builder']:<25}  2025={b['yearly'][2025]}")
