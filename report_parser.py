"""
report_parser.py — Parses the Ember Dashboard Excel file and extracts
Consolidated Project Returns and Loan Capacities & Debt Schedules into
structured JSON.
"""

import io
from datetime import date, datetime
from typing import Any

import openpyxl


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _num(val: Any, precision: int | None = 2) -> float | int:
    """Convert a cell value to a number. None -> 0. Round if precision given."""
    if val is None:
        return 0
    try:
        n = float(val)
    except (TypeError, ValueError):
        return 0
    if precision is not None:
        n = round(n, precision)
    # Return int when the value is whole
    if precision is not None and n == int(n) and abs(n) < 1e15:
        return int(n)
    return n


def _str(val: Any) -> str:
    """Convert a cell value to a string. None -> ''."""
    if val is None:
        return ""
    return str(val).strip()


def _date_iso(val: Any) -> str:
    """Convert a cell value to an ISO-format date string."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    s = str(val).strip()
    return s


def _row_yearly(ws, row: int, col_start: int = 7, col_end: int = 28,
                precision: int | None = 2) -> list:
    """Read yearly values from columns G(7) through AB(28)."""
    return [_num(ws.cell(row=row, column=c).value, precision)
            for c in range(col_start, col_end + 1)]


def _row_monthly(ws, row: int, col_start: int, col_end: int,
                 precision: int | None = 2) -> list:
    """Read monthly values from a contiguous range of columns.

    Used for the new "Monthly Cashflows" block on the Consolidated
    Project Returns tab — one cell per month from `col_start`
    (inclusive) through `col_end` (inclusive). Empty / non-numeric
    cells return 0 so consumers can sum without checking for None.
    """
    return [_num(ws.cell(row=row, column=c).value, precision)
            for c in range(col_start, col_end + 1)]


# ---------------------------------------------------------------------------
# Consolidated Project Returns
# ---------------------------------------------------------------------------

_PROJECT_STARTS = [4, 15, 26, 37, 48, 59, 70, 81, 92]

_METRIC_OFFSETS = [
    # (row offset from block start, label, has_status flag, precision)
    (2, "Preferred Return", True, 2),
    (3, "Return of Capital", True, 2),
    (4, "Excess Cash Flow", True, 2),
    (5, "Total LP Distributions", False, 2),
    (6, "Total LP Contributions", False, 2),
    (7, "Total LP Profit", False, 2),
    (8, "LP IRR", False, None),        # full precision for IRR
    (9, "LP Equity Multiple", False, None),  # full precision
    (10, "Promote", True, 2),
]


def _parse_returns(ws) -> dict:
    """Parse the 'Consolidated Project Returns' tab.

    Newer Ember dashboard uploads include a Monthly Cashflows block to
    the right of the yearly grid (starting at column AD/30). When
    present, we attach a `monthly` array to each metric (the same
    row-offsets as the yearly block, just sourced from the monthly
    columns) plus a top-level `months` array of ISO date strings so
    consumers can compute "to date through today's month" without
    pro-rating from yearly totals.
    """
    title = _str(ws.cell(row=3, column=3).value)  # C3
    # "Data From" date — the label "Date Updated:" lives in C2, the date in E2.
    # Older workbooks used Z3 (which is `=TODAY()`); we now prefer E2 because
    # it reflects the actual underwriting cutoff, not the upload day. Falls
    # back to Z3 → today() so an older workbook still gets a usable date.
    report_date = (
        _date_iso(ws.cell(row=2, column=5).value)   # E2 (new convention)
        or _date_iso(ws.cell(row=3, column=26).value)  # Z3 (legacy)
    )

    # Year headers from first project block row, columns G-AB
    years = []
    for c in range(7, 29):
        v = ws.cell(row=3, column=c).value
        if v is None:
            # Try row 4 (the first project header row) as fallback
            v = ws.cell(row=4, column=c).value
        try:
            yr = int(v)
            years.append(yr)
        except (TypeError, ValueError):
            # Fill gaps — likely a merged cell; infer from neighbors
            if years:
                years.append(years[-1] + 1)
            else:
                years.append(0)

    # Monthly date headers — read row 4 starting at col 30 (AD) until
    # the first empty cell. Only ACTUAL date cells count; the project
    # name string at col 30 is skipped. Empty list when this workbook
    # predates the monthly block.
    months: list[str] = []
    monthly_col_start = 0
    monthly_col_end   = 0
    for c in range(30, ws.max_column + 1):
        v = ws.cell(row=4, column=c).value
        if v is None:
            # Empty cell ends the run — but only after we've seen at
            # least one date (skips a possible project-name header).
            if months:
                break
            continue
        if not isinstance(v, (datetime, date)):
            # Non-date cell (e.g. project name at col 30) — skip until
            # we find the first real date.
            if months:
                break
            continue
        if not months:
            monthly_col_start = c
        months.append(_date_iso(v))
        monthly_col_end = c

    # --- Projects ---
    # Detect each project block dynamically instead of using a fixed list of
    # start rows: a block begins at a row whose column C holds the project
    # name and whose NEXT row's column C is the "LP Returns Metrics"
    # sub-header. This imports ANY number of projects — and adding a project
    # (which pushes every block below it down) no longer drops the last one
    # or misaligns the summary. Falls back to the legacy fixed rows only if
    # no sub-header markers are found (older/odd workbooks).
    project_starts = []
    for r in range(4, ws.max_row + 1):
        if not _str(ws.cell(row=r, column=3).value):
            continue
        if _str(ws.cell(row=r + 1, column=3).value).strip().lower() == "lp returns metrics":
            project_starts.append(r)
    if not project_starts:
        project_starts = _PROJECT_STARTS

    projects = []
    for start_row in project_starts:
        name = _str(ws.cell(row=start_row, column=3).value)  # C column
        if not name:
            continue

        metrics = []
        for offset, label, has_status, prec in _METRIC_OFFSETS:
            r = start_row + offset
            metric: dict[str, Any] = {"label": _str(ws.cell(row=r, column=3).value) or label}
            if has_status:
                metric["status"] = _str(ws.cell(row=r, column=4).value)
            else:
                metric["status"] = _num(ws.cell(row=r, column=4).value, 0)
            metric["total"] = _num(ws.cell(row=r, column=5).value, prec)
            metric["yearly"] = _row_yearly(ws, r, precision=prec)
            # Monthly array — same row, sourced from the AD-onward block
            # when this workbook ships one. Length matches `months`.
            if monthly_col_start and monthly_col_end:
                metric["monthly"] = _row_monthly(
                    ws, r, monthly_col_start, monthly_col_end, precision=prec,
                )

            metrics.append(metric)

        projects.append({"name": name, "metrics": metrics})

    # --- Summary section ---
    # The summary block sits below the projects, so its rows shift whenever a
    # project is added. Locate each line by its (unique) label in column C
    # rather than a fixed row; fall back to the legacy rows if a label can't
    # be found.
    _legacy_summary_rows = {
        "MPC Contributions": 105, "MPC Distributions": 106, "MPC Net Cashflow": 107,
        "Vertical Contributions": 108, "Vertical Distributions": 109,
        "Vertical Net Cashflow": 110, "Total Assets Net Cashflow": 111,
    }
    summary_order = [
        "MPC Contributions", "MPC Distributions", "MPC Net Cashflow",
        "Vertical Contributions", "Vertical Distributions",
        "Vertical Net Cashflow", "Total Assets Net Cashflow",
    ]
    after = (project_starts[-1] + 11) if project_starts else 0  # scan below the last block
    label_row = {}
    for r in range(max(after, 1), ws.max_row + 1):
        lbl = _str(ws.cell(row=r, column=3).value)
        if lbl in _legacy_summary_rows and lbl not in label_row:
            label_row[lbl] = r
    summary = []
    for label in summary_order:
        r = label_row.get(label) or _legacy_summary_rows[label]
        total = _num(ws.cell(row=r, column=5).value, 2)
        yearly = _row_yearly(ws, r, precision=2)
        summary.append({"label": _str(ws.cell(row=r, column=3).value) or label,
                        "total": total, "yearly": yearly})

    return {
        "title": title or "Consolidated Ember Project Returns",
        "date": report_date,
        "data_from": report_date,  # canonical name; "date" kept for back-compat
        "years": years,
        "months": months,    # ISO YYYY-MM-DD per monthly column; [] when the workbook has no monthly block
        "projects": projects,
        "summary": summary,
    }


# ---------------------------------------------------------------------------
# Loan Capacities & Debt Schedules
# ---------------------------------------------------------------------------

_LOAN_HEADERS = [
    "Community", "Lender", "Collateral", "Recourse",
    "Loan Origination", "Loan Term Date", "Months Remaining",
    "Rem. Interest Reserve", "Monthly Interest Burn",
    "Remaining Mos. of IR", "IR Health", "Index + Spread",
    "Today's Rate", "Extensions Remaining", "Extension Cost",
    "Loan Amount", "Drawn", "Balance", "Utilization",
    "Remaining", "Forecasted Thru Term", "Capacity Health",
]

# Column mapping: B=2, D=4, E=5, F=6, G=7, H=8, ... X=24
_LOAN_COLS = [2, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24]

# Fields that should be treated as dates
_DATE_FIELDS = {"Loan Origination", "Loan Term Date"}
# Fields that are percentages / rates — keep full precision
_RATE_FIELDS = {"Index + Spread", "Today's Rate", "Utilization"}
# Fields that are strings
_STR_FIELDS = {"Community", "Lender", "Collateral", "Recourse", "IR Health",
               "Capacity Health"}


def _read_loan_cell(header: str, val: Any) -> Any:
    if header in _DATE_FIELDS:
        return _date_iso(val)
    if header in _STR_FIELDS:
        return _str(val)
    if header in _RATE_FIELDS:
        return _num(val, precision=None)
    # Numeric by default
    return _num(val, 2)


def _read_loan_row(ws, row: int) -> dict:
    """Read a single loan data row into a dict keyed by header name."""
    result = {}
    for header, col in zip(_LOAN_HEADERS, _LOAN_COLS):
        result[header] = _read_loan_cell(header, ws.cell(row=row, column=col).value)
    return result


def _read_totals_row(ws, row: int) -> dict:
    """Read a totals row — only non-empty/non-zero fields."""
    totals = {}
    for header, col in zip(_LOAN_HEADERS, _LOAN_COLS):
        val = ws.cell(row=row, column=col).value
        if val is not None and val != "" and val != 0:
            totals[header] = _read_loan_cell(header, val)
    return totals


def _parse_loans(ws) -> dict:
    """Parse the 'Loan Capacities & DS' tab."""

    # "Data From" date — label "Date Updated:" in T3, date in U3.
    # Falls back to today() so older workbooks without the cell still parse.
    report_date = (
        _date_iso(ws.cell(row=3, column=21).value)  # U3
        or date.today().isoformat()
    )

    # --- MPC Loans (rows 7-10, totals row 11) ---
    mpc_rows = []
    for r in range(7, 11):
        val = ws.cell(row=r, column=2).value  # B column = Community
        if val is None or _str(val) == "":
            continue
        mpc_rows.append(_read_loan_row(ws, r))
    mpc_totals = _read_totals_row(ws, 11)

    # --- Vertical Loans (row 15 data, row 16 totals, row 17 footnote) ---
    vert_rows = []
    r = 15
    while True:
        val = ws.cell(row=r, column=2).value
        label = _str(val)
        if label == "" or label.lower() == "totals":
            break
        vert_rows.append(_read_loan_row(ws, r))
        r += 1
    vert_totals = _read_totals_row(ws, 16)
    footnote = _str(ws.cell(row=17, column=2).value)

    # --- Debt Schedules (starting at row 56) ---
    debt_schedules = _parse_debt_schedules(ws)

    return {
        "date": report_date,
        "data_from": report_date,  # canonical name; "date" kept for back-compat
        "mpc_loans": {
            "headers": list(_LOAN_HEADERS),
            "rows": mpc_rows,
            "totals": mpc_totals,
        },
        "vertical_loans": {
            "headers": list(_LOAN_HEADERS),
            "rows": vert_rows,
            "totals": vert_totals,
            "footnote": footnote,
        },
        "debt_schedules": debt_schedules,
    }


def _parse_debt_schedules(ws) -> list[dict]:
    """Parse debt schedule blocks starting at row 56."""
    schedules = []

    # First (and possibly only) debt schedule block
    project_name = _str(ws.cell(row=57, column=2).value)  # B57
    if not project_name:
        project_name = _str(ws.cell(row=57, column=3).value)  # C57 fallback
    # Clean project name — might contain extra text like "Associated Revenues"
    if project_name:
        project_name = project_name.split("\n")[0].strip()

    # Monthly date headers in L57-W57 (cols 12-23)
    months = []
    for c in range(12, 24):
        val = ws.cell(row=57, column=c).value
        months.append(_date_iso(val))

    # Payment rows 59-63
    payments = []
    for r in range(59, 64):
        dt = ws.cell(row=r, column=2).value   # B = date
        lender = ws.cell(row=r, column=3).value  # C = lender
        amount = ws.cell(row=r, column=4).value  # D = amount
        covered = ws.cell(row=r, column=5).value  # E = covered status
        if dt is None and lender is None and amount is None:
            continue
        payments.append({
            "date": _date_iso(dt),
            "lender": _str(lender),
            "amount": _num(amount, 2),
            "covered": _str(covered),
        })

    # Payment total — row 64, column D
    payment_total = _num(ws.cell(row=64, column=4).value, 2)

    # Revenue rows 60-67 (H=type, J=pct, K=total, L-W=monthly)
    # These overlap with payment rows in different columns
    revenues = []
    for r in range(60, 68):
        rev_type = ws.cell(row=r, column=8).value  # H
        if rev_type is None or _str(rev_type) == "":
            continue
        pct = _num(ws.cell(row=r, column=10).value, precision=None)  # J — keep full
        total = _num(ws.cell(row=r, column=11).value, 2)  # K
        monthly = [_num(ws.cell(row=r, column=c).value, 2) for c in range(12, 24)]
        revenues.append({
            "type": _str(rev_type),
            "pct": pct,
            "total": total,
            "monthly": monthly,
        })

    # Total Revenues — row 68
    total_rev_pct = _num(ws.cell(row=68, column=10).value, precision=None)
    total_rev_total = _num(ws.cell(row=68, column=11).value, 2)
    total_rev_monthly = [_num(ws.cell(row=68, column=c).value, 2) for c in range(12, 24)]
    total_revenues = {
        "pct": total_rev_pct,
        "total": total_rev_total,
        "monthly": total_rev_monthly,
    }

    # Cumulative Revenues — row 69, cols L-W
    cumulative_revenues = [_num(ws.cell(row=69, column=c).value, 2) for c in range(12, 24)]

    # Cumulative Payments — row 70, cols L-W
    cumulative_payments = [_num(ws.cell(row=70, column=c).value, 2) for c in range(12, 24)]

    schedules.append({
        "project": project_name,
        "months": months,
        "payments": payments,
        "payment_total": payment_total,
        "revenues": revenues,
        "total_revenues": total_revenues,
        "cumulative_revenues": cumulative_revenues,
        "cumulative_payments": cumulative_payments,
    })

    return schedules


# ---------------------------------------------------------------------------
# Operations (Ember Operating Revenues)
# ---------------------------------------------------------------------------

_OPS_CATEGORIES = [
    "Development Fees",
    "Project Personnel",
    "Bookkeeping",
    "Receivables & Bond Fees",
    "EB Fees (Lots)",
    "EB Fees (Pods & Commercial)",
]

def _parse_operations(ws) -> dict:
    """Parse the 'Operations' tab.

    Layout-robust: every section is located by a LABEL ANCHOR rather than a
    hardcoded row number, and the yearly / quarterly / next-12-month rollups
    are COMPUTED from the per-project monthly rows. This means adding a
    project to the tab — which pushes the whole monthly section down and adds
    another 6-row block — is handled automatically, with no code change.

    Anchors used:
      * "Model Dates" row (col D)  -> the monthly date axis (cols 5+).
      * Each project block         -> project name in col C, then 6 category
                                      rows (col D = category), until the
                                      "Totals" row in col C.
      * "Expected ..." rows (col D)-> headline KPIs (label in D, value in E).
    """
    from datetime import date as _date

    # "Data From" date — label "Date Updated" in C1, date in D1.
    report_date = (
        _date_iso(ws.cell(row=1, column=4).value)  # D1
        or _date.today().isoformat()
    )

    max_row = ws.max_row or 300
    max_col = ws.max_column or 300

    # --- Locate the monthly date axis by the "Model Dates" label (col D) ---
    model_row = None
    for r in range(1, max_row + 1):
        if _str(ws.cell(row=r, column=4).value) == "Model Dates":
            model_row = r
            break
    if model_row is None:
        return {}

    dates = []       # ISO date strings, one per monthly column
    date_cols = []   # corresponding sheet column indices
    for c in range(5, max_col + 1):
        v = ws.cell(row=model_row, column=c).value
        if v is None:
            break
        d = v.date() if hasattr(v, "date") else v
        dates.append(d.isoformat() if hasattr(d, "isoformat") else str(d))
        date_cols.append(c)
    if not dates:
        return {}

    def _ym(iso):
        try:
            parts = iso.split("-")
            return int(parts[0]), int(parts[1])
        except Exception:
            return None, None

    # --- Per-project monthly blocks: 6 category rows each, starting just
    #     below the Model Dates row, until the "Totals" row in col C. ---
    monthly_rows = []
    r = model_row + 1
    while r + 5 <= max_row:
        name = _str(ws.cell(row=r, column=3).value)
        if name == "" or name.lower() == "totals":
            break
        for offset in range(6):
            cat = _str(ws.cell(row=r + offset, column=4).value)
            values = [_num(ws.cell(row=r + offset, column=c).value, 2) for c in date_cols]
            monthly_rows.append({"project": name, "category": cat, "values": values})
        r += 6

    n = len(date_cols)
    # Monthly totals per column = sum of every project/category value.
    monthly_totals = [
        round(sum(mr["values"][i] for mr in monthly_rows), 2) for i in range(n)
    ]

    # Category totals per column, summed across all projects (the rollup basis).
    cat_series = {cat: [0.0] * n for cat in _OPS_CATEGORIES}
    for mr in monthly_rows:
        series = cat_series.get(mr["category"])
        if series is None:
            continue
        for i, v in enumerate(mr["values"]):
            series[i] += v

    today = _date.today()
    # Index (into date_cols/dates) of the current calendar month, if present.
    today_col = None
    for i, iso in enumerate(dates):
        y, m = _ym(iso)
        if y == today.year and m == today.month:
            today_col = i
            break

    # --- Headline KPIs: every "Expected ..." row (label in D, value in E) ---
    kpis = []
    for r2 in range(model_row, max_row + 1):
        label = _str(ws.cell(row=r2, column=4).value)
        if label.startswith("Expected"):
            kpis.append({"label": label, "value": _num(ws.cell(row=r2, column=5).value, 2)})

    # --- Yearly rollup: next 5 calendar years from the current year ---
    years_present = sorted({_ym(iso)[0] for iso in dates if _ym(iso)[0] is not None})
    yearly_years = [y for y in years_present if y >= today.year][:5]
    year_idx = {y: [i for i, iso in enumerate(dates) if _ym(iso)[0] == y] for y in yearly_years}
    yearly_rows = []
    for cat in _OPS_CATEGORIES:
        vals = [round(sum(cat_series[cat][i] for i in year_idx[y]), 2) for y in yearly_years]
        yearly_rows.append({"label": cat, "values": vals})
    yearly_totals = [round(sum(monthly_totals[i] for i in year_idx[y]), 2) for y in yearly_years]

    # --- Quarterly rollup: next 12 quarters from the current quarter ---
    def _q(iso):
        y, m = _ym(iso)
        return f"Q{(m - 1) // 3 + 1} {y}" if y else None

    quarter_order = []
    quarter_idx = {}
    for i, iso in enumerate(dates):
        ql = _q(iso)
        if ql is None:
            continue
        if ql not in quarter_idx:
            quarter_idx[ql] = []
            quarter_order.append(ql)
        quarter_idx[ql].append(i)
    current_q = f"Q{(today.month - 1) // 3 + 1} {today.year}"
    try:
        start_idx = quarter_order.index(current_q)
    except ValueError:
        start_idx = 0
    next_12_quarters = quarter_order[start_idx:start_idx + 12]
    quarterly_rows = [
        {"label": cat,
         "values": [round(sum(cat_series[cat][i] for i in quarter_idx[ql]), 2) for ql in next_12_quarters]}
        for cat in _OPS_CATEGORIES
    ]
    quarterly_totals = [
        round(sum(monthly_totals[i] for i in quarter_idx[ql]), 2) for ql in next_12_quarters
    ]

    # --- Next 12 months from the current month ---
    next_12_dates = []
    next_12_month_rows = []
    n12_totals = []
    if today_col is not None:
        cols = list(range(today_col, min(today_col + 12, n)))
        next_12_dates = [dates[i] for i in cols]
        for cat in _OPS_CATEGORIES:
            next_12_month_rows.append({"label": cat, "values": [round(cat_series[cat][i], 2) for i in cols]})
        n12_totals = [monthly_totals[i] for i in cols]

    # Computed KPI appended for parity with the historical output shape.
    kpis.append({"label": "Expected Next 12 Months", "value": round(sum(n12_totals), 2)})

    return {
        "date": report_date,
        "data_from": report_date,  # canonical name for "Data From" UI
        "kpis": kpis,
        "yearly_rollup": {
            "years": yearly_years,
            "rows": yearly_rows,
            "totals": yearly_totals,
        },
        "monthly": {
            "dates": dates,
            "rows": monthly_rows,
            "totals": monthly_totals,
        },
        "next_12_months": {
            "dates": next_12_dates,
            "rows": next_12_month_rows,
            "totals": n12_totals,
        },
        "quarterly_rollup": {
            "quarters": next_12_quarters,
            "rows": quarterly_rows,
            "totals": quarterly_totals,
        },
    }


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse_dashboard(file_bytes: bytes) -> dict:
    """
    Parse an Ember Dashboard Excel file and return structured JSON with
    Consolidated Project Returns and Loan Capacities & Debt Schedules.

    Parameters
    ----------
    file_bytes : bytes
        Raw bytes of the .xlsx file.

    Returns
    -------
    dict
        Nested dictionary with "returns" and "loans" top-level keys.
    """
    wb = openpyxl.load_workbook(
        filename=io.BytesIO(file_bytes),
        data_only=True,
        read_only=False,
    )

    # --- Returns tab ---
    returns_ws = None
    for name in wb.sheetnames:
        if "consolidated" in name.lower() and "return" in name.lower():
            returns_ws = wb[name]
            break
    if returns_ws is None:
        # Fallback: try exact name
        returns_ws = wb.get("Consolidated Project Returns")

    returns_data = _parse_returns(returns_ws) if returns_ws else {}

    # --- Loans tab ---
    loans_ws = None
    for name in wb.sheetnames:
        if "loan" in name.lower() and ("capacit" in name.lower() or "ds" in name.lower()):
            loans_ws = wb[name]
            break
    if loans_ws is None:
        loans_ws = wb.get("Loan Capacities & DS")

    loans_data = _parse_loans(loans_ws) if loans_ws else {}

    # --- Operations tab ---
    ops_ws = None
    for name in wb.sheetnames:
        if "operation" in name.lower():
            ops_ws = wb[name]
            break
    if ops_ws is None:
        ops_ws = wb.get("Operations")

    ops_data = _parse_operations(ops_ws) if ops_ws else {}

    # Legacy override: older workbooks didn't have a date cell on the
    # loans tab, so we copied the returns date over. Now that we read
    # U3 on the Loan Capacities & DS tab directly, only fall back to
    # the returns date when the loans tab genuinely has no date.
    if loans_data and not loans_data.get("data_from") and returns_data.get("date"):
        loans_data["date"]      = returns_data["date"]
        loans_data["data_from"] = returns_data["date"]

    # Debt tab — the Loan Capacities & DS sheet pulls debt-schedule rows
    # from here via formulas, so Debt!D1's "Date Updated" is the as-of
    # date for the debt schedules section (separate from the loan
    # capacities cutoff in U3 on the loans tab). Surface it as a second
    # date field on the Loans dashboard footer.
    if "Debt" in wb.sheetnames and loans_data is not None:
        debt_ws = wb["Debt"]
        debt_data_from = _date_iso(debt_ws.cell(row=1, column=4).value)  # D1
        if debt_data_from:
            loans_data["debt_data_from"] = debt_data_from

    wb.close()

    return {
        "returns": returns_data,
        "loans": loans_data,
        "operations": ops_data,
    }
