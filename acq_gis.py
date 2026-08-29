"""GIS engine for the Acquisitions tab.

Lifted from the standalone Acquisitions GIS app (csaldierna1-bot/-ember-acquisitions-gis).
Everything here is pure logic - live ArcGIS/REST layer queries, the response
caches, geometry helpers, HCAD/MCAD owner overlays and the spatial enrichment
that decorates a tract with flood, wetlands, utility and school-district facts.
It touches no Flask request state, so app.py can call it directly.

Two things carry over that look odd but are load-bearing, both learned the hard
way against the live services:

  * _safe_union repairs self-intersecting rings before unioning. NWI and
    FEMA both ship them, unary_union raises TopologyException on them, and
    a bare except upstream used to drop an entire constraint layer silently -
    which reported a site as more developable than it is.
  * Layer fetches retry rather than fail. FEMA NFHL drops roughly one
    connection in four under load and the USFWS wetlands service 503s; a single
    attempt reports a live layer as missing.

Do not add an Overpass mirror without checking it can find Houston -
overpass.osm.ch answers fast and is a Switzerland-only extract that returns
HTTP 200 with zero elements for every Texas query.
"""

import io
import json
import os
import secrets
import threading
import time
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from functools import wraps
from pathlib import Path

# Where this tab's file-backed data lives: the parcel cache and the
# special-district rate sheet. On Railway this is a mounted volume; locally it
# falls back to ./storage. acq_parcels reads the same variable.
_ACQ_DATA_DIR = Path(os.environ.get("ACQ_DATA_DIR") or (Path(__file__).parent / "storage"))

import requests
from werkzeug.security import generate_password_hash, check_password_hash
from shapely.geometry import Point
from shapely.ops import transform
from pyproj import Transformer

# --------------------------------------------------------------------------
# Endpoints
# --------------------------------------------------------------------------
ENDPOINTS = {
    "parcels":       "https://services1.arcgis.com/1mtXwieMId59thmg/arcgis/rest/services/2019_Texas_Parcels_StratMap/FeatureServer/0/query",
    "counties":      "https://tigerweb.geo.census.gov/arcgis/rest/services/TIGERweb/State_County/MapServer/1/query",
    "fema_flood":    "https://hazards.fema.gov/arcgis/rest/services/public/NFHL/MapServer/28/query",
    "rrc_wells":     "https://www.gis.hctx.net/arcgishcpid/rest/services/TXRRC/Wells/MapServer/0/query",
    "rrc_pipelines": "https://www.gis.hctx.net/arcgishcpid/rest/services/TXRRC/Pipelines/MapServer/0/query",
    "tea_schools":   "https://services2.arcgis.com/5MVN2jsqIrNZD4tP/arcgis/rest/services/Districts1920/FeatureServer/0/query",
    "tceq_water":    "https://harcags.harcresearch.org/arcgisserver/rest/services/Boundaries/TCEQ_Water_Districts/MapServer/0/query",
    "puc_ccn":       "https://services.twdb.texas.gov/arcgis/rest/services/PWS/Public_Utility_Commission_CCN_Water/MapServer/0/query",
    "wetlands":      "https://fwspublicservices.wim.usgs.gov/wetlandsmapservice/rest/services/Wetlands/MapServer/0/query",
    # NOTE: the old wetlands_alt (www.fws.gov/wetlands/arcgis/...) was removed —
    # it 404s on every request now, so it only added latency before each failure.
    "streams":       "https://hydro.nationalmap.gov/arcgis/rest/services/nhd/MapServer/6/query",
    "transmission":  "https://services2.arcgis.com/FiaPA4ga0iQKduv3/arcgis/rest/services/US_Electric_Power_Transmission_Lines/FeatureServer/0/query",
    "etj":           "https://gis.h-gac.com/arcgis/rest/services/Open_Data/Boundaries/MapServer/9/query",
    # HCAD live — for Harris County, refreshed weekly (fresher owner names)
    "hcad":          "https://www.gis.hctx.net/arcgis/rest/services/HCAD/Parcels/MapServer/0/query",
    # MCAD live — for Montgomery County, refreshed monthly
    "mcad":          "https://services1.arcgis.com/PRoAPGnMSUqvTrzq/arcgis/rest/services/Tax_Parcel_view/FeatureServer/0/query",
    # Electric Transmission & Distribution Utilities (TEPRI / PUC) — CenterPoint, Oncor, AEP, TNMP
    "electric":      "https://services2.arcgis.com/LYMgRMwHfrWWEg3s/arcgis/rest/services/TEPRI_TDU_Boundary/FeatureServer/0/query",
    # USGS National Elevation Point Query Service — per-point elevation in ft (no auth)
    "usgs_elevation": "https://epqs.nationalmap.gov/v1/json",
    # TxDOT planned and active highway projects
    "txdot_projects": "https://services.arcgis.com/KTcxiTD9dsQw4r7Z/arcgis/rest/services/TxDOT_Projects/FeatureServer/0/query",
}

# --------------------------------------------------------------------------
# Published property-tax rates per $100 valuation.
# Source: Tax-year 2024 truth-in-taxation rates published by each jurisdiction
# (county tax offices, ISD adoption notices, city budget documents).
# UPDATE ANNUALLY: replace this dict when each entity adopts its TY2025 rate.
# Final tax bills are computed by Harris County Tax Office; these are reference
# rates for estimating combined exposure.
# --------------------------------------------------------------------------
TAX_YEAR = 2024
TAX_RATES = {
    # County-level — applies to every Harris County parcel.
    "Harris County (general)":                ("0.350772", "County operations"),
    "Harris County Flood Control District":   ("0.027583", "Drainage / flood mitigation"),
    "Harris County Hospital District":        ("0.135159", "Public health system"),
    "Port of Houston Authority":              ("0.005107", "Port infrastructure"),

    # Community college districts — at most one applies per parcel by geography
    "Lone Star College System":               ("0.107080", "Community college"),
    "Houston Community College System":       ("0.099091", "Community college"),
    "San Jacinto Community College":          ("0.157400", "Community college"),

    # School districts in / around Harris County (TEA name -> rate)
    "Aldine ISD":                             ("1.0936",  "School district"),
    "Alief ISD":                              ("1.1075",  "School district"),
    "Channelview ISD":                        ("1.1075",  "School district"),
    "Crosby ISD":                             ("1.1075",  "School district"),
    "Cypress-Fairbanks ISD":                  ("1.0834",  "School district"),
    "Cy-Fair ISD":                            ("1.0834",  "School district"),
    "Deer Park ISD":                          ("1.0817",  "School district"),
    "Galena Park ISD":                        ("1.1075",  "School district"),
    "Goose Creek CISD":                       ("1.1075",  "School district"),
    "Houston ISD":                            ("0.8683",  "School district"),
    "Huffman ISD":                            ("1.0834",  "School district"),
    "Humble ISD":                             ("1.1146",  "School district"),
    "Katy ISD":                               ("1.0934",  "School district"),
    "Klein ISD":                              ("1.1075",  "School district"),
    "La Porte ISD":                           ("1.0817",  "School district"),
    "Magnolia ISD":                           ("1.0734",  "School district"),
    "Pasadena ISD":                           ("1.1075",  "School district"),
    "Sheldon ISD":                            ("1.1075",  "School district"),
    "Spring Branch ISD":                      ("1.0584",  "School district"),
    "Spring ISD":                             ("1.2575",  "School district"),
    "Stafford MSD":                           ("0.8809",  "School district"),
    "Tomball ISD":                            ("1.1075",  "School district"),
    "Waller ISD":                             ("1.0934",  "School district"),

    # Cities (only relevant if the tract is inside city limits, NOT just ETJ)
    "Houston":                                ("0.51919", "City property tax"),
    "Pasadena":                               ("0.6000",  "City property tax"),
    "Baytown":                                ("0.7637",  "City property tax"),
    "Pearland":                               ("0.6260",  "City property tax"),
    "Tomball":                                ("0.2790",  "City property tax"),
    "Katy":                                   ("0.4770",  "City property tax"),
    "La Porte":                               ("0.6900",  "City property tax"),
    "Humble":                                 ("0.5500",  "City property tax"),
    "Bellaire":                               ("0.4673",  "City property tax"),
    "Deer Park":                              ("0.7000",  "City property tax"),
    "Friendswood":                            ("0.4873",  "City property tax"),
    "Webster":                                ("0.4900",  "City property tax"),
    "Stafford":                               ("0.0000",  "City property tax (Stafford has no city ad-valorem)"),
}


# Special-district tax rates ship as a spreadsheet next to the parcel cache.
# Same directory as everything else this tab persists - see acq_parcels.
STORAGE_DIR = _ACQ_DATA_DIR


def _load_special_district_rates():
    """Parse the TX Comptroller's annual Special District Rates & Levies XLSX.
    Returns a dict keyed by NORMALIZED district name -> (rate_per_$100, note).

    Source: https://comptroller.texas.gov/taxes/property-tax/rates/
    The file ships with rates for every MUD, ESD, WCID, LID, hospital district,
    junior college, etc. in Texas. ~2,500 entries. Update annually by replacing
    storage/{TAX_YEAR}-special-district-rates.xlsx.
    """
    from openpyxl import load_workbook
    rates = {}
    xlsx = STORAGE_DIR / f"{TAX_YEAR}-special-district-rates.xlsx"
    if not xlsx.exists():
        print(f"[tax-rates] {xlsx.name} not found — special-district rates unavailable", flush=True)
        return rates
    try:
        wb = load_workbook(xlsx, read_only=True, data_only=True)
        ws = wb["Detail"] if "Detail" in wb.sheetnames else wb[wb.sheetnames[0]]
        rows = ws.iter_rows(min_row=4, values_only=True)   # skip header rows 1-3
        for r in rows:
            if not r or not r[5]:   # TU NAME column
                continue
            tu_name = str(r[5]).strip()
            county  = str(r[3] or "").strip()
            total_rate = r[15] if len(r) > 15 else None    # TOTAL TAX RATE column
            if total_rate is None or total_rate == "":
                continue
            try:
                rate = float(total_rate)
            except (TypeError, ValueError):
                continue
            note = f"Special district ({county} County)"
            # Index by both the raw name and a normalized form for fuzzy matching
            rates[tu_name] = (f"{rate:.6f}", note)
            norm = _normalize_district_name(tu_name)
            if norm and norm != tu_name and norm not in rates:
                rates[norm] = (f"{rate:.6f}", note)
        print(f"[tax-rates] loaded {len(rates)} special-district rates from {xlsx.name}", flush=True)
    except Exception as e:
        print(f"[tax-rates] failed to parse {xlsx.name}: {e}", flush=True)
    return rates


def _normalize_district_name(name):
    """Build a canonical form so HCAD's 'HARRIS COUNTY MUD 319' matches the
    Comptroller's 'Harris Co MUD #319', etc. Returns UPPER, stripped, with
    common variants collapsed."""
    if not name:
        return ""
    n = str(name).upper().strip()
    # Common abbreviations
    n = n.replace("MUNICIPAL UTILITY DISTRICT", "MUD")
    n = n.replace("EMERGENCY SERVICES DISTRICT", "ESD")
    n = n.replace("WATER CONTROL AND IMPROVEMENT DISTRICT", "WCID")
    n = n.replace("WATER CONTROL & IMPROVEMENT DISTRICT", "WCID")
    n = n.replace("LEVEE IMPROVEMENT DISTRICT", "LID")
    n = n.replace("FRESH WATER SUPPLY DISTRICT", "FWSD")
    n = n.replace(" CO ", " COUNTY ")
    n = n.replace(" CO. ", " COUNTY ")
    n = n.replace("#", "")
    n = n.replace("NO.", "")
    n = n.replace("NUMBER", "")
    # Collapse whitespace
    return " ".join(n.split())


# Lazy cache — load on first request after the app is fully initialized.
_special_district_rates_cache = None
def _all_tax_rates():
    """Curated rates (county, ISD, city, community college) PLUS the Comptroller's
    annual statewide special-district file (MUDs, ESDs, WCIDs, etc.)."""
    global _special_district_rates_cache
    if _special_district_rates_cache is None:
        _special_district_rates_cache = _load_special_district_rates()
    # Curated entries win on key collisions (we have explicit data for those)
    merged = dict(_special_district_rates_cache)
    merged.update(TAX_RATES)
    return merged

# --------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------
to_utm = Transformer.from_crs("EPSG:4326", "EPSG:32614", always_xy=True).transform
to_wgs = Transformer.from_crs("EPSG:32614", "EPSG:4326", always_xy=True).transform


def to_float(x):
    if x is None:
        return None
    try:
        return float(str(x).replace(",", "").strip())
    except (ValueError, TypeError):
        return None




def _build_params(geometry_polygon=None, bbox=None, where="1=1", out_fields="*",
                  paginate=True, offset=0, page_size=1000, return_geometry=True,
                  max_allowable_offset=None):
    data = {
        "f": "geojson",
        "where": where,
        "outFields": out_fields,
        "outSR": 4326,
        "returnGeometry": "true" if return_geometry else "false",
    }
    # Degrees of tolerance for server-side generalisation. Dense polygon layers
    # (NWI wetlands especially) send far more vertices than a map needs.
    if max_allowable_offset:
        data["maxAllowableOffset"] = max_allowable_offset
    if paginate:
        data["resultOffset"] = offset
        data["resultRecordCount"] = page_size
    if bbox is not None:
        data["geometry"] = f"{bbox[0]},{bbox[1]},{bbox[2]},{bbox[3]}"
        data["geometryType"] = "esriGeometryEnvelope"
        data["inSR"] = 4326
        data["spatialRel"] = "esriSpatialRelIntersects"
    elif geometry_polygon is not None:
        # Esri REST accepts a list of rings — disjoint rings are read as a multipolygon.
        # Handle both Polygon (single exterior + optional holes) and MultiPolygon.
        rings = []
        if geometry_polygon.geom_type == "Polygon":
            rings.append(list(geometry_polygon.exterior.coords))
            for interior in geometry_polygon.interiors:
                rings.append(list(interior.coords))
        elif geometry_polygon.geom_type == "MultiPolygon":
            for sub in geometry_polygon.geoms:
                rings.append(list(sub.exterior.coords))
                for interior in sub.interiors:
                    rings.append(list(interior.coords))
        elif geometry_polygon.geom_type == "GeometryCollection":
            for sub in geometry_polygon.geoms:
                if sub.geom_type == "Polygon":
                    rings.append(list(sub.exterior.coords))
                    for interior in sub.interiors:
                        rings.append(list(interior.coords))
        else:
            raise ValueError(f"_build_params: unsupported geometry {geometry_polygon.geom_type}")
        esri_geom = {
            "rings": rings,
            "spatialReference": {"wkid": 4326},
        }
        data["geometry"] = json.dumps(esri_geom)
        data["geometryType"] = "esriGeometryPolygon"
        data["inSR"] = 4326
        data["spatialRel"] = "esriSpatialRelIntersects"
    return data


def _request(url, data, timeout, method="POST"):
    if method == "POST":
        r = requests.post(url, data=data, timeout=timeout)
    else:
        r = requests.get(url, params=data, timeout=timeout)
    r.raise_for_status()
    try:
        payload = r.json()
    except Exception as e:
        snippet = (r.text or "")[:300].replace("\n", " ")
        raise RuntimeError(f"Non-JSON ({r.status_code}): {snippet!r}") from e
    if "error" in payload:
        raise RuntimeError(f"ArcGIS error: {payload['error']}")
    return payload


def _esri_to_geojson_features(payload):
    feats = payload.get("features", [])
    out = []
    for f in feats:
        attrs = f.get("attributes", {})
        geom = f.get("geometry", {})
        gj = None
        if "rings" in geom:
            gj = {"type": "Polygon" if len(geom["rings"]) == 1 else "MultiPolygon",
                  "coordinates": geom["rings"] if len(geom["rings"]) == 1 else [[r] for r in geom["rings"]]}
        elif "x" in geom and "y" in geom:
            gj = {"type": "Point", "coordinates": [geom["x"], geom["y"]]}
        elif "paths" in geom:
            gj = {"type": "MultiLineString", "coordinates": geom["paths"]}
        out.append({"type": "Feature", "geometry": gj, "properties": attrs})
    return out


# Errors that mean "the server is busy / dropped us", not "the request is wrong".
# FEMA's NFHL in particular resets connections under concurrent load — retrying
# the same request after a short wait succeeds where changing HTTP method does not.
_TRANSIENT_MARKERS = (
    "connection reset", "connection aborted", "connectionreseterror",
    "remotedisconnected", "connection broken", "read timed out",
    "timed out", "temporarily unavailable", "max retries exceeded",
    "502", "503", "504", "429",
)

def _is_transient(err) -> bool:
    s = str(err).lower()
    return any(m in s for m in _TRANSIENT_MARKERS)


def _post_query(url, data, timeout, single_attempt=False, max_retries=3):
    """POST an ArcGIS query with transient-failure retries.

    Strategy order:
      1. POST, retried with exponential backoff on transient network errors
         (FEMA NFHL resets connections when hit concurrently).
      2. GET — some Esri deployments reject long POST bodies.
      3. GET with f=json + manual Esri->GeoJSON conversion.
    """
    import time as _t
    last_exc = None
    for attempt in range(max_retries):
        try:
            return _request(url, data, timeout, method="POST")
        except Exception as e:
            last_exc = e
            # A real ArcGIS error (bad field, bad geometry) won't fix itself.
            if "ArcGIS error" in str(e) or single_attempt:
                raise
            if _is_transient(e) and attempt < max_retries - 1:
                _t.sleep(0.6 * (2 ** attempt))    # 0.6s, 1.2s
                continue
            break

    # Fall back to other request shapes
    try:
        return _request(url, data, timeout, method="GET")
    except Exception:
        pass
    try:
        data_json = dict(data); data_json["f"] = "json"
        payload = _request(url, data_json, timeout, method="GET")
        payload["features"] = _esri_to_geojson_features(payload)
        return payload
    except Exception as e2:
        raise RuntimeError(f"All strategies failed for {url}. Last: {e2}") from last_exc


def _count_query(url, geometry_polygon, bbox, where, timeout):
    data = _build_params(geometry_polygon=geometry_polygon, bbox=bbox,
                         where=where, paginate=False, return_geometry=False)
    data["f"] = "json"
    data["returnCountOnly"] = "true"
    try:
        payload = _request(url, data, timeout, method="POST")
        return int(payload.get("count", 0))
    except Exception:
        return None


def arcgis_query(url, *, geometry_polygon=None, bbox=None, where="1=1",
                 out_fields="*", page_size=1000, timeout=120, max_pages=50,
                 return_geometry=True, parallel_pagination=True, single_attempt=False,
                 max_allowable_offset=None):
    def fetch_page(offset, paginate=True):
        data = _build_params(geometry_polygon=geometry_polygon, bbox=bbox,
                             where=where, out_fields=out_fields,
                             paginate=paginate, offset=offset, page_size=page_size,
                             return_geometry=return_geometry,
                             max_allowable_offset=max_allowable_offset)
        try:
            return _post_query(url, data, timeout, single_attempt=single_attempt), paginate
        except RuntimeError as e:
            if paginate and "Pagination is not supported" in str(e):
                data2 = _build_params(geometry_polygon=geometry_polygon, bbox=bbox,
                                      where=where, out_fields=out_fields,
                                      paginate=False, offset=0, page_size=page_size,
                                      return_geometry=return_geometry,
                                      max_allowable_offset=max_allowable_offset)
                return _post_query(url, data2, timeout, single_attempt=single_attempt), False
            raise

    total = None
    if parallel_pagination:
        total = _count_query(url, geometry_polygon, bbox, where, timeout)

    if total is not None and total > page_size:
        needed_pages = (total + page_size - 1) // page_size
        n_pages = min(max_pages, needed_pages)
        offsets = [i * page_size for i in range(n_pages)]
        all_features = []
        # Some public Esri hosts drop connections when hit hard. FEMA's NFHL is
        # the worst offender — it resets at 12 concurrent page requests but is
        # stable at 3. Throttle per-host rather than globally so the fast
        # services (TxGIO, TIGERweb) keep their parallelism.
        host = (url or "").lower()
        RATE_LIMITED_HOSTS = ("hazards.fema.gov", "fwspublicservices", "fws.gov")
        cap = 3 if any(h in host for h in RATE_LIMITED_HOSTS) else 12
        with ThreadPoolExecutor(max_workers=min(cap, n_pages)) as pool:
            page_futs = {pool.submit(fetch_page, off, True): off for off in offsets}
            for fut in as_completed(page_futs):
                payload, _ = fut.result()
                all_features.extend(payload.get("features", []))
        out = {"type": "FeatureCollection", "features": all_features}
        if needed_pages > n_pages:
            out["_truncated"] = True
            out["_total_count"] = total
            out["_fetched"] = len(all_features)
        return out

    all_features, offset, pages, paginate = [], 0, 0, True
    while True:
        payload, paginate = fetch_page(offset, paginate)
        feats = payload.get("features", [])
        all_features.extend(feats)
        if not paginate or len(feats) < page_size:
            break
        offset += page_size
        pages += 1
        if pages >= max_pages:
            break
    return {"type": "FeatureCollection", "features": all_features}


# --------------------------------------------------------------------------
# Wetlands — USFWS is flaky, so try a backup URL + cache 24h
# --------------------------------------------------------------------------
_wetlands_cache = {}            # key -> (timestamp, result)
_wetlands_cache_lock = threading.Lock()
WETLANDS_CACHE_TTL = 24 * 60 * 60   # 24 hours


# --------------------------------------------------------------------------
# Generic per-layer in-memory cache with TTL.
# Keys are (layer_name, quantized_bbox, out_fields, where) so iterative searches
# in the same area (Hockley refinement, etc.) hit the cache. Stable layers
# (counties, schools, water districts) refresh once per day; live layers
# (parcels, flood, wells, HCAD) stay live every search.
# --------------------------------------------------------------------------
_LAYER_CACHE = {}
_LAYER_CACHE_LOCK = threading.Lock()
LAYER_CACHE_TTL_SEC = 24 * 60 * 60   # 24h

def _quantize_bbox(buffer_wgs, step=0.1):
    """Round bbox to step-degree tiles so nearby searches share cache entries."""
    minx, miny, maxx, maxy = buffer_wgs.bounds
    return (round(minx / step) * step, round(miny / step) * step,
            round(maxx / step) * step, round(maxy / step) * step)


def cached_layer_query(layer_name, buffer_wgs, **kwargs):
    """Fetch a layer via arcgis_query, caching the result for 24h keyed on
    a quantized bbox + the kwargs (out_fields, where). Subsequent calls in
    the same area return the cached FeatureCollection."""
    qbox = _quantize_bbox(buffer_wgs)
    cache_key = (layer_name, qbox,
                  kwargs.get("out_fields"),
                  kwargs.get("where"))
    now = time.time()
    with _LAYER_CACHE_LOCK:
        cached = _LAYER_CACHE.get(cache_key)
        if cached and cached[1] > now:
            return cached[0]
    data = arcgis_query(ENDPOINTS[layer_name], geometry_polygon=buffer_wgs, **kwargs)
    with _LAYER_CACHE_LOCK:
        _LAYER_CACHE[cache_key] = (data, now + LAYER_CACHE_TTL_SEC)
        # Cap cache size to prevent unbounded growth
        if len(_LAYER_CACHE) > 200:
            # Evict oldest 50
            oldest = sorted(_LAYER_CACHE.items(), key=lambda kv: kv[1][1])[:50]
            for k, _ in oldest:
                _LAYER_CACHE.pop(k, None)
    return data


def cached_wetlands(buffer_wgs, lat, lon, radius_mi):
    """Try primary USFWS endpoint, then alt, then stale cache, then empty. Always returns a dict."""
    key = (round(lat, 2), round(lon, 2), round(radius_mi, 1))
    now = time.time()

    with _wetlands_cache_lock:
        cached = _wetlands_cache.get(key)
        if cached and (now - cached[0]) < WETLANDS_CACHE_TTL:
            print(f"[wetlands] cache hit ({len(cached[1].get('features', []))} features)", flush=True)
            return cached[1]

    # The USFWS service is alive and returns real data - a count query over a
    # 10-mile box comes back with 2,527 wetland polygons - but it is slow, and a
    # single 10-second attempt with no retry could never finish a fetch that
    # size. That is why this layer read as "unavailable" while the endpoint was
    # fine. 45 seconds with the normal retry path lets it complete.
    #
    # wetlands_alt is dropped: www.fws.gov/wetlands/arcgis/... now 404s on every
    # request, so trying it only added latency to each failure.
    for url_key in ("wetlands",):
        try:
            # The USFWS NWI service is genuinely slow - measured at ~58s for a
            # single 200-feature page over a 10-mile box, and it returns HTTP 503
            # "Wait time expired" under any sustained load. maxAllowableOffset
            # generalises geometry server-side (~11 m, invisible at map zoom) to
            # cut the payload, and the page cap bounds the worst case so a toggle
            # can never hang the UI. Result may be partial in a dense area; the
            # 6-hour cache means you pay this once.
            result = arcgis_query(
                ENDPOINTS[url_key], geometry_polygon=buffer_wgs,
                timeout=45, parallel_pagination=False,
                page_size=500, max_pages=3,
                max_allowable_offset=0.0001)
            if result and result.get("features") is not None:
                print(f"[wetlands] {url_key} OK ({len(result['features'])} features)", flush=True)
                with _wetlands_cache_lock:
                    _wetlands_cache[key] = (now, result)
                return result
        except Exception as e:
            print(f"[wetlands] {url_key} failed: {str(e)[:80]}", flush=True)

    # Fetch failed — fall back to stale cache if any
    with _wetlands_cache_lock:
        stale = _wetlands_cache.get(key)
        if stale:
            print(f"[wetlands] using stale cache from {datetime.fromtimestamp(stale[0]).isoformat(timespec='seconds')}", flush=True)
            stale[1]["_stale"] = True
            return stale[1]
    print("[wetlands] no data available", flush=True)
    return {"type": "FeatureCollection", "features": [], "error": "wetlands service unavailable"}


# --------------------------------------------------------------------------
# HCAD live overlay — fresher owner names for Harris tracts.
# Per-Prop_ID 24h cache: HCAD refreshes weekly upstream, so 24h cache stays
# well within their freshness window AND eliminates ~6k network calls per
# search (parcel cache alone is 25-30s wasted re-fetching the same owners).
# --------------------------------------------------------------------------
_HCAD_PID_CACHE = {}              # pid -> (props_dict, expires_at)
_HCAD_CACHE_LOCK = threading.Lock()
HCAD_CACHE_TTL_SEC = 24 * 60 * 60
HCAD_REQUIRED_FIELDS = ("total_market_val", "total_appraised_val", "tax_value")


def _hcad_cache_get_bulk(pids):
    """Return {pid: cached_props} for pids whose cache is still fresh.
    Pids missing or expired are NOT in the returned dict."""
    now = time.time()
    out = {}
    with _HCAD_CACHE_LOCK:
        for pid in pids:
            entry = _HCAD_PID_CACHE.get(pid)
            if entry and entry[1] > now:
                out[pid] = entry[0]
    return out


def _hcad_cache_put_bulk(fresh):
    expires_at = time.time() + HCAD_CACHE_TTL_SEC
    with _HCAD_CACHE_LOCK:
        for pid, data in fresh.items():
            _HCAD_PID_CACHE[pid] = (data, expires_at)
        # Cap cache size — Harris has 1.5M parcels, we shouldn't grow unbounded.
        # 200k entries × ~500 bytes each = ~100 MB; evict in batches.
        if len(_HCAD_PID_CACHE) > 200_000:
            oldest = sorted(_HCAD_PID_CACHE.items(), key=lambda kv: kv[1][1])[:50_000]
            for k, _ in oldest:
                _HCAD_PID_CACHE.pop(k, None)


def hcad_live_overlay(tracts):
    """For each Harris County tract, fetch the live HCAD record and merge in
    the fresh owner_name / mail_addr. HCAD updates weekly. Chunks run in
    parallel so this stays fast even for thousands of Harris tracts.

    Previously this looked at ANY numeric Prop_ID, but Fort Bend / Waller /
    Brazos all have numeric prop_ids too — they'd return nothing from HCAD
    but still cost a round-trip. Filter to county='Harris' so multi-county
    corridor searches don't waste calls."""
    harris_pids = [
        str(f["properties"].get("Prop_ID") or "").strip()
        for f in tracts["features"]
        if (f["properties"].get("_county") or "").strip().lower() == "harris"
        and str(f["properties"].get("Prop_ID") or "").strip().isdigit()
    ]
    if not harris_pids:
        return 0

    # Check the per-Prop_ID cache first — within 24h, HCAD data is fresh enough
    # (HCAD's own refresh is weekly). Only fetch cache-misses live.
    cached_all = _hcad_cache_get_bulk(harris_pids)
    cached = {
        pid: data for pid, data in cached_all.items()
        if all(k in data for k in HCAD_REQUIRED_FIELDS)
    }
    miss_pids = [p for p in harris_pids if p not in cached]
    print(f"  [hcad] {len(cached):,} cached, {len(miss_pids):,} need live fetch", flush=True)

    chunk_size = 50
    chunks = [miss_pids[i:i + chunk_size] for i in range(0, len(miss_pids), chunk_size)]

    def fetch_chunk(chunk):
        if not chunk:
            return []
        in_clause = ",".join(f"'{p}'" for p in chunk)
        try:
            fc = arcgis_query(
                ENDPOINTS["hcad"],
                where=f"HCAD_NUM IN ({in_clause})",
                # Also fetch new_owner_date + tax_year (freshness markers) and the
                # legal_dscr_* + site_str_* fields so we can rebuild a clean site address
                # for parcels where StratMap's SITUS_ADDR is empty.
                out_fields=("HCAD_NUM,owner_name_1,mail_addr_1,mail_city,mail_state,mail_zip,"
                            "total_market_val,total_appraised_val,tax_value,"
                            "Acreage,new_owner_date,tax_year,"
                            "legal_dscr_1,legal_dscr_2,legal_dscr_3,"
                            "site_str_num,site_str_pfx,site_str_name,site_str_sfx,"
                            "site_str_sfx_dir,site_city,site_zip"),
                page_size=200, max_pages=5, return_geometry=False, parallel_pagination=False,
            )
            return fc.get("features", [])
        except Exception as e:
            print(f"  HCAD overlay chunk failed: {e}", flush=True)
            return []

    fresh = dict(cached)   # start with what's in cache, add what we fetch
    newly_fetched = {}
    if chunks:
        with ThreadPoolExecutor(max_workers=10) as pool:
            for features in pool.map(fetch_chunk, chunks):
                for f in features:
                    p = f["properties"] or {}
                    pid = str(p.get("HCAD_NUM") or "").strip()
                    if pid:
                        fresh[pid] = p
                        newly_fetched[pid] = p
    # Stash the newly-fetched in cache for next time
    if newly_fetched:
        _hcad_cache_put_bulk(newly_fetched)

    # Merge fresh fields. Critically also OVERWRITE Acres from HCAD when available —
    # StratMap's GIS_AREA can be 5–10x off from reality on re-platted tracts (a 444 ac tract
    # shows as 2,374 in StratMap because StratMap still has the pre-subdivision parent parcel).
    def _parse_hcad_acres(s):
        if not s:
            return None
        try:
            return float(str(s).replace(" AC", "").replace(",", "").strip())
        except (ValueError, TypeError):
            return None

    def _epoch_ms_to_date(v):
        """HCAD ships new_owner_date as a Unix ms epoch (e.g. 1609459200000 → 2021-01-01)."""
        if v in (None, 0, "", "0"):
            return None
        try:
            from datetime import datetime as _dt, timezone as _tz
            return _dt.fromtimestamp(int(v) / 1000, tz=_tz.utc).strftime("%Y-%m-%d")
        except (ValueError, TypeError, OSError):
            return None

    def _clean(s):
        s = (s or "")
        if not isinstance(s, str):
            s = str(s)
        return s.strip()

    def _addr_has_street(s):
        r"""True only if the address has real content in the first segment (before any comma).
        Catches StratMap's ', , TX' garbage so we know to rebuild from HCAD components."""
        if not s:
            return False
        first = str(s).split(",")[0].strip()
        return bool(first) and any(ch.isalnum() for ch in first)

    merged = 0
    for f in tracts["features"]:
        pid = str(f["properties"].get("Prop_ID") or "").strip()
        if pid in fresh:
            h = fresh[pid]
            props = f["properties"]
            if h.get("owner_name_1"):
                props["OWNER_NAME"] = h["owner_name_1"]
                props["_hcad_owner_verified"] = True
            mail_parts = [h.get("mail_addr_1"), h.get("mail_city"), h.get("mail_state"), str(h.get("mail_zip") or "")]
            mail = ", ".join(p for p in mail_parts if p)
            if mail:
                props["MAIL_ADDR"] = mail
            if h.get("total_market_val"):
                props["MARKET_VAL"] = h["total_market_val"]
            if h.get("total_appraised_val"):
                props["total_appraised_val"] = h["total_appraised_val"]
            if h.get("tax_value"):
                props["tax_value"] = h["tax_value"]
            hcad_acres = _parse_hcad_acres(h.get("Acreage"))
            if hcad_acres is not None and hcad_acres > 0:
                props["_hcad_acres"] = hcad_acres
                # Use HCAD's acreage as the authoritative value when present
                props["Acres"] = round(hcad_acres, 1)
            # Freshness markers — exposes HCAD's staleness window in the popup
            since = _epoch_ms_to_date(h.get("new_owner_date"))
            if since:
                props["_owner_since"] = since
            if h.get("tax_year"):
                props["_hcad_tax_year"] = h.get("tax_year")
            # Rebuild a usable site address from HCAD components when StratMap's SITUS_ADDR
            # has no real street content (StratMap often ships ', , TX' or similar garbage).
            if not _addr_has_street(props.get("SITUS_ADDR")):
                num    = _clean(h.get("site_str_num"))
                pfx    = _clean(h.get("site_str_pfx"))
                name   = _clean(h.get("site_str_name"))
                sfx    = _clean(h.get("site_str_sfx"))
                sfxdir = _clean(h.get("site_str_sfx_dir"))
                city   = _clean(h.get("site_city"))
                zipc   = _clean(h.get("site_zip"))
                street = " ".join(p for p in [num if num and num != "0" else "", pfx, name, sfx, sfxdir] if p).strip()
                site = ", ".join(p for p in [street, city, "TX", zipc] if p).strip(", ")
                if street:
                    props["SITUS_ADDR"] = site
            # Build a clean HCAD legal description (preferred over StratMap's LEGAL_DESC)
            legal_parts = [_clean(h.get(k)) for k in ("legal_dscr_1", "legal_dscr_2", "legal_dscr_3")]
            legal = " ".join(p for p in legal_parts if p).strip()
            if legal:
                props["LEGAL_DESC"] = legal
            merged += 1
    return merged


def mcad_live_overlay(tracts):
    """Same idea as HCAD overlay, but for Montgomery County. MCAD's PIN field
    is the parcel key (small integer). Filter to county='Montgomery' so we
    don't waste round-trips for tracts in other counties that happen to have
    short-numeric Prop_IDs (Fort Bend, Galveston, etc. all match the same
    length pattern)."""
    candidates = []
    for f in tracts["features"]:
        if (f["properties"].get("_county") or "").strip().lower() != "montgomery":
            continue
        p = f["properties"].get("Prop_ID")
        if not p:
            continue
        ps = str(p).strip()
        # MCAD PINs are pure-digit integers, usually 4-7 chars. Filter loosely.
        if ps.isdigit() and 3 < len(ps) < 9:
            candidates.append(ps)
    if not candidates:
        return 0

    chunk_size = 80
    chunks = [candidates[i:i + chunk_size] for i in range(0, len(candidates), chunk_size)]

    def fetch_chunk(chunk):
        # MCAD PIN is integer — IN clause without quotes
        in_clause = ",".join(chunk)
        try:
            fc = arcgis_query(
                ENDPOINTS["mcad"],
                where=f"PIN IN ({in_clause})",
                out_fields="PIN,ownerName,ownerAddress,situs",
                page_size=500, max_pages=5, return_geometry=False, parallel_pagination=False,
            )
            return fc.get("features", [])
        except Exception as e:
            print(f"  MCAD overlay chunk failed: {e}", flush=True)
            return []

    fresh = {}
    with ThreadPoolExecutor(max_workers=10) as pool:
        for features in pool.map(fetch_chunk, chunks):
            for f in features:
                p = f["properties"] or {}
                pid = str(p.get("PIN") or "").strip()
                if pid:
                    fresh[pid] = p

    merged = 0
    for f in tracts["features"]:
        pid = str(f["properties"].get("Prop_ID") or "").strip()
        if pid in fresh:
            m = fresh[pid]
            props = f["properties"]
            # Only override if HCAD didn't already verify (HCAD is fresher for Harris)
            if not props.get("_hcad_owner_verified") and m.get("ownerName"):
                props["OWNER_NAME"] = m["ownerName"]
                props["_mcad_owner_verified"] = True
                if m.get("ownerAddress"):
                    props["MAIL_ADDR"] = m["ownerAddress"]
                if m.get("situs"):
                    props["SITUS_ADDR"] = m["situs"]
                merged += 1
    return merged


# --------------------------------------------------------------------------
# Main search pipeline
# --------------------------------------------------------------------------
def _fetch_matching_tracts(buffer_wgs, min_acres, max_acres, radius_mi=None):
    """Pull parcels in the buffer matching the acreage band.

    Strategy (in priority order):
      1. If every county the buffer touches is fully cached, query SQLite locally
         (sub-second; same data as StratMap, refreshed weekly).
      2. Otherwise fall back to a LIVE StratMap query, paginated.
      3. HCAD / MCAD overlays still run live downstream regardless, so owner
         names stay current even when parcels come from the cache.
    """
    # --- Cache fast-path ---------------------------------------------------
    try:
        import acq_parcels as parcel_cache
        coverage = parcel_cache.coverage_for_polygon(buffer_wgs)
        if parcel_cache.is_fully_cached(coverage):
            fc = parcel_cache.query_parcels_in_polygon(buffer_wgs, min_acres, max_acres)
            names = ",".join(c["county_name"] for c in coverage)
            print(f"[cache] HIT — {len(fc['features']):,} parcels from local DB ({names})", flush=True)
            return fc
        elif coverage:
            uncached = [c["county_name"] for c in coverage if not c["cached"]]
            print(f"[cache] MISS — uncached counties in buffer: {uncached}; using live StratMap",
                  flush=True)
    except Exception as e:
        print(f"[cache] lookup error, falling back to live: {e}", flush=True)
    # --- Live StratMap fallback -------------------------------------------
    # Only filter server-side by a generous LOWER Shape__Area bound.
    # SAFETY_LOW is very small (5% of min_acres) because StratMap's GIS_AREA can be wildly
    # wrong for re-platted tracts — a tract that's actually 444 ac may have StratMap GIS_AREA
    # of 2,374 (or vice versa). We need to pull it through so HCAD overlay can correct the acres.
    SQDEG_PER_ACRE = 3.8e-7
    SAFETY_LOW = 0.05
    low = max(SAFETY_LOW * min_acres * SQDEG_PER_ACRE, 1e-9)
    where = f"Shape__Area > {low}"

    # Scale pagination cap with the search area so a 100-mi radius doesn't truncate.
    # 40 pages × 2,000 = 80,000 records — fine for ~25 mi metro; insufficient for 100 mi
    # which covers ~10 counties and hundreds of thousands of parcels.
    if radius_mi and radius_mi >= 50:
        max_pages = 300   # up to 600,000 records — covers 100 mi metro Houston
    elif radius_mi and radius_mi >= 25:
        max_pages = 120
    else:
        max_pages = 40
    out_fields = "Prop_ID,OWNER_NAME,LEGAL_DESC,SITUS_ADDR,MAIL_ADDR,LEGAL_AREA,GIS_AREA"
    fc = arcgis_query(
        ENDPOINTS["parcels"], geometry_polygon=buffer_wgs,
        where=where,
        out_fields=out_fields, page_size=2000, max_pages=max_pages,
    )
    # If the response was capped, surface that to the caller so we can warn in the UI.
    truncated = fc.get("_truncated", False)
    total_in_buffer = fc.get("_total_count")

    # NO client-side acreage band here. StratMap's GIS_AREA / LEGAL_AREA are unreliable
    # for re-platted parcels (a 444-ac tract can show as 50 OR 5000 in StratMap), so any
    # client-side acreage filter at this stage would drop legitimate tracts before HCAD
    # has a chance to correct the acreage. The server-side `Shape__Area > low` already
    # gates on the polygon's actual geometric area (which is reliable), and the post-
    # overlay re-filter in run_search() applies the user's exact range AFTER HCAD/MCAD
    # have corrected the acreage. So this stage just dedupes by Prop_ID.
    #
    # Dedup ONLY by Prop_ID. The earlier (owner, rounded_acres) key was over-aggressive
    # — it dropped legitimate same-owner same-size adjacent tracts (assemblage holdings).
    matching, seen_pid = [], set()
    for f in fc["features"]:
        props = f["properties"] or {}
        acres = to_float(props.get("GIS_AREA")) or to_float(props.get("LEGAL_AREA")) or 0
        pid = str(props.get("Prop_ID") or "").strip()
        if pid and pid in seen_pid:
            continue
        if pid:
            seen_pid.add(pid)
        props["Acres"] = round(acres, 1)
        f["properties"] = props
        matching.append(f)
    # Sort by acres descending (biggest first — most useful for screening)
    matching.sort(key=lambda f: -(f["properties"].get("Acres") or 0))

    # NO CAP. Every matching tract is returned. Only truncation is the upstream parcel-fetch
    # ceiling (80k records from StratMap per buffer).
    result = {"type": "FeatureCollection", "features": matching}
    if truncated:
        result["_truncated"] = True
        result["_total_in_buffer"] = total_in_buffer
        result["_fetched"] = fc.get("_fetched")
        result["_total_matched"] = len(matching)
    return result


def run_search(lat, lon, radius_mi, min_acres, max_acres, use_hcad_overlay=True,
               polygon_geojson=None, tracts_only=False):
    """If polygon_geojson is provided, use that polygon as the search buffer.
    Otherwise build a circular buffer from (lat, lon, radius_mi).

    `tracts_only=True` skips the 13 ancillary layer queries (flood, wetlands,
    wells, pipelines, schools, etc.) and just returns the matching tracts.
    Big speed win for the /corridors workspace which only needs tract data."""
    from shapely.geometry import shape as shp_shape
    t0 = time.time()
    if polygon_geojson:
        buffer_wgs = shp_shape(polygon_geojson)
        # If user provided a polygon but not a center, use the centroid for HCAD/MCAD overlays
        if not lat or not lon:
            c = buffer_wgs.centroid
            lat, lon = c.y, c.x
    else:
        center_wgs = Point(lon, lat)
        center_utm = transform(to_utm, center_wgs)
        buffer_utm = center_utm.buffer(radius_mi * 1609.344)
        buffer_wgs = transform(to_wgs, buffer_utm)

    # Static layers (refresh ~yearly): cache for 24h on a quantized bbox so iterative
    # searches in the same area don't re-fetch. Live layers (parcels, flood, wells,
    # pipelines) stay direct so they always reflect today's data.
    if tracts_only:
        # Counties is included even in tracts_only mode — it's a tiny cached
        # query (10 features for the Houston metro) and lets us populate the
        # `_county` field on every tract via the spatial join. Without it the
        # corridors-page "Top counties" KPI just shows "Unknown".
        jobs = {
            "tracts":   lambda: _fetch_matching_tracts(buffer_wgs, min_acres, max_acres, radius_mi),
            "counties": lambda: cached_layer_query("counties", buffer_wgs),
        }
    else:
        jobs = {
            "tracts":       lambda: _fetch_matching_tracts(buffer_wgs, min_acres, max_acres, radius_mi),
            "counties":     lambda: cached_layer_query("counties",    buffer_wgs),
            "schools":      lambda: cached_layer_query("tea_schools", buffer_wgs),
            # TCEQ water districts: filter to ACCURACY='P' (validated boundaries) + STATUS='A' (active)
            "water_dist":   lambda: cached_layer_query("tceq_water",  buffer_wgs,
                                                  where="ACCURACY='P' AND STATUS='A'",
                                                  out_fields="NAME,TYPE,TYPE_DESCRIPTION,COUNTY,DISTRICT_ID,Area_Acres"),
            # PUC CCN: keep only essential identification fields
            "ccn":          lambda: cached_layer_query("puc_ccn",     buffer_wgs,
                                                  out_fields="UTILITY,CCN_NO,COUNTY,CCN_TYPE"),
            # FEMA flood panels change at most annually — cache 24h
            "flood":        lambda: cached_layer_query("fema_flood",  buffer_wgs,
                                                  where="FLD_ZONE IN ('A','AE','AH','AO','AR','A99','V','VE')",
                                                  out_fields="FLD_ZONE,ZONE_SUBTY"),
            # Wells & pipelines from RRC change with permit filings — keep live
            "wells":        lambda: arcgis_query(ENDPOINTS["rrc_wells"],     geometry_polygon=buffer_wgs),
            "pipelines":    lambda: arcgis_query(ENDPOINTS["rrc_pipelines"], geometry_polygon=buffer_wgs),
            "wetlands":     lambda: cached_wetlands(buffer_wgs, lat, lon, radius_mi),
            # Static — cached
            "streams":      lambda: cached_layer_query("streams",     buffer_wgs),
            "transmission": lambda: cached_layer_query("transmission", buffer_wgs),
            "etj":          lambda: arcgis_query(ENDPOINTS["etj"],           bbox=buffer_wgs.bounds),
            "electric":     lambda: cached_layer_query("electric",    buffer_wgs,
                                                  out_fields="Utility_Name"),
            # TxDOT planned/active projects — verified working
            "txdot_projects": lambda: cached_layer_query("txdot_projects", buffer_wgs,
                                                  page_size=500, max_pages=4),
            # NOTE: easements, superfund, soils, glo_grants temporarily disabled —
            # their public endpoints either require an auth token (USDA soils via
            # ArcGIS Online), return 404/400 (NCED, EPA Superfund — service moved),
            # or 500 with HTML (TX GLO grants — service moved). Replacing them
            # with working sources is on the backlog. Keeping the layer toggles in
            # the UI so the user knows the slots exist; they'll just return 0
            # features until the new endpoints are wired.
        }

    results = {}
    print(f"[search] {radius_mi}mi @ ({lat:.4f},{lon:.4f}) acres {min_acres}-{max_acres}", flush=True)
    with ThreadPoolExecutor(max_workers=12) as pool:
        futures = {pool.submit(fn): name for name, fn in jobs.items()}
        for fut in as_completed(futures):
            name = futures[fut]
            try:
                results[name] = fut.result()
            except Exception as e:
                results[name] = {"type": "FeatureCollection", "features": [], "error": str(e)}

    # MUD subset (only when ancillary layers were queried)
    if "water_dist" in results:
        muds_features = []
        for f in results["water_dist"]["features"]:
            for v in (f["properties"] or {}).values():
                if v and "MUD" in str(v).upper():
                    muds_features.append(f); break
        results["muds"] = {"type": "FeatureCollection", "features": muds_features}

    # Live CAD overlays — refresh owner names AND acreage with fresher data per-county.
    hcad_count = 0
    mcad_count = 0
    if use_hcad_overlay:
        try:
            hcad_count = hcad_live_overlay(results["tracts"])
        except Exception as e:
            print(f"  HCAD overlay error: {e}", flush=True)
        try:
            mcad_count = mcad_live_overlay(results["tracts"])
        except Exception as e:
            print(f"  MCAD overlay error: {e}", flush=True)

    # FINAL FILTER — now that HCAD/MCAD may have corrected acreage, drop tracts that
    # fall outside the user's actual range. Keep tracts with no acreage info at all
    # (HCAD/MCAD didn't supply it either — better to surface than hide).
    pre_filter = len(results["tracts"]["features"])
    kept = []
    for f in results["tracts"]["features"]:
        acres = f["properties"].get("Acres") or 0
        if acres == 0:
            kept.append(f)   # acreage unknown — keep it visible
            continue
        if min_acres <= acres <= max_acres:
            kept.append(f)
    results["tracts"]["features"] = kept
    if pre_filter != len(kept):
        print(f"  post-overlay re-filter: {pre_filter} -> {len(kept)} tracts", flush=True)

    summary = {k: len(v["features"]) for k, v in results.items()}
    summary["search_center"] = [lat, lon]
    summary["search_radius_mi"] = radius_mi
    summary["acreage_band"] = [min_acres, max_acres]
    summary["elapsed_sec"] = round(time.time() - t0, 1)
    summary["hcad_verified"] = hcad_count
    summary["mcad_verified"] = mcad_count
    summary["data_source_dates"] = {
        "tracts": "TxGIO StratMap (statewide) + HCAD live (Harris, weekly) + MCAD live (Montgomery, monthly)",
        "fema_flood": "FEMA NFHL (current effective)",
        "schools": "TEA 2019-20",
        "wells_pipelines": "TX Railroad Commission live",
    }
    # Surface truncation so the front-end can warn the user to refine
    if results["tracts"].get("_truncated"):
        summary["truncated"] = True
        summary["total_in_buffer"] = results["tracts"].get("_total_in_buffer")
        summary["fetched"] = results["tracts"].get("_fetched")
    print(f"[search] DONE {summary['elapsed_sec']}s  tracts={summary['tracts']}  hcad={hcad_count}  mcad={mcad_count}", flush=True)
    return {"layers": results, "summary": summary}


# --------------------------------------------------------------------------
# Export builders
# --------------------------------------------------------------------------
def build_excel(tracts):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Tracts"
    headers = ["#", "Owner", "Acres", "County", "Site Address", "Mailing", "Flood %",
               "Wells", "Pipelines", "Transmission", "Prop ID", "Centroid Lat", "Centroid Lon",
               "Source Verified"]
    ws.append(headers)
    for i, f in enumerate(tracts.get("features", []), start=1):
        p = f["properties"]
        # centroid (rough — use bounding box midpoint, good enough)
        try:
            coords = f["geometry"]["coordinates"][0]
            if isinstance(coords[0][0], list): coords = coords[0]
            xs = [c[0] for c in coords]; ys = [c[1] for c in coords]
            clat, clon = sum(ys) / len(ys), sum(xs) / len(xs)
        except Exception:
            clat, clon = "", ""
        ws.append([
            i, p.get("OWNER_NAME"), p.get("Acres"),
            p.get("_county"), p.get("SITUS_ADDR"), p.get("MAIL_ADDR"),
            p.get("_flood_pct"), p.get("_wells_n"), p.get("_pipelines_n"),
            p.get("_transmission_n"), p.get("Prop_ID"), clat, clon,
            _tract_source_label(p),
        ])
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 16
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf


def _build_kml_doc(tracts):
    """Return the simplekml.Kml object — shared by KML and KMZ exporters."""
    import simplekml
    kml = simplekml.Kml()
    for i, f in enumerate(tracts.get("features", []), start=1):
        p = f["properties"]
        desc = (
            f"<b>Owner:</b> {p.get('OWNER_NAME','')}<br>"
            f"<b>Acres:</b> {p.get('Acres','')}<br>"
            f"<b>County:</b> {p.get('_county','')}<br>"
            f"<b>Site:</b> {p.get('SITUS_ADDR','')}<br>"
            f"<b>Mailing:</b> {p.get('MAIL_ADDR','')}<br>"
            f"<b>Flood %:</b> {p.get('_flood_pct','')}<br>"
            f"<b>Wells:</b> {p.get('_wells_n','')}<br>"
            f"<b>Pipelines:</b> {p.get('_pipelines_n','')}<br>"
            f"<b>Prop ID:</b> {p.get('Prop_ID','')}<br>"
        )
        nm = f"#{i} {p.get('OWNER_NAME','?')} - {p.get('Acres','?')} ac"
        g = f["geometry"]
        rings = []
        if g["type"] == "Polygon":
            rings = [g["coordinates"]]
        elif g["type"] == "MultiPolygon":
            rings = g["coordinates"]
        for poly in rings:
            outer = poly[0]
            pol = kml.newpolygon(name=nm, description=desc)
            pol.outerboundaryis = [(c[0], c[1]) for c in outer]
            for inner in poly[1:]:
                pol.innerboundaryis = [(c[0], c[1]) for c in inner]
            pol.style.linestyle.width = 2
            pol.style.linestyle.color = simplekml.Color.red
            pol.style.polystyle.color = simplekml.Color.changealphaint(80, simplekml.Color.yellow)
    return kml


def build_kml(tracts):
    buf = io.BytesIO()
    buf.write(_build_kml_doc(tracts).kml().encode("utf-8"))
    buf.seek(0)
    return buf


def build_kmz(tracts):
    """KMZ = ZIP-compressed KML. Google Earth opens KMZ natively; the file is
    self-contained (one .kml inside) so it's the preferred upload format."""
    import zipfile
    kml_str = _build_kml_doc(tracts).kml()
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("doc.kml", kml_str)
    buf.seek(0)
    return buf


def _walk_coords(geojson_geom):
    """Yield every [lon, lat] in any GeoJSON Polygon/MultiPolygon/LineString geometry."""
    if not geojson_geom:
        return
    t = geojson_geom.get("type")
    coords = geojson_geom.get("coordinates")
    if not coords:
        return
    if t == "Point":
        yield coords
    elif t in ("LineString", "MultiPoint"):
        yield from coords
    elif t in ("Polygon", "MultiLineString"):
        for ring in coords:
            yield from ring
    elif t == "MultiPolygon":
        for poly in coords:
            for ring in poly:
                yield from ring


def _build_search_map_png(tracts, search_meta, width=1180, height=720):
    """Render a satellite map of the search area with tract polygons overlaid.
    Uses Esri's World_Imagery free static-export endpoint (no key required).
    Returns PNG bytes, or None if rendering fails (PDF still builds without map)."""
    try:
        from PIL import Image, ImageDraw
        from pyproj import Transformer

        features = (tracts or {}).get("features") or []
        # Compute the bbox from tract geometries first; if empty, fall back to the
        # imported polygon or center+radius.
        minx, miny, maxx, maxy = float("inf"), float("inf"), float("-inf"), float("-inf")
        for f in features:
            for lon, lat in _walk_coords(f.get("geometry")):
                if lon < minx: minx = lon
                if lon > maxx: maxx = lon
                if lat < miny: miny = lat
                if lat > maxy: maxy = lat
        polygon_meta = (search_meta or {}).get("polygon")
        if minx == float("inf") and polygon_meta:
            for lon, lat in _walk_coords(polygon_meta):
                if lon < minx: minx = lon
                if lon > maxx: maxx = lon
                if lat < miny: miny = lat
                if lat > maxy: maxy = lat
        if minx == float("inf") and search_meta and search_meta.get("center"):
            lat, lon = search_meta["center"]
            r_mi = search_meta.get("radius_mi") or 10
            dlat = r_mi / 69.0
            dlon = r_mi / (69.0 * max(0.1, abs(__import__("math").cos(__import__("math").radians(lat)))))
            minx, maxx, miny, maxy = lon - dlon, lon + dlon, lat - dlat, lat + dlat
        if minx == float("inf"):
            return None

        # Padding (10%) so polygons don't crowd the frame
        dx = (maxx - minx) * 0.08 or 0.005
        dy = (maxy - miny) * 0.08 or 0.005
        minx -= dx; maxx += dx; miny -= dy; maxy += dy

        # Preserve image aspect ratio so the satellite tile isn't squished
        bbox_aspect = (maxx - minx) / max(0.0001, (maxy - miny))
        img_aspect = width / height
        if bbox_aspect > img_aspect:
            need_dy = ((maxx - minx) / img_aspect - (maxy - miny)) / 2
            miny -= need_dy; maxy += need_dy
        else:
            need_dx = ((maxy - miny) * img_aspect - (maxx - minx)) / 2
            minx -= need_dx; maxx += need_dx

        # Esri free static map export — World_Imagery is the satellite basemap
        export_url = "https://services.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/export"
        r = requests.get(export_url, params={
            "bbox": f"{minx},{miny},{maxx},{maxy}",
            "bboxSR": "4326", "imageSR": "3857",
            "size": f"{width},{height}",
            "format": "png", "f": "image",
        }, timeout=30)
        r.raise_for_status()
        img = Image.open(io.BytesIO(r.content)).convert("RGBA")

        # Convert lon/lat -> Web Mercator -> pixel coords for accurate overlay
        to_merc = Transformer.from_crs(4326, 3857, always_xy=True).transform
        merc_minx, merc_miny = to_merc(minx, miny)
        merc_maxx, merc_maxy = to_merc(maxx, maxy)
        merc_w = merc_maxx - merc_minx
        merc_h = merc_maxy - merc_miny
        def to_px(lon, lat):
            mx, my = to_merc(lon, lat)
            px = (mx - merc_minx) / merc_w * width
            py = height - (my - merc_miny) / merc_h * height
            return px, py

        overlay = Image.new("RGBA", img.size, (0, 0, 0, 0))
        draw = ImageDraw.Draw(overlay, "RGBA")

        # Optional: outline the imported polygon (if any) in Ember orange
        if polygon_meta:
            for ring in (polygon_meta.get("coordinates") or []):
                if polygon_meta.get("type") == "MultiPolygon":
                    for inner_ring in ring:
                        pts = [to_px(c[0], c[1]) for c in inner_ring]
                        if len(pts) >= 3:
                            draw.line(pts + [pts[0]], fill=(242, 89, 41, 220), width=4)
                else:
                    pts = [to_px(c[0], c[1]) for c in ring]
                    if len(pts) >= 3:
                        draw.line(pts + [pts[0]], fill=(242, 89, 41, 220), width=4)

        # Load a font for tract-number labels. Smaller (10 px) so labels don't dominate.
        # If the search returned 40+ tracts, labels would visually swamp the map at PDF
        # resolution — skip labels entirely on dense searches and let the table on page 2
        # serve as the index. Otherwise small white pills sit on each polygon.
        from PIL import ImageFont
        TRACT_COUNT = len(features)
        SHOW_LABELS = TRACT_COUNT <= 40
        label_font = None
        if SHOW_LABELS:
            for font_path in (
                "C:/Windows/Fonts/arialbd.ttf",
                "C:/Windows/Fonts/arial.ttf",
                "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
                "/System/Library/Fonts/Helvetica.ttc",
            ):
                try:
                    label_font = ImageFont.truetype(font_path, 10)
                    break
                except (OSError, IOError):
                    continue
            if label_font is None:
                label_font = ImageFont.load_default()

        # Draw each tract polygon, colored by acreage (matches the on-screen legend).
        # GeoJSON Polygon coords are [outer_ring, hole1, hole2, ...]; MultiPolygon
        # is [polygon1_rings, polygon2_rings, ...]. Flatten to one list-of-rings.
        # Track centroid pixel-coords so we can layer numbers ON TOP of all polygons
        # (otherwise later polygons would draw over earlier numbers).
        label_positions = []
        for idx, f in enumerate(features, start=1):
            g = f.get("geometry")
            if not g: continue
            if g["type"] == "Polygon":
                rings = g["coordinates"]
            elif g["type"] == "MultiPolygon":
                rings = [r for poly in g["coordinates"] for r in poly]
            else:
                continue
            acres = (f.get("properties") or {}).get("Acres") or 0
            if   acres < 500:  fill = (230, 57, 70, 140)
            elif acres < 1000: fill = (244, 162, 97, 140)
            else:              fill = (233, 196, 106, 140)
            outer_px = None
            for ring_i, ring in enumerate(rings):
                pts = [to_px(c[0], c[1]) for c in ring]
                if len(pts) >= 3:
                    draw.polygon(pts, fill=fill, outline=(123, 18, 22, 255))
                    if outer_px is None:
                        outer_px = pts
            # Compute polygon centroid from outer ring (simple average is fine for labels)
            if outer_px and len(outer_px) >= 3:
                cx = sum(p[0] for p in outer_px) / len(outer_px)
                cy = sum(p[1] for p in outer_px) / len(outer_px)
                label_positions.append((idx, cx, cy))

        # Render tract numbers (only when there are ≤40 — dense searches skip labels
        # to keep the map readable; the table on page 2 is the index in that case).
        if SHOW_LABELS:
            for idx, cx, cy in label_positions:
                label = str(idx)
                try:
                    bbox = draw.textbbox((0, 0), label, font=label_font)
                    tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
                except AttributeError:
                    tw, th = draw.textsize(label, font=label_font)
                pad = 2   # tighter pill
                # White pill with dark border — small enough to not dominate
                draw.rounded_rectangle(
                    [cx - tw/2 - pad, cy - th/2 - pad, cx + tw/2 + pad, cy + th/2 + pad],
                    radius=3, fill=(255, 255, 255, 230), outline=(19, 52, 78, 220), width=1,
                )
                draw.text((cx - tw/2, cy - th/2 - 1), label,
                          fill=(19, 52, 78, 255), font=label_font)

        composite = Image.alpha_composite(img, overlay).convert("RGB")
        out = io.BytesIO()
        composite.save(out, format="PNG", optimize=True)
        out.seek(0)
        return out.getvalue()
    except Exception as e:
        import traceback
        print(f"[pdf] map render failed: {e}", flush=True)
        traceback.print_exc()
        return None


def _tract_source_label(props):
    """Human-readable source label, matching the live overlay state."""
    if props.get("_hcad_owner_verified"):
        return "HCAD live"
    if props.get("_mcad_owner_verified"):
        return "MCAD live"
    return "StratMap"


def build_pdf(tracts, search_meta):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter), leftMargin=0.4*inch,
                            rightMargin=0.4*inch, topMargin=0.4*inch, bottomMargin=0.4*inch)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading1"], fontName="Helvetica-Bold",
                                  fontSize=18, textColor=colors.HexColor("#13344E"), spaceAfter=4)
    sub_style = ParagraphStyle("sub", parent=styles["Normal"], fontName="Helvetica",
                                fontSize=10, textColor=colors.HexColor("#58595B"), spaceAfter=10)
    accent_style = ParagraphStyle("accent", parent=styles["Normal"], fontName="Helvetica-Bold",
                                   fontSize=10, textColor=colors.HexColor("#F25929"))

    # --- Title (uses the search label + actual location, no longer generic) ---
    raw_label = (search_meta.get("label") or "Search").strip() or "Search"
    title_text = f"Acquisitions GIS — {raw_label}"

    # --- Location subtitle (handles polygon-search vs radius-search) ---
    center = search_meta.get("center") or [0, 0]
    radius_mi = search_meta.get("radius_mi") or 0
    has_polygon = bool(search_meta.get("polygon"))
    if has_polygon:
        loc_str = "Inside imported boundary"
        if center and (center[0] or center[1]):
            loc_str += f" (centroid {center[0]:.4f}, {center[1]:.4f})"
    elif center and (center[0] or center[1]):
        loc_str = f"Center {center[0]:.4f}, {center[1]:.4f} &nbsp;·&nbsp; Radius {radius_mi} mi"
    else:
        loc_str = "—"

    min_a = search_meta.get("min_acres", 0)
    max_a = search_meta.get("max_acres", 0)
    meta_line = (f"{loc_str} &nbsp;|&nbsp; "
                 f"Acres {min_a:,.0f}–{max_a:,.0f} &nbsp;|&nbsp; "
                 f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}")

    n_tracts = len(tracts.get("features", []))
    # Source-mix counts so the user sees data freshness at a glance
    hcad_n = sum(1 for f in tracts.get("features", []) if (f.get("properties") or {}).get("_hcad_owner_verified"))
    mcad_n = sum(1 for f in tracts.get("features", []) if (f.get("properties") or {}).get("_mcad_owner_verified"))
    stratmap_n = n_tracts - hcad_n - mcad_n
    src_line = (f"<b>{n_tracts}</b> tracts &nbsp;·&nbsp; "
                f"{hcad_n} HCAD-live &nbsp;·&nbsp; {mcad_n} MCAD-live &nbsp;·&nbsp; "
                f"{stratmap_n} StratMap")

    story = []
    story.append(Paragraph(title_text, title_style))
    story.append(Paragraph(meta_line, sub_style))
    story.append(Paragraph(src_line, accent_style))
    story.append(Spacer(1, 8))

    # --- Map image — satellite basemap + tract overlay ---
    map_png = _build_search_map_png(tracts, search_meta)
    if map_png:
        from reportlab.platypus import Image as RLImage
        # Fit on landscape letter (10 in wide x ~6 in tall available)
        img = RLImage(io.BytesIO(map_png), width=10*inch, height=6.1*inch)
        img.hAlign = "CENTER"
        story.append(img)
        story.append(Spacer(1, 8))
        story.append(PageBreak())   # table starts on a fresh page

    # --- Tract table (page 2+) ---
    rows = [["#", "Owner", "Acres", "County", "Flood %", "Wells", "Pipes", "Tx Lines", "Prop ID", "Source"]]
    for i, f in enumerate(tracts.get("features", []), start=1):
        p = f["properties"] or {}
        rows.append([
            str(i),
            (p.get("OWNER_NAME") or "")[:38],
            f"{p.get('Acres','')}",
            (p.get("_county") or "").replace(" County", ""),
            f"{p.get('_flood_pct','')}",
            str(p.get("_wells_n", "")),
            str(p.get("_pipelines_n", "")),
            str(p.get("_transmission_n", "")),
            (p.get("Prop_ID") or "")[:16],
            _tract_source_label(p),
        ])

    t = Table(rows, repeatRows=1, colWidths=[0.4*inch, 2.8*inch, 0.6*inch, 0.9*inch,
                                              0.6*inch, 0.5*inch, 0.5*inch, 0.6*inch,
                                              1.3*inch, 1.1*inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#13344E")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, 0), 9),
        ("ALIGN",      (0, 0), (-1, 0), "LEFT"),
        ("FONTSIZE",   (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F2F3")]),
        ("GRID",       (0, 0), (-1, -1), 0.25, colors.HexColor("#D1D3D4")),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING",(0, 0), (-1, -1), 5),
        ("RIGHTPADDING",(0, 0), (-1, -1), 5),
    ]))
    story.append(t)

    # ----- Definitions & data-source footer ---------------------------------
    story.append(Spacer(1, 14))
    section_style = ParagraphStyle("section", parent=styles["Normal"], fontName="Helvetica-Bold",
                                    fontSize=10, textColor=colors.HexColor("#F25929"),
                                    spaceAfter=6, spaceBefore=2)
    def_style = ParagraphStyle("def", parent=styles["Normal"], fontName="Helvetica",
                                fontSize=8.5, textColor=colors.HexColor("#13344E"),
                                leading=11)
    story.append(Paragraph("COLUMN DEFINITIONS", section_style))

    defs = [
        ("Owner",      "Recorded owner of the parcel. Always live from HCAD when available (Harris) "
                       "or MCAD (Montgomery); annual StratMap snapshot for other counties."),
        ("Acres",      "Parcel area in acres. For Harris (HCAD) and Montgomery (MCAD) parcels this "
                       "is the appraisal district's official figure; for other counties it's "
                       "computed from StratMap geometry."),
        ("County",     "County the parcel falls in (per Census TIGER boundaries)."),
        ("Flood %",    "Percentage of the parcel inside FEMA's 100-year floodplain (Special Flood "
                       "Hazard Area, Zone A/AE). 0 means none mapped; values close to 100 indicate "
                       "the parcel is almost entirely in the floodplain."),
        ("Wells",      "Count of oil/gas well bores located within the parcel boundary "
                       "(TX Railroad Commission live database)."),
        ("Pipes",      "Count of pipelines crossing the parcel (TX Railroad Commission, "
                       "any commodity: gas, crude, products)."),
        ("Tx Lines",   "Count of high-voltage electric transmission lines crossing the parcel "
                       "(HIFLD national dataset). Lower-voltage distribution lines are NOT included."),
        ("Prop ID",    "The county appraisal district account number (HCAD 13-digit format in Harris). "
                       "Use this number directly on hcad.org for the full account page."),
        ("Source",     "Where the owner name + acreage came from. \"HCAD live\" = refreshed from "
                       "HCAD's weekly feed at search time. \"MCAD live\" = same, Montgomery's monthly "
                       "feed. \"StratMap\" = TxGIO statewide annual parcel layer (other counties)."),
    ]
    for label, text in defs:
        story.append(Paragraph(f"<b>{label}</b> &nbsp; {text}", def_style))

    story.append(Spacer(1, 10))
    story.append(Paragraph("DATA SOURCES &amp; REFRESH", section_style))
    src_text = (
        "<b>Parcels &amp; ownership.</b> Harris County: HCAD's live GIS feed at "
        "<i>gis.hctx.net/arcgis/rest/services/HCAD</i>, refreshed weekly by HCAD; this report "
        "pulled it at search time. Montgomery County: MCAD's live FeatureServer, refreshed monthly. "
        "Other Texas counties: TxGIO StratMap statewide parcel layer (most recent annual snapshot). "
        "Note that HCAD/MCAD themselves lag actual deed recordings by ~2–8 weeks plus their annual "
        "appraisal-roll cycle — for the freshest possible ownership, check the County Clerk's deed "
        "records directly (link in each tract's popup in-app).<br/>"
        "<b>Floodplain.</b> FEMA National Flood Hazard Layer (NFHL), current effective panels.<br/>"
        "<b>Wells &amp; pipelines.</b> Texas Railroad Commission live data (gis.hctx.net/arcgishcpid).<br/>"
        "<b>Transmission lines.</b> HIFLD (Homeland Infrastructure Foundation-Level Data).<br/>"
        "<b>School districts.</b> Texas Education Agency (TEA) district boundaries.<br/>"
        "<b>Water / MUDs.</b> TCEQ water-district boundaries + Texas Comptroller's annual "
        f"special-district tax-rate report (TY{TAX_YEAR}, " + str(_special_district_rates_cache and len(_special_district_rates_cache) or "—") +
        " rates loaded).<br/>"
        "<b>This report.</b> Generated " + datetime.now().strftime("%Y-%m-%d %H:%M") +
        " from Ember Acquisitions GIS. Live ownership / acreage as of this timestamp."
    )
    story.append(Paragraph(src_text, def_style))

    doc.build(story)
    buf.seek(0)
    return buf


def annotate_tracts_with_enrichment(layers):
    """Attach county / school / water / flood% / wetlands% / wells / pipelines counts onto
    each tract's properties. Uses shapely STRtree spatial indexes so the math scales to
    tens of thousands of tracts without blowing up to O(n×m)."""
    from shapely.geometry import shape as shp_shape
    try:
        from shapely.strtree import STRtree
    except ImportError:
        STRtree = None
    tracts = layers["tracts"]["features"]
    counties = [shp_shape(f["geometry"]) if f.get("geometry") else None for f in layers.get("counties", {}).get("features", [])]
    county_names = [(f["properties"] or {}).get("BASENAME") or (f["properties"] or {}).get("NAME") for f in layers.get("counties", {}).get("features", [])]
    etjs = [shp_shape(f["geometry"]) if f.get("geometry") else None for f in layers.get("etj", {}).get("features", [])]
    etj_names = []
    etj_statuses = []
    for f in layers.get("etj", {}).get("features", []):
        p = f.get("properties") or {}
        etj_names.append(p.get("CityName") or p.get("NAME") or p.get("CITY") or "")
        # ETJ_Status is "City" / "ETJ" — we use this to decide whether city property tax applies
        etj_statuses.append((p.get("ETJ_Status") or p.get("Status") or "").upper())
    flood_shapes = [shp_shape(f["geometry"]) for f in layers.get("flood", {}).get("features", []) if f.get("geometry")]
    wetland_shapes = [shp_shape(f["geometry"]) for f in layers.get("wetlands", {}).get("features", []) if f.get("geometry")]
    well_points = [shp_shape(f["geometry"]) for f in layers.get("wells", {}).get("features", []) if f.get("geometry")]
    pipeline_lines = [shp_shape(f["geometry"]) for f in layers.get("pipelines", {}).get("features", []) if f.get("geometry")]
    transmission_lines = [shp_shape(f["geometry"]) for f in layers.get("transmission", {}).get("features", []) if f.get("geometry")]

    # Build spatial indexes so per-tract math is O(log n) instead of O(n) per layer.
    # Without these, 10k tracts × 439 flood polygons = 4.4M ops; with these, ~50k ops.
    def _build_tree(geoms):
        if STRtree is None or not geoms:
            return None
        try:
            return STRtree(geoms)
        except Exception:
            return None

    flood_tree = _build_tree(flood_shapes)
    wetland_tree = _build_tree(wetland_shapes)
    wells_tree = _build_tree(well_points)
    pipelines_tree = _build_tree(pipeline_lines)
    transmission_tree = _build_tree(transmission_lines)

    def _candidates(tree, all_geoms, g):
        """Return geometries from all_geoms whose bbox potentially intersects g.
        Handles both shapely 2.x (returns ndarray of int indices) and 1.x (returns geometries)."""
        if tree is None or not all_geoms:
            return all_geoms
        try:
            result = tree.query(g)
            out = []
            for item in result:
                if hasattr(item, "intersection"):
                    out.append(item)             # 1.x — already a geometry
                else:
                    out.append(all_geoms[int(item)])   # 2.x — integer index (may be numpy int64)
            return out
        except Exception:
            return all_geoms

    # Water provider polygons (PUC CCN) — for each tract, find the utility serving it
    ccn_shapes = []
    for f in layers.get("ccn", {}).get("features", []):
        if not f.get("geometry"):
            continue
        p = f.get("properties") or {}
        ccn_shapes.append((shp_shape(f["geometry"]),
                           p.get("UTILITY"), p.get("CCN_NO"), p.get("CCN_TYPE")))

    # Water district polygons (TCEQ) — for each tract, find the MUD/WCID/etc.
    water_dist_shapes = []
    for f in layers.get("water_dist", {}).get("features", []):
        if not f.get("geometry"):
            continue
        p = f.get("properties") or {}
        water_dist_shapes.append((shp_shape(f["geometry"]),
                                  p.get("NAME"), p.get("TYPE"), p.get("TYPE_DESCRIPTION")))

    # Electric TDU polygons (TEPRI / PUC)
    electric_shapes = []
    for f in layers.get("electric", {}).get("features", []):
        if not f.get("geometry"):
            continue
        p = f.get("properties") or {}
        electric_shapes.append((shp_shape(f["geometry"]), p.get("Utility_Name")))

    # School district polygons (TEA) — for per-tract ISD assignment.
    # TEA layer's name field can be NAME, DISTNAME, or DISTRICT_N depending on the schema version.
    school_shapes = []
    for f in layers.get("schools", {}).get("features", []):
        if not f.get("geometry"):
            continue
        p = f.get("properties") or {}
        name = p.get("NAME") or p.get("DISTNAME") or p.get("DISTRICT_N") or ""
        if name:
            school_shapes.append((shp_shape(f["geometry"]), name))

    def overlap_pct(g, g_utm, shapes):
        if not shapes or g_utm.area <= 0:
            return 0
        # Pre-filter to bbox candidates via the parent layer's spatial index when possible
        inter_area = 0
        for s in shapes:
            try:
                ig = s.intersection(g)
                if not ig.is_empty:
                    inter_area += transform(to_utm, ig).area
            except Exception:
                pass
        return round(100 * inter_area / g_utm.area, 1)

    for f in tracts:
        if not f.get("geometry"):
            continue
        g = shp_shape(f["geometry"])
        g_utm = transform(to_utm, g)
        props = f["properties"]
        # county — prefer the result of the spatial join, but PRESERVE any
        # existing _county value (e.g. set by the parcel-cache lookup) when
        # the join finds nothing. Without this guard, a tracts_only search
        # (no counties layer fetched) would clobber every tract's county
        # back to None and the UI would show "Unknown" everywhere.
        county = None
        for i, c in enumerate(counties):
            if c and c.contains(g.centroid):
                county = county_names[i]; break
        if county or not props.get("_county"):
            props["_county"] = county
        # ETJ / city — also capture whether the tract is INSIDE city limits or just in ETJ.
        # H-GAC's etj layer publishes separate polygons for each ("City" vs "ETJ"); when both
        # contain the centroid, the in-city polygon wins because city tax only applies in-city.
        city_etj = None
        in_city = False
        for i, c in enumerate(etjs):
            if c and c.contains(g.centroid):
                if etj_statuses[i] == "CITY":
                    city_etj = etj_names[i] or city_etj
                    in_city = True
                    break    # in-city wins; stop looking
                elif not city_etj:
                    city_etj = etj_names[i] or city_etj
        if not city_etj:
            for i, c in enumerate(etjs):
                if c and c.intersects(g):
                    city_etj = etj_names[i] or city_etj
                    if etj_statuses[i] == "CITY":
                        in_city = True
                    break
        props["_city_etj"] = city_etj
        props["_in_city"] = in_city
        # Overlaps + counts — use spatial indexes so this is fast even for tens of thousands of tracts
        props["_flood_pct"] = overlap_pct(g, g_utm, _candidates(flood_tree, flood_shapes, g))
        props["_wetlands_pct"] = overlap_pct(g, g_utm, _candidates(wetland_tree, wetland_shapes, g))
        well_cands = _candidates(wells_tree, well_points, g)
        props["_wells_n"] = sum(1 for w in well_cands if w.within(g))
        pipe_cands = _candidates(pipelines_tree, pipeline_lines, g)
        props["_pipelines_n"] = sum(1 for pl in pipe_cands if pl.intersects(g))
        tx_cands = _candidates(transmission_tree, transmission_lines, g)
        props["_transmission_n"] = sum(1 for tl in tx_cands if tl.intersects(g))

        # Water provider — pick the CCN polygon that contains the tract centroid (or first intersect)
        props["_water_provider"] = None
        props["_water_ccn"] = None
        for shp, util, ccn_no, ccn_type in ccn_shapes:
            if shp.contains(g.centroid) or shp.intersects(g):
                props["_water_provider"] = util
                props["_water_ccn"] = ccn_no
                break

        # Electric TDU — same logic
        props["_electric_provider"] = None
        for shp, util in electric_shapes:
            if shp.contains(g.centroid) or shp.intersects(g):
                props["_electric_provider"] = util
                break

        # Water district / MUD — pick the polygon containing the tract centroid
        props["_water_dist_name"] = None
        props["_water_dist_type"] = None
        props["_in_mud"] = False
        for shp, name, type_code, type_desc in water_dist_shapes:
            if shp.contains(g.centroid) or shp.intersects(g):
                props["_water_dist_name"] = name
                props["_water_dist_type"] = type_desc or type_code
                if (type_code or "").upper() == "MUD":
                    props["_in_mud"] = True
                break

        # School district (ISD) — point-in-polygon on tract centroid against TEA layer
        props["_school_dist"] = None
        for shp, name in school_shapes:
            if shp.contains(g.centroid) or shp.intersects(g):
                props["_school_dist"] = name
                break


# --------------------------------------------------------------------------
# Geocoding and single-parcel lookup
#
# These lived inline in the standalone app's route handlers. They are pure
# functions of their arguments, so they belong here rather than in app.py.
# --------------------------------------------------------------------------

_GEO_UA = {"User-Agent": "EmberApps-Acquisitions/1.0"}


def _photon_format(features):
    """Photon GeoJSON features -> {label, lat, lon} suggestions."""
    out = []
    for f in features or []:
        coords = (f.get("geometry") or {}).get("coordinates") or []
        if len(coords) < 2:
            continue
        p = f.get("properties") or {}
        prefix = " ".join(filter(None, [p.get("housenumber"), p.get("street")])).strip()
        primary = prefix or p.get("name") or "?"
        context = ", ".join(filter(None, [p.get("city"), p.get("county"),
                                          p.get("state"), p.get("country")]))
        out.append({"label": primary + (f" — {context}" if context else ""),
                    "lat": coords[1], "lon": coords[0]})
    return out


def geocode_suggest(q, lat=None, lon=None):
    """Autocomplete as the user types. Photon: OSM-based, free, no key."""
    q = (q or "").strip()
    if len(q) < 2:
        return []
    params = {"q": q, "limit": 6}
    if lat is not None and lon is not None:
        params["lat"], params["lon"] = lat, lon
    r = requests.get("https://photon.komoot.io/api/", params=params,
                     headers=_GEO_UA, timeout=10)
    return _photon_format((r.json() or {}).get("features", []))


def geocode(q, lat_bias=None, lon_bias=None):
    """Resolve one query to a point, or None.

    Three providers in descending order of how well they handle the thing the
    user actually typed: Photon for place names, Census for precise US street
    addresses, Nominatim as a backstop. Each is allowed to fail quietly - a
    geocoder being down should demote to the next one, not fail the search.
    """
    q = (q or "").strip()
    if not q:
        return None

    try:
        params = {"q": q, "limit": 1}
        if lat_bias is not None and lon_bias is not None:
            params["lat"], params["lon"] = lat_bias, lon_bias
        r = requests.get("https://photon.komoot.io/api/", params=params,
                         headers=_GEO_UA, timeout=10)
        feats = (r.json() or {}).get("features", [])
        if feats:
            s = _photon_format(feats)[0]
            return {"lat": s["lat"], "lon": s["lon"],
                    "address": s["label"], "source": "photon"}
    except Exception as e:
        print(f"[acq-geocode] photon failed: {e}", flush=True)

    try:
        r = requests.get(
            "https://geocoding.geo.census.gov/geocoder/locations/onelineaddress",
            params={"address": q, "benchmark": "Public_AR_Current", "format": "json"},
            timeout=8)
        matches = ((r.json() or {}).get("result") or {}).get("addressMatches") or []
        if matches:
            m = matches[0]; c = m["coordinates"]
            return {"lat": c["y"], "lon": c["x"],
                    "address": m["matchedAddress"], "source": "census"}
    except Exception as e:
        print(f"[acq-geocode] census failed: {e}", flush=True)

    try:
        r = requests.get("https://nominatim.openstreetmap.org/search",
                         params={"q": q, "format": "json", "limit": 1},
                         headers=_GEO_UA, timeout=8)
        results = r.json() or []
        if results:
            m = results[0]
            return {"lat": float(m["lat"]), "lon": float(m["lon"]),
                    "address": m.get("display_name", q), "source": "osm"}
    except Exception as e:
        print(f"[acq-geocode] nominatim failed: {e}", flush=True)

    return None


def parcel_detail(prop_id):
    """Every field on one parcel, refreshed live.

    Local cache first - it answers instantly by Prop_ID - then an HCAD/MCAD
    overlay so the owner and appraisal values are current rather than as of the
    last statewide StratMap refresh.
    """
    import acq_parcels as _pc

    pid = str(prop_id or "").strip()
    if not pid:
        return {"error": "prop_id required"}

    feat = _pc.find_parcel_by_pid(pid)
    if not feat:
        return {"error": f"No parcel found for Prop_ID {pid}"}

    fc = {"type": "FeatureCollection", "features": [feat]}
    try:
        hcad_live_overlay(fc)
    except Exception as e:
        print(f"[acq-detail] hcad overlay failed for {pid}: {e}", flush=True)
    try:
        mcad_live_overlay(fc)
    except Exception as e:
        print(f"[acq-detail] mcad overlay failed for {pid}: {e}", flush=True)

    props = dict(feat.get("properties") or {})
    props["_source"] = _tract_source_label(props)
    return {"prop_id": pid, "properties": props, "geometry": feat.get("geometry")}


# --------------------------------------------------------------------------
# The constraint / yield engine
#
# Lifted out of the standalone app's /api/projects/<pid>/analyze handler, which
# had the whole thing inline. It is a pure function of the project dict, so it
# belongs here rather than in a route.
#
# The order matters and is not arbitrary:
#   gross -> physical constraints (unioned, so overlaps are not double-counted)
#         -> net developable
#         -> infrastructure and landscaping
#         -> net saleable, which is what yield runs off
#
# Each constraint can be switched off through proj["netout_assumptions"];
# wetlands in particular are often mitigable or already permitted, so a blanket
# deduction understates a site. Defaults stay conservative - everything
# deducted, infrastructure at 30%.
#
# A layer that fails to load is reported as unavailable rather than as zero
# acres. Those two states look identical in a total and mean opposite things:
# one of them silently overstates how much of the site you can sell.
# --------------------------------------------------------------------------

def run_analysis(proj):
    """Analyse one project dict. Returns the analysis, or raises ValueError
    when the project carries no usable tract geometry."""
    from shapely.geometry import shape as shp_shape, mapping as shp_mapping
    from shapely.ops import unary_union, transform

    tracts = proj.get("tracts") or []
    geoms = []
    for t in tracts:
        g = t.get("geometry")
        if not g:
            continue
        try:
            geoms.append(shp_shape(g))
        except Exception:
            continue
    if not geoms:
        raise ValueError("no tract geometries - re-create the project with tract polygons")

    project_union = unary_union(geoms)
    project_union_utm = transform(to_utm, project_union)
    gross_acres_m2 = project_union_utm.area
    gross_acres = gross_acres_m2 / 4046.8564224

    # Build a SIMPLIFIED query polygon for the Esri spatial filter — complex
    # parcel geometry (thousands of vertices) gets rejected by some Esri layers
    # (FEMA NFHL especially). Buffer it slightly so the simplification doesn't
    # exclude any real floodplain coverage at the edges.
    try:
        # tolerance in degrees — 0.0002° ≈ 22m, good detail for ancillary layers.
        simplified_query_poly = project_union.simplify(0.0002, preserve_topology=True)
        # Small outward buffer so simplification + tiny gaps don't miss edge features.
        # This buffer is ONLY for the QUERY; we still intersect with the real polygon later.
        bbox_query_poly = project_union.envelope.buffer(0.005)   # ~500m envelope buffer
    except Exception:
        simplified_query_poly = project_union
        bbox_query_poly = project_union.envelope

    # Pull constraint layers from cached/live sources, clipped to project polygon.
    #
    # Every layer variable is bound to an empty FeatureCollection FIRST. A failed
    # fetch used to leave the name unbound entirely, and the net-out loop below
    # then swallowed a NameError - so the constraint silently did not apply and
    # net developable came back overstated with no clear reason why.
    constraints = {}
    flood_fc = {"type": "FeatureCollection", "features": []}
    wet_fc = {"type": "FeatureCollection", "features": []}
    flood_union = None
    wet_union = None
    # The three line layers need the same treatment. Each is assigned inside its
    # own try, and the net-out loop below reads all three from a list literal
    # that is evaluated BEFORE the loop body's except can catch anything - so a
    # single failed layer used to take the whole analysis down with an
    # UnboundLocalError 500 rather than degrading to "layer unavailable".
    trans_lines = []
    stream_lines = []
    pipe_lines = []

    def _safe_union(geoms):
        """Union layer polygons, repairing the invalid ones first.

        NWI and NFHL both ship self-intersecting rings. unary_union raises
        TopologyException ("side location conflict") on those, and the callers
        below used to swallow it - so the constraint silently did not apply and
        the acreage came back overstated. make_valid repairs the ring; buffer(0)
        is the fallback, and anything still broken is dropped rather than taking
        the whole layer down with it."""
        from shapely.validation import make_valid
        from shapely.geometry import Polygon, MultiPolygon, GeometryCollection
        cleaned = []
        for g in geoms or []:
            if g is None or g.is_empty:
                continue
            if not g.is_valid:
                try:
                    g = make_valid(g)
                except Exception:
                    try:
                        g = g.buffer(0)
                    except Exception:
                        continue
            if g is None or g.is_empty:
                continue
            # make_valid can hand back a collection with stray lines in it;
            # only the polygonal parts have area worth deducting.
            if isinstance(g, GeometryCollection):
                parts = [p for p in g.geoms if isinstance(p, (Polygon, MultiPolygon))]
                if not parts:
                    continue
                g = parts[0] if len(parts) == 1 else MultiPolygon(
                    [q for p in parts for q in (p.geoms if isinstance(p, MultiPolygon) else [p])])
            cleaned.append(g)
        if not cleaned:
            return None
        try:
            return unary_union(cleaned)
        except Exception:
            acc = None                     # union what we can, drop what we cannot
            for c in cleaned:
                try:
                    acc = c if acc is None else acc.union(c)
                except Exception:
                    continue
            return acc

    def _fetch_flood(attempts=3):
        """FEMA NFHL answers in under a second but drops roughly one connection
        in four under load. It is worth retrying rather than reporting the
        floodplain as unavailable and overstating the developable acreage."""
        last = None
        for i in range(attempts):
            if i:
                time.sleep(1.2 * i)
            try:
                return cached_layer_query(
                    "fema_flood", bbox_query_poly,
                    where="FLD_ZONE IN ('A','AE','AH','AO','AR','A99','V','VE')",
                    out_fields="FLD_ZONE"), None
            except Exception as e:
                last = e
                print(f"[analyze] floodplain attempt {i + 1}/{attempts} failed: "
                      f"{type(e).__name__}: {str(e)[:90]}", flush=True)
        return None, last

    try:
        # FEMA NFHL is sensitive to polygon complexity — use the simplified
        # bbox-envelope polygon for the query, then intersect with the real
        # project polygon client-side.
        _ff, _ferr = _fetch_flood()
        if _ff is None:
            raise RuntimeError(f"FEMA NFHL unavailable after 3 attempts: {_ferr}")
        flood_fc = _ff
        flood_geoms = [shp_shape(f["geometry"]) for f in flood_fc.get("features", [])
                        if f.get("geometry")]
        print(f"[analyze] floodplain features returned: {len(flood_geoms)}", flush=True)
        flood_union = _safe_union(flood_geoms)
        if flood_union:
            flood_in_project = transform(to_utm, flood_union.intersection(project_union))
            constraints["floodplain"] = {
                "acres":   round(flood_in_project.area / 4046.8564224, 2),
                "pct":     round(flood_in_project.area / gross_acres_m2 * 100, 1) if gross_acres_m2 > 0 else 0,
                "feature_count": len(flood_geoms),
            }
        else:
            constraints["floodplain"] = {"acres": 0, "pct": 0, "feature_count": 0}
    except Exception as e:
        print(f"[analyze] floodplain ERROR: {e}", flush=True)
        constraints["floodplain"] = {"error": str(e)}

    try:
        _wf = cached_wetlands(bbox_query_poly, project_union.centroid.y,
                              project_union.centroid.x, 1.0)
        if _wf and _wf.get("features") is not None:
            wet_fc = _wf
        wet_geoms = [shp_shape(f["geometry"]) for f in wet_fc.get("features", [])
                      if f.get("geometry")]
        print(f"[analyze] wetlands features returned: {len(wet_geoms)}", flush=True)
        wet_union = _safe_union(wet_geoms)
        if wet_union:
            wet_in_project = transform(to_utm, wet_union.intersection(project_union))
            constraints["wetlands"] = {
                "acres":   round(wet_in_project.area / 4046.8564224, 2),
                "pct":     round(wet_in_project.area / gross_acres_m2 * 100, 1) if gross_acres_m2 > 0 else 0,
                "feature_count": len(wet_geoms),
            }
        else:
            constraints["wetlands"] = {"acres": 0, "pct": 0, "feature_count": 0}
    except Exception as e:
        constraints["wetlands"] = {"error": str(e)}

    # Transmission lines — assume 75ft each side (150ft ROW) standard
    try:
        trans_fc = cached_layer_query("transmission", bbox_query_poly)
        trans_lines = [shp_shape(f["geometry"]) for f in trans_fc.get("features", [])
                        if f.get("geometry")]
        print(f"[analyze] transmission features returned: {len(trans_lines)}", flush=True)
        if trans_lines:
            # Buffer each line in UTM by 75ft = ~22.86m
            trans_lines_utm = [transform(to_utm, l) for l in trans_lines]
            buffered = unary_union([l.buffer(22.86) for l in trans_lines_utm])
            project_utm = transform(to_utm, project_union)
            trans_in_project = buffered.intersection(project_utm)
            constraints["transmission_row"] = {
                "acres":         round(trans_in_project.area / 4046.8564224, 2),
                "pct":           round(trans_in_project.area / gross_acres_m2 * 100, 1) if gross_acres_m2 > 0 else 0,
                "line_count":    len(trans_lines),
                "row_width_ft":  150,
            }
        else:
            constraints["transmission_row"] = {"acres": 0, "pct": 0, "line_count": 0,
                                                 "row_width_ft": 150}
    except Exception as e:
        constraints["transmission_row"] = {"error": str(e)}

    # Streams — 50ft buffer each side (riparian setback)
    try:
        streams_fc = cached_layer_query("streams", bbox_query_poly)
        stream_lines = [shp_shape(f["geometry"]) for f in streams_fc.get("features", [])
                        if f.get("geometry")]
        print(f"[analyze] stream features returned: {len(stream_lines)}", flush=True)
        if stream_lines:
            stream_lines_utm = [transform(to_utm, l) for l in stream_lines]
            buffered = unary_union([l.buffer(15.24) for l in stream_lines_utm])  # 50ft = 15.24m
            project_utm = transform(to_utm, project_union)
            streams_in_project = buffered.intersection(project_utm)
            constraints["stream_buffers"] = {
                "acres":         round(streams_in_project.area / 4046.8564224, 2),
                "pct":           round(streams_in_project.area / gross_acres_m2 * 100, 1) if gross_acres_m2 > 0 else 0,
                "stream_count":  len(stream_lines),
                "buffer_ft":     50,
            }
        else:
            constraints["stream_buffers"] = {"acres": 0, "pct": 0, "stream_count": 0,
                                              "buffer_ft": 50}
    except Exception as e:
        constraints["stream_buffers"] = {"error": str(e)}

    # Pipelines — 50ft each side (standard pipeline easement). Query by bbox
    # envelope to avoid Esri polygon-complexity rejection.
    try:
        pipe_fc = arcgis_query(ENDPOINTS["rrc_pipelines"], bbox=bbox_query_poly.bounds)
        pipe_lines = [shp_shape(f["geometry"]) for f in pipe_fc.get("features", [])
                       if f.get("geometry")]
        print(f"[analyze] pipeline features returned: {len(pipe_lines)}", flush=True)
        if pipe_lines:
            pipe_lines_utm = [transform(to_utm, l) for l in pipe_lines]
            buffered = unary_union([l.buffer(15.24) for l in pipe_lines_utm])
            project_utm = transform(to_utm, project_union)
            pipes_in_project = buffered.intersection(project_utm)
            constraints["pipeline_easements"] = {
                "acres":         round(pipes_in_project.area / 4046.8564224, 2),
                "pct":           round(pipes_in_project.area / gross_acres_m2 * 100, 1) if gross_acres_m2 > 0 else 0,
                "pipeline_count": len(pipe_lines),
                "buffer_ft":     50,
            }
        else:
            constraints["pipeline_easements"] = {"acres": 0, "pct": 0,
                                                   "pipeline_count": 0, "buffer_ft": 50}
    except Exception as e:
        constraints["pipeline_easements"] = {"error": str(e)}

    # Recompute carefully — use individual unions.
    #
    # Every net-out is individually switchable. Wetlands in particular are often
    # mitigable or already permitted, so a blanket deduction understates a site;
    # the same applies to a pipeline easement that runs along a boundary you were
    # never going to build on. Defaults stay conservative (everything netted out)
    # and the project's `netout_assumptions` overrides them.
    netout = proj.get("netout_assumptions") or {}
    def _on(key):
        return netout.get(key, True) is not False

    project_utm = transform(to_utm, project_union)
    all_constraints_utm = []
    netout_detail = []          # per-constraint acreage, so the UI can show the ladder

    def _record(key, label, geom_utm, enabled):
        """Measure a constraint's own footprint, then include it only if enabled."""
        try:
            ac = geom_utm.area / 4046.8564224 if geom_utm else 0.0
        except Exception:
            ac = 0.0
        netout_detail.append({"key": key, "label": label,
                              "acres": round(ac, 2), "applied": bool(enabled and ac > 0)})
        if enabled and geom_utm is not None and not geom_utm.is_empty:
            all_constraints_utm.append(geom_utm)

    # Reuse the unions built above rather than repeating the work - on a large
    # tract against a dense NWI area that union is the expensive part.
    for key, label, geom_wgs, ckey in [
        ("flood", "Floodplain (100-yr)", flood_union, "floodplain"),
        ("wetlands", "Wetlands (NWI)", wet_union, "wetlands"),
    ]:
        layer_err = (constraints.get(ckey) or {}).get("error")
        if layer_err:
            # The layer never loaded. That is not the same as a zero-acre
            # constraint, and the UI has to be able to tell them apart.
            netout_detail.append({"key": key, "label": label, "acres": 0.0,
                                  "applied": False, "error": True,
                                  "reason": str(layer_err)[:120]})
            continue
        try:
            g = transform(to_utm, geom_wgs.intersection(project_union)) if geom_wgs is not None else None
            _record(key, label, g, _on(key))
        except Exception as e:
            netout_detail.append({"key": key, "label": label, "acres": 0.0,
                                  "applied": False, "error": True, "reason": str(e)[:120]})

    for key, label, lines, buf, ckey in [
        ("transmission", "Transmission easement", trans_lines, 22.86, "transmission_row"),
        ("streams", "Stream buffer", stream_lines, 15.24, "stream_buffers"),
        ("pipelines", "Pipeline easement", pipe_lines, 15.24, "pipeline_easements"),
    ]:
        try:
            if lines:
                ls = [transform(to_utm, l) for l in lines]
                g = unary_union([l.buffer(buf) for l in ls]).intersection(project_utm)
                _record(key, label, g, _on(key))
            else:
                # An empty list means either "nothing here" or "the fetch died".
                # Those are opposite facts about the site - one of them means the
                # acreage is overstated - so read which it was off the constraints
                # dict rather than reporting a confident zero either way.
                _err = (constraints.get(ckey) or {}).get("error")
                netout_detail.append({"key": key, "label": label, "acres": 0.0,
                                      "applied": False,
                                      **({"error": True, "reason": str(_err)[:120]} if _err else {})})
        except Exception as e:
            netout_detail.append({"key": key, "label": label, "acres": 0.0,
                                  "applied": False, "error": True, "reason": str(e)[:120]})

    if all_constraints_utm:
        constraints_union = unary_union(all_constraints_utm)
        net_dev_utm = project_utm.difference(constraints_union)
        net_dev_acres = net_dev_utm.area / 4046.8564224
    else:
        net_dev_acres = gross_acres

    # Infrastructure and landscaping. What survives the physical constraints is
    # still raw land: roads, detention, utility corridors, amenity and open space
    # all come out of it before a single lot can be platted. 30% is the planning
    # default here; it is an input, not a constant, so it can be dialled per deal.
    try:
        infra_pct = float(netout.get("infrastructure_pct", 30))
    except (TypeError, ValueError):
        infra_pct = 30.0
    infra_pct = max(0.0, min(infra_pct, 90.0))
    infra_acres = net_dev_acres * (infra_pct / 100.0)
    net_saleable_acres = max(0.0, net_dev_acres - infra_acres)

    # Yield — pro-forma with multi-product mix. Each lot type has:
    #   - allocation_pct (% of net developable assigned to this product)
    #   - units_per_acre (density, gross — roads/drainage already factored in)
    # Total lots = sum of (net_dev × allocation% × density).
    assumptions = proj.get("yield_assumptions") or {}
    lot_types_in = assumptions.get("lot_types") or []
    # Back-compat: if an old project has just units_per_acre, treat as single-type
    if not lot_types_in and assumptions.get("units_per_acre"):
        lot_types_in = [{
            "label": "Default",
            "units_per_acre": float(assumptions["units_per_acre"]),
            "allocation_pct": 100,
        }]
    if not lot_types_in:
        lot_types_in = [
            {"label": "40 FF", "units_per_acre": 6.0, "allocation_pct": 25},
            {"label": "50 FF", "units_per_acre": 5.0, "allocation_pct": 25},
            {"label": "60 FF", "units_per_acre": 4.0, "allocation_pct": 25},
            {"label": "70 FF", "units_per_acre": 3.5, "allocation_pct": 25},
        ]
    # Normalize: clamp values + compute per-type yield + totals
    breakdown = []
    total_lots = 0
    total_alloc_pct = 0
    for lt in lot_types_in:
        try:
            upa  = max(0.0, float(lt.get("units_per_acre") or 0))
            alloc = max(0.0, min(100.0, float(lt.get("allocation_pct") or 0)))
        except (TypeError, ValueError):
            upa, alloc = 0.0, 0.0
        ac_for_type = net_saleable_acres * (alloc / 100.0)
        lots = int(round(ac_for_type * upa))
        breakdown.append({
            "label":           lt.get("label") or "Lot type",
            "units_per_acre":  upa,
            "allocation_pct":  alloc,
            "acres":           round(ac_for_type, 2),
            "lots":            lots,
        })
        total_lots += lots
        total_alloc_pct += alloc
    weighted_density = (total_lots / net_saleable_acres) if net_saleable_acres > 0 else 0
    yields = {
        "total_lots":         total_lots,
        "weighted_density":   round(weighted_density, 2),
        "total_allocation_pct": round(total_alloc_pct, 1),
        "breakdown":          breakdown,
        "assumptions": {
            "lot_types": breakdown,
        },
    }

    # Build constraint GEOMETRIES (clipped to project) for client-side visualization.
    # Each entry is a GeoJSON FeatureCollection ready to drop into Leaflet.
    constraint_geoms = {}
    def _constraint_geom_failed(key, err):
        print(f"[analyze] constraint geometry '{key}' skipped: {type(err).__name__}: {err}",
              flush=True)

    try:
        if flood_union:
            clipped = flood_union.intersection(project_union)
            if not clipped.is_empty:
                constraint_geoms["floodplain"] = {"type": "Feature",
                                                    "geometry": shp_mapping(clipped),
                                                    "properties": {"acres": constraints["floodplain"]["acres"]}}
    except Exception as e:
        _constraint_geom_failed("floodplain", e)
    try:
        if wet_union:
            clipped = wet_union.intersection(project_union)
            if not clipped.is_empty:
                constraint_geoms["wetlands"] = {"type": "Feature",
                                                   "geometry": shp_mapping(clipped),
                                                   "properties": {"acres": constraints["wetlands"]["acres"]}}
    except Exception as e:
        _constraint_geom_failed("wetlands", e)
    try:
        # Transmission and pipelines are LINES — return them so client can render with the standard line style
        if trans_lines:
            constraint_geoms["transmission"] = {
                "type": "FeatureCollection",
                "features": [{"type": "Feature", "geometry": shp_mapping(l), "properties": {}}
                              for l in trans_lines],
            }
    except Exception as e:
        _constraint_geom_failed("transmission", e)
    try:
        if stream_lines:
            constraint_geoms["streams"] = {
                "type": "FeatureCollection",
                "features": [{"type": "Feature", "geometry": shp_mapping(l), "properties": {}}
                              for l in stream_lines],
            }
    except Exception as e:
        _constraint_geom_failed("streams", e)
    try:
        if pipe_lines:
            constraint_geoms["pipelines"] = {
                "type": "FeatureCollection",
                "features": [{"type": "Feature", "geometry": shp_mapping(l), "properties": {}}
                              for l in pipe_lines],
            }
    except Exception as e:
        _constraint_geom_failed("pipelines", e)

    analysis = {
        "computed_at":         datetime.now().isoformat(timespec="seconds"),
        "gross_acres":         round(gross_acres, 2),
        "net_developable_acres": round(net_dev_acres, 2),
        "net_developable_pct": round(net_dev_acres / gross_acres * 100, 1) if gross_acres > 0 else 0,
        "netout_detail": netout_detail,
        "infrastructure_pct": round(infra_pct, 1),
        "infrastructure_acres": round(infra_acres, 2),
        "net_saleable_acres": round(net_saleable_acres, 2),
        "net_saleable_pct": round(net_saleable_acres / gross_acres * 100, 1) if gross_acres > 0 else 0,
        "netout_assumptions": {
            **{k: _on(k) for k in ("flood", "wetlands", "transmission", "streams", "pipelines")},
            "infrastructure_pct": round(infra_pct, 1),
        },
        "constraints":         constraints,
        "constraint_geoms":    constraint_geoms,   # for client-side map overlay
        "yield_estimates":     yields,
        "tract_count":         len(tracts),
        "union_geometry":      shp_mapping(project_union),
    }
    return analysis


# ══════════════════════════════════════════════════════════════════════════
# Market and context helpers
#
# The rest of the standalone app's module-level machinery: CBAS subdivision
# data, Overpass amenity lookups, ACS demographics, FRED series, KML parsing,
# builder-name classification, and the tract-sheet PDF builder.
#
# Its Flask app, session auth, JSON storage, audit log, backup thread and cache
# refresh thread are all deliberately left behind - EmberApps has its own, and
# carrying them across would give the tab a second user table and a second
# scheduler, which is the thing integrating is supposed to remove.
# ══════════════════════════════════════════════════════════════════════════

# Cache last search result for export endpoints (single-user local app)
_last_search_cache = {"data": None, "meta": None, "saved_at": None}


# Outreach status options — used for both the picker UI and validation.
# Terminal "out of scope" statuses (not_interested, lost, passed) are hidden
# by default in the corridors list so they don't clutter active workflow.
OUTREACH_STATUSES = [
    ("lead",              "Lead",              "#6B7B8B"),
    ("contacted",         "Contacted",         "#1976D2"),
    ("negotiating",       "Negotiating",       "#F25929"),
    ("under_contract",    "Under contract",    "#FFA000"),
    ("closed",            "Closed",            "#2E7D32"),
    ("not_interested",    "Not interested",    "#7B1FA2"),
    ("lost",              "Lost",              "#9E9E9E"),
    ("passed",            "Passed",            "#5D4037"),
]


# Statuses considered "out of scope" / "don't show me again" — hidden in the
# corridor list by default; a toggle brings them back.
OUTREACH_OUT_OF_SCOPE = {"not_interested", "lost", "passed"}
OUTREACH_METHODS = ["call", "email", "letter", "text", "meeting", "site visit", "other"]


# In-memory cache for corridor tract searches. Keyed by
# (uid, corridor_id, buffer_mi, min_ac, max_ac). 30-minute TTL.
_CORRIDOR_TRACTS_CACHE = {}
_CORRIDOR_CACHE_TTL = 30 * 60


# --------------------------------------------------------------------------
# Parcel-cache admin endpoints (Carlos / admin only). Used to bootstrap the
# local SQLite cache county-by-county, see status, and force refresh.
# --------------------------------------------------------------------------
_cache_bootstrap_lock = threading.Lock()
_cache_bootstrap_status = {"running": False, "county": None, "pct": 0, "msg": ""}


def _export_filename(label, ext):
    safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in (label or "search"))
    return f"ember_gis_{safe}_{datetime.now().strftime('%Y%m%d_%H%M')}.{ext}"


# --------------------------------------------------------------------------
# Single-tract PDF — comprehensive one-page report for a specific Prop_ID.
# Used to send a broker / seller / partner as an offer cover sheet.
# --------------------------------------------------------------------------
def build_tract_sheet_pdf(tract_feature, search_meta=None,
                          outreach_record=None, user_lookup=None,
                          notes_records=None):
    """One-page tract report. tract_feature is one GeoJSON Feature with all the
    enrichment props already attached."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph,
                                       Spacer, Image as RLImage, KeepInFrame)

    props = tract_feature.get("properties") or {}
    geom  = tract_feature.get("geometry") or {}

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, leftMargin=0.4*inch,
                            rightMargin=0.4*inch, topMargin=0.4*inch, bottomMargin=0.4*inch)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading1"], fontName="Helvetica-Bold",
                                  fontSize=16, textColor=colors.HexColor("#13344E"), spaceAfter=2)
    sub_style = ParagraphStyle("sub", parent=styles["Normal"], fontName="Helvetica",
                                fontSize=9, textColor=colors.HexColor("#58595B"), spaceAfter=8)
    h_style = ParagraphStyle("h", parent=styles["Normal"], fontName="Helvetica-Bold",
                              fontSize=10, textColor=colors.HexColor("#F25929"), spaceBefore=4,
                              spaceAfter=4, leading=12)
    body_style = ParagraphStyle("body", parent=styles["Normal"], fontName="Helvetica",
                                 fontSize=9, textColor=colors.HexColor("#13344E"), leading=11)

    story = []
    owner = props.get("OWNER_NAME") or "Unknown owner"
    acres = props.get("Acres", "?")
    pid   = props.get("Prop_ID", "")
    story.append(Paragraph(f"{owner}", title_style))
    story.append(Paragraph(
        f"<b>{acres} ac</b> &nbsp;·&nbsp; "
        f"{props.get('_county', '')} County &nbsp;·&nbsp; "
        f"Prop ID {pid} &nbsp;·&nbsp; "
        f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        sub_style))

    # ----- Map snippet — same Esri World_Imagery export, just the tract -----
    tracts_fc = {"type": "FeatureCollection", "features": [tract_feature]}
    map_png = _build_search_map_png(tracts_fc, search_meta or {}, width=1100, height=520)
    if map_png:
        img = RLImage(io.BytesIO(map_png), width=7.5*inch, height=3.55*inch)
        img.hAlign = "CENTER"
        story.append(img)
        story.append(Spacer(1, 4))

    # ----- Two-column body: left = property/owner/tax, right = location/contact/utilities ---
    def kv(label, value):
        return [Paragraph(f"<b>{label}</b>", body_style),
                Paragraph(str(value) if value not in (None, "", 0) else "—", body_style)]

    mkt   = to_float(props.get("MARKET_VAL") or props.get("total_market_val"))
    appr  = to_float(props.get("total_appraised_val"))
    taxv  = to_float(props.get("tax_value"))
    flood_pct = props.get("_flood_pct")
    if flood_pct not in (None, ""):
        flood_pct = f"{flood_pct}%"
    util_counts = (
        f"{props.get('_wells_n', '—')} / "
        f"{props.get('_pipelines_n', '—')} / "
        f"{props.get('_transmission_n', '—')}"
    )
    since = props.get("_owner_since") or "—"
    tyr   = props.get("_hcad_tax_year") or "—"
    src   = _tract_source_label(props)

    left = [
        kv("Owner",          owner),
        kv("Mailing",        props.get("MAIL_ADDR") or ""),
        kv("Site address",   props.get("SITUS_ADDR") or "(no street address)"),
        kv("Legal desc",     (props.get("LEGAL_DESC") or "")[:120]),
        kv("Acres",          acres),
        kv("County",         props.get("_county") or ""),
        kv("ETJ / city",     props.get("_city_etj") or "—"),
        kv("School dist",    props.get("_school_dist") or "—"),
    ]
    right = [
        kv("Source",          src),
        kv("Owner since",     since),
        kv("HCAD tax year",   tyr),
        kv("Market value",    f"${mkt:,.0f}" if mkt is not None else "—"),
        kv("Appraised",       f"${appr:,.0f}" if appr is not None else "—"),
        kv("Taxable",         f"${taxv:,.0f}" if taxv is not None else "—"),
        kv("Flood %",         flood_pct or "—"),
        kv("Wells / Pipes / Tx", util_counts),
    ]

    def col(rows):
        t = Table(rows, colWidths=[1.2*inch, 2.3*inch])
        t.setStyle(TableStyle([
            ("VALIGN", (0,0), (-1,-1), "TOP"),
            ("LEFTPADDING", (0,0), (-1,-1), 0),
            ("RIGHTPADDING", (0,0), (-1,-1), 4),
            ("BOTTOMPADDING", (0,0), (-1,-1), 1),
            ("TOPPADDING", (0,0), (-1,-1), 1),
        ]))
        return t

    two_col = Table([[col(left), col(right)]], colWidths=[3.8*inch, 3.8*inch])
    two_col.setStyle(TableStyle([("VALIGN", (0,0), (-1,-1), "TOP")]))
    story.append(two_col)
    story.append(Spacer(1, 8))

    # ----- Outreach status block (if present) -----
    # Passed in rather than read from storage: persistence is Postgres now and
    # lives in acq_store, which this module deliberately does not import.
    outreach = outreach_record or None
    if outreach:
        user_lookup = user_lookup or {}
        status = outreach.get("status", "lead")
        status_label = next((l for v, l, c in OUTREACH_STATUSES if v == status), status)
        story.append(Paragraph(f"OUTREACH — {status_label.upper()}", h_style))
        meta_bits = []
        if outreach.get("broker_name"): meta_bits.append(f"Broker: <b>{outreach['broker_name']}</b>")
        if outreach.get("next_action"): meta_bits.append(f"Next: <b>{outreach['next_action']}</b>"
                                                          + (f" by {outreach.get('next_action_date','')}" if outreach.get('next_action_date') else ""))
        if meta_bits:
            story.append(Paragraph(" &nbsp;·&nbsp; ".join(meta_bits), body_style))
        log = (outreach.get("log") or [])[-3:]   # last 3 entries
        for e in reversed(log):
            who = user_lookup.get(e.get("by"))
            who_name = who["name"] if who else "Someone"
            when = e.get("at", "")[:16].replace("T", " ")
            story.append(Paragraph(
                f"<font color='#9a9a9a'>[{when} · {who_name} · {e.get('method','')}]</font> "
                f"{e.get('notes','')[:200]}",
                body_style))

    # ----- Notes thread (team-shared, append-only) — newest first -----
    notes = [n for n in (notes_records or [])
             if (n.get("text") or "").strip()]
    notes.sort(key=lambda n: n.get("created_at") or n.get("updated_at") or "", reverse=True)
    if notes:
        story.append(Paragraph(f"NOTES ({len(notes)} entries — newest first)", h_style))
        for n in notes:
            u = user_lookup.get(n.get("owner_id"))
            who = u.get("name") if u else "?"
            when = (n.get("created_at") or n.get("updated_at") or "")[:16].replace("T", " ")
            story.append(Paragraph(
                f"<font color='#9a9a9a'>[{when} · {who}]</font> {(n.get('text') or '')[:400]}",
                body_style))

    story.append(Spacer(1, 8))
    story.append(Paragraph(
        f"<font size=7 color='#9a9a9a'><i>Generated by Ember Acquisitions GIS. "
        f"Owner data: {src} (HCAD lags actual deed recordings by 2-8 weeks + annual roll). "
        f"Floodplain: FEMA NFHL current effective.</i></font>", body_style))

    doc.build(story)
    buf.seek(0)
    return buf


def _enrich_tract_sheet_feature(tract_feature):
    """Attach search-style spatial fields needed by the one-tract PDF.

    Search results already carry these fields, but a tract sheet can also be
    opened from a cold tract page or after restart. Keep failures as missing
    values so the PDF shows a dash instead of a false zero.
    """
    from shapely.geometry import box as shp_box, shape as shp_shape

    props = tract_feature.get("properties") or {}
    needed = (
        "_city_etj", "_school_dist", "_flood_pct",
        "_wells_n", "_pipelines_n", "_transmission_n",
    )
    if all(k in props and props.get(k) not in (None, "") for k in needed):
        return []
    if not tract_feature.get("geometry"):
        return ["geometry missing"]

    try:
        g = shp_shape(tract_feature["geometry"])
        minx, miny, maxx, maxy = g.bounds
    except Exception as e:
        return [f"bad geometry: {e}"]

    pad_x = max((maxx - minx) * 0.08, 0.002)
    pad_y = max((maxy - miny) * 0.08, 0.002)
    bbox_poly = shp_box(minx - pad_x, miny - pad_y, maxx + pad_x, maxy + pad_y)
    layers = {"tracts": {"type": "FeatureCollection", "features": [tract_feature]}}
    original = dict(props)

    fetchers = {
        "counties":     lambda: cached_layer_query("counties", bbox_poly),
        "etj":          lambda: arcgis_query(ENDPOINTS["etj"], bbox=bbox_poly.bounds),
        "schools":      lambda: cached_layer_query("tea_schools", bbox_poly),
        "flood":        lambda: cached_layer_query(
            "fema_flood", bbox_poly,
            where="FLD_ZONE IN ('A','AE','AH','AO','AR','A99','V','VE')",
            out_fields="FLD_ZONE,ZONE_SUBTY",
        ),
        "wells":        lambda: arcgis_query(ENDPOINTS["rrc_wells"], geometry_polygon=bbox_poly),
        "pipelines":    lambda: arcgis_query(ENDPOINTS["rrc_pipelines"], geometry_polygon=bbox_poly),
        "transmission": lambda: cached_layer_query("transmission", bbox_poly),
    }

    failed = set()
    errors = []
    with ThreadPoolExecutor(max_workers=7) as pool:
        futures = {pool.submit(fn): key for key, fn in fetchers.items()}
        for fut in as_completed(futures):
            key = futures[fut]
            try:
                fc = fut.result()
                if fc.get("error"):
                    raise RuntimeError(fc.get("error"))
                layers[key] = fc
            except Exception as e:
                failed.add(key)
                errors.append(f"{key}: {e}")
                layers[key] = {"type": "FeatureCollection", "features": []}

    annotate_tracts_with_enrichment(layers)

    failed_outputs = {
        "counties":     ("_county",),
        "etj":          ("_city_etj", "_in_city"),
        "schools":      ("_school_dist",),
        "flood":        ("_flood_pct",),
        "wells":        ("_wells_n",),
        "pipelines":    ("_pipelines_n",),
        "transmission": ("_transmission_n",),
    }
    for layer_key in failed:
        for prop_key in failed_outputs.get(layer_key, ()):
            if prop_key in original:
                props[prop_key] = original[prop_key]
            else:
                props.pop(prop_key, None)

    if errors:
        props["_tract_sheet_layer_errors"] = errors
        print(f"[tract-sheet] spatial enrichment partial: {errors}", flush=True)
    return errors


# --------------------------------------------------------------------------
# Import KMZ / KML — accept a Google-Earth boundary file and parse out the
# polygon(s) so the user can search inside them.
# --------------------------------------------------------------------------
def _parse_kml_coords(text):
    """KML <coordinates> is whitespace-separated 'lon,lat,alt' triplets.
    Return [[lon, lat], ...] (alt dropped — we're 2D)."""
    out = []
    if not text:
        return out
    for tok in text.split():
        parts = tok.split(",")
        if len(parts) < 2:
            continue
        try:
            lon = float(parts[0]); lat = float(parts[1])
        except ValueError:
            continue
        out.append([lon, lat])
    return out


def _kml_to_polygons(kml_bytes):
    """Parse KML XML, return a list of polygons in GeoJSON-coord form:
    each polygon = [outer_ring, inner_ring1, ...] where each ring is
    [[lon, lat], ...].  Handles both <Polygon> and closed <LineString>
    (a polyline drawn in Google Earth that the user closed)."""
    import xml.etree.ElementTree as ET
    kml_bytes = _sanitize_kml_bytes(kml_bytes)
    try:
        root = ET.fromstring(kml_bytes)
    except ET.ParseError as e:
        raise ValueError(f"Invalid KML XML: {e}")
    polygons = []
    # KML 2.2 namespace
    NS = "{http://www.opengis.net/kml/2.2}"

    # 1. Native <Polygon> elements (Google Earth's standard polygon draw)
    for poly_el in root.iter(f"{NS}Polygon"):
        outer_coords_el = poly_el.find(f".//{NS}outerBoundaryIs/{NS}LinearRing/{NS}coordinates")
        if outer_coords_el is None:
            continue
        outer = _parse_kml_coords(outer_coords_el.text)
        if len(outer) < 3:
            continue
        # Close the ring if not already closed
        if outer[0] != outer[-1]:
            outer.append(outer[0])
        rings = [outer]
        for inner_el in poly_el.findall(f".//{NS}innerBoundaryIs/{NS}LinearRing/{NS}coordinates"):
            inner = _parse_kml_coords(inner_el.text)
            if len(inner) >= 3:
                if inner[0] != inner[-1]:
                    inner.append(inner[0])
                rings.append(inner)
        polygons.append(rings)

    # 2. Also handle LineStrings that LOOK closed (start == end) — those are
    # polyline-drawn boundaries from Google Earth's "Add Path" tool.
    for line_el in root.iter(f"{NS}LineString"):
        coords_el = line_el.find(f"{NS}coordinates")
        if coords_el is None:
            continue
        coords = _parse_kml_coords(coords_el.text)
        if len(coords) < 3:
            continue
        # If start and end coincide (closed line), treat as a polygon
        first, last = coords[0], coords[-1]
        if abs(first[0] - last[0]) < 1e-7 and abs(first[1] - last[1]) < 1e-7:
            polygons.append([coords])

    # Try namespace-less parse as a last resort (some KML files drop the xmlns)
    if not polygons:
        for poly_el in root.iter("Polygon"):
            outer_coords_el = poly_el.find(".//outerBoundaryIs/LinearRing/coordinates")
            if outer_coords_el is None:
                continue
            outer = _parse_kml_coords(outer_coords_el.text)
            if len(outer) >= 3:
                if outer[0] != outer[-1]:
                    outer.append(outer[0])
                polygons.append([outer])

    return polygons


def _kml_to_linestrings(kml_bytes):
    """Extract OPEN LineStrings from KML (closed ones already become polygons
    via _kml_to_polygons). Returns list of [[lon, lat], ...]."""
    import xml.etree.ElementTree as ET
    kml_bytes = _sanitize_kml_bytes(kml_bytes)
    try:
        root = ET.fromstring(kml_bytes)
    except ET.ParseError as e:
        raise ValueError(f"Invalid KML XML: {e}")
    NS = "{http://www.opengis.net/kml/2.2}"
    lines = []
    def _collect(iter_fn):
        for line_el in iter_fn():
            coords_el = line_el.find(f"{NS}coordinates") if NS in (line_el.tag or "") \
                        else line_el.find("coordinates")
            if coords_el is None:
                # Try both NS and non-NS regardless
                coords_el = line_el.find(f"{NS}coordinates") or line_el.find("coordinates")
            if coords_el is None: continue
            coords = _parse_kml_coords(coords_el.text)
            if len(coords) < 2: continue
            first, last = coords[0], coords[-1]
            is_closed = (abs(first[0] - last[0]) < 1e-7 and abs(first[1] - last[1]) < 1e-7)
            if not is_closed:
                lines.append(coords)
    _collect(lambda: root.iter(f"{NS}LineString"))
    if not lines:
        _collect(lambda: root.iter("LineString"))
    return lines


def _buffer_line_to_polygon(coords, meters):
    """Buffer a LineString (lon/lat coords) by `meters` to produce a corridor
    polygon. Projects to local UTM for accurate metric buffering, then back to
    WGS84. Returns the outer-ring coords [[lon, lat], ...] or None."""
    from shapely.geometry import LineString
    from pyproj import Transformer
    if len(coords) < 2:
        return None
    avg_lon = sum(c[0] for c in coords) / len(coords)
    avg_lat = sum(c[1] for c in coords) / len(coords)
    zone = int((avg_lon + 180) // 6) + 1
    epsg = 32600 + zone if avg_lat >= 0 else 32700 + zone
    try:
        to_utm = Transformer.from_crs("EPSG:4326", f"EPSG:{epsg}", always_xy=True)
        to_wgs = Transformer.from_crs(f"EPSG:{epsg}", "EPSG:4326", always_xy=True)
        line_utm = LineString([to_utm.transform(c[0], c[1]) for c in coords])
        poly_utm = line_utm.buffer(meters, cap_style=1, join_style=1)   # rounded caps + joins
        if poly_utm.is_empty:
            return None
        if poly_utm.geom_type == "Polygon":
            target = poly_utm
        elif poly_utm.geom_type == "MultiPolygon":
            target = max(poly_utm.geoms, key=lambda g: g.area)
        else:
            return None
        return [
            [round(to_wgs.transform(x, y)[0], 6),
             round(to_wgs.transform(x, y)[1], 6)]
            for x, y in target.exterior.coords
        ]
    except Exception as e:
        print(f"[buffer] failed: {e}", flush=True)
        return None


# --------------------------------------------------------------------------
# Land plans — existing / planned communities as toggleable polygon overlays.
# Differs from corridors: these are about VISUALIZING context (where current
# developments are) rather than DRIVING a search. Each plan stores rich
# metadata (developer, status, lot counts, contact, notes).
# --------------------------------------------------------------------------
def _sanitize_kml_bytes(kml_bytes):
    """Defensive pre-processing for KMLs exported by Google Earth Pro and other
    tools that occasionally produce technically-invalid XML. Common issue:
    a namespace prefix (xsi:, gx:, atom:) is used somewhere in the document
    (often on <Document>'s schemaLocation attribute) but `xmlns:prefix=...`
    is never declared on the root <kml> element. Patch the <kml> tag to add
    the missing decl(s) so the standard XML parser doesn't throw on it."""
    import re
    # Locate the <kml ...> tag (case-insensitive)
    head = kml_bytes[:8192]
    m = re.search(rb'<kml\b([^>]*)>', head, re.IGNORECASE)
    if not m: return kml_bytes
    attrs = m.group(1)
    # Scan a generous chunk of the document (not just the kml tag's attrs) for
    # any `prefix:something` usage so we can detect prefixes referenced on
    # nested elements like <Document xsi:schemaLocation=...>.
    scan = kml_bytes[:65536]
    additions = b""
    for prefix, ns in [
        (b"xsi",  b"http://www.w3.org/2001/XMLSchema-instance"),
        (b"gx",   b"http://www.google.com/kml/ext/2.2"),
        (b"atom", b"http://www.w3.org/2005/Atom"),
    ]:
        decl = b"xmlns:" + prefix + b"="
        if decl in attrs:
            continue
        # Match `<Foo prefix:attr=` or ` prefix:attr=` inside any tag — that means
        # the prefix is being used somewhere and needs a matching xmlns: decl.
        if re.search(rb'\b' + prefix + rb':[A-Za-z_][\w.-]*\s*=', scan):
            additions += b' xmlns:' + prefix + b'="' + ns + b'"'
    if not additions: return kml_bytes
    fixed_tag = b"<kml " + attrs.strip() + additions + b">"
    return kml_bytes.replace(m.group(0), fixed_tag, 1)


def _strip_buffer_suffix(label):
    """Remove a trailing ' ±Nmi' suffix from a corridor label so we can
    rebuild it cleanly when the buffer changes."""
    import re
    return re.sub(r"\s*±[\d.]+mi\s*$", "", label or "").rstrip()


# --------------------------------------------------------------------------
# Folders
# --------------------------------------------------------------------------
FOLDER_COLORS = ["#F25929", "#1976d2", "#2e7d32", "#7b1fa2",
                 "#c62828", "#00838f", "#ef6c00", "#5d4037"]


import re as _sub_parse_re

_SUBDIVISION_CUT_RE = _sub_parse_re.compile(
    r"^([A-Z0-9][A-Z0-9 &/'\-\.]{2,60}?)"
    r"(?:\s+(?:SEC|SECTION|PHASE|PH|BLK|BLOCK|LT|LOT|PT|RES|RESERVE|SUB|TRACT|TR|ACRES?|U/R|UNRECORDED)\b|\s*\(|,|$)"
)
# Leading lot/block token (Harris/HCAD "lot-first" convention):
#   "LT 22 BLK 3 CYPRESS CREEK LAKES SEC 5" — name comes AFTER the tokens.
_LOTFIRST_TOKEN_RE = _sub_parse_re.compile(r"^(?:LTS?|LOTS?|BLKS?|BLOCK)\s+")
# The designator(s) after a token: "22", "C1", "5A", "1-3", "1 & 2", "18 19 & 20"
_LOTFIRST_DESIG_RE = _sub_parse_re.compile(
    r"^(?:[A-Z]?\d+[A-Z]?(?:-[A-Z0-9]+)?|[A-Z])"
    r"(?:\s*(?:&|,|AND)?\s*(?:[A-Z]?\d+[A-Z]?(?:-[A-Z0-9]+)?))*\s+"
)


def _parse_subdivision_name(legal: str):
    """Extract a residential-subdivision name from a CAD legal description.

    Handles both major Texas CAD conventions (verified against real cache data):
      • name-first (Waller/Ft Bend/MCAD): "S778100 SKY LAKES 1 BLK 7 LOT 3"
      • lot-first  (Harris/HCAD):        "LT 22 BLK 3 CYPRESS CREEK LAKES SEC 5"
    Returns None for rural abstracts, street ROW, reserves, and raw dev tracts.
    """
    if not legal:
        return None
    u = legal.strip().upper()
    # --- Hard rejects: not residential platted lots -------------------------
    # Street right-of-way dedications ("ROW-STREET WIDENING (DEDICATED...)")
    if u.startswith("ROW"):
        return None
    # Standalone reserves (landscape/utility/drainage) — not homes
    if u.startswith(("RES ", "RESERVE")):
        return None
    # Raw tracts / development land ("TR 5Z PROPOSED DIRECTORS LOT...", "TRS O1 P1...")
    if u.startswith(("TR ", "TRS ", "TRACT")):
        return None
    # Rural abstract land (unplatted survey acreage)
    if u.startswith(("ABS ", "AB ", "ABST", "A-", "ABSTRACT")):
        return None

    # Strip ALL parenthesized noise: "(NM)", "(PURE ACCT*...)", "(OPEN SPACE...)"
    u = _sub_parse_re.sub(r"\([^)]*\)", " ", u)
    u = _sub_parse_re.sub(r"\s+", " ", u).strip()

    # Strip CAD plat-code prefixes: Waller "S778100 ", MCAD "S9235 - "
    u = _sub_parse_re.sub(r"^[SR]\d{3,8}\s*-?\s+", "", u)

    # HARRIS lot-first: iteratively strip leading LT/BLK token+designator runs
    # so the subdivision name floats to the front. Loop handles "LT 1 BLK 2 &
    # RES C1 BLK 3 VILLAGE OF ..." style compound prefixes too.
    guard = 0
    while guard < 8:
        guard += 1
        m_tok = _LOTFIRST_TOKEN_RE.match(u)
        if m_tok:
            u = u[m_tok.end():]
            m_desig = _LOTFIRST_DESIG_RE.match(u)
            if m_desig:
                u = u[m_desig.end():]
            continue
        # Mid-stream reserve refs in compound prefixes: "& RES C1 BLK 3 NAME"
        m_join = _sub_parse_re.match(r"^(?:&|,|AND)\s+(?:RES(?:ERVE)?S?\s+[A-Z0-9\-]+\s+)?", u)
        if m_join and m_join.end() > 0:
            u = u[m_join.end():]
            continue
        break
    u = u.strip(" -.,&")

    # After stripping, reject if what's left is still abstract/rural
    if u.startswith(("ABS ", "ABST", "ABSTRACT", "U/R")):
        return None
    if not u:
        return None

    m = _SUBDIVISION_CUT_RE.match(u)
    if not m:
        return None
    name = m.group(1).strip(" -.,")
    # Kill trailing noise the cut regex may leave
    name = _sub_parse_re.sub(r"\s+ACRES?\s*[\d\.]*$", "", name)
    name = _sub_parse_re.sub(r"\s+TRACTS?\s*[\d\-]*$", "", name)
    name = name.strip(" -.,")
    if len(name) < 4:
        return None
    # Reject leftovers that are clearly not subdivision names
    if _sub_parse_re.match(r"^[SR]?\d+$", name):
        return None
    if _sub_parse_re.match(r"^(?:LT|LOT|BLK|BLOCK|TR|TRACT|SEC|SECTION|PT|RES|RESERVE)\b", name):
        return None
    if name in ("TRACT", "PARTIAL", "RESERVE", "UNRECORDED", "RURAL"):
        return None
    return name


# Known Texas / Houston homebuilders — matched against parcel OWNER_NAME.
# Shared by the /builders endpoint (who's active nearby) and the /communities
# endpoint (builder-held % per subdivision = absorption proxy).
_BUILDER_PATTERNS = [
    # National publics
    ("D.R. Horton",           [r"\bD\.?\s?R\.?\s?HORTON\b", r"\bDRH\b(?!\w)", r"\bDR HORTON\b"]),
    ("Lennar",                [r"\bLENNAR\b"]),
    ("PulteGroup",            [r"\bPULTE\b", r"\bDEL WEBB\b", r"\bCENTEX HOMES\b"]),
    ("KB Home",               [r"\bKB HOME\b", r"\bKBS HOME\b"]),
    ("Meritage Homes",        [r"\bMERITAGE\b"]),
    ("Taylor Morrison",       [r"\bTAYLOR MORRISON\b"]),
    ("Toll Brothers",         [r"\bTOLL BROS\b", r"\bTOLL BROTHERS\b"]),
    ("Beazer Homes",          [r"\bBEAZER\b"]),
    ("LGI Homes",             [r"\bLGI HOMES\b", r"\bLGI HOME\b"]),
    ("Century Communities",   [r"\bCENTURY COMMUNITIES\b", r"\bCCS COMMUNITIES\b"]),
    ("M/I Homes",             [r"\bM/I HOMES\b", r"\bMI HOMES\b"]),
    ("Ashton Woods",          [r"\bASHTON WOODS\b", r"\bASHTON WOOD\b"]),
    ("David Weekley Homes",   [r"\bDAVID WEEKLEY\b", r"\bWEEKLEY HOMES\b"]),
    ("Dream Finders Homes",   [r"\bDREAM FINDERS\b"]),
    ("Tri Pointe Homes",      [r"\bTRI POINTE\b", r"\bTRIPOINTE\b"]),
    # Texas/Houston strong regionals
    ("Perry Homes",           [r"\bPERRY HOMES\b"]),
    ("Highland Homes",        [r"\bHIGHLAND HOMES\b"]),
    ("Trendmaker Homes",      [r"\bTRENDMAKER\b"]),
    ("Chesmar Homes",         [r"\bCHESMAR\b"]),
    ("Coventry Homes",        [r"\bCOVENTRY HOMES\b"]),
    ("Westin Homes",          [r"\bWESTIN HOMES\b"]),
    ("Empire Communities",    [r"\bEMPIRE COMMUNITIES\b"]),
    ("Village Builders",      [r"\bVILLAGE BUILDERS\b"]),
    ("Ryland/CalAtlantic",    [r"\bRYLAND\b", r"\bCALATLANTIC\b"]),
    ("Newmark Homes",         [r"\bNEWMARK HOMES\b"]),
    ("Anglia Homes",          [r"\bANGLIA HOMES\b"]),
    ("Long Lake Ltd",         [r"\bLONG LAKE\b"]),
    ("Sitterle Homes",        [r"\bSITTERLE\b"]),
    ("Plantation Homes",      [r"\bPLANTATION HOMES\b"]),
    ("Riverside Homebuilders",[r"\bRIVERSIDE HOMEBUILDERS\b", r"\bRIVERSIDE HOMES\b"]),
    # Master-planned / land developers
    ("Land Tejas",            [r"\bLAND TEJAS\b"]),
    ("Johnson Development",   [r"\bJOHNSON DEVELOPMENT\b"]),
    ("Signorelli",            [r"\bSIGNORELLI\b"]),
    ("Caldwell Companies",    [r"\bCALDWELL COMPANIES\b"]),
    ("Newland",               [r"\bNEWLAND COMMUNITIES\b", r"\bNEWLAND ASSOCIATES\b"]),
    # Common holding-company/LLC patterns builders use
    ("Generic 'HOMES' entity",[r"\b(?:HOMES?|HOMEBUILDERS?)\s+(?:INC|LLC|LTD|LP|LP\.|CORP)\b"]),
]
_BUILDER_REGEX = [(name, [_sub_parse_re.compile(p) for p in patterns])
                    for name, patterns in _BUILDER_PATTERNS]


def _classify_builder(owner: str):
    """Return the builder name if OWNER_NAME matches a known homebuilder, else None."""
    if not owner: return None
    u = owner.upper()
    for name, patterns in _BUILDER_REGEX:
        if any(p.search(u) for p in patterns):
            return name
    return None


# --------------------------------------------------------------------------
# Overpass (OpenStreetMap) query helper — shared by amenities + communities.
# UA header required or .de returns 406 Not Acceptable.
# --------------------------------------------------------------------------
# Mirror reality, re-measured live 2026-08 with our actual amenities query:
#   overpass-api.de   planet data, 22-33s, HTTP 504 roughly 1 run in 3
#   maps.mail.ru      planet data, 17-21s, HTTP 504 roughly 1 run in 2
#   kumi.systems      read-timeout at 36s+  — effectively down
#   private.coffee    read-timeout at 36s+  — effectively down
#   osm.jp            bad TLS cert;  nchc.org.tw / comeuppance  DNS dead
#   overpass.osm.ch   answers in 1.3s but is a SWITZERLAND-ONLY extract — it
#                     returns HTTP 200 with elements=[] for every Texas query,
#                     including "find Houston". Fast and silently wrong is worse
#                     than slow: never add a mirror without first checking it
#                     can actually find something in Texas.
#
# Only two mirrors work and both fail often, but they fail INDEPENDENTLY. So we
# race them instead of trying them in series — fire both at once and take the
# first response that parses. Serial-with-backoff is what kept producing
# "Overpass query failed on all mirrors": a 504 from .de burned 30s before the
# fallback even started, and the fallback then timed out on its own leash.
#
# Throttled mirrors signal it two ways — a 504, or (worse) an HTML error page
# under HTTP 200 that makes r.json() raise. Both are transient: lose the race,
# don't fail the request.
_OVERPASS_MIRRORS = [
    "https://overpass-api.de/api/interpreter",
    "https://maps.mail.ru/osm/tools/overpass/api/interpreter",
]


# Texas airports with FAA commercial passenger service. OSM cannot be used to
# derive this: only IAH tags aerodrome=international, while Hobby carries nothing
# beyond iata/icao — identical to a private grass strip — so any tag-based filter
# either drops Hobby or lets every GA field through.
_TX_COMMERCIAL_AIRPORTS = {
    "IAH", "HOU", "DFW", "DAL", "AUS", "SAT", "ELP", "LBB", "MAF", "AMA",
    "CRP", "HRL", "BRO", "MFE", "LRD", "GRK", "SPS", "TYR", "ACT", "ABI",
    "SJT", "VCT", "CLL", "BPT", "GGG", "DRT",
}
_OVERPASS_HEADERS = {
    "User-Agent": "EmberAcquisitionsGIS/1.0 (internal real-estate analysis tool; contact: carlos@ember-grp.com)",
}


# Overpass answers never change minute-to-minute; cache them so a page reload
# (or a second card needing the same query) doesn't re-hit a throttled server.
_OVERPASS_CACHE = {}
_OVERPASS_CACHE_LOCK = threading.Lock()
_OVERPASS_CACHE_TTL = 6 * 60 * 60      # 6 hours


def _overpass_query(query: str, read_timeout: int = 45, attempts: int = 3):
    """POST an Overpass QL query, racing every mirror. Returns (json, error).

    Each round fires all mirrors concurrently and takes the first response that
    parses as JSON carrying an `elements` key; throttle responses (504, or an
    HTML page under a 200) simply lose the race. Up to `attempts` rounds with
    backoff, then the last error is returned. Results are cached for 6h.
    """
    import requests as _orq
    import time as _time
    import hashlib as _hl
    from concurrent.futures import ThreadPoolExecutor as _TP, as_completed as _ac

    key = _hl.sha1(query.encode("utf-8")).hexdigest()
    now = _time.time()
    with _OVERPASS_CACHE_LOCK:
        hit = _OVERPASS_CACHE.get(key)
        if hit and hit[1] > now:
            return hit[0], None

    def _store(data):
        # Cache any well-formed response, including a legitimately empty one —
        # a rural site really can have zero supermarkets, and re-querying that
        # every page load burns the retry budget for ~60s. Only refuse to cache
        # malformed payloads (no `elements` key), which is what a throttled or
        # truncated response looks like.
        if not isinstance(data, dict) or "elements" not in data:
            return data
        with _OVERPASS_CACHE_LOCK:
            _OVERPASS_CACHE[key] = (data, _time.time() + _OVERPASS_CACHE_TTL)
            if len(_OVERPASS_CACHE) > 200:
                for k, _ in sorted(_OVERPASS_CACHE.items(), key=lambda kv: kv[1][1])[:50]:
                    _OVERPASS_CACHE.pop(k, None)
        return data

    def _fetch(url):
        r = _orq.post(url, data={"data": query}, headers=_OVERPASS_HEADERS,
                      timeout=read_timeout)
        if r.status_code in (429, 502, 503, 504):
            raise RuntimeError(f"HTTP {r.status_code} (throttled)")
        r.raise_for_status()
        d = r.json()          # ValueError on the HTML throttle page
        if not isinstance(d, dict) or "elements" not in d:
            raise RuntimeError("malformed payload (no `elements`)")
        return d

    last = None
    for i in range(attempts):
        if i:
            _time.sleep(1.5 * (2 ** (i - 1)))          # 1.5s, 3s
        pool = _TP(max_workers=len(_OVERPASS_MIRRORS))
        winner = None
        try:
            futs = {pool.submit(_fetch, m): m for m in _OVERPASS_MIRRORS}
            for f in _ac(futs):
                try:
                    winner = f.result()
                    break                              # first good answer wins
                except Exception as e:
                    host = futs[f].split("/")[2]
                    last = RuntimeError(f"{host}: {str(e)[:90]}")
        finally:
            # Don't block on the losing mirror — it may sit there for 45s.
            pool.shutdown(wait=False)
        if winner is not None:
            return _store(winner), None
    return None, last


def _osm_communities_near(lat: float, lon: float, radius_mi: float):
    """Named residential communities near a point, keyed by normalized name.

    Returns {NORMALIZEDNAME: {lat, lon, distance_mi}}. Radius is rounded to
    whole miles and coordinates to 3dp so repeat calls from different cards
    hit the same Overpass cache entry instead of issuing new queries — the
    server throttles aggressively and every avoided query matters.
    """
    import math as _m
    lat_r = round(lat, 3)
    lon_r = round(lon, 3)
    rad = int(round(max(1.0, min(radius_mi, 25.0))))
    rad_m = int(rad * 1609.34)
    # Kept deliberately lean: place nodes are cheap and cover most named
    # subdivisions; named residential ways add the rest. Relations were dropped
    # (rare here, and they roughly doubled query time into throttle territory).
    q = f"""
[out:json][timeout:25];
(
  node["place"~"^(neighbourhood|suburb|quarter|village|hamlet)$"](around:{rad_m},{lat_r},{lon_r});
  way["landuse"="residential"]["name"](around:{rad_m},{lat_r},{lon_r});
);
out center 400;
"""
    data, _err = _overpass_query(q, 30)
    out = {}
    for el in (data or {}).get("elements", []):
        tags = el.get("tags") or {}
        nm = (tags.get("name") or "").strip()
        if not nm:
            continue
        elat = el.get("lat") or (el.get("center") or {}).get("lat")
        elon = el.get("lon") or (el.get("center") or {}).get("lon")
        if elat is None or elon is None:
            continue
        key = "".join(ch for ch in nm.upper() if ch.isalnum())
        d_mi = _m.hypot((elon - lon) * 69.0 * _m.cos(_m.radians(lat)),
                         (elat - lat) * 69.0)
        prev = out.get(key)
        if not prev or d_mi < prev["distance_mi"]:
            out[key] = {"name": nm, "lat": elat, "lon": elon,
                        "distance_mi": round(d_mi, 2)}
    return out


# ---------------------------------------------------------------------------
# Census ACS vintage
# ---------------------------------------------------------------------------
# Census publishes a new ACS 5-year vintage every December, so a hardcoded year
# goes quietly stale — this app was still reporting 2022 data in mid-2026, three
# vintages behind. Probe newest-first instead and cache the answer for a day.
#
# The probe has to request real data, not metadata: api.census.gov returns a 200
# for /data/<year>/acs/acs5.json on vintages that have no rows behind them yet.
_ACS_YEAR_CACHE = {"year": None, "at": 0.0}
_ACS_YEAR_LOCK = threading.Lock()


def _acs_radius_demographics(lat, lon, radius_mi):
    """Population / households / avg household size for tracts around a point.

    Compact companion to the full submarket endpoint — just the household
    numbers the retail-gap maths needs. Returns None when Census is unavailable
    (tract-level ACS requires CENSUS_API_KEY; county-level sometimes answers
    without one, tract-level does not), so every caller must have a fallback.
    """
    import requests as _rq
    from concurrent.futures import ThreadPoolExecutor as _DTP

    key = os.environ.get("CENSUS_API_KEY", "").strip()
    year = _acs_year()
    deg = radius_mi / 69.0
    bbox = f"{lon-deg},{lat-deg},{lon+deg},{lat+deg}"
    try:
        r = _rq.get("https://tigerweb.geo.census.gov/arcgis/rest/services/TIGERweb/"
                    "Tracts_Blocks/MapServer/0/query",
                    params={"geometry": bbox, "geometryType": "esriGeometryEnvelope",
                            "inSR": "4326", "spatialRel": "esriSpatialRelIntersects",
                            "outFields": "GEOID,STATE,COUNTY,TRACT", "returnGeometry": "false",
                            "f": "json"}, timeout=25)
        feats = r.json().get("features") or []
    except Exception:
        return None
    tracts = [(f["attributes"]["STATE"], f["attributes"]["COUNTY"], f["attributes"]["TRACT"])
              for f in feats if f.get("attributes")]
    if not tracts:
        return None

    VARS = "B01003_001E,B11001_001E,B25010_001E,B19013_001E,B25077_001E,B01002_001E"

    def _one(t):
        st, co, tr = t
        p = {"get": VARS, "for": f"tract:{tr}", "in": f"state:{st} county:{co}"}
        if key:
            p["key"] = key
        try:
            rr = _rq.get(f"https://api.census.gov/data/{year}/acs/acs5", params=p, timeout=20)
            if rr.status_code != 200 or not rr.text.lstrip().startswith("["):
                return None
            rows = rr.json()
            return dict(zip(rows[0], rows[1]))
        except Exception:
            return None

    with _DTP(max_workers=6) as pool:
        got = [g for g in pool.map(_one, tracts[:25]) if g]
    if not got:
        return None

    def _i(d, k):
        try:
            v = int(float(d.get(k)))
            return v if v > -10000 else None      # Census uses -666666666 for nulls
        except (TypeError, ValueError):
            return None

    pop = sum(_i(d, "B01003_001E") or 0 for d in got)
    hh = sum(_i(d, "B11001_001E") or 0 for d in got)
    incomes = [_i(d, "B19013_001E") for d in got]
    incomes = [v for v in incomes if v]
    values = [_i(d, "B25077_001E") for d in got]
    values = [v for v in values if v]
    ages = [_i(d, "B01002_001E") for d in got]
    ages = [v for v in ages if v]
    return {
        "year": year, "tracts": len(got),
        "population": pop, "households": hh,
        "avg_household_size": round(pop / hh, 2) if hh else None,
        "median_household_income": round(sum(incomes) / len(incomes)) if incomes else None,
        "median_home_value": round(sum(values) / len(values)) if values else None,
        "median_age": round(sum(ages) / len(ages), 1) if ages else None,
    }


def _acs_year(default=2023, probe_back=4):
    """Newest ACS 5-year vintage that actually returns data."""
    import requests as _rq
    import time as _t
    import datetime as _dt

    with _ACS_YEAR_LOCK:
        c = _ACS_YEAR_CACHE
        if c["year"] and c["at"] + 24 * 3600 > _t.time():
            return c["year"]
    key = os.environ.get("CENSUS_API_KEY", "").strip()
    newest = _dt.date.today().year - 1
    for yr in range(newest, newest - probe_back, -1):
        try:
            p = {"get": "NAME,B01003_001E", "for": "county:201", "in": "state:48"}
            if key:
                p["key"] = key
            r = _rq.get(f"https://api.census.gov/data/{yr}/acs/acs5", params=p, timeout=15)
            if r.status_code == 200 and r.text.lstrip().startswith("["):
                with _ACS_YEAR_LOCK:
                    _ACS_YEAR_CACHE.update({"year": yr, "at": _t.time()})
                return yr
        except Exception:
            continue
    return default


# ---------------------------------------------------------------------------
# CBAS new-home market survey
# ---------------------------------------------------------------------------
# The public GraphQL endpoint (cv-server.cbashome.com/graphql) exposes only four
# queries, and its survey entries carry NO builder names and NO coordinates.
# That forced geo-scoping to be approximated by ZIP polygons and left every
# competitor labelled "Builder #14".
#
# The CommunityVision portal bundle revealed the REST endpoints the app itself
# uses. Two of them do everything we need:
#
#   POST /getData     -> every subdivision CBAS tracks (1,186 in Houston) with
#                        lat/lng, builders + builders_names, lot_types (FF),
#                        prices, startsByQuarter / closingsByQuarter, vdls,
#                        futures, districtsNames, schoolNames, developerName,
#                        sections[].lot_types_builders (lots by builder by lot
#                        width) and latestFloorplanPricing (plan-level price,
#                        sqft and $/sf carrying builderName). The request body is
#                        ignored — it always returns the full set.
#   POST /buildExcel  -> report generator; `topBuilderRanking` returns named
#                        builders with annual starts/closings/VDL/market share.
#
# So: fetch /getData once, cache it, and filter by true great-circle distance
# from the project centroid. That removes ~90 network round-trips per request
# (80 builder probes + 5 lot-band queries + TIGERweb ZIP lookups + Overpass
# geocoding) and replaces every "Builder #14" with a real name.
#
# These are undocumented internal endpoints, so they can change when CBAS
# redeploys the portal. Everything below degrades to an error message rather
# than a traceback if the shape shifts.
_CBAS_BASE = "https://cv-server.cbashome.com"
_CBAS_DATA_CACHE = {"at": 0.0, "entries": None, "latest": None}
_CBAS_CACHE_LOCK = threading.Lock()
_CBAS_CACHE_TTL = 6 * 60 * 60      # 6 hours; the survey moves quarterly


def _cbas_headers(token):
    return {"Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
            "User-Agent": "EmberAcquisitionsGIS/1.0",
            "Origin": "https://cv.cbashome.com",
            "Referer": "https://cv.cbashome.com/"}


def _cbas_get_data(token, force=False):
    """Every subdivision CBAS tracks, cached. -> (entries, latest_quarter, error)."""
    import requests as _rq
    import time as _t

    now = _t.time()
    with _CBAS_CACHE_LOCK:
        c = _CBAS_DATA_CACHE
        if not force and c["entries"] and c["at"] + _CBAS_CACHE_TTL > now:
            return c["entries"], c["latest"], None
    try:
        r = _rq.post(f"{_CBAS_BASE}/getData", headers=_cbas_headers(token),
                     json={"table": "subdivisions"}, timeout=120)
        r.raise_for_status()
        d = r.json()
    except Exception as e:
        return None, None, f"CBAS /getData failed: {str(e)[:200]}"
    entries = d.get("entries") if isinstance(d, dict) else None
    if not entries:
        return None, None, "CBAS /getData returned no entries"
    latest = d.get("latestQuarter")
    with _CBAS_CACHE_LOCK:
        _CBAS_DATA_CACHE.update({"at": now, "entries": entries, "latest": latest})
    return entries, latest, None


def _cbas_builder_map(entries):
    """builder id -> name, read off the aligned builders / builders_names lists.

    CBAS has no builders table in the API, but every subdivision carries both
    the id list and the name list in the same order, so the union across all
    subdivisions reconstructs the full roster (171 builders in Houston).
    """
    m = {}
    for e in entries:
        ids = e.get("builders") or []
        nms = e.get("builders_names") or []
        if isinstance(ids, list) and isinstance(nms, list) and len(ids) == len(nms):
            for bid, nm in zip(ids, nms):
                if isinstance(nm, str) and nm.strip():
                    m[bid] = nm.strip()
    return m


# Lot widths in frontage feet — how the acquisitions team actually thinks about
# product. CBAS `lot_type` is the width in FF already (0 = unspecified).
_CBAS_LOT_BANDS = [(20, 40, "Under 40 FF"), (40, 50, "40–50 FF"), (50, 60, "50–60 FF"),
                   (60, 70, "60–70 FF"), (70, 200, "70+ FF")]


# Builder id 0 (and the literal name "Builder TBD") is CBAS's placeholder for
# "lots exist, builder not yet assigned" — it is not a competitor.
_CBAS_PLACEHOLDER_BUILDERS = {0}


# CBAS lot_type is mostly a clean frontage width, but the field also carries
# placeholders (0) and stray values (135, 1000). Anything outside a plausible
# single-family frontage is dropped rather than rendered as a lot width.
_CBAS_FF_MIN, _CBAS_FF_MAX = 20, 200


def _cbas_ff(v):
    """A lot width in frontage feet, or None if it isn't a plausible one."""
    try:
        f = int(v)
    except (TypeError, ValueError):
        return None
    return f if _CBAS_FF_MIN <= f <= _CBAS_FF_MAX else None


def _cbas_band(ft):
    f = _cbas_ff(ft)
    if f is None:
        return None
    for lo, hi, label in _CBAS_LOT_BANDS:
        if lo <= f < hi:
            return label
    return None


# FRED series — verified-working codes only. National series stripped.
# Per Carlos's note: TRUE submarket data (down to school district + community level)
# requires Zonda or scraping; FRED + Census tract-level cover Texas + Houston MSA
# and the area immediately around the project (census tract).
_FRED_HOUSTON_SERIES = [
    # --- HOUSTON METRO ---
    ("ATNHPIUS26420Q",       "Houston MSA home price index (FHFA, quarterly)",  "index"),

    # --- TEXAS STATE ---
    ("TXSTHPI",              "Texas home price index (FHFA)",                   "index"),
    ("MEHOINUSTXA646N",      "Texas median household income",                   "dollars"),
    ("TXUR",                 "Texas unemployment rate",                         "percent"),
    ("TXNA",                 "Texas non-farm employment",                       "thousands"),
    ("TXPOP",                "Texas resident population",                       "thousands"),
    ("TXPCPI",               "Texas per capita personal income",                "dollars"),
    ("TXBPPRIVSA",           "Texas total private permits (SA)",                "count"),
    ("TXNQGSP",              "Texas GDP (nominal, $B)",                         "dollars_b"),
    ("TXSLIND",              "Texas leading economic index",                    "index"),

    # --- FINANCING CONTEXT (national, but financing cost is national) ---
    ("MORTGAGE30US",         "30-yr fixed mortgage rate",                       "percent"),
    ("MORTGAGE15US",         "15-yr fixed mortgage rate",                       "percent"),
    ("DGS10",                "10-yr US Treasury yield",                         "percent"),
]
