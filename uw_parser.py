"""
uw_parser.py — Parses the "UW Performance Export" xlsx produced from the
GPD underwriting model and returns structured JSON for the Underwriting
dashboard tab.

Expected workbook layout (3 sheets, roughly):

  • 'GPD UW Home Sales'    — section + lot size + monthly home-sale counts
  • 'GPD UW Lot Takedowns' — section + lot size + monthly lot-takedown counts
  • 'GPD UW BEM Timing'    — section + lot size + ($-BEM) + (BEM Date) + monthly BEM $

Home Sales / Lot Takedowns share the same column layout:
    col 0 = "Section N", col 1 = lot size, cols 2.. = monthly date headers

BEM sheet is similar but has 2 extra metadata columns before the monthly data:
    col 0 = section, col 1 = lot size, col 2 = '$ - BEM',
    col 3 = 'BEM Date', cols 4.. = monthly date headers

Values are summed across lot-size rows so the final shape is
{section: {YYYY-MM: value}} matching the coworker's dashboard.

Ported from coworker's _parse_uw_sheet() in update_dashboard.py.
"""

from __future__ import annotations

import io
import re
from datetime import datetime
from typing import Any

import openpyxl


_SECTION_RE = re.compile(r"(?i)^section\s+\d+")


def _s(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _num(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _find_sheet(wb, name_contains: str):
    """Return the first worksheet whose name contains the given substring
    (case-insensitive). Chartsheets are skipped."""
    needle = name_contains.lower()
    for sn in wb.sheetnames:
        if needle in sn.lower():
            ws = wb[sn]
            if hasattr(ws, "iter_rows"):
                return ws
    return None


def _parse_uw_sheet(ws, data_col_start: int) -> tuple[dict, dict]:
    """
    Parse one UW sheet into:
      result  = {"Section 1": {"YYYY-MM": value, ...}, ...}
      lot_map = {"Section 1": "40", "Section 2": "45", ...}

    `data_col_start` is 0-indexed — 2 for Home Sales / Lot Takedowns,
    4 for BEM Timing (which has $-BEM + BEM Date before the months).
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}, {}

    header = rows[0]

    # Build {col_idx -> 'YYYY-MM'} for date headers
    month_cols: dict[int, str] = {}
    for c in range(data_col_start, len(header)):
        v = header[c]
        if isinstance(v, datetime):
            month_cols[c] = f"{v.year}-{v.month:02d}"

    result: dict[str, dict[str, float]] = {}
    lot_map: dict[str, str] = {}

    for r in rows[1:]:
        if not r or r[0] is None:
            continue
        sec = _s(r[0])
        if not _SECTION_RE.match(sec):
            continue

        # Lot size — capture first non-date value seen per section
        if sec not in lot_map and len(r) > 1 and r[1] is not None:
            lot = r[1]
            if not isinstance(lot, datetime):
                try:
                    lot_map[sec] = str(int(float(str(lot).strip())))
                except (TypeError, ValueError):
                    lot_map[sec] = _s(lot)

        bucket = result.setdefault(sec, {})
        for c, ym in month_cols.items():
            if c >= len(r):
                break
            val = _num(r[c])
            if val is None or val == 0:
                continue
            bucket[ym] = round(bucket.get(ym, 0) + val, 2)

    return result, lot_map


def parse_uw(file_bytes: bytes) -> dict:
    """
    Parse a UW Performance Export xlsx and return:
      {
        "home_sales":    {"Section N": {"YYYY-MM": count, ...}, ...},
        "lot_takedowns": {"Section N": {"YYYY-MM": count, ...}, ...},
        "bem":           {"Section N": {"YYYY-MM": dollars, ...}, ...},
        "section_lots":  {"Section N": "40", ...},   # inferred from Home Sales sheet
        "sheet_names":   [...],                      # raw workbook sheet names
      }
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)

    data: dict[str, Any] = {
        "home_sales":    {},
        "lot_takedowns": {},
        "bem":           {},
        "section_lots":  {},
        "sheet_names":   list(wb.sheetnames),
    }

    hs_ws = _find_sheet(wb, "home sales")
    if hs_ws is not None:
        home_sales, section_lots = _parse_uw_sheet(hs_ws, data_col_start=2)
        data["home_sales"] = home_sales
        data["section_lots"] = section_lots

    lt_ws = _find_sheet(wb, "lot takedowns")
    if lt_ws is not None:
        lot_takedowns, lt_lots = _parse_uw_sheet(lt_ws, data_col_start=2)
        data["lot_takedowns"] = lot_takedowns
        # Fall back to LT sheet lot map if Home Sales was missing
        if not data["section_lots"]:
            data["section_lots"] = lt_lots

    bem_ws = _find_sheet(wb, "bem")
    if bem_ws is not None:
        bem, _ = _parse_uw_sheet(bem_ws, data_col_start=4)
        data["bem"] = bem

    wb.close()
    return data


if __name__ == "__main__":
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else \
        r"C:\Users\Javier\Documents\EmberAcquisitions\_sales_dashboard_ref\UW Performance Export - GPD.xlsx"
    with open(path, "rb") as f:
        out = parse_uw(f.read())
    print(f"sheet_names: {out['sheet_names']}")
    print(f"home_sales sections: {len(out['home_sales'])}")
    print(f"lot_takedowns sections: {len(out['lot_takedowns'])}")
    print(f"bem sections: {len(out['bem'])}")
    print(f"section_lots: {len(out['section_lots'])}  sample={dict(list(out['section_lots'].items())[:5])}")
    # Dump one section's home sales to verify
    if out["home_sales"]:
        first_sec = next(iter(out["home_sales"]))
        months = out["home_sales"][first_sec]
        print(f"\n{first_sec} home sales: {len(months)} monthly buckets")
        non_zero = {k: v for k, v in sorted(months.items()) if v}
        for k, v in list(non_zero.items())[:8]:
            print(f"  {k}: {v}")
    # Annual totals across all sections
    print("\nHome Sales — annual totals:")
    annual = {}
    for sec, mo in out["home_sales"].items():
        for ym, v in mo.items():
            y = ym.split("-")[0]
            annual[y] = annual.get(y, 0) + v
    for y in sorted(annual):
        print(f"  {y}: {annual[y]:.0f}")
    # BEM annual totals
    print("\nBEM — annual totals ($):")
    annual_bem = {}
    for sec, mo in out["bem"].items():
        for ym, v in mo.items():
            y = ym.split("-")[0]
            annual_bem[y] = annual_bem.get(y, 0) + v
    for y in sorted(annual_bem):
        print(f"  {y}: ${annual_bem[y]:,.0f}")
