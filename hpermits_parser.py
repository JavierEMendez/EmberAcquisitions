"""
hpermits_parser.py — Parses the raw HPermits Houston-wide monthly permit
workbook (e.g. "HPermits - February 2026.xlsx") and returns structured JSON
for the Houston Permits dashboard tab.

Sheets we read (data sheets only; chartsheets are skipped):

  • MKT      — market totals.        Row 1 = Grand Total (Houston-wide).
               Rows 2+ = per-market (CENTRAL, FAR NORTH, NORTHEAST, ...).
  • SUB      — submarket totals.     Same column layout as MKT.
  • BLD      — builder totals.       Top ~290 builders, same column layout.
  • BLD-CO   — parent-company totals (builder parent companies).
  • PROJ data — pre-computed top 10 projects (community-level) with
                previous TTM + current TTM.

All monthly-layout sheets share this column map:
    col 0          : name
    cols 1-12      : 12 prior-year monthly values  (MAR 24 → FEB 25)
    col 13         : prior-year TTM total          "MAR 24 - FEB 25"
    cols 14-25     : 12 current-year monthly values (MAR 25 → FEB 26)
    col 26         : current-year TTM total        "MAR 25 - FEB 26"
    col 27         : RANK (current TTM)
    col 28         : YoY % change (as fraction, e.g. -0.0864)
"""

from __future__ import annotations

import io
from typing import Any

import openpyxl


PREV_MONTH_COLS = list(range(1, 13))    # MAR 24 → FEB 25
CURR_MONTH_COLS = list(range(14, 26))   # MAR 25 → FEB 26
PREV_TTM_COL = 13
CURR_TTM_COL = 26
RANK_COL = 27
YOY_COL = 28


def _s(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _num(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _int0(v: Any) -> int:
    n = _num(v)
    return int(round(n)) if n else 0


def _int_or_none(v: Any) -> int | None:
    n = _num(v)
    return None if n is None else int(round(n))


def _round_pct(v: Any) -> float | None:
    """Convert fraction (-0.0864) → percent rounded (-8.64)."""
    n = _num(v)
    return None if n is None else round(n * 100, 2)


def _extract_header_labels(row: tuple, cols: list[int]) -> list[str]:
    return [_s(row[c]) if len(row) > c else "" for c in cols]


def _monthly(r: tuple, cols: list[int]) -> list[int]:
    return [_int0(r[c]) if len(r) > c else 0 for c in cols]


def _parse_monthly_sheet(ws, name_label: str) -> dict:
    """
    Parse a sheet with the standard MKT/BLD/SUB/BLD-CO layout.
    Returns: {grand_total: {...} | None, rows: [{name, monthly_prev,
              monthly_curr, ttm_prev, ttm_curr, rank, yoy_pct}, ...]}
    """
    rows = list(ws.iter_rows(values_only=True))
    result = {"grand_total": None, "rows": []}
    if len(rows) < 2:
        return result

    # Row 1 = Grand Total
    gr = rows[1]
    if _s(gr[0]).upper().startswith("GRAND TOTAL"):
        result["grand_total"] = {
            "monthly_prev": _monthly(gr, PREV_MONTH_COLS),
            "monthly_curr": _monthly(gr, CURR_MONTH_COLS),
            "ttm_prev":     _int0(gr[PREV_TTM_COL]) if len(gr) > PREV_TTM_COL else 0,
            "ttm_curr":     _int0(gr[CURR_TTM_COL]) if len(gr) > CURR_TTM_COL else 0,
            "yoy_pct":      _round_pct(gr[YOY_COL]) if len(gr) > YOY_COL else None,
        }

    for i in range(2, len(rows)):
        r = rows[i]
        if not r:
            continue
        nm = _s(r[0])
        if not nm:
            continue
        ttm_curr = _int_or_none(r[CURR_TTM_COL]) if len(r) > CURR_TTM_COL else None
        ttm_prev = _int_or_none(r[PREV_TTM_COL]) if len(r) > PREV_TTM_COL else None
        if not ttm_curr and not ttm_prev:
            continue
        result["rows"].append({
            name_label:     nm,
            "monthly_prev": _monthly(r, PREV_MONTH_COLS),
            "monthly_curr": _monthly(r, CURR_MONTH_COLS),
            "ttm_prev":     ttm_prev or 0,
            "ttm_curr":     ttm_curr or 0,
            "rank":         _int_or_none(r[RANK_COL]) if len(r) > RANK_COL else None,
            "yoy_pct":      _round_pct(r[YOY_COL])   if len(r) > YOY_COL  else None,
        })

    # Sort by current TTM desc
    result["rows"].sort(key=lambda x: x["ttm_curr"], reverse=True)
    return result


def _parse_proj_data(ws) -> list[dict]:
    """
    PROJ data sheet — top N communities, already ranked.
      col 0: community name
      col 1: prior TTM
      col 2: current TTM
    """
    rows = list(ws.iter_rows(values_only=True))
    out = []
    for i in range(1, len(rows)):
        r = rows[i]
        if not r:
            continue
        nm = _s(r[0])
        if not nm or nm.upper() == "COMMUNITY":
            continue
        tp = _int_or_none(r[1]) if len(r) > 1 else None
        tc = _int_or_none(r[2]) if len(r) > 2 else None
        if tc is None and tp is None:
            continue
        yoy = None
        if tp and tc is not None:
            yoy = round((tc - tp) / tp * 100, 2)
        out.append({
            "community": nm,
            "ttm_prev":  tp or 0,
            "ttm_curr":  tc or 0,
            "yoy_pct":   yoy,
        })
    out.sort(key=lambda x: x["ttm_curr"], reverse=True)
    return out


def parse_hpermits(file_bytes: bytes) -> dict:
    """
    Parse HPermits xlsx and return:
      {
        "months_prev":    12 labels (MAR 24..FEB 25),
        "months_curr":    12 labels (MAR 25..FEB 26),
        "ttm_prev_label": "MAR 24 - FEB 25",
        "ttm_curr_label": "MAR 25 - FEB 26",
        "grand_total":    {monthly_prev, monthly_curr, ttm_prev, ttm_curr, yoy_pct},
        "markets":        [ {market, monthly_prev, monthly_curr, ttm_prev, ttm_curr, rank, yoy_pct}, ... ],
        "submarkets":     [ {submarket, ...} ],
        "builders":       [ {builder, ...} ],
        "companies":      [ {company, ...} ],
        "top_projects":   [ {community, ttm_prev, ttm_curr, yoy_pct}, ... ],
      }
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)

    data: dict[str, Any] = {
        "months_prev":    [],
        "months_curr":    [],
        "ttm_prev_label": "",
        "ttm_curr_label": "",
        "grand_total":    None,
        "markets":        [],
        "submarkets":     [],
        "builders":       [],
        "companies":      [],
        "top_projects":   [],
    }

    # Pull header labels off the MKT sheet (all sheets share the same layout)
    if "MKT" in wb.sheetnames:
        ws = wb["MKT"]
        if hasattr(ws, "iter_rows"):
            hdr = next(ws.iter_rows(values_only=True), None)
            if hdr:
                data["months_prev"]    = _extract_header_labels(hdr, PREV_MONTH_COLS)
                data["months_curr"]    = _extract_header_labels(hdr, CURR_MONTH_COLS)
                data["ttm_prev_label"] = _s(hdr[PREV_TTM_COL]) if len(hdr) > PREV_TTM_COL else ""
                data["ttm_curr_label"] = _s(hdr[CURR_TTM_COL]) if len(hdr) > CURR_TTM_COL else ""

    sheet_map = [
        ("MKT",    "market",    "markets",    "grand_total"),
        ("SUB",    "submarket", "submarkets", None),
        ("BLD",    "builder",   "builders",   None),
        ("BLD-CO", "company",   "companies",  None),
    ]
    for sheet_name, name_key, list_key, gt_key in sheet_map:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        if not hasattr(ws, "iter_rows"):
            continue
        parsed = _parse_monthly_sheet(ws, name_key)
        data[list_key] = parsed["rows"]
        if gt_key and parsed["grand_total"]:
            data[gt_key] = parsed["grand_total"]

    if "PROJ data" in wb.sheetnames:
        ws = wb["PROJ data"]
        if hasattr(ws, "iter_rows"):
            data["top_projects"] = _parse_proj_data(ws)

    wb.close()
    return data


if __name__ == "__main__":
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else \
        r"C:\Users\Javier\Downloads\HPermits - February 2026.xlsx"
    with open(path, "rb") as f:
        out = parse_hpermits(f.read())
    print(f"months_prev: {out['months_prev'][0]}..{out['months_prev'][-1]}  ({len(out['months_prev'])})")
    print(f"months_curr: {out['months_curr'][0]}..{out['months_curr'][-1]}  ({len(out['months_curr'])})")
    print(f"TTM labels : {out['ttm_prev_label']!r}  ->  {out['ttm_curr_label']!r}")
    gt = out["grand_total"]
    print(f"Grand Total: prev={gt['ttm_prev']}  curr={gt['ttm_curr']}  YoY={gt['yoy_pct']}%")
    print(f"Markets: {len(out['markets'])}  | Submarkets: {len(out['submarkets'])}  | "
          f"Builders: {len(out['builders'])}  | Companies: {len(out['companies'])}  | "
          f"Top projects: {len(out['top_projects'])}")
    print("\nTop 5 markets (by curr TTM):")
    for m in out["markets"][:5]:
        print(f"  {m['market']:<18} prev={m['ttm_prev']:>6}  curr={m['ttm_curr']:>6}  "
              f"rank={m['rank']}  YoY={m['yoy_pct']}%")
    print("\nTop 5 builders:")
    for b in out["builders"][:5]:
        print(f"  {b['builder']:<25} prev={b['ttm_prev']:>6}  curr={b['ttm_curr']:>6}  "
              f"rank={b['rank']}  YoY={b['yoy_pct']}%")
    print("\nTop projects:")
    for p in out["top_projects"]:
        print(f"  {p['community']:<25} prev={p['ttm_prev']:>6}  curr={p['ttm_curr']:>6}  YoY={p['yoy_pct']}%")
