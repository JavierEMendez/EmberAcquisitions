"""
partners_cf_parser.py — Parses the "EMBER Partners CF" workbook (partner-level
sources & uses ledger) into structured JSON.

The sheet is one wide monthly timeline: a title row carrying year markers, a
label row of Jan..Dec + a per-year "Total" column, then category blocks down
column B — each a header (INVESTMENTS/DIVIDENDS, OPEX, PRE-DEVELOPMENT,
GP CO-INVEST, PROMOTES), its line items (partners / entities / projects),
and "Total <category>" / "Cumm <category>" rows — ending with the grand
"Total CF" / "Cumm CF" rows.

Everything is located by label: categories, line items, years, and months are
all dynamic between versions. Per-year "Total" columns and all "Cumm" rows are
skipped and recomputed downstream, so a stray formula in the workbook can't
poison the dashboard.

Sign convention (preserved as-is): partner contributions and cash received
are positive; dividends out and capital deployed are negative. The grand
Total CF nets to ~zero because the ledger balances sources against uses.
"""

import io
import re
from typing import Any

import openpyxl

_TITLE = "EMBER PARTNERS CF"
_MONTHS = {"jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
           "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12}


def _str(val: Any) -> str:
    return "" if val is None else str(val).strip()


def _num(val: Any):
    if val is None or isinstance(val, str):
        return None
    try:
        n = float(val)
    except (TypeError, ValueError):
        return None
    return round(n, 2)


def parse_partners_cf(file_bytes: bytes) -> dict:
    """Parse the Partners CF workbook. Raises ValueError when the title cell
    or the month axis can't be found."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    # Locate the sheet + title cell ("EMBER PARTNERS CF").
    grid, title_row, title_col = None, None, None
    for name in wb.sheetnames:
        ws = wb[name]
        g = {}
        rn = 0
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 400),
                                max_col=min(ws.max_column, 400)):
            rn += 1
            for i, c in enumerate(row):
                v = getattr(c, "value", None)
                if v is not None:
                    g[(rn, i + 1)] = v
        for (r, c), v in g.items():
            if _str(v).upper() == _TITLE:
                grid, title_row, title_col = g, r, c
                break
        if grid:
            break
    if grid is None:
        raise ValueError('Workbook has no "%s" title cell' % _TITLE)
    max_row = max(r for r, _c in grid)
    max_col = max(c for _r, c in grid)

    # Month axis: the label row sits under the title row. Year markers live on
    # the title row over each year's Jan column; per-year "Total" columns are
    # skipped (recomputed downstream).
    label_row = title_row + 1
    month_cols = []          # (col, 'YYYY-MM')
    year = None
    for c in range(title_col + 1, max_col + 1):
        marker = grid.get((title_row, c))
        if isinstance(marker, (int, float)) and 1990 < marker < 2100:
            mon = _MONTHS.get(_str(grid.get((label_row, c))).lower()[:3])
            if mon:                       # year over Jan, not over its Total col
                year = int(marker)
        label = _str(grid.get((label_row, c))).lower()[:3]
        mon = _MONTHS.get(label)
        if mon and year:
            month_cols.append((c, f"{year}-{mon:02d}"))
    if not month_cols:
        raise ValueError("No month columns found under the title row")
    months = [m for _c, m in month_cols]

    def _row_values(r: int) -> dict:
        out = {}
        for c, key in month_cols:
            v = _num(grid.get((r, c)))
            if v:
                out[key] = v
        return out

    # Category blocks down the label column.
    categories = []
    current = None
    grand_total = {}
    for r in range(label_row + 1, max_row + 1):
        label = _str(grid.get((r, title_col)))
        if not label:
            continue
        low = label.lower()
        if low.startswith("cumm"):
            continue                                  # recomputed downstream
        if low == "total cf":
            grand_total = _row_values(r)
            current = None
            continue
        if low.startswith("total "):
            if current is not None:
                current["total"] = _row_values(r)
                current = None                        # next label = new category
            continue
        if current is None:
            current = {"name": label, "items": [], "total": {}}
            categories.append(current)
        else:
            current["items"].append({"name": label, "months": _row_values(r)})
    categories = [c for c in categories if c["items"]]
    if not categories:
        raise ValueError("No category blocks found under the title cell")

    # "Actuals through" month from a date in the filename-style title or the
    # workbook itself is not carried on-sheet; the caller stamps it.
    return {
        "months": months,
        "categories": categories,
        "grand_total": grand_total,
    }


_FILENAME_DATE_RE = re.compile(r"(\d{4})[_\-. ](\d{1,2})(?:[_\-. ](\d{1,2}))?")


def actuals_through_from_filename(filename: str):
    """'EMBER Partners CF - 2026_08_12.xlsx' → '2026-08'. None when absent."""
    m = _FILENAME_DATE_RE.search(filename or "")
    if not m:
        return None
    y, mo = int(m.group(1)), int(m.group(2))
    if 1990 < y < 2100 and 1 <= mo <= 12:
        return f"{y}-{mo:02d}"
    return None
