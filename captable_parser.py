"""
captable_parser.py — Parses the Ember "All Cap Tables" workbook into a flat
list of per-project ownership positions, one row per investment vehicle.

Source of truth = the full per-project cap-table tabs (NOT the curated
"VEXA Positions" subset). Each project's authoritative block (tab name, row
range, entity/percentage columns) is encoded in CAPTABLE_SPEC below. The
project names here are the canonical display names; PROJECT_ALIASES maps them
to however the project appears in the Project Returns data so the Investors
view can scale returns by ownership %.

Output row shape (one per vehicle per project; LightHaven yields two rows per
vehicle — one common, one preferred):

    {
        "project":      "Hawthorne",        # canonical project name
        "vehicle":      "Aljosan USA, LLC", # investment vehicle (Entity col)
        "contribution": 2726667.0,          # dollars contributed
        "pct":          0.1340841,          # ownership fraction (0..1)
        "equity_class": "common",           # "common" for everything except
                                            # LightHaven preferred ("preferred")
    }
"""

from typing import Any

import openpyxl


# ---------------------------------------------------------------------------
# Per-project authoritative blocks. Confirmed by Carlos against
# "2026.05.06 All Cap Tables.xlsx" (12 tabs).
#
#   tab          — worksheet name
#   project      — canonical display name (matched to returns via PROJECT_ALIASES)
#   entity_col   — column letter holding the vehicle name
#   pct_col      — column letter holding the ownership fraction (0..1)
#   contrib_col  — column letter holding the dollar contribution
#   start / end  — inclusive 1-based data row range (header + Total excluded)
#   equity_class — "common" (default) or "preferred" (LightHaven pref block)
#   exclude      — vehicle names to drop (e.g. holding-entity sub-members that
#                  would double-count). Compared case-insensitively, trimmed.
#
# No look-through anywhere: holding entities (5656 San Felipe, Mid Main East
# Management) stay as single lines. The San Felipe sub-members are excluded
# because they are San Felipe's internal breakdown and summing them with the
# San Felipe line itself would double-count.
# ---------------------------------------------------------------------------
CAPTABLE_SPEC: list[dict] = [
    {
        "project": "Grand Prairie", "tab": "GPD (Grand Prairie)",
        "entity_col": "B", "pct_col": "D", "contrib_col": "C",
        "start": 7, "end": 15,
    },
    {
        "project": "Windrose Green", "tab": "WRG (Windrose)",
        "entity_col": "B", "pct_col": "D", "contrib_col": "C",
        "start": 7, "end": 11,
    },
    {
        "project": "Dennison", "tab": "WRRD (Dennison)",
        "entity_col": "B", "pct_col": "D", "contrib_col": "C",
        "start": 7, "end": 14,
    },
    {
        "project": "Lexington Land", "tab": "LLP (Lexington)",
        "entity_col": "B", "pct_col": "D", "contrib_col": "C",
        "start": 8, "end": 12,
    },
    {
        # "Final Hawthorne Cap Table (January 2025)" block. 5656 San Felipe
        # stays one line at ~20.78%; its four indented sub-members are
        # dropped (their %s sum to San Felipe's line — keeping both = double
        # count). Remaining 8 vehicles sum to exactly 100%.
        "project": "Hawthorne", "tab": "SFC (Hawthorne)",
        "entity_col": "B", "pct_col": "D", "contrib_col": "C",
        "start": 25, "end": 36,
        "exclude": [
            "Alfonso Mancilla Leal",
            "Eduardo Rangel Abella",
            "Jesus Rodriguez (Baldomero)",
            "Maria Cecilia Mancilla Leal",
        ],
    },
    {
        # FIRST/Sept-2024 snapshot (15 vehicles, $12M total). NOT the
        # June-2026 block lower on the same tab. Mid Main East Management
        # stays one line at 2.50% (no look-through to its CCEMB/RHS table).
        "project": "Mid Main", "tab": "MMEC (Mid Main)",
        "entity_col": "B", "pct_col": "D", "contrib_col": "C",
        "start": 7, "end": 21,
    },
    {
        # LightHaven Common Equity block. Ownership = within-class "LP Share"
        # (col E), not the blended "Ending Capital %" (col D). Wires to the
        # standalone "LightHaven (Common Equity)" returns project.
        "project": "LightHaven Common", "tab": "LHDW (Lighthaven)",
        "entity_col": "B", "pct_col": "E", "contrib_col": "C",
        "start": 7, "end": 22, "equity_class": "common",
    },
    {
        # LightHaven Preferred Equity block — wires to the standalone
        # "LightHaven (Preferred Equity)" returns project (its own deal with
        # its own Total LP Distributions).
        "project": "LightHaven Preferred", "tab": "LHDW (Lighthaven)",
        "entity_col": "B", "pct_col": "E", "contrib_col": "C",
        "start": 26, "end": 29, "equity_class": "preferred",
    },
]


# Maps a canonical cap-table project name to the candidate names it may appear
# under in the Project Returns data. Matching is case-insensitive on a
# normalized (alnum-only) form; the first candidate that resolves wins. This is
# config-driven on purpose: the exact live returns names are verified on
# deploy, and adding/renaming a project is a one-line edit here.
# Full Returns names go FIRST so the exact-match pass wins before the loose
# containment fallback can grab a same-token sibling. Notably the cap-table
# "GPD" tab is Grand Prairie *Development*, NOT "Grand Prairie East (CCI)" —
# those are two separate deals and only Development has a cap table.
PROJECT_ALIASES: dict[str, list[str]] = {
    "Grand Prairie":  ["Grand Prairie Development", "GPD"],
    "Windrose Green": ["Windrose Green", "WRG"],
    "Dennison":       ["Dennison", "WRRD"],
    "Lexington Land": ["Lexington Land Partners", "Lexington Land", "Lexington", "LLP"],
    "Hawthorne":      ["The Hawthorne", "Hawthorne", "SFC"],
    "Mid Main":       ["Mid Main East Commons (Land)", "Mid Main East Commons", "Mid Main East", "Mid Main", "MMEC"],
    # LightHaven is two standalone returns deals now; wire each cap-table block
    # to its own project (common -> Common Equity, preferred -> Preferred Equity).
    "LightHaven Common":    ["LightHaven (Common Equity)", "Lighthaven Common"],
    "LightHaven Preferred": ["LightHaven (Preferred Equity)", "Lighthaven Preferred"],
}


def _num(val: Any) -> float:
    """Cell value -> float. None / non-numeric -> 0.0."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0


def _str(val: Any) -> str:
    """Cell value -> trimmed string. None -> ''."""
    return "" if val is None else str(val).strip()


def normalize_project(name: str) -> str:
    """Lowercase, alnum-only key for fuzzy project-name matching."""
    return "".join(ch for ch in (name or "").lower() if ch.isalnum())


def resolve_returns_project(captable_project: str, returns_names: list[str]) -> str | None:
    """Best-effort match of a cap-table project to a Project Returns name.

    Tries each alias (and the canonical name) against the normalized returns
    names. Returns the matching returns name, or None when nothing matches
    (those positions are ignored per spec — match by project name only).
    """
    candidates = [captable_project] + PROJECT_ALIASES.get(captable_project, [])
    norm_returns = {normalize_project(n): n for n in returns_names}
    for cand in candidates:
        key = normalize_project(cand)
        if key in norm_returns:
            return norm_returns[key]
    # Loose containment fallback (e.g. "Mid Main" inside "Mid Main Phase I").
    for cand in candidates:
        key = normalize_project(cand)
        if not key:
            continue
        for nk, original in norm_returns.items():
            if key in nk or nk in key:
                return original
    return None


def parse_captable_workbook(source: Any) -> list[dict]:
    """Parse the cap-table workbook into a flat list of position rows.

    `source` is anything openpyxl.load_workbook accepts (path or file-like).
    Rows whose entity cell is blank, or starts with "Total", or appears in the
    block's exclude list, are skipped. Percentages are kept as fractions (0..1).
    """
    wb = openpyxl.load_workbook(source, data_only=True)
    positions: list[dict] = []

    for spec in CAPTABLE_SPEC:
        tab = spec["tab"]
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        ecol, pcol, ccol = spec["entity_col"], spec["pct_col"], spec["contrib_col"]
        equity_class = spec.get("equity_class", "common")
        exclude = {e.strip().lower() for e in spec.get("exclude", [])}

        for r in range(spec["start"], spec["end"] + 1):
            vehicle = _str(ws["%s%d" % (ecol, r)].value)
            if not vehicle or vehicle.lower().startswith("total"):
                continue
            if vehicle.lower() in exclude:
                continue
            pct = _num(ws["%s%d" % (pcol, r)].value)
            contribution = _num(ws["%s%d" % (ccol, r)].value)
            # A position with neither ownership nor contribution is a stray /
            # spacer row — skip it.
            if pct == 0.0 and contribution == 0.0:
                continue
            positions.append({
                "project": spec["project"],
                "vehicle": vehicle,
                "contribution": round(contribution, 2),
                "pct": pct,
                "equity_class": equity_class,
            })

    return positions


if __name__ == "__main__":  # pragma: no cover — manual offline check
    import sys
    src = sys.argv[1] if len(sys.argv) > 1 else \
        r"C:\Users\CarlosSaldierna\Downloads\2026.05.06 All Cap Tables.xlsx"
    rows = parse_captable_workbook(src)
    by_proj: dict[str, list[dict]] = {}
    for row in rows:
        by_proj.setdefault(row["project"], []).append(row)
    for proj, rws in by_proj.items():
        tot = sum(r["pct"] for r in rws)
        print("\n{} - {} positions, pct sum={:.4f}".format(proj, len(rws), tot))
        for r in rws:
            print("   {:<40} {:8.4f}%  ${:>14,.0f}  [{}]".format(
                r["vehicle"], r["pct"] * 100, r["contribution"], r["equity_class"]))
    print("\nTOTAL positions:", len(rows))
