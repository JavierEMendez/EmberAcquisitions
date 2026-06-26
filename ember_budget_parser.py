"""
ember_budget_parser.py — Parse the "EMBER Budget" workbook (the firm's
corporate cash-flow forecast) into the JSON that drives the /budget page and,
via the Ember DB `reports` table (report_type='ember_budget'), the Maquina
dashboard's Ember company page.

The workbook is a single sheet ("EMBER CF - FORECAST …") laid out as a monthly
forecast. Section 1 ("EMBER COMPANIES CASH FLOW") is the firm P&L / cashflow we
care about:

    Revenue            WRG · GPD · DEN · CCI            → Total Revenue
    Costs
      People (E + M)                                    → Total People
      Operations       Company Insurance · Rent · …     → Total Operations
                                                        → Total Costs
    Net Income   ( = Total Revenue − Total Costs )
    Cash Flow
    Cummulative Cash Flow

Months run across the columns (year on the row above the month names). We read
the axis dynamically (find the month-name row, forward-fill the year above it),
so a re-upload covering a different window still imports. Broken model cells
(#REF! / #VALUE! / blanks) collapse to 0.0 via _num.

We only parse the first section (stop at "EMBER Partners") — partner capital
flows are a separate concern and would pollute the firm P&L.
"""
from __future__ import annotations

import io
import re
from typing import Optional

import openpyxl

_MONTHS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}
_MABBR = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _num(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        try:
            f = float(v)
            return f if f == f else 0.0
        except (TypeError, ValueError, OverflowError):
            return 0.0
    s = str(v).strip()
    if not s or s.startswith("#"):
        return 0.0
    s = s.replace(",", "").replace("$", "")
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    try:
        return float(s)
    except ValueError:
        return 0.0


def _label(v) -> str:
    return "" if v is None else str(v).strip()


def _month_num(v) -> Optional[int]:
    key = _label(v).lower()[:3]
    return _MONTHS.get(key)


def _find_month_row(ws, scan_rows: int = 8):
    """Return (month_row, {col: (year, month)}). The month row is the first row
    holding >= 6 month names; the year row is the one directly above it, with
    years forward-filled left→right across the month columns."""
    for r in range(1, scan_rows + 1):
        cols = {}
        for c in range(1, ws.max_column + 1):
            mn = _month_num(ws.cell(row=r, column=c).value)
            if mn:
                cols[c] = mn
        if len(cols) >= 6:
            # forward-fill years from the row above
            year_row = r - 1
            cur_year = None
            col_period = {}
            for c in range(1, ws.max_column + 1):
                yv = _label(ws.cell(row=year_row, column=c).value)
                ym = re.search(r"(20\d{2})", yv)
                if ym:
                    cur_year = int(ym.group(1))
                if c in cols and cur_year:
                    col_period[c] = (cur_year, cols[c])
            # Fallback: if a month col has no year (forward-fill missed), infer
            # by continuity — a month number <= the previous one rolls the year.
            if col_period:
                ordered = sorted(cols)
                prev_m = None
                yr = None
                for c in ordered:
                    if c in col_period:
                        yr, prev_m = col_period[c][0], col_period[c][1]
                        break
                for c in ordered:
                    if c in col_period:
                        yr, prev_m = col_period[c]
                        continue
                    if yr is None:
                        continue
                    m = cols[c]
                    if prev_m is not None and m <= prev_m:
                        yr += 1
                    col_period[c] = (yr, m)
                    prev_m = m
            return r, col_period
    return None, {}


def _series(ws, row: int, col_period: dict) -> dict:
    """{months: {'YYYY-MM': v}, by_year: {year: total}} for one sheet row."""
    months, by_year = {}, {}
    for c, (yr, mo) in col_period.items():
        v = _num(ws.cell(row=row, column=c).value)
        if v:
            months[f"{yr}-{mo:02d}"] = round(v, 2)
        by_year[yr] = round(by_year.get(yr, 0.0) + v, 2)
    return {"months": months, "by_year": by_year}


def parse_ember_budget(file_bytes: bytes, filename: str = "") -> dict:
    """Parse the workbook bytes → budget JSON. Raises ValueError if the sheet
    layout isn't recognized."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
    ws = wb[wb.sheetnames[0]]   # single-sheet forecast workbook

    month_row, col_period = _find_month_row(ws)
    if not col_period:
        raise ValueError("Couldn't find the monthly header row — is this the "
                         "standard EMBER Budget forecast file?")

    # Ordered month period keys + display labels.
    ordered_cols = sorted(col_period)
    months = [f"{col_period[c][0]}-{col_period[c][1]:02d}" for c in ordered_cols]
    month_labels = [f"{_MABBR[col_period[c][1]]} '{str(col_period[c][0])[2:]}"
                    for c in ordered_cols]
    years = sorted({yr for (yr, _m) in col_period.values()})

    # Walk the first section (rows after the month header, stop at partners).
    # Label can be in col B (sections/totals) or col C (operations sub-lines).
    def lbl(r, col):
        return _label(ws.cell(row=r, column=col).value)

    start = month_row + 1
    end = ws.max_row
    for r in range(start, ws.max_row + 1):
        if lbl(r, 2).lower().startswith("ember partner"):
            end = r - 1
            break

    revenue_lines, ops_lines = [], []
    revenue_total = people = ops_total = total_costs = None
    net_income = cash_flow = cumulative = None
    section = None
    for r in range(start, end + 1):
        b = lbl(r, 2)
        cc = lbl(r, 3)
        bl = b.lower()
        if bl == "revenue":
            section = "revenue"; continue
        if bl == "costs":
            section = "costs"; continue
        if bl.startswith("operations"):
            section = "operations"; continue
        if bl.startswith("total revenue"):
            revenue_total = _series(ws, r, col_period); section = None; continue
        if bl.startswith("total people"):
            people = _series(ws, r, col_period); continue
        if bl.startswith("total operations"):
            ops_total = _series(ws, r, col_period); section = None; continue
        if bl.startswith("total costs"):
            total_costs = _series(ws, r, col_period); continue
        if bl.startswith("net income"):
            net_income = _series(ws, r, col_period); continue
        if bl.startswith("cummulative") or bl.startswith("cumulative"):
            if cumulative is None:                 # first occurrence = company CF
                cumulative = _series(ws, r, col_period)
            continue
        if bl.startswith("cash flow"):
            if cash_flow is None:
                cash_flow = _series(ws, r, col_period)
            continue
        # detail lines
        if section == "revenue" and b:
            revenue_lines.append({"name": b, **_series(ws, r, col_period)})
        elif section == "operations" and cc:
            ops_lines.append({"name": cc, **_series(ws, r, col_period)})

    def _by(series, yr):
        return round((series or {}).get("by_year", {}).get(yr, 0.0), 2) if series else 0.0

    def _life(series):
        return round(sum((series or {}).get("by_year", {}).values()), 2) if series else 0.0

    last_year = years[-1] if years else None
    first_year = years[0] if years else None
    kpis = {
        "years": years,
        "horizon_label": (f"{month_labels[0]} – {month_labels[-1]}" if month_labels else ""),
        "revenue_life": _life(revenue_total),
        "people_life": _life(people),
        "operations_life": _life(ops_total),
        "costs_life": _life(total_costs),
        "net_income_life": _life(net_income),
        "cash_flow_life": _life(cash_flow),
        "revenue_by_year": (revenue_total or {}).get("by_year", {}),
        "net_income_by_year": (net_income or {}).get("by_year", {}),
        "cash_flow_by_year": (cash_flow or {}).get("by_year", {}),
        "people_by_year": (people or {}).get("by_year", {}),
        "operations_by_year": (ops_total or {}).get("by_year", {}),
        "first_year": first_year,
        "last_year": last_year,
    }

    # Report date from filename (e.g. "EMBER Budget May 2026").
    report_date = ""
    m = re.search(r"([A-Za-z]{3,9})\s+(\d{4})", filename or "")
    if m:
        report_date = f"{m.group(1)} {m.group(2)}"

    empty = {"months": {}, "by_year": {}}
    return {
        "meta": {
            "currency": "USD", "filename": filename, "report_date": report_date,
            "months": months, "month_labels": month_labels, "years": years,
        },
        "revenue": {"lines": revenue_lines, "total": revenue_total or empty},
        "people": people or empty,
        "operations": {"lines": ops_lines, "total": ops_total or empty},
        "total_costs": total_costs or empty,
        "net_income": net_income or empty,
        "cash_flow": cash_flow or empty,
        "cumulative": cumulative or empty,
        "kpis": kpis,
    }


if __name__ == "__main__":
    import sys, json
    path = sys.argv[1] if len(sys.argv) > 1 else \
        r"C:/Users/Javier/Downloads/EMBER Budget May 2026.xlsx"
    with open(path, "rb") as f:
        d = parse_ember_budget(f.read(), filename=path.split("/")[-1])
    m = d["meta"]
    print("report:", m["report_date"], "| currency:", m["currency"])
    print("months:", m["month_labels"][0], "→", m["month_labels"][-1], f"({len(m['months'])})")
    print("years:", m["years"])
    print("\nrevenue lines:", [r["name"] for r in d["revenue"]["lines"]])
    print("operations lines:", [r["name"] for r in d["operations"]["lines"]])
    print("\n-- by year --")
    for lbl_, s in [("Revenue", d["revenue"]["total"]), ("People", d["people"]),
                    ("Operations", d["operations"]["total"]), ("Total Costs", d["total_costs"]),
                    ("Net Income", d["net_income"]), ("Cash Flow", d["cash_flow"])]:
        print(f"  {lbl_:14s} " + " | ".join(f"{y}:{s['by_year'].get(y,0):>12,.0f}" for y in m["years"]))
    print("\nkpis horizon:", d["kpis"]["horizon_label"])
    print("net_income_by_year:", d["kpis"]["net_income_by_year"])
