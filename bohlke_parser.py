"""
bohlke_parser.py — Parses the Bohlke competitive-set Excel file
(e.g. "Ember - The Grand Prairie Competitive Set thru Q4-2025.xlsx")
and returns structured JSON for the Bohlke Report dashboard tab.

The file contains three sheets we care about:

  • YEARLY TOTALS — annual Waller ISD permit totals, avg size/price/PPSF,
    30yr rate and notable events.
  • COMPS         — community-level permit counts and 2025 size/price/PPSF.
                    Communities are grouped under ISD headers (e.g.
                    "WALLER ISD", "KATY ISD"). We only keep rows under
                    the "WALLER ISD" header.
  • MPCS          — builder-level permit counts grouped under MPC headers
                    (e.g. "GRAND PRAIRIE", "BRIDGELAND"). We only keep
                    builder rows under the "GRAND PRAIRIE" header.

Ported from the coworker's bulkImportExcel() JS at
_sales_dashboard_ref/index.html lines 2825-2913.
"""

from __future__ import annotations

import io
from typing import Any

import openpyxl


# ── Helpers ─────────────────────────────────────────────────────────────────

def _s(val: Any) -> str:
    return "" if val is None else str(val).strip()


def _num(val: Any) -> float | None:
    """Coerce to float; return None for blanks / non-numeric."""
    if val is None or val == "":
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _int_or_none(val: Any) -> int | None:
    n = _num(val)
    return None if n is None else int(round(n))


def _round2(val: Any) -> float | None:
    n = _num(val)
    return None if n is None else round(n, 2)


# ── Sheet parsers ───────────────────────────────────────────────────────────

def _parse_yearly(ws) -> list[dict]:
    """
    YEARLY TOTALS — row 3+ contains per-year data.
      col A (0): year
      col B (1): total permits
      col D (3): avg home size
      col F (5): avg base price
      col H (7): avg PPSF
      col J (9): 30yr rate (stored as fraction, e.g. 0.065)
      col K (10): events
    """
    rows = list(ws.iter_rows(values_only=True))
    out: list[dict] = []
    # Match JS: "for (let i = 2; i < rows.length; i++)" — skip first two rows
    for i in range(2, len(rows)):
        r = rows[i]
        if not r or not r[0]:
            continue
        year = _int_or_none(r[0])
        if year is None:
            continue
        rate = _num(r[9]) if len(r) > 9 else None
        out.append({
            "year":     year,
            "total":    _int_or_none(r[1])               if len(r) > 1  else None,
            "avgSize":  _int_or_none(r[3])               if len(r) > 3  else None,
            "avgPrice": _int_or_none(r[5])               if len(r) > 5  else None,
            "avgPPSF":  _round2(r[7])                    if len(r) > 7  else None,
            "rate30yr": round(rate * 100, 2) if rate else None,
            "events":   _s(r[10]) or None                if len(r) > 10 else None,
        })
    return out


def _parse_comps(ws) -> list[dict]:
    """
    COMPS — community rows grouped by ISD headers.
      col A (0): community name OR section marker like "WALLER ISD"
      col B (1): MPC type (defaults to "NON")
      col C (2): permits 2023
      col D (3): permits 2024
      col E (4): permits 2025
      col H (7): avg size 2025
      col K (10): avg base price 2025
      col N (13): avg PPSF 2025
    """
    rows = list(ws.iter_rows(values_only=True))
    out: list[dict] = []
    in_waller = False
    for i in range(2, len(rows)):
        r = rows[i]
        if not r:
            continue
        name = _s(r[0]) if len(r) > 0 else ""
        # Section transitions
        if name == "WALLER ISD":
            in_waller = True
            continue
        if " ISD" in name and name != "WALLER ISD":
            in_waller = False
            continue
        if not in_waller or not name:
            continue
        # Skip rows with no permit data across 2023-2025
        p23 = _num(r[2]) if len(r) > 2 else None
        p24 = _num(r[3]) if len(r) > 3 else None
        p25 = _num(r[4]) if len(r) > 4 else None
        if not p23 and not p24 and not p25:
            continue
        out.append({
            "community": name,
            "mpc":       (_s(r[1]) if len(r) > 1 and _s(r[1]) else "NON"),
            "p23":       int(p23) if p23 else 0,
            "p24":       int(p24) if p24 else 0,
            "p25":       int(p25) if p25 else 0,
            "size25":    _int_or_none(r[7])  if len(r) > 7  else None,
            "price25":   _int_or_none(r[10]) if len(r) > 10 else None,
            "ppsf25":    _round2(r[13])      if len(r) > 13 else None,
        })
    return out


def _parse_tgp_builders(ws) -> list[dict]:
    """
    MPCS — builder rows grouped by MPC name in col A.
      When col A contains "GRAND PRAIRIE" we enter the TGP section.
      Inside the section:
        col D (3): builder name
        col G (6): 2025 permits
        col P (15): PPSF
      Section ends when we hit a new non-empty col A with no col D value.
    Permits are summed per builder and PPSF is permit-weighted.
    """
    rows = list(ws.iter_rows(values_only=True))
    in_tgp = False
    bt: dict[str, dict] = {}
    for i in range(2, len(rows)):
        r = rows[i]
        if not r:
            continue
        name = _s(r[0]).upper() if len(r) > 0 else ""
        builder = _s(r[3]) if len(r) > 3 else ""
        # Enter TGP section
        if "GRAND PRAIRIE" in name:
            in_tgp = True
            continue
        # Exit when a new section header appears (col A set, col D blank)
        if in_tgp and name and not builder:
            in_tgp = False
            continue
        if not in_tgp or not builder or "Total" in builder:
            continue
        p25 = _num(r[6]) if len(r) > 6 else None
        ppsf = _num(r[15]) if len(r) > 15 else None
        if not p25 or p25 <= 0:
            continue
        if builder not in bt:
            bt[builder] = {"builder": builder, "permits": 0, "_ps": 0.0, "_pc": 0}
        bt[builder]["permits"] += int(p25)
        if ppsf:
            bt[builder]["_ps"] += ppsf * p25
            bt[builder]["_pc"] += p25

    result = []
    for b in bt.values():
        avg = round(b["_ps"] / b["_pc"], 2) if b["_pc"] else None
        result.append({"builder": b["builder"], "permits": b["permits"], "avgPPSF": avg})
    result.sort(key=lambda x: x["permits"], reverse=True)
    return result


# ── Entry point ─────────────────────────────────────────────────────────────

def parse_bohlke(file_bytes: bytes) -> dict:
    """
    Parse a Bohlke competitive-set xlsx and return:
      {
        "yearly":       [ {year, total, avgSize, avgPrice, avgPPSF, rate30yr, events}, ... ],
        "comps":        [ {community, mpc, p23, p24, p25, size25, price25, ppsf25}, ... ],
        "tgp_builders": [ {builder, permits, avgPPSF}, ... ],
        # ── Phase 4a extended fields (added when sheets are present) ──
        "all_isd_comps":     [ {isd, community, mpc, p23/p24/p25, size/price/ppsf for each year}, ... ],
        "isd_totals":        [ {isd, "2010": int, "2011": int, ..., "2025": int}, ... ],
        "waller_comps":      [ {community, mpc, builder, p23/24/25, size/price/ppsf for each}, ... ],
        "tgp_plans":         [ {builder, plan, sizeBand, homeSize, config, homeWidth,
                                avgPrice, avgPPSF, total2025, q1, q2, q3, q4}, ... ],
        "all_mpcs":          [ {isd, community, lotSize, projType, builder, p23/24/25, ...}, ... ],
        "waller_size_dist":  [ {community, mpc, builder, total, bands: {bandLabel: count}}, ... ],
        "waller_price_dist": [ {community, mpc, builder, total, bands: {bandLabel: count}}, ... ],
        "waller_top_config": [ {sizeBand, config, total, priceBands: {bandLabel: count}}, ... ],
      }

    The core 3 fields (yearly / comps / tgp_builders) are kept by the
    original openpyxl path for backward compatibility. The extended
    fields use pandas to mirror the coworker's extract_bulk_report
    extraction. Each extended sheet is wrapped in try/except so an
    older xlsx without those sheets still parses successfully.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)

    data: dict[str, Any] = {"yearly": [], "comps": [], "tgp_builders": []}

    if "YEARLY TOTALS" in wb.sheetnames:
        data["yearly"] = _parse_yearly(wb["YEARLY TOTALS"])
    if "COMPS" in wb.sheetnames:
        data["comps"] = _parse_comps(wb["COMPS"])
    if "MPCS" in wb.sheetnames:
        data["tgp_builders"] = _parse_tgp_builders(wb["MPCS"])

    wb.close()

    # ── Extended extraction (pandas) ─────────────────────────────────
    # Mirrors update_dashboard.extract_bulk_report logic. pandas is
    # already in requirements.txt. Per-sheet try/except so a thinner
    # xlsx (e.g. one without WALLER 2025 Size) still returns the
    # base 3 datasets without erroring.
    try:
        import pandas as pd  # noqa: F401
        _enrich_with_pandas(file_bytes, data)
    except Exception as e:  # pragma: no cover
        # Pandas missing or unexpected error — keep base data only.
        import logging
        logging.warning("bohlke_parser: extended extraction skipped: %s", e)

    return data


def _enrich_with_pandas(file_bytes: bytes, result: dict) -> None:
    """Add the extended datasets (ISD roll-ups, dist bands, etc.).
    Mutates `result` in place. Each sheet wrapped in try/except so
    a missing/renamed sheet doesn't kill the whole extraction."""
    import pandas as pd
    buf = io.BytesIO(file_bytes)

    # ALL ISD COMPS — community-level for every ISD (not just Waller).
    # Drives the cross-ISD search table on the Bohlke tab.
    result.setdefault("all_isd_comps", [])
    try:
        df = pd.read_excel(buf, sheet_name="COMPS", header=None)
        current_isd = None
        for i in range(2, len(df)):
            name = str(df.iloc[i][0]).strip() if pd.notna(df.iloc[i][0]) else ""
            if name.endswith(" ISD") and "Total" not in name:
                current_isd = name; continue
            if "Total" in name:
                continue
            if current_isd and name:
                row = df.iloc[i]
                # Skip rows with no permit counts at all
                if not any(pd.notna(row[j]) for j in (2, 3, 4)):
                    continue
                result["all_isd_comps"].append({
                    "isd":       current_isd,
                    "community": name,
                    "mpc":       (str(row[1]).strip() if pd.notna(row[1]) else "NON"),
                    "p23":       int(row[2]) if pd.notna(row[2]) else 0,
                    "p24":       int(row[3]) if pd.notna(row[3]) else 0,
                    "p25":       int(row[4]) if pd.notna(row[4]) else 0,
                    "size23":    _safe_round(row[5], 0),
                    "size24":    _safe_round(row[6], 0),
                    "size25":    _safe_round(row[7], 0),
                    "price23":   _safe_round(row[8], 0),
                    "price24":   _safe_round(row[9], 0),
                    "price25":   _safe_round(row[10], 0),
                    "ppsf23":    _safe_round(row[11], 2),
                    "ppsf24":    _safe_round(row[12], 2),
                    "ppsf25":    _safe_round(row[13], 2),
                })
    except Exception:
        pass

    # ISD TOTALS — per-ISD yearly permit trends. Drives the ISD-vs-ISD chart.
    result.setdefault("isd_totals", [])
    try:
        buf.seek(0)
        df = pd.read_excel(buf, sheet_name="ISD TOTALS", header=None)
        for i in range(2, len(df)):
            isd_name = str(df.iloc[i][0]).strip() if pd.notna(df.iloc[i][0]) else ""
            if (not isd_name) or "Total" in isd_name or "PPSF" in isd_name:
                continue
            row = df.iloc[i]
            entry: dict = {"isd": isd_name}
            for j in range(1, min(17, len(row))):
                yr = 2009 + j
                entry[str(yr)] = int(row[j]) if pd.notna(row[j]) else 0
            result["isd_totals"].append(entry)
    except Exception:
        pass

    # WALLER COMPS — per-community/builder breakdown for Waller ISD only.
    result.setdefault("waller_comps", [])
    try:
        buf.seek(0)
        df = pd.read_excel(buf, sheet_name="WALLER COMPS", header=None)
        for i in range(2, len(df)):
            row = df.iloc[i]
            comm = str(row[0]).strip() if pd.notna(row[0]) else ""
            if (not comm) or "Total" in comm or "ISD" in comm:
                continue
            builder = str(row[2]).strip() if pd.notna(row[2]) else ""
            if not builder:
                continue
            result["waller_comps"].append({
                "community": comm,
                "mpc":       (str(row[1]).strip() if pd.notna(row[1]) else "NON"),
                "builder":   builder,
                "p23":       int(row[3]) if pd.notna(row[3]) else 0,
                "p24":       int(row[4]) if pd.notna(row[4]) else 0,
                "p25":       int(row[5]) if pd.notna(row[5]) else 0,
                "size23":    _safe_round(row[6], 0),
                "size24":    _safe_round(row[7], 0),
                "size25":    _safe_round(row[8], 0),
                "price23":   _safe_round(row[9], 0),
                "price24":   _safe_round(row[10], 0),
                "price25":   _safe_round(row[11], 0),
                "ppsf23":    _safe_round(row[12], 2),
                "ppsf24":    _safe_round(row[13], 2),
                "ppsf25":    _safe_round(row[14], 2),
            })
    except Exception:
        pass

    # TGP 2025 Plans — builder × plan with size/price/quarterly counts.
    result.setdefault("tgp_plans", [])
    try:
        buf.seek(0)
        df = pd.read_excel(buf, sheet_name="TGP 2025 Plans", header=None)
        for i in range(2, len(df)):
            row = df.iloc[i]
            builder = str(row[0]).strip() if pd.notna(row[0]) else ""
            plan    = str(row[1]).strip() if pd.notna(row[1]) else ""
            if not builder or "Total" in builder or not plan:
                continue
            result["tgp_plans"].append({
                "builder":   builder,
                "plan":      plan,
                "sizeBand":  str(row[2]).strip() if pd.notna(row[2]) else "",
                "homeSize":  _safe_round(row[3], 0),
                "config":    str(row[4]).strip() if pd.notna(row[4]) else "",
                "homeWidth": str(row[5]).strip() if pd.notna(row[5]) else "",
                "avgPrice":  _safe_round(row[6], 0),
                "avgPPSF":   _safe_round(row[7], 2),
                "total2025": int(row[8]) if pd.notna(row[8]) else 0,
                "q1":        int(row[9]) if pd.notna(row[9]) else 0,
                "q2":        int(row[10]) if pd.notna(row[10]) else 0,
                "q3":        int(row[11]) if pd.notna(row[11]) else 0,
                "q4":        int(row[12]) if pd.notna(row[12]) else 0,
            })
    except Exception:
        pass

    # WALLER 2025 Size — home-size distribution bands by community+builder.
    result.setdefault("waller_size_dist", [])
    try:
        buf.seek(0)
        df = pd.read_excel(buf, sheet_name="WALLER 2025 Size", header=None)
        size_bands = [
            str(df.iloc[0][j]).strip() if pd.notna(df.iloc[0][j]) else ""
            for j in range(4, min(20, len(df.columns)))
        ]
        for i in range(1, len(df)):
            row = df.iloc[i]
            comm = str(row[0]).strip() if pd.notna(row[0]) else ""
            if (not comm) or "Total" in comm or "SFR" in comm or " TH" in comm:
                continue
            builder = str(row[2]).strip() if pd.notna(row[2]) else ""
            total   = int(row[3]) if pd.notna(row[3]) else 0
            if total == 0:
                continue
            bands: dict = {}
            for j, band in enumerate(size_bands):
                if band:
                    val = int(row[4 + j]) if pd.notna(row[4 + j]) else 0
                    if val:
                        bands[band] = val
            result["waller_size_dist"].append({
                "community": comm,
                "mpc":       str(row[1]).strip() if pd.notna(row[1]) else "",
                "builder":   builder,
                "total":     total,
                "bands":     bands,
            })
    except Exception:
        pass

    # WALLER 2025 Price — price-distribution bands by community+builder.
    result.setdefault("waller_price_dist", [])
    try:
        buf.seek(0)
        df = pd.read_excel(buf, sheet_name="WALLER 2025 Price", header=None)
        price_bands = [
            str(df.iloc[0][j]).strip() if pd.notna(df.iloc[0][j]) else ""
            for j in range(4, min(23, len(df.columns)))
        ]
        for i in range(1, len(df)):
            row = df.iloc[i]
            comm = str(row[0]).strip() if pd.notna(row[0]) else ""
            if (not comm) or "Total" in comm or "SFR" in comm or " TH" in comm:
                continue
            builder = str(row[2]).strip() if pd.notna(row[2]) else ""
            total   = int(row[3]) if pd.notna(row[3]) else 0
            if total == 0:
                continue
            bands: dict = {}
            for j, band in enumerate(price_bands):
                if band:
                    val = int(row[4 + j]) if pd.notna(row[4 + j]) else 0
                    if val:
                        bands[band] = val
            result["waller_price_dist"].append({
                "community": comm,
                "mpc":       str(row[1]).strip() if pd.notna(row[1]) else "",
                "builder":   builder,
                "total":     total,
                "bands":     bands,
            })
    except Exception:
        pass

    # WALLER 2025 Top Config — sizeBand × config × priceBands matrix.
    result.setdefault("waller_top_config", [])
    try:
        buf.seek(0)
        df = pd.read_excel(buf, sheet_name="WALLER 2025 Top Config", header=None)
        price_bands = [
            str(df.iloc[0][j]).strip() if pd.notna(df.iloc[0][j]) else ""
            for j in range(3, min(28, len(df.columns)))
        ]
        for i in range(1, len(df)):
            row = df.iloc[i]
            size_band = str(row[0]).strip() if pd.notna(row[0]) else ""
            if not size_band:
                continue
            config = str(row[1]).strip() if pd.notna(row[1]) else ""
            total  = int(row[2]) if pd.notna(row[2]) else 0
            if total == 0:
                continue
            bands: dict = {}
            for j, band in enumerate(price_bands):
                if band:
                    val = int(row[3 + j]) if pd.notna(row[3 + j]) else 0
                    if val:
                        bands[band] = val
            result["waller_top_config"].append({
                "sizeBand":   size_band,
                "config":     config,
                "total":      total,
                "priceBands": bands,
            })
    except Exception:
        pass

    # ALL MPCS — full builder-level community breakdown across every ISD.
    # Used for both the TGP builder summary (already extracted by openpyxl)
    # and the cross-ISD builder roll-ups in the Bohlke tab.
    result.setdefault("all_mpcs", [])
    try:
        buf.seek(0)
        df = pd.read_excel(buf, sheet_name="MPCS", header=None)
        current_isd: str | None = None
        current_community: str | None = None
        for i in range(2, len(df)):
            row = df.iloc[i]
            name = str(row[0]).strip() if pd.notna(row[0]) else ""
            if name.endswith(" ISD") and "Total" not in name:
                current_isd = name; current_community = None; continue
            if "Total" in name:
                continue
            if name and current_isd:
                current_community = name; continue
            if current_community and current_isd:
                builder = str(row[3]).strip() if pd.notna(row[3]) else ""
                lot_size = str(row[1]).strip() if pd.notna(row[1]) else ""
                proj_type = str(row[2]).strip() if pd.notna(row[2]) else ""
                if (not builder) or "Total" in builder:
                    continue
                p23 = int(row[4]) if pd.notna(row[4]) else 0
                p24 = int(row[5]) if pd.notna(row[5]) else 0
                p25 = int(row[6]) if pd.notna(row[6]) else 0
                if p23 == 0 and p24 == 0 and p25 == 0:
                    continue
                result["all_mpcs"].append({
                    "isd":       current_isd,
                    "community": current_community,
                    "lotSize":   lot_size,
                    "projType":  proj_type,
                    "builder":   builder,
                    "p23":       p23, "p24": p24, "p25": p25,
                    "size23":    _safe_round(row[7], 0),
                    "size24":    _safe_round(row[8], 0),
                    "size25":    _safe_round(row[9], 0),
                    "price23":   _safe_round(row[10], 0),
                    "price24":   _safe_round(row[11], 0),
                    "price25":   _safe_round(row[12], 0),
                    "ppsf23":    _safe_round(row[13], 2),
                    "ppsf24":    _safe_round(row[14], 2),
                    "ppsf25":    _safe_round(row[15], 2),
                })
    except Exception:
        pass


def _safe_round(val, ndigits):
    """pandas-safe rounder: returns None on NaN/missing instead of raising."""
    try:
        import pandas as pd
        if pd.isna(val):
            return None
        return round(float(val), ndigits)
    except (TypeError, ValueError):
        return None
