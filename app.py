"""
Ember Tract Underwriting Web App
Flask + PostgreSQL + Flask-Login — no Excel required
"""
import os, re, html, json, datetime, io, base64, requests, threading, concurrent.futures
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail, Attachment, FileContent, FileName, FileType, Disposition, Content
from functools import wraps
from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file, Response
import psycopg2
import psycopg2.extras
from werkzeug.security import generate_password_hash, check_password_hash
from calc import calculate
from report_parser import parse_dashboard
from macro_parser import parse_macro
from data_puller import run_pull
from sales_parser import (
    get_sales_dashboard_data,
    pipsy_home_sales_by_section,
    pipsy_lot_takedowns_by_section,
)
from bohlke_parser import parse_bohlke
from waller_parser import parse_waller_monthly
from hpermits_parser import parse_hpermits
from uw_parser import parse_uw
import sage_intacct
from frp_mapping import roll_up_balance_sheet, summarize_bs
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "ember-dev-secret-change-in-production")

# ── Macro data refresh state ───────────────────────────────────────────────────
_refresh_lock = threading.Lock()
_refresh_state = {"running": False, "last_started": None, "last_finished": None, "last_error": None}

def _do_macro_refresh():
    """Background thread: pull all macro data, parse, and store in DB."""
    with _refresh_lock:
        if _refresh_state["running"]:
            return
        _refresh_state["running"] = True
        _refresh_state["last_started"] = datetime.datetime.utcnow().isoformat()
        _refresh_state["last_error"] = None

    print("[macro_refresh] Starting data pull…", flush=True)
    try:
        excel_bytes = run_pull()
        print("[macro_refresh] Pull complete, parsing…", flush=True)
        data = parse_macro(excel_bytes)
        conn = get_db()
        cur = conn.cursor()
        cur.execute("DELETE FROM reports WHERE report_type = 'macro'")
        cur.execute("INSERT INTO reports (report_type, data, uploaded_by) VALUES (%s, %s, %s)",
                    ("macro", json.dumps(data), None))
        conn.commit()
        cur.close()
        conn.close()
        with _refresh_lock:
            _refresh_state["last_finished"] = datetime.datetime.utcnow().isoformat()
            _refresh_state["running"] = False
        print("[macro_refresh] Done — data stored.", flush=True)
    except Exception as e:
        print(f"[macro_refresh] ERROR: {e}", flush=True)
        with _refresh_lock:
            _refresh_state["last_error"] = str(e)
            _refresh_state["running"] = False

# Schedule: 1st of every month at 02:00 UTC
_scheduler = BackgroundScheduler(daemon=True)
_scheduler.add_job(_do_macro_refresh, CronTrigger(day=1, hour=2, minute=0), id="macro_monthly")

def _purge_old_activity():
    """Delete activity_log rows older than 12 months. Runs daily.
    Bounded retention keeps the table small (small team × ~50 page
    views/day = a few hundred thousand rows max) and matches the
    privacy intent of internal-only analytics."""
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("DELETE FROM activity_log WHERE ts < NOW() - INTERVAL '12 months'")
        conn.commit()
        cur.close(); conn.close()
    except Exception as e:
        print(f"[activity_purge] ERROR: {e}", flush=True)

_scheduler.add_job(_purge_old_activity, CronTrigger(hour=3, minute=15), id="activity_retention_daily")
_scheduler.start()

# Auto-initialize DB on first request
_db_initialized = False

@app.before_request
def auto_init():
    global _db_initialized
    if not _db_initialized:
        try:
            init_db()
            _db_initialized = True
        except Exception as e:
            print(f"DB init error: {e}")

# ─── DATABASE ────────────────────────────────────────────────────────────────
def get_db():
    conn = psycopg2.connect(os.environ["DATABASE_URL"], cursor_factory=psycopg2.extras.RealDictCursor)
    return conn

def init_db():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            is_admin BOOLEAN DEFAULT FALSE,
            page_access JSONB NOT NULL DEFAULT '{"mpc_underwriting":true,"returns":true,"loans":true,"operations":true}'::jsonb,
            created_at TIMESTAMP DEFAULT NOW()
        );
        -- Add columns if upgrading from older schema
        ALTER TABLE users ADD COLUMN IF NOT EXISTS page_access JSONB NOT NULL DEFAULT '{"mpc_underwriting":true,"returns":true,"loans":true,"operations":true}'::jsonb;
        ALTER TABLE users ADD COLUMN IF NOT EXISTS email TEXT;
        ALTER TABLE users ADD COLUMN IF NOT EXISTS report_opt_in BOOLEAN DEFAULT FALSE;
        ALTER TABLE users ADD COLUMN IF NOT EXISTS report_format TEXT DEFAULT 'pdf';
        -- Per-report subscriptions: {report_key: 'pdf'|'excel'|null}.
        -- null/missing = not subscribed; 'pdf'|'excel' = subscribed in
        -- that format. Replaces the legacy `report_opt_in`+`report_format`
        -- one-flag-fits-all model. Admin manages this per user from the
        -- Team Management → Reports panel.
        ALTER TABLE users ADD COLUMN IF NOT EXISTS report_subscriptions JSONB NOT NULL DEFAULT '{}'::jsonb;
        ALTER TABLE users ADD COLUMN IF NOT EXISTS first_name TEXT;
        ALTER TABLE users ADD COLUMN IF NOT EXISTS last_name TEXT;
        CREATE TABLE IF NOT EXISTS report_sends (
            id SERIAL PRIMARY KEY,
            period TEXT UNIQUE NOT NULL,
            sent_at TIMESTAMP DEFAULT NOW()
        );
        ALTER TABLE projects ADD COLUMN IF NOT EXISTS scenarios JSONB DEFAULT '[]'::jsonb;
        ALTER TABLE projects ADD COLUMN IF NOT EXISTS status TEXT DEFAULT 'Initial UW';
        -- Existing DBs had the column with DEFAULT 'Active'. New deals
        -- should land in "Initial UW" first; existing rows keep their
        -- current status (this ALTER only changes the column default).
        ALTER TABLE projects ALTER COLUMN status SET DEFAULT 'Initial UW';
        ALTER TABLE projects ADD COLUMN IF NOT EXISTS change_log JSONB DEFAULT '[]'::jsonb;
        CREATE TABLE IF NOT EXISTS projects (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL,
            address TEXT,
            created_by INTEGER REFERENCES users(id),
            created_at TIMESTAMP DEFAULT NOW(),
            updated_at TIMESTAMP DEFAULT NOW(),
            inputs JSONB NOT NULL DEFAULT '{}'::jsonb,
            outputs JSONB NOT NULL DEFAULT '{}'::jsonb,
            archived BOOLEAN DEFAULT FALSE
        );
        CREATE TABLE IF NOT EXISTS reports (
            id SERIAL PRIMARY KEY,
            report_type TEXT NOT NULL,
            data JSONB NOT NULL DEFAULT '{}'::jsonb,
            uploaded_by INTEGER REFERENCES users(id),
            uploaded_at TIMESTAMP DEFAULT NOW()
        );
        -- Per-user activity ledger powering the admin Team Activity
        -- dashboard. Captures successful logins, logouts, and page views
        -- (no IPs, no user agents — internal-team usage analytics only).
        -- Username is denormalized so the dashboard keeps working even
        -- if a user is renamed or deleted later. Retention is 12 months,
        -- enforced by a daily cleanup job in the BackgroundScheduler.
        CREATE TABLE IF NOT EXISTS activity_log (
            id BIGSERIAL PRIMARY KEY,
            user_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
            username TEXT NOT NULL,
            event_type TEXT NOT NULL,
            path TEXT,
            ts TIMESTAMPTZ NOT NULL DEFAULT NOW()
        );
        CREATE INDEX IF NOT EXISTS activity_log_user_ts ON activity_log (user_id, ts DESC);
        CREATE INDEX IF NOT EXISTS activity_log_ts ON activity_log (ts DESC);

        -- ── Vertical Dashboard ──────────────────────────────────────
        -- Ported from the standalone Node/Express service; every table
        -- carries a `vertical` discriminator (lighthaven / hawthorne /
        -- tgp) so a single set of tables handles all properties.
        CREATE TABLE IF NOT EXISTS vd_units (
          id SERIAL PRIMARY KEY,
          vertical TEXT NOT NULL DEFAULT 'lighthaven',
          addr TEXT NOT NULL,
          unit_num TEXT,
          floorplan TEXT,
          type TEXT,
          bed TEXT,
          sf INTEGER,
          phase INTEGER,
          status TEXT,
          grid_row INTEGER,
          grid_col INTEGER,
          tenant_name TEXT,
          lease_executed_date DATE,
          move_in_date DATE,
          move_out_date DATE,
          updated_at TIMESTAMPTZ DEFAULT NOW()
        );
        CREATE INDEX IF NOT EXISTS vd_units_vertical_idx ON vd_units(vertical);
        CREATE INDEX IF NOT EXISTS vd_units_status_idx   ON vd_units(status);
        CREATE INDEX IF NOT EXISTS vd_units_phase_idx    ON vd_units(phase);

        CREATE TABLE IF NOT EXISTS vd_noi_data (
          id SERIAL PRIMARY KEY, vertical TEXT NOT NULL DEFAULT 'lighthaven',
          data JSONB NOT NULL, updated_by TEXT, created_at TIMESTAMPTZ DEFAULT NOW()
        );
        CREATE INDEX IF NOT EXISTS vd_noi_vertical_idx ON vd_noi_data(vertical);

        CREATE TABLE IF NOT EXISTS vd_leasing_pace_data (
          id SERIAL PRIMARY KEY, vertical TEXT NOT NULL DEFAULT 'lighthaven',
          data JSONB NOT NULL, updated_by TEXT, created_at TIMESTAMPTZ DEFAULT NOW()
        );
        CREATE INDEX IF NOT EXISTS vd_leasing_pace_vertical_idx ON vd_leasing_pace_data(vertical);

        CREATE TABLE IF NOT EXISTS vd_traffic_data (
          id SERIAL PRIMARY KEY, vertical TEXT NOT NULL DEFAULT 'lighthaven',
          data JSONB NOT NULL, updated_by TEXT, created_at TIMESTAMPTZ DEFAULT NOW()
        );
        CREATE INDEX IF NOT EXISTS vd_traffic_vertical_idx ON vd_traffic_data(vertical);

        CREATE TABLE IF NOT EXISTS vd_rents_data (
          id SERIAL PRIMARY KEY, vertical TEXT NOT NULL DEFAULT 'lighthaven',
          data JSONB NOT NULL, updated_by TEXT, created_at TIMESTAMPTZ DEFAULT NOW()
        );
        CREATE INDEX IF NOT EXISTS vd_rents_vertical_idx ON vd_rents_data(vertical);

        CREATE TABLE IF NOT EXISTS vd_phase_rollup_data (
          id SERIAL PRIMARY KEY, vertical TEXT NOT NULL DEFAULT 'lighthaven',
          data JSONB NOT NULL, updated_by TEXT, created_at TIMESTAMPTZ DEFAULT NOW()
        );
        CREATE INDEX IF NOT EXISTS vd_phase_rollup_vertical_idx ON vd_phase_rollup_data(vertical);

        CREATE TABLE IF NOT EXISTS vd_psr_data (
          id SERIAL PRIMARY KEY, vertical TEXT NOT NULL DEFAULT 'lighthaven',
          data JSONB NOT NULL, updated_by TEXT, created_at TIMESTAMPTZ DEFAULT NOW()
        );
        CREATE INDEX IF NOT EXISTS vd_psr_vertical_idx ON vd_psr_data(vertical);

        CREATE TABLE IF NOT EXISTS vd_hawthorne_data (
          id SERIAL PRIMARY KEY, vertical TEXT NOT NULL DEFAULT 'lighthaven',
          data JSONB NOT NULL, updated_by TEXT, created_at TIMESTAMPTZ DEFAULT NOW()
        );
        CREATE INDEX IF NOT EXISTS vd_hawthorne_vertical_idx ON vd_hawthorne_data(vertical);

        CREATE TABLE IF NOT EXISTS vd_comments (
          vertical TEXT NOT NULL DEFAULT 'lighthaven',
          section_key TEXT NOT NULL,
          body TEXT NOT NULL DEFAULT '',
          updated_by TEXT,
          updated_at TIMESTAMPTZ DEFAULT NOW(),
          PRIMARY KEY (vertical, section_key)
        );

        CREATE TABLE IF NOT EXISTS vd_imports (
          id SERIAL PRIMARY KEY,
          vertical TEXT NOT NULL DEFAULT 'lighthaven',
          source_file TEXT,
          imported_by TEXT,
          payload JSONB,
          created_at TIMESTAMPTZ DEFAULT NOW()
        );
        CREATE INDEX IF NOT EXISTS vd_imports_vertical_idx ON vd_imports(vertical);

        -- ── Invoice Dashboard ──────────────────────────────────────
        -- Bi-weekly Stampli HTML uploads. Unlike the `reports` table
        -- (latest-snapshot pattern), every upload here is a permanent
        -- entry the user can browse from the Period Archive rail.
        -- Period key format: YYYY-MM-A (days 1-15) or YYYY-MM-B (16-EOM).
        CREATE TABLE IF NOT EXISTS invoice_periods (
            id              BIGSERIAL PRIMARY KEY,
            period_key      TEXT       NOT NULL UNIQUE,
            period_start    DATE       NOT NULL,
            period_end      DATE       NOT NULL,
            report_date     DATE       NOT NULL,
            entities        TEXT[]     NOT NULL DEFAULT ARRAY['Ember Group LLC','CCDL Ventures LLC'],
            total_records   INTEGER    NOT NULL,
            total_amount    NUMERIC(14,2) NOT NULL,
            data            JSONB      NOT NULL,
            source_filename TEXT,
            source_blob     BYTEA,
            uploaded_by     TEXT,
            uploaded_at     TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            CONSTRAINT invoice_periods_period_key_chk
                CHECK (period_key ~ '^[0-9]{4}-[0-9]{2}-[AB]$'),
            CONSTRAINT invoice_periods_window_chk
                CHECK (period_end >= period_start)
        );
        CREATE INDEX IF NOT EXISTS invoice_periods_period_start_idx
            ON invoice_periods (period_start DESC);

        -- ── Financial Statements (Sage Intacct cache) ──────────────
        -- Caches the parsed trial balance per (entity, period) so the
        -- /financials dashboard renders without re-hitting Sage on
        -- every page load. Refresh-from-Sage button overwrites the row.
        -- `accounts` is the array of TB rows from sage_intacct.get_trial_balance.
        CREATE TABLE IF NOT EXISTS intacct_tb_cache (
            entity_id    TEXT          NOT NULL,
            period_name  TEXT          NOT NULL,
            fetched_at   TIMESTAMPTZ   NOT NULL DEFAULT NOW(),
            fetched_by   INTEGER       REFERENCES users(id),
            accounts     JSONB         NOT NULL,
            accounts_ytd JSONB,
            PRIMARY KEY (entity_id, period_name)
        );
        -- Backfill `accounts_ytd` column for existing schemas that
        -- pre-date Phase 2 (IS/SCF need YTD-TB data alongside monthly).
        ALTER TABLE intacct_tb_cache
            ADD COLUMN IF NOT EXISTS accounts_ytd JSONB;
        CREATE INDEX IF NOT EXISTS intacct_tb_cache_fetched_idx
            ON intacct_tb_cache (fetched_at DESC);
        -- Lightweight summary view for the archive rail so the page
        -- route can render snapshot cards without pulling the full
        -- JSONB blob from every row.
        CREATE OR REPLACE VIEW invoice_periods_summary AS
        SELECT
            period_key,
            period_start,
            period_end,
            report_date,
            total_records,
            total_amount,
            COALESCE((SELECT COUNT(*)::int FROM jsonb_array_elements(data->'records') r
                      WHERE r->>'entity' = 'Ember Group LLC'),  0) AS ember_count,
            COALESCE((SELECT COUNT(*)::int FROM jsonb_array_elements(data->'records') r
                      WHERE r->>'entity' = 'CCDL Ventures LLC'), 0) AS ccdl_count,
            uploaded_at
        FROM invoice_periods
        ORDER BY period_start DESC;
    """)
    # Backfill `verticals` page_access for existing users so the new
    # dashboard is visible by default — same convention used for
    # portfolio / macro / sales when those shipped.
    cur.execute("UPDATE users SET page_access = page_access || '{\"verticals\": true}'::jsonb WHERE page_access->>'verticals' IS NULL")
    # Backfill `invoice_dashboard` page_access — new Stampli analytics page.
    cur.execute("UPDATE users SET page_access = page_access || '{\"invoice_dashboard\": true}'::jsonb WHERE page_access->>'invoice_dashboard' IS NULL")
    # Per-user "can edit comments on the Vertical Dashboard" flag.
    # Default true to match the legacy behavior (anyone with verticals
    # access could edit, gated only by a shared admin password). Admins
    # always get edit rights server-side regardless of this flag.
    cur.execute("UPDATE users SET page_access = page_access || '{\"verticals_comment\": true}'::jsonb WHERE page_access->>'verticals_comment' IS NULL")
    # Backfill portfolio access for existing users
    cur.execute("UPDATE users SET page_access = page_access || '{\"portfolio\": true}'::jsonb WHERE page_access->>'portfolio' IS NULL")
    cur.execute("UPDATE users SET page_access = page_access || '{\"macro\": true}'::jsonb WHERE page_access->>'macro' IS NULL")
    cur.execute("UPDATE users SET page_access = page_access || '{\"sales\": true}'::jsonb WHERE page_access->>'sales' IS NULL")
    # `financials` gates the Sage-Intacct-backed Financial Statements
    # dashboard. Defaults to FALSE — financial statements are sensitive,
    # so we want explicit admin grants rather than backfilling true.
    cur.execute("UPDATE users SET page_access = page_access || '{\"financials\": false}'::jsonb WHERE page_access->>'financials' IS NULL")
    # `reports` toggle gates download/export of executive PDFs and Excels.
    # Admin-controllable per user; defaults to true so we don't yank the
    # ability away from existing accounts on deploy.
    cur.execute("UPDATE users SET page_access = page_access || '{\"reports\": true}'::jsonb WHERE page_access->>'reports' IS NULL")
    # `captable_edit` gates uploading/editing investor cap tables (money data).
    # Defaults to FALSE for everyone; granted explicitly to the owner account
    # so it can manage cap tables without opening the control to all users.
    cur.execute("UPDATE users SET page_access = page_access || '{\"captable_edit\": true}'::jsonb WHERE username = %s", ("carlossaldierna",))
    # `bva` gates the Budget vs Actuals dashboard (dev-team financial data).
    # Defaults FALSE; admins always pass and grant it to the dev team per user.
    cur.execute("UPDATE users SET page_access = page_access || '{\"bva\": false}'::jsonb WHERE page_access->>'bva' IS NULL")
    # Owner account is a full admin (manage users, all dashboards, Sage tools).
    cur.execute("UPDATE users SET is_admin = TRUE WHERE username = %s", ("carlossaldierna",))

    # Backfill report_subscriptions for users created before the column
    # existed. If they were opted-in under the legacy single-flag model,
    # subscribe them to every report in their preferred format.
    cur.execute("""
        UPDATE users
        SET report_subscriptions = jsonb_build_object(
            'returns',       COALESCE(report_format, 'pdf'),
            'ember_capital', COALESCE(report_format, 'pdf'),
            'operations',    COALESCE(report_format, 'pdf'),
            'loans',         COALESCE(report_format, 'pdf')
        )
        WHERE report_opt_in = TRUE
          AND (report_subscriptions IS NULL OR report_subscriptions = '{}'::jsonb)
    """)
    # Create default admin if no users exist
    cur.execute("SELECT COUNT(*) as cnt FROM users")
    row = cur.fetchone()
    if row["cnt"] == 0:
        cur.execute(
            "INSERT INTO users (username, password_hash, is_admin) VALUES (%s, %s, TRUE)",
            ("admin", generate_password_hash("ember2024"))
        )
    conn.commit()
    cur.close()
    conn.close()

# ─── AUTH HELPERS ─────────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            if request.is_json:
                return jsonify({"error": "Unauthorized"}), 401
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("is_admin"):
            return jsonify({"error": "Admin required"}), 403
        return f(*args, **kwargs)
    return decorated

# ─── ACTIVITY TRACKING ────────────────────────────────────────────────────────
# Page-view + login/logout capture for the admin Team Activity dashboard.
# Per spec: user-level only (no IPs, no user agents), 12-month retention.

# Path prefixes that should NEVER be logged: static assets, AJAX/JSON
# endpoints, healthchecks, internal Flask routes. Everything that's a
# real "I am looking at this dashboard" page view falls through to the
# logger.
_ACTIVITY_DENY_PREFIXES = (
    "/static/", "/api/", "/health", "/favicon",
    "/login", "/logout",  # captured explicitly so we record success only
)
# File extensions that indicate downloads / asset fetches rather than
# page views. We still log the visit to the page that triggered the
# download — the download URL itself is noise.
_ACTIVITY_DENY_SUFFIXES = (".json", ".css", ".js", ".ico", ".png", ".svg",
                            ".jpg", ".jpeg", ".gif", ".woff", ".woff2", ".map")

def _log_activity(event_type, path=None, user_id=None, username=None):
    """Insert one row into activity_log. Swallows all errors — logging
    failure must never break a real request."""
    try:
        uid = user_id if user_id is not None else session.get("user_id")
        uname = username or session.get("username")
        if not uid or not uname:
            return
        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO activity_log (user_id, username, event_type, path) "
            "VALUES (%s, %s, %s, %s)",
            (uid, uname, event_type, path)
        )
        conn.commit()
        cur.close(); conn.close()
    except Exception as e:
        # Don't let a logging hiccup take down a page load.
        print(f"[activity] log failed ({event_type}): {e}", flush=True)

@app.before_request
def _capture_page_view():
    """Log GET requests to authenticated dashboard pages. Login and
    logout are captured by their own handlers; POSTs and AJAX traffic
    are intentionally ignored to keep the log focused on 'who looked at
    what'."""
    if request.method != "GET":
        return
    if not session.get("user_id"):
        return
    p = request.path or ""
    if p.startswith(_ACTIVITY_DENY_PREFIXES):
        return
    if p.endswith(_ACTIVITY_DENY_SUFFIXES):
        return
    _log_activity("page_view", path=p)

# ─── AUTH ROUTES ─────────────────────────────────────────────────────────────
@app.route("/health")
def health():
    return "ok", 200

@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT * FROM users WHERE username = %s", (username,))
        user = cur.fetchone()
        cur.close(); conn.close()
        if user and check_password_hash(user["password_hash"], password):
            session["user_id"] = user["id"]
            session["username"] = user["username"]
            session["is_admin"] = user["is_admin"]
            session["page_access"] = user.get("page_access") or {"mpc_underwriting": True, "returns": True, "loans": True, "operations": True}
            fn = (user.get("first_name") or "").strip()
            ln = (user.get("last_name") or "").strip()
            session["display_name"] = f"{fn} {ln}".strip() or user["username"]
            _log_activity("login", user_id=user["id"], username=user["username"])
            return redirect(url_for("home"))
        error = "Invalid username or password."
    return render_template("login.html", error=error)

@app.route("/logout")
def logout():
    # Read identity off the session before clearing it so the logout
    # event is attributed to the right user.
    uid = session.get("user_id")
    uname = session.get("username")
    if uid and uname:
        _log_activity("logout", user_id=uid, username=uname)
    session.clear()
    return redirect(url_for("login"))

# ─── MAIN APP ─────────────────────────────────────────────────────────────────
def _fmt_data_date(val) -> str | None:
    """Format an ISO yyyy-mm-dd string or date/datetime as '10 May 2026'.

    Returns None when the input is falsy or unparseable. Used by dashboard
    routes to render a consistent "Data From" / "Uploaded" date in the
    page-footer partial (_partials/_data_dates_footer.html).
    """
    if not val:
        return None
    if isinstance(val, str):
        try:
            d = datetime.datetime.strptime(val[:10], "%Y-%m-%d").date()
        except (TypeError, ValueError):
            return None
    elif isinstance(val, datetime.datetime):
        d = val.date()
    elif isinstance(val, datetime.date):
        d = val
    else:
        return None
    return d.strftime("%-d %b %Y")


def _home_portfolio_summary():
    """Compute the at-a-glance portfolio numbers shown on the home hero.

    The total equity figure mirrors the Ember Capital page's "Totals" row
    so the home and Ember Capital pages always agree:
        total LP equity = sum across projects of |Total LP Contributions|
    Values in the returns report are stored in $K (thousands of dollars,
    matching the Excel pro-forma convention), so we multiply by 1000 to
    get raw dollars before formatting.

    Returns a dict with:
      - lp_equity_dollars (float)       — total LP equity in raw dollars
      - lp_equity_label (str)           — '$94.4M', '$1.2B', or '—'
      - portfolio_irr (float|None)      — LP-contribution-weighted IRR
      - portfolio_irr_label (str)       — '22%', or '—'
      - active_project_count (int)      — projects with any non-zero metric
      - report_dates (dict)             — upload-date string per type
      - data_from_dates (dict)          — "Data From" date string per dashboard
      - reports_updated_today (int)     — types updated in last ~36h
    """
    out = {
        "lp_equity_dollars": 0.0,
        "lp_equity_label": "—",
        "portfolio_irr": None,
        "portfolio_irr_label": "—",
        "active_project_count": 0,
        "report_dates": {},
        "data_from_dates": {},
        "reports_updated_today": 0,
    }
    try:
        conn = get_db()
        cur = conn.cursor()
        # Latest row per report_type — pulls both upload time and the
        # in-data `data_from` field that the parsers now populate.
        cur.execute("""
            SELECT DISTINCT ON (report_type)
                   report_type,
                   uploaded_at,
                   data->>'data_from' AS data_from
            FROM reports
            ORDER BY report_type, uploaded_at DESC
        """)
        rows = cur.fetchall()
        out["report_dates"] = {
            r["report_type"]: r["uploaded_at"].strftime("%-d %b %Y")
            for r in rows if r["uploaded_at"]
        }
        # Ember Capital reads from reports[returns], so its upload date
        # mirrors the returns row. Without this, the home card has no
        # date at all when the returns blob predates `data_from`.
        if "returns" in out["report_dates"]:
            out["report_dates"].setdefault(
                "ember_capital", out["report_dates"]["returns"]
            )
        # `data_from` is an ISO yyyy-mm-dd string (the parser calls
        # _date_iso). Format it the same way as report_dates.
        data_from_by_type: dict[str, str] = {}
        for r in rows:
            df = r.get("data_from") if isinstance(r, dict) else r["data_from"]
            if not df:
                continue
            try:
                d = datetime.datetime.strptime(df, "%Y-%m-%d").date()
                data_from_by_type[r["report_type"]] = d.strftime("%-d %b %Y")
            except (TypeError, ValueError):
                continue
        # Ember Capital reads from the returns blob, so its "Data From"
        # is the same as returns. Mirror it so the home card shows a date.
        if "returns" in data_from_by_type:
            data_from_by_type.setdefault("ember_capital", data_from_by_type["returns"])
        cutoff = datetime.datetime.now() - datetime.timedelta(hours=36)
        out["reports_updated_today"] = sum(
            1 for r in rows if r["uploaded_at"] and r["uploaded_at"].replace(tzinfo=None) >= cutoff
        )

        # Invoice dashboard lives in a separate table (invoice_periods).
        # Use the latest period's report_date as "Data From".
        try:
            cur.execute(
                "SELECT report_date FROM invoice_periods "
                "ORDER BY period_start DESC LIMIT 1"
            )
            inv = cur.fetchone()
            if inv and inv["report_date"]:
                data_from_by_type["invoice_dashboard"] = (
                    inv["report_date"].strftime("%-d %b %Y")
                )
        except Exception:
            pass

        out["data_from_dates"] = data_from_by_type

        # Pull the latest returns report — same source the Ember Capital page
        # reads from.
        cur.execute(
            "SELECT data FROM reports WHERE report_type = 'returns' "
            "ORDER BY uploaded_at DESC LIMIT 1"
        )
        row = cur.fetchone()
        cur.close(); conn.close()
        if row and row["data"]:
            data = row["data"]
            projects = data.get("projects") or []
            total_equity_k = 0.0      # in $K, matches Ember Capital totals row
            active = 0
            irr_weighted = 0.0
            irr_weight   = 0.0
            for p in projects:
                by_label = {m.get("label"): m for m in (p.get("metrics") or [])}
                def _f(label):
                    v = (by_label.get(label) or {}).get("total")
                    try: return float(v) if v is not None else 0.0
                    except (TypeError, ValueError): return 0.0
                contrib_abs = abs(_f("Total LP Contributions"))
                distrib     = _f("Total LP Distributions")
                irr         = _f("LP IRR")
                if contrib_abs > 0 or distrib > 0:
                    active += 1
                total_equity_k += contrib_abs
                # LP-contribution-weighted IRR — bigger checks weigh more.
                if irr and contrib_abs > 0:
                    irr_weighted += irr * contrib_abs
                    irr_weight   += contrib_abs

            out["active_project_count"] = active
            dollars = total_equity_k * 1000.0   # $K -> $
            out["lp_equity_dollars"] = dollars
            # Round DOWN (truncate) to one decimal — finance convention so the
            # headline never overstates capital. $94,454,000 -> "$94.4M",
            # not "$94.5M".
            def _trunc1(x): return int(x * 10) / 10.0
            if dollars >= 1e9:
                out["lp_equity_label"] = f"${_trunc1(dollars/1e9):.1f}B"
            elif dollars >= 1e6:
                out["lp_equity_label"] = f"${_trunc1(dollars/1e6):.1f}M"
            elif dollars >= 1e3:
                out["lp_equity_label"] = f"${dollars/1e3:.0f}K"
            elif dollars > 0:
                out["lp_equity_label"] = f"${dollars:,.0f}"

            if irr_weight > 0:
                irr_blended = irr_weighted / irr_weight
                out["portfolio_irr"] = irr_blended
                out["portfolio_irr_label"] = f"{irr_blended*100:.0f}%"
    except Exception:
        # Home page must never blow up on a dashboard summary glitch.
        pass
    return out


@app.route("/home")
@login_required
def home():
    pa = session.get("page_access") or {"mpc_underwriting": True, "returns": True, "loans": True, "operations": True}
    if session.get("is_admin"):
        pa = {"mpc_underwriting": True, "returns": True, "loans": True, "operations": True}
    summary = _home_portfolio_summary()
    return render_template("home.html",
        username=session.get("username"),
        display_name=session.get("display_name", session.get("username")),
        is_admin=session.get("is_admin"),
        page_access=pa,
        report_dates=summary["report_dates"],
        data_from_dates=summary["data_from_dates"],
        lp_equity_label=summary["lp_equity_label"],
        portfolio_irr_label=summary["portfolio_irr_label"],
        active_project_count=summary["active_project_count"],
        reports_updated_today=summary["reports_updated_today"])

@app.route("/")
@login_required
def index():
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("mpc_underwriting", True):
        return redirect(url_for("home"))
    pa = session.get("page_access") or {"mpc_underwriting": True, "returns": True, "loans": True, "operations": True}
    if session.get("is_admin"):
        pa = {"mpc_underwriting": True, "returns": True, "loans": True, "operations": True}
    return render_template("app.html", username=session.get("username"), is_admin=session.get("is_admin"), page_access=pa)

# ─── PROJECT API ─────────────────────────────────────────────────────────────
@app.route("/api/projects", methods=["GET"])
@login_required
def list_projects():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT p.id, p.name, p.address, p.updated_at,
               u.username as created_by,
               COALESCE(
                   CASE WHEN jsonb_array_length(COALESCE(p.scenarios,'[]'::jsonb)) > 0
                        THEN p.scenarios->0->'outputs'->>'total_revenue' END,
                   p.outputs->>'total_revenue'
               ) as total_revenue,
               COALESCE(
                   CASE WHEN jsonb_array_length(COALESCE(p.scenarios,'[]'::jsonb)) > 0
                        THEN p.scenarios->0->'outputs'->>'gross_margin_pct' END,
                   p.outputs->>'gross_margin_pct'
               ) as gross_margin_pct,
               COALESCE(
                   CASE WHEN jsonb_array_length(COALESCE(p.scenarios,'[]'::jsonb)) > 0
                        THEN p.scenarios->0->'outputs'->>'total_lots' END,
                   p.outputs->>'total_lots'
               ) as total_lots,
               COALESCE(
                   CASE WHEN jsonb_array_length(COALESCE(p.scenarios,'[]'::jsonb)) > 0
                        THEN p.scenarios->0->'outputs'->>'unlevered_irr' END,
                   p.outputs->>'unlevered_irr'
               ) as unlevered_irr,
               COALESCE(
                   CASE WHEN jsonb_array_length(COALESCE(p.scenarios,'[]'::jsonb)) > 0
                        THEN p.scenarios->0->'outputs'->>'project_length_years' END,
                   p.outputs->>'project_length_years'
               ) as project_length_years,
               p.archived,
               COALESCE(p.status, 'Active') as status
        FROM projects p
        LEFT JOIN users u ON p.created_by = u.id
        WHERE p.archived = FALSE
        ORDER BY p.updated_at DESC
    """)
    rows = cur.fetchall()
    cur.close(); conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/projects", methods=["POST"])
@login_required
def create_project():
    data = request.json or {}
    name = data.get("name", "New Project")
    inputs = default_inputs(name)
    try:
        outputs = calculate(inputs)
    except Exception:
        outputs = {}
    conn = get_db()
    cur = conn.cursor()
    # New deals start in "Initial UW" — they only flow through to the
    # Ember Capital pipeline once an underwriter promotes them to
    # "Active" (right-click → Set Status, or click the badge on the
    # project list). Setting it explicitly here rather than relying on
    # the schema default makes the intent obvious in the code path.
    cur.execute(
        "INSERT INTO projects (name, address, created_by, inputs, outputs, status) "
        "VALUES (%s, %s, %s, %s, %s, %s) RETURNING id",
        (name, data.get("address", ""), session["user_id"],
         json.dumps(inputs), json.dumps(outputs), "Initial UW")
    )
    pid = cur.fetchone()["id"]
    conn.commit(); cur.close(); conn.close()
    return jsonify({"id": pid, "name": name})

@app.route("/api/projects/<int:pid>", methods=["GET"])
@login_required
def get_project(pid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM projects WHERE id = %s", (pid,))
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row:
        return jsonify({"error": "Not found"}), 404
    d = dict(row)
    # Prefer first scenario's inputs/outputs so callers always get current data
    scens = list(d.get("scenarios") or [])
    if scens:
        d["inputs"]  = scens[0].get("inputs",  d.get("inputs",  {}))
        d["outputs"] = scens[0].get("outputs", d.get("outputs", {}))
    return jsonify(d)

@app.route("/api/projects/<int:pid>", methods=["PUT"])
@login_required
def save_project(pid):
    data = request.json or {}
    inputs = data.get("inputs", {})
    try:
        outputs = calculate(inputs)
    except Exception as e:
        return jsonify({"error": f"Calculation error: {e}"}), 500
    conn = get_db()
    cur = conn.cursor()
    # Build change log entry
    cur.execute("SELECT inputs, change_log FROM projects WHERE id = %s", (pid,))
    row = cur.fetchone()
    old_log = list(row["change_log"] or []) if row else []
    changes = _compare_inputs(row["inputs"] or {} if row else {}, inputs)
    if changes:
        old_log.append({
            "ts": datetime.datetime.utcnow().isoformat() + "Z",
            "user": session.get("username", "unknown"),
            "changes": changes
        })
        old_log = old_log[-200:]
    cur.execute("""
        UPDATE projects
        SET inputs = %s, outputs = %s, name = %s, address = %s,
            change_log = %s, updated_at = NOW()
        WHERE id = %s
    """, (
        json.dumps(inputs),
        json.dumps(outputs),
        inputs.get("project_name", "Unnamed"),
        inputs.get("address", ""),
        json.dumps(old_log),
        pid
    ))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True, "outputs": outputs})

@app.route("/api/projects/<int:pid>", methods=["DELETE"])
@login_required
def delete_project(pid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE projects SET archived = TRUE WHERE id = %s", (pid,))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/projects/<int:pid>/status", methods=["PATCH"])
@login_required
def set_project_status(pid):
    data = request.json or {}
    status = data.get("status", "Active")
    if status not in {"Initial UW", "Active", "Under Contract", "Closed", "Dead"}:
        return jsonify({"error": "Invalid status"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute("UPDATE projects SET status = %s WHERE id = %s", (status, pid))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/projects/<int:pid>/changelog", methods=["GET"])
@login_required
def get_changelog(pid):
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT change_log FROM projects WHERE id = %s", (pid,))
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row: return jsonify({"error": "Not found"}), 404
    return jsonify(list(reversed(row["change_log"] or [])))

@app.route("/api/projects/<int:pid>/sensitivity", methods=["POST"])
@login_required
def sensitivity(pid):
    data = request.json or {}
    axis_x = data.get("axis_x", {})
    axis_y = data.get("axis_y", {})
    base_inputs = data.get("base_inputs", {})
    x_field, x_vals = axis_x.get("field"), axis_x.get("values", [])
    y_field, y_vals = axis_y.get("field"), axis_y.get("values", [])
    if not x_field or not y_field or not x_vals or not y_vals:
        return jsonify({"error": "Missing axis config"}), 400
    # Cap grid size for performance
    x_vals = x_vals[:7]
    y_vals = y_vals[:7]
    matrix = []
    for yv in y_vals:
        row_results = []
        for xv in x_vals:
            inp = _apply_sensitivity_override(base_inputs, x_field, xv)
            inp = _apply_sensitivity_override(inp, y_field, yv)
            try:
                out = calculate(inp)
                row_results.append({
                    "irr": out.get("unlevered_irr"),
                    "gm_pct": out.get("gross_margin_pct"),
                })
            except Exception:
                row_results.append({"irr": None, "gm_pct": None})
        matrix.append(row_results)
    return jsonify({"ok": True, "matrix": matrix, "x_values": x_vals, "y_values": y_vals})

# ─── SCENARIO API ─────────────────────────────────────────────────────────────
@app.route("/api/projects/<int:pid>/scenarios", methods=["GET"])
@login_required
def list_scenarios(pid):
    import uuid
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT inputs, outputs, scenarios FROM projects WHERE id = %s", (pid,))
    row = cur.fetchone()
    if not row: cur.close(); conn.close(); return jsonify({"error": "Not found"}), 404
    scenarios = list(row["scenarios"] or [])
    if not scenarios:
        sid = str(uuid.uuid4())[:8]
        scenarios = [{"id": sid, "name": "Main",
                      "inputs": row["inputs"] or {}, "outputs": row["outputs"] or {}}]
        cur.execute("UPDATE projects SET scenarios = %s WHERE id = %s", (json.dumps(scenarios), pid))
        conn.commit()
    cur.close(); conn.close()
    return jsonify(scenarios)

@app.route("/api/projects/<int:pid>/scenarios", methods=["POST"])
@login_required
def create_scenario(pid):
    import uuid
    data = request.json or {}
    name = data.get("name", "New Scenario").strip() or "New Scenario"
    clone_id = data.get("clone_from")
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT inputs, outputs, scenarios FROM projects WHERE id = %s", (pid,))
    row = cur.fetchone()
    if not row: cur.close(); conn.close(); return jsonify({"error": "Not found"}), 404
    scenarios = list(row["scenarios"] or [])
    sid = str(uuid.uuid4())[:8]
    if clone_id:
        src = next((s for s in scenarios if s["id"] == clone_id), None)
        inp = dict(src["inputs"]) if src else dict(row["inputs"] or {})
        out = dict(src["outputs"]) if src else dict(row["outputs"] or {})
    else:
        # Seed from first scenario when available so new scenarios start from current inputs
        base = scenarios[0] if scenarios else None
        inp = dict(base["inputs"]) if base else dict(row["inputs"] or {})
        out = dict(base["outputs"]) if base else dict(row["outputs"] or {})
    new_scen = {"id": sid, "name": name, "inputs": inp, "outputs": out}
    scenarios.append(new_scen)
    cur.execute("UPDATE projects SET scenarios = %s WHERE id = %s", (json.dumps(scenarios), pid))
    conn.commit(); cur.close(); conn.close()
    return jsonify(new_scen)

@app.route("/api/projects/<int:pid>/scenarios/<sid>", methods=["PUT"])
@login_required
def save_scenario(pid, sid):
    data = request.json or {}
    inp = data.get("inputs", {})
    try:
        out = calculate(inp)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT scenarios FROM projects WHERE id = %s", (pid,))
    row = cur.fetchone()
    if not row: cur.close(); conn.close(); return jsonify({"error": "Not found"}), 404
    scenarios = list(row["scenarios"] or [])
    idx = next((i for i, s in enumerate(scenarios) if s["id"] == sid), None)
    if idx is None: cur.close(); conn.close(); return jsonify({"error": "Scenario not found"}), 404
    scenarios[idx]["inputs"] = inp
    scenarios[idx]["outputs"] = out
    # Keep projects.inputs/outputs in sync with the first (Main) scenario so
    # the sidebar IRR and portfolio view always reflect current inputs.
    if idx == 0:
        cur.execute(
            "UPDATE projects SET scenarios = %s, inputs = %s, outputs = %s WHERE id = %s",
            (json.dumps(scenarios), json.dumps(inp), json.dumps(out), pid)
        )
    else:
        cur.execute("UPDATE projects SET scenarios = %s WHERE id = %s", (json.dumps(scenarios), pid))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True, "outputs": out})

@app.route("/api/projects/<int:pid>/scenarios/<sid>", methods=["DELETE"])
@login_required
def delete_scenario(pid, sid):
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT scenarios FROM projects WHERE id = %s", (pid,))
    row = cur.fetchone()
    if not row: cur.close(); conn.close(); return jsonify({"error": "Not found"}), 404
    scenarios = [s for s in (row["scenarios"] or []) if s["id"] != sid]
    cur.execute("UPDATE projects SET scenarios = %s WHERE id = %s", (json.dumps(scenarios), pid))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/projects/<int:pid>/scenarios/<sid>/promote", methods=["POST"])
@login_required
def promote_scenario(pid, sid):
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT scenarios FROM projects WHERE id = %s", (pid,))
    row = cur.fetchone()
    if not row: cur.close(); conn.close(); return jsonify({"error": "Not found"}), 404
    scen = next((s for s in (row["scenarios"] or []) if s["id"] == sid), None)
    if not scen: cur.close(); conn.close(); return jsonify({"error": "Scenario not found"}), 404
    inp = scen["inputs"]; out = scen["outputs"]
    cur.execute("""UPDATE projects SET inputs=%s, outputs=%s, name=%s, address=%s, updated_at=NOW()
                   WHERE id=%s""",
                (json.dumps(inp), json.dumps(out),
                 inp.get("project_name", "Unnamed"), inp.get("address", ""), pid))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True, "outputs": out})

@app.route("/api/projects/<int:pid>/scenarios/<sid>/name", methods=["PATCH"])
@login_required
def rename_scenario(pid, sid):
    data = request.json or {}
    name = data.get("name", "").strip()
    if not name: return jsonify({"error": "Name required"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT scenarios FROM projects WHERE id = %s", (pid,))
    row = cur.fetchone()
    if not row: cur.close(); conn.close(); return jsonify({"error": "Not found"}), 404
    scenarios = list(row["scenarios"] or [])
    idx = next((i for i, s in enumerate(scenarios) if s["id"] == sid), None)
    if idx is None: cur.close(); conn.close(); return jsonify({"error": "Scenario not found"}), 404
    scenarios[idx]["name"] = name
    cur.execute("UPDATE projects SET scenarios = %s WHERE id = %s", (json.dumps(scenarios), pid))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/projects/<int:pid>/calculate", methods=["POST"])
@login_required
def recalculate(pid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT inputs FROM projects WHERE id = %s", (pid,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        return jsonify({"error": "Not found"}), 404
    inputs = row["inputs"]
    try:
        outputs = calculate(inputs)
    except Exception as e:
        cur.close(); conn.close()
        return jsonify({"error": str(e)}), 500
    cur.execute("UPDATE projects SET outputs = %s, updated_at = NOW() WHERE id = %s",
                (json.dumps(outputs), pid))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True, "outputs": outputs})

# ─── VERTICAL DASHBOARD ──────────────────────────────────────────────────────
# Multi-property leasing/operations dashboard ported from a standalone
# Node/Express service (csaldierna1-bot/Vertical-Dashboard). The frontend
# is a static SPA living in static/verticals/; these routes serve the
# JSON it consumes. Schema lives in init_db() under `vd_*` tables.
#
# Conventions kept from the Node implementation:
#   - Every read/write is scoped by ?vertical=<slug>; KNOWN_VERTICALS
#     gatekeeps the slugs.
#   - Tab snapshots (noi, leasing pace, etc.) use a DELETE-then-INSERT
#     pattern so each vertical has exactly one current row per table.
#   - Status normalization: 'UNRELEASED' → 'AVAILABLE' (legacy data).
#
# Auth: write endpoints require admin (replaces the Node app's
# X-API-Key shim); read endpoints require login. Hawthorne POST is the
# one exception — its UI uploads the master xlsx in-browser, so any
# logged-in user can write to it (mirrors Node behavior).

_VD_KNOWN_VERTICALS = ('lighthaven', 'hawthorne', 'tgp')

def _vd_vertical(req=None):
    """Pull `vertical` from request args or JSON body. Returns None if
    missing/invalid so the caller can 400."""
    if req is None:
        req = request
    v = (req.args.get("vertical")
         or (req.get_json(silent=True) or {}).get("vertical")
         or "lighthaven")
    v = str(v).lower()
    return v if v in _VD_KNOWN_VERTICALS else None

def _vd_normalize_status(s):
    """Mirror server.js normalizeStatus: collapse legacy UNRELEASED → AVAILABLE."""
    if not s:
        return s
    return "AVAILABLE" if str(s).upper() == "UNRELEASED" else s

def _vd_pct(n, d):
    return round((n / d) * 1000) / 10 if d else 0

# ── UNITS ────────────────────────────────────────────────────────────
@app.route("/api/verticals/units", methods=["GET"])
@login_required
def vd_get_units():
    v = _vd_vertical()
    if not v:
        return jsonify({"error": "Unknown vertical"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute("""
        SELECT addr, unit_num AS num, floorplan AS fp, type, bed, sf, phase, status,
               grid_row, grid_col,
               tenant_name, lease_executed_date, move_in_date, move_out_date, updated_at
        FROM vd_units WHERE vertical = %s ORDER BY phase, addr
    """, (v,))
    rows = [dict(r) for r in cur.fetchall()]
    cur.close(); conn.close()
    return jsonify({"hasData": bool(rows), "vertical": v, "units": rows})

# ── JSONB tab stores (psr) ──────────────────────────────────────
# Single GET (latest snapshot) + POST (DELETE-then-INSERT replace) per vertical.
# The other tab tables (noi, leasing-pace, traffic, rents, phase-rollup)
# are written by /api/verticals/upload and read internally by the
# snapshot endpoint — they don't need standalone HTTP routes.
def _vd_jsonb_get(table):
    v = _vd_vertical()
    if not v:
        return jsonify({"error": "Unknown vertical"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute(
        f"SELECT data, updated_by, created_at FROM {table} "
        "WHERE vertical = %s ORDER BY created_at DESC LIMIT 1",
        (v,)
    )
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row:
        return jsonify({"hasData": False, "vertical": v})
    return jsonify({"hasData": True, "vertical": v, **dict(row)})

def _vd_jsonb_post(table):
    v = _vd_vertical()
    if not v:
        return jsonify({"error": "Unknown vertical"}), 400
    body = request.get_json(silent=True) or {}
    data = body.get("data")
    if not isinstance(data, dict):
        return jsonify({"error": "Missing data object"}), 400
    updated_by = body.get("updatedBy") or session.get("username") or "anonymous"
    conn = get_db(); cur = conn.cursor()
    try:
        cur.execute(f"DELETE FROM {table} WHERE vertical = %s", (v,))
        cur.execute(
            f"INSERT INTO {table} (vertical, data, updated_by) VALUES (%s, %s, %s) "
            "RETURNING id, updated_by, created_at",
            (v, json.dumps(data), updated_by)
        )
        rec = dict(cur.fetchone())
        conn.commit()
        return jsonify({"success": True, "vertical": v, "record": rec})
    except Exception as e:
        conn.rollback()
        return jsonify({"error": f"Failed to save {table}", "detail": str(e)}), 500
    finally:
        cur.close(); conn.close()

# Generic JSONB endpoint — write requires admin. Only psr_data has a
# direct HTTP route since the TGP page reads/writes the PSR blob
# inline; the other tab tables are populated by /api/verticals/upload
# and read internally by /api/verticals/snapshot.
for _path, _table in [
    ("/api/verticals/psr",          "vd_psr_data"),
]:
    # Bind table name in default arg to avoid late-binding closure bug.
    def _vd_get(table=_table): return _vd_jsonb_get(table)
    def _vd_post(table=_table):
        if not session.get("is_admin"):
            return jsonify({"error": "Admin required"}), 403
        return _vd_jsonb_post(table)
    _vd_get.__name__  = f"vd_get_{_table}"
    _vd_post.__name__ = f"vd_post_{_table}"
    app.add_url_rule(_path, view_func=login_required(_vd_get),  methods=["GET"])
    app.add_url_rule(_path, view_func=login_required(_vd_post), methods=["POST"])

# ── HAWTHORNE (public POST in original — kept for parity, login-only here) ──
@app.route("/api/verticals/hawthorne", methods=["GET"])
@login_required
def vd_get_hawthorne():
    return _vd_jsonb_get("vd_hawthorne_data")

@app.route("/api/verticals/hawthorne", methods=["POST"])
@login_required
def vd_post_hawthorne():
    # Mirrors the Node app's quirk: any logged-in user can update the
    # Hawthorne master xlsx blob — the in-browser UI is what writes.
    return _vd_jsonb_post("vd_hawthorne_data")

# ── PSR raw xlsx round-trip ──
# Node app wrote raw bytes to public/tgp_psr_template.xlsx so the next
# download is byte-identical. We do the same against static/verticals/.
_VD_PSR_RAW_FILES = {"tgp": "tgp_psr_template.xlsx"}

@app.route("/api/verticals/psr/raw", methods=["POST"])
@login_required
@admin_required
def vd_post_psr_raw():
    v = _vd_vertical()
    if not v:
        return jsonify({"error": "Unknown vertical"}), 400
    filename = _VD_PSR_RAW_FILES.get(v)
    if not filename:
        return jsonify({"error": f"No PSR raw slot for vertical: {v}"}), 404
    raw = request.get_data()
    if not raw:
        return jsonify({"error": "No body bytes received"}), 400
    dest = os.path.join(os.path.dirname(__file__), "static", "verticals", filename)
    try:
        with open(dest, "wb") as f:
            f.write(raw)
    except Exception as e:
        return jsonify({"error": f"Save failed: {e}"}), 500
    return jsonify({"success": True, "vertical": v, "bytes": len(raw), "filename": filename})

# ── COMMENTS (vertical, section_key) composite-PK upsert ──
@app.route("/api/verticals/comments", methods=["GET"])
@login_required
def vd_get_comments():
    v = _vd_vertical()
    if not v:
        return jsonify({"error": "Unknown vertical"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute(
        "SELECT section_key, body, updated_by, updated_at FROM vd_comments WHERE vertical = %s",
        (v,)
    )
    out = {}
    for r in cur.fetchall():
        out[r["section_key"]] = dict(r)
    cur.close(); conn.close()
    return jsonify(out)

@app.route("/api/verticals/comments/<key>", methods=["GET"])
@login_required
def vd_get_comment(key):
    v = _vd_vertical()
    if not v:
        return jsonify({"error": "Unknown vertical"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute(
        "SELECT section_key, body, updated_by, updated_at FROM vd_comments "
        "WHERE vertical = %s AND section_key = %s",
        (v, key)
    )
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row:
        return jsonify({"section_key": key, "body": "", "updated_at": None})
    return jsonify(dict(row))

def _refresh_page_access_from_db():
    """Pull the *current* page_access for the logged-in user from the
    DB and update the session snapshot so subsequent reads see fresh
    values.

    `session["page_access"]` is populated at login and only refreshed
    on subsequent logins — so when an admin grants a new permission to
    a user who's already logged in, that user is stuck until they log
    out and back in. For permissions that default to False with no
    init-time backfill (e.g. `verticals_upload`), this manifests as
    "I gave Carlos the permission but the upload button is still
    hidden." Re-reading from the DB inside the permission helpers
    avoids that.

    Returns the latest page_access dict, or the existing session
    snapshot if the lookup fails (so a transient DB error doesn't
    randomly deny access).
    """
    uid = session.get("user_id")
    if not uid:
        return session.get("page_access") or {}
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("SELECT page_access FROM users WHERE id = %s", (uid,))
        row = cur.fetchone(); cur.close(); conn.close()
        if row and row["page_access"] is not None:
            pa = row["page_access"]
            session["page_access"] = pa
            return pa
    except Exception:
        # Defensive: any DB hiccup falls back to the session snapshot.
        pass
    return session.get("page_access") or {}


def _vd_can_edit_comments():
    """Server-side gate for vertical comment writes. Admin overrides;
    otherwise the user must hold page_access.verticals_comment. Default
    true on the flag so existing accounts (backfilled in init_db) keep
    the legacy behavior. Re-reads from DB so admin grants apply without
    requiring the user to log out and back in."""
    if session.get("is_admin"):
        return True
    pa = _refresh_page_access_from_db()
    return bool(pa.get("verticals_comment", True))

def _vd_can_upload():
    """Server-side gate for vertical *data* uploads (Master Tracker /
    Master xlsx / PSR xlsx). Mirrors `_vd_can_edit_comments` but
    **defaults to False** — uploads replace project data, so we want
    explicit grants rather than implicit access. Admins always pass.
    Re-reads from DB so admin grants apply without requiring the user
    to log out and back in (this was reported in the wild — Carlos
    granted the perm, button stayed hidden until he re-logged in)."""
    if session.get("is_admin"):
        return True
    pa = _refresh_page_access_from_db()
    return bool(pa.get("verticals_upload", False))

def _capital_can_edit_captable():
    """Server-side gate for Investor cap-table writes (manual edits and
    Excel uploads on the Ember Capital → Investors tab). Mirrors
    `_vd_can_upload`: admins always pass, otherwise the user must hold
    page_access.captable_edit. **Defaults to False** — cap tables drive
    investor-facing returns, so writes need an explicit grant. Re-reads
    from the DB so admin grants apply without a re-login. Viewing the
    Investors tab is gated by the page's existing `portfolio` access;
    only editing requires this flag."""
    if session.get("is_admin"):
        return True
    pa = _refresh_page_access_from_db()
    return bool(pa.get("captable_edit", False))

@app.route("/api/verticals/me", methods=["GET"])
@login_required
def vd_me():
    """Lightweight permission probe used by the Vertical Dashboard SPA
    on load to decide whether to show the Edit-comments dock and the
    per-vertical upload buttons. The dashboard itself enforces nothing
    — the server gates writes — but hiding affordances the user can't
    act on keeps the UI clean."""
    return jsonify({
        "username":          session.get("username"),
        "displayName":       session.get("display_name") or session.get("username"),
        "isAdmin":           bool(session.get("is_admin")),
        "canEditComments":   _vd_can_edit_comments(),
        "canUpload":         _vd_can_upload(),
    })

@app.route("/api/verticals/comments/<key>", methods=["PUT"])
@login_required
def vd_put_comment(key):
    if not _vd_can_edit_comments():
        return jsonify({"error": "Comment editing not permitted for this user"}), 403
    v = _vd_vertical()
    if not v:
        return jsonify({"error": "Unknown vertical"}), 400
    body = request.get_json(silent=True) or {}
    text = body.get("body")
    if not isinstance(text, str):
        return jsonify({"error": "body (string) required"}), 400
    updated_by = body.get("updatedBy") or session.get("username") or "anonymous"
    conn = get_db(); cur = conn.cursor()
    cur.execute("""
        INSERT INTO vd_comments (vertical, section_key, body, updated_by, updated_at)
        VALUES (%s, %s, %s, %s, NOW())
        ON CONFLICT (vertical, section_key) DO UPDATE
          SET body = EXCLUDED.body,
              updated_by = EXCLUDED.updated_by,
              updated_at = NOW()
        RETURNING vertical, section_key, body, updated_by, updated_at
    """, (v, key, text, updated_by))
    row = dict(cur.fetchone())
    conn.commit(); cur.close(); conn.close()
    return jsonify({"success": True, "comment": row})

# ── BULK UPLOAD (called by update_dashboard.py) ──
@app.route("/api/verticals/upload", methods=["POST"])
@login_required
def vd_post_upload():
    # Was @admin_required; now any user holding page_access.verticals_upload
    # can upload (admins always pass). UI in static/verticals/index.html
    # hides the Upload buttons for users who fail this check.
    if not _vd_can_upload():
        return jsonify({"error": "Upload not permitted for this user"}), 403
    v = _vd_vertical()
    if not v:
        return jsonify({"error": "Unknown vertical"}), 400
    body = request.get_json(silent=True) or {}
    units      = body.get("units")
    source     = body.get("sourceFile")
    imported_by = body.get("importedBy") or session.get("username") or "uploader"
    conn = get_db(); cur = conn.cursor()
    try:
        # Only replace the unit roster when the client actually sent
        # some — an empty list used to wipe vd_units (the DELETE ran
        # even with zero rows to insert), which masked a partial-import
        # bug on the LightHaven client: when one sub-importer (Data
        # Table) threw but the other sheets succeeded, units came over
        # as `[]` and the previous data got cleared without replacement.
        # The vd_imports row still landed (and the "Report as of" badge
        # bumped), making the upload look like it "only updated the
        # date." Now: skip the DELETE unless we have rows to put back.
        if isinstance(units, list) and units:
            cur.execute("DELETE FROM vd_units WHERE vertical = %s", (v,))
            for u in units:
                cur.execute("""
                    INSERT INTO vd_units
                      (vertical, addr, unit_num, floorplan, type, bed, sf, phase, status,
                       grid_row, grid_col, tenant_name,
                       lease_executed_date, move_in_date, move_out_date)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    v, u.get("addr"), u.get("num"), u.get("fp"), u.get("type"),
                    u.get("bed"), u.get("sf"), u.get("phase"),
                    _vd_normalize_status(u.get("status")),
                    u.get("grid_row"), u.get("grid_col"), u.get("tenant_name"),
                    u.get("lease_executed_date"), u.get("move_in_date"), u.get("move_out_date"),
                ))
        jsonb_puts = [
            ("vd_noi_data",          body.get("noi")),
            ("vd_leasing_pace_data", body.get("leasingPace")),
            ("vd_traffic_data",      body.get("traffic")),
            ("vd_rents_data",        body.get("rents")),
            ("vd_phase_rollup_data", body.get("phaseRollup")),
            ("vd_psr_data",          body.get("psr")),
        ]
        tabs_updated = []
        for tbl, payload in jsonb_puts:
            if payload is None:
                continue
            cur.execute(f"DELETE FROM {tbl} WHERE vertical = %s", (v,))
            cur.execute(
                f"INSERT INTO {tbl} (vertical, data, updated_by) VALUES (%s, %s, %s)",
                (v, json.dumps(payload), imported_by)
            )
            tabs_updated.append(tbl)
        cur.execute(
            "INSERT INTO vd_imports (vertical, source_file, imported_by, payload) "
            "VALUES (%s, %s, %s, %s)",
            (v, source, imported_by, json.dumps(body))
        )
        conn.commit()
        return jsonify({
            "success": True, "vertical": v,
            "unitsImported": len(units) if isinstance(units, list) else 0,
            "tabsUpdated": tabs_updated,
        })
    except Exception as e:
        conn.rollback()
        return jsonify({"error": "Upload failed", "detail": str(e)}), 500
    finally:
        cur.close(); conn.close()

@app.route("/api/verticals/last-import", methods=["GET"])
@login_required
def vd_last_import():
    v = _vd_vertical()
    if not v:
        return jsonify({"error": "Unknown vertical"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute(
        "SELECT created_at, source_file FROM vd_imports WHERE vertical = %s "
        "ORDER BY created_at DESC LIMIT 1",
        (v,)
    )
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row:
        return jsonify({"vertical": v, "lastImportAt": None, "sourceFile": None})
    return jsonify({"vertical": v, "lastImportAt": row["created_at"].isoformat() if row["created_at"] else None,
                    "sourceFile": row["source_file"]})

@app.route("/api/verticals/imports", methods=["GET"])
@login_required
@admin_required
def vd_imports_history():
    v = _vd_vertical()
    if not v:
        return jsonify({"error": "Unknown vertical"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute(
        "SELECT id, source_file, imported_by, created_at FROM vd_imports "
        "WHERE vertical = %s ORDER BY created_at DESC LIMIT 25",
        (v,)
    )
    rows = []
    for r in cur.fetchall():
        d = dict(r)
        if d.get("created_at"):
            d["created_at"] = d["created_at"].isoformat()
        rows.append(d)
    cur.close(); conn.close()
    return jsonify({"vertical": v, "imports": rows})

# ── SNAPSHOT ────────────────────────────────────────────────────────
# Server-side derivation of every number the Executive Report consumes.
# Mirrors the per-vertical functions in server.js (lightHavenSnapshot /
# hawthorneSnapshot / tgpSnapshot) — kept faithful to those because the
# frontend report templates already speak this exact shape.

def _vd_fetch_comments_map(vertical):
    conn = get_db(); cur = conn.cursor()
    cur.execute(
        "SELECT section_key, body, updated_by, updated_at FROM vd_comments WHERE vertical = %s",
        (vertical,)
    )
    out = {}
    for r in cur.fetchall():
        out[r["section_key"]] = {
            "body": r.get("body") or "",
            "updatedBy": r.get("updated_by"),
            "updatedAt": r["updated_at"].isoformat() if r.get("updated_at") else None,
        }
    cur.close(); conn.close()
    return out

def _vd_lighthaven_snapshot():
    conn = get_db(); cur = conn.cursor()
    cur.execute(
        "SELECT created_at, source_file FROM vd_imports WHERE vertical = 'lighthaven' "
        "ORDER BY created_at DESC LIMIT 1"
    )
    last = cur.fetchone()
    cur.execute("""
        SELECT addr, unit_num AS num, floorplan AS fp, type, bed, sf, phase, status,
               tenant_name, lease_executed_date, move_in_date, move_out_date
        FROM vd_units WHERE vertical = 'lighthaven' ORDER BY phase, addr
    """)
    units = [dict(r) for r in cur.fetchall()]
    if not last and not units:
        cur.close(); conn.close()
        return {"vertical": "lighthaven", "hasData": False,
                "generatedAt": None, "sourceFile": None, "data": None}

    counts = {}
    for u in units:
        k = u.get("status") or "AVAILABLE"
        counts[k] = counts.get(k, 0) + 1
    total = len(units)
    occupied = counts.get("Tenant Occupied", 0) + counts.get("Renewal", 0)
    leased   = occupied + counts.get("Leased", 0) + counts.get("Model", 0)
    occupancy = {
        "total": total, "counts": counts,
        "occupiedPct":  _vd_pct(occupied, total),
        "leasedPct":    _vd_pct(leased, total),
        "availablePct": _vd_pct(counts.get("AVAILABLE", 0), total),
    }

    LEASED_STATUSES = {"Tenant Occupied", "Leased", "Renewal", "Model"}
    fp_map = {}
    for u in units:
        key = u.get("fp") or "Unknown"
        row = fp_map.setdefault(key, {"byPhase": {}, "total": 0, "leased": 0})
        ph = u.get("phase") or 0
        row["byPhase"][ph] = row["byPhase"].get(ph, 0) + 1
        row["total"] += 1
        if u.get("status") in LEASED_STATUSES:
            row["leased"] += 1
    phase_rollup = sorted([
        {"fp": fp, "byPhase": r["byPhase"], "total": r["total"], "leased": r["leased"],
         "pct": _vd_pct(r["leased"], r["total"])}
        for fp, r in fp_map.items()
    ], key=lambda x: x["fp"])

    def _latest_jsonb(table):
        cur.execute(
            f"SELECT data FROM {table} WHERE vertical = 'lighthaven' "
            "ORDER BY created_at DESC LIMIT 1"
        )
        row = cur.fetchone()
        return row["data"] if row else None
    leasing_pace = _latest_jsonb("vd_leasing_pace_data")
    noi          = _latest_jsonb("vd_noi_data")
    rents        = _latest_jsonb("vd_rents_data")
    traffic      = _latest_jsonb("vd_traffic_data")
    cur.close(); conn.close()

    # Synthesize leasing_pace from lease_executed_date if no JSONB record exists.
    if not leasing_pace:
        today = datetime.date.today()
        months_lp, month_keys = [], []
        for i in range(11, -1, -1):
            m = today.month - i
            y = today.year + (m - 1) // 12
            mm = ((m - 1) % 12) + 1
            d = datetime.date(y, mm, 1)
            months_lp.append(d.strftime("%b"))
            month_keys.append(f"{y:04d}-{mm:02d}")
        monthly = [0] * len(month_keys)
        prior = 0
        for u in units:
            led = u.get("lease_executed_date")
            if not led:
                continue
            k = str(led)[:7]
            if k in month_keys:
                monthly[month_keys.index(k)] += 1
            elif k < month_keys[0]:
                prior += 1
        cum = prior
        actual_series = []
        for n in monthly:
            cum += n
            actual_series.append(cum)
        target_budget = max(round(len(units) * 0.85), cum)
        budget_series = [round(target_budget * (i + 1) / len(months_lp))
                         for i in range(len(months_lp))]
        leasing_pace = {"months": months_lp, "budget": budget_series, "actual": actual_series}

    comments = _vd_fetch_comments_map("lighthaven")
    return {
        "vertical": "lighthaven", "hasData": True,
        "generatedAt": last["created_at"].isoformat() if last and last.get("created_at") else None,
        "sourceFile":  last["source_file"] if last else None,
        "data": {
            "units": units, "occupancy": occupancy, "phaseRollup": phase_rollup,
            "leasingPace": leasing_pace or {"months": [], "budget": [], "actual": []},
            "noi":         noi         or {"year": datetime.date.today().year,
                                            "months": [], "budget": [], "forecast": []},
            "rents":       rents       or {"byFloorplan": {}},
            "traffic":     traffic     or {"weeks": [], "visitors": [], "leads": [], "sources": {}},
            "comments":    comments,
        },
    }

def _vd_hawthorne_snapshot():
    conn = get_db(); cur = conn.cursor()
    cur.execute(
        "SELECT data, created_at FROM vd_hawthorne_data WHERE vertical = 'hawthorne' "
        "ORDER BY created_at DESC LIMIT 1"
    )
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row:
        return {"vertical": "hawthorne", "hasData": False,
                "generatedAt": None, "sourceFile": None, "data": None}
    blob = row.get("data") or {}
    units = blob.get("units") or []
    bp = blob.get("bp") or {}
    as_of = blob.get("savedAt") or row["created_at"].isoformat()
    as_of_date = str(as_of)[:10]

    def categorize(u):
        if u.get("status") == "SOLD":
            cd = u.get("closingDate")
            if not cd:
                return "snc"
            return "closed" if str(cd)[:10] <= as_of_date else "snc"
        return "avail"

    def unit_dollar(u, c):
        if c in ("closed", "snc"):
            return u.get("purchasePrice") or u.get("listPrice") or 0
        return u.get("listPrice") or 0

    total = closed = avail = snc = 0
    closed_dollar = 0
    closed_price_sum = 0; closed_price_n = 0
    for u in units:
        total += 1
        c = categorize(u); d = unit_dollar(u, c)
        if c == "closed":
            closed += 1; closed_dollar += d
            cp = u.get("purchasePrice") or u.get("listPrice") or 0
            if cp > 0:
                closed_price_sum += cp; closed_price_n += 1
        elif c == "snc":
            snc += 1
        else:
            avail += 1

    t = bp.get("targets") or {}
    def t_count(k): return ((t.get(k) or {}).get("count")) or 0
    sales = {
        "total":         {"actual": total,  "budget": t_count("total") or total},
        "closed":        {"actual": closed, "budget": t_count("closed")},
        "available":     {"actual": avail,  "budget": t_count("available")},
        "soldNotClosed": {"actual": snc,    "budget": t_count("soldNotClosed")},
        "avgPrice": {
            "actual": round(closed_price_sum / closed_price_n) if closed_price_n else 0,
            "budget": round((t.get("closed") or {}).get("dollar", 0) /
                            ((t.get("closed") or {}).get("count") or 1))
                      if (t.get("closed") or {}).get("count") else 0,
        },
        "gci": {
            "actual": closed_dollar / 1e6,
            "budget": ((bp.get("budget") or {}).get("totalSales", 0)) / 1e6,
        },
    }

    floor_map = {}
    for u in units:
        lvl = u.get("level") or 1
        floor_map.setdefault(lvl, []).append(u)
    stacking = []
    for lvl in sorted(floor_map.keys(), reverse=True):
        units_on_floor = sorted(floor_map[lvl], key=lambda u: str(u.get("title") or ""))
        stacking.append([categorize(u) for u in units_on_floor])

    months, month_keys = [], []
    try:
        as_of_d = datetime.date.fromisoformat(as_of_date)
    except Exception:
        as_of_d = datetime.date.today()
    for i in range(11, -1, -1):
        m = as_of_d.month - i
        y = as_of_d.year + (m - 1) // 12
        mm = ((m - 1) % 12) + 1
        months.append(datetime.date(y, mm, 1).strftime("%b"))
        month_keys.append(f"{y:04d}-{mm:02d}")
    closed_by_month = [0] * 12
    snc_by_month    = [0] * 12
    for u in units:
        if u.get("status") != "SOLD" or not u.get("closingDate"):
            continue
        cd = str(u["closingDate"])[:10]
        k = cd[:7]
        if k not in month_keys:
            continue
        idx = month_keys.index(k)
        if cd <= as_of_date:
            closed_by_month[idx] += 1
        else:
            snc_by_month[idx] += 1
    sales_by_month = {"months": months, "closed": closed_by_month, "snc": snc_by_month}

    fp_map = {}
    for u in units:
        key = u.get("fp") or "Unknown"
        r = fp_map.setdefault(key, {"fp": key, "total": 0, "closed": 0, "snc": 0,
                                     "available": 0, "_priceSum": 0, "_priceN": 0})
        r["total"] += 1
        c = categorize(u)
        if c == "closed":
            r["closed"] += 1
            p = u.get("purchasePrice") or u.get("listPrice") or 0
            if p > 0:
                r["_priceSum"] += p; r["_priceN"] += 1
        elif c == "snc":
            r["snc"] += 1
        else:
            r["available"] += 1
    floorplans = sorted([
        {"fp": r["fp"], "total": r["total"], "closed": r["closed"], "snc": r["snc"],
         "available": r["available"],
         "avgPrice": round(r["_priceSum"] / r["_priceN"]) if r["_priceN"] else 0}
        for r in fp_map.values()
    ], key=lambda x: x["fp"])

    bg = bp.get("budget") or {}; pf = bp.get("proforma") or {}
    proforma = [
        {"line": "Gross Sales Revenue",
         "budget":   bg.get("totalSales", 0) / 1e6,
         "actual":   closed_dollar / 1e6,
         "forecast": (pf.get("totalSales") or bg.get("totalSales", 0)) / 1e6},
        {"line": "Avg Revenue / Unit",
         "budget":   bg.get("revenuePerUnit", 0) / 1e6,
         "actual":   ((closed_price_sum / closed_price_n) if closed_price_n else 0) / 1e6,
         "forecast": pf.get("revenuePerUnit", 0) / 1e6},
        {"line": "Revenue per SF",
         "budget":   bg.get("revenuePerSF", 0) / 1e3,
         "actual":   0,
         "forecast": pf.get("revenuePerSF", 0) / 1e3},
    ]

    comments = _vd_fetch_comments_map("hawthorne")
    return {
        "vertical": "hawthorne", "hasData": True,
        "generatedAt": row["created_at"].isoformat() if row.get("created_at") else None,
        "sourceFile":  blob.get("sourceFile"),
        "data": {"sales": sales, "stacking": stacking, "salesByMonth": sales_by_month,
                  "floorplans": floorplans, "proforma": proforma, "comments": comments},
    }

def _vd_tgp_snapshot():
    conn = get_db(); cur = conn.cursor()
    cur.execute(
        "SELECT data, created_at FROM vd_psr_data WHERE vertical = 'tgp' "
        "ORDER BY created_at DESC LIMIT 1"
    )
    row = cur.fetchone()
    cur.execute(
        "SELECT created_at, source_file FROM vd_imports WHERE vertical = 'tgp' "
        "ORDER BY created_at DESC LIMIT 1"
    )
    last = cur.fetchone()
    cur.close(); conn.close()
    if not row:
        return {"vertical": "tgp", "hasData": False, "generatedAt": None,
                "sourceFile": None, "data": None}
    raw = row.get("data") or {}
    comments = _vd_fetch_comments_map("tgp")

    import re as _re
    def status_color(s):
        if not s:
            return "gray"
        t = str(s).lower()
        if _re.search(r"complete|done|on track|ahead|favorable", t): return "green"
        if _re.search(r"at risk|behind|delay|weather|amber|caution|slip", t): return "amber"
        if _re.search(r"critical|red|over|breach|missed", t): return "red"
        if _re.search(r"in progress|active|started", t): return "blue"
        return "gray"

    metrics = raw.get("metrics") or {}
    budget_total = ((raw.get("budget") or {}).get("total")) or {}
    overview = {
        "totalBudget": budget_total.get("adjusted") or budget_total.get("total") or 0,
        "nrsf":        metrics.get("rentableSF", 0),
        "padAcres":    metrics.get("padSitesAcres", 0),
        "costPerNrsf": metrics.get("costPerNRSF", 0),
        "costPerAcre": metrics.get("costPerAcre", 0),
        "status":      raw.get("overallStatus") or "—",
        "statusColor": status_color(raw.get("overallStatus")),
    }

    i = raw.get("info") or {}
    info = [x for x in [
        {"k": "Project Name", "v": i.get("projectName") or "—"},
        {"k": "Project Type", "v": i.get("projectType") or "—"},
        {"k": "Address",      "v": i.get("address")     or "—"},
        {"k": "City / State", "v": i.get("cityState")   or "—"},
        {"k": "MUD",          "v": i.get("mud")         or "—"},
        {"k": "Structure",    "v": i.get("structure")   or "—"},
    ] if x["v"] != "—"]
    tm = raw.get("team") or {}
    team = [x for x in [
        {"role": "Ember Lead",     "name": tm.get("emberLead")     or "—"},
        {"role": "Developer Lead", "name": tm.get("developerLead") or "—"},
        {"role": "GC",             "name": tm.get("gc")            or "—"},
        {"role": "Architect",      "name": tm.get("architect")     or "—"},
        {"role": "Civil Engineer", "name": tm.get("civilEng")      or "—"},
    ] if x["name"] != "—"]

    def pct_norm(v):
        v = v or 0
        return round(v * (1 if v > 1 else 100))
    budget = []
    for l in ((raw.get("budget") or {}).get("lines") or []):
        budget.append({
            "line":    l.get("item"),
            "retail":  l.get("retail") or 0,
            "pads":    l.get("pad") or 0,
            "realloc": l.get("reallocations") or 0,
            "spent":   l.get("spent") or 0,
            "balance": l.get("balanceToFinish") or 0,
            "pct":     pct_norm(l.get("pctComplete")),
        })

    all_schedule = [{
        "milestone": m.get("milestone"),
        "target":    m.get("target"),
        "actual":    m.get("actual"),
        "variance":  m.get("variance") or 0,
        "status":    status_color(m.get("status")),
        "note":      m.get("notes") or m.get("status") or "—",
        "group":     m.get("group") or "main",
    } for m in (raw.get("schedule") or [])]
    schedule           = [{k: v for k, v in m.items() if k != "group"}
                          for m in all_schedule if m["group"] == "main"]
    leasing_milestones = [{k: v for k, v in m.items() if k != "group"}
                          for m in all_schedule if m["group"] == "leasing"]

    gmp_total = ((raw.get("gmp") or {}).get("total")) or {}
    co_rows   = ((raw.get("changeOrders") or {}).get("rows")) or []
    gmp = {
        "gmpDate": None,
        "gmpAmount": gmp_total.get("adjustedGMP") or gmp_total.get("originalGMP") or 0,
        "buyoutVariance": (gmp_total.get("committed") or 0) - (gmp_total.get("adjustedGMP") or 0),
        "buyoutPct": pct_norm(gmp_total.get("pctBoughtOut")),
        "contingencyOpening": 0,
        "contingencyUsed": (raw.get("changeOrders") or {}).get("total") or 0,
        "contingencyProjected": 0,
        "changeOrders": [{
            "num": str(co.get("co")) if co.get("co") is not None else "",
            "desc": co.get("description") or "",
            "amount": co.get("amount") or 0,
            "status": "Approved" if _re.match(r"(?i)^(y|approved|yes)", str(co.get("approved") or "")) else "In Review",
        } for co in co_rows],
    }

    lu_rows = ((raw.get("leaseUp") or {}).get("rows")) or []
    is_executed = lambda s: bool(_re.match(r"(?i)^(executed|lease executed|signed)", str(s or "")))
    is_loi      = lambda s: bool(_re.search(r"(?i)(loi|negot|in negotiation|in review)", str(s or "")))
    tenants = [{
        "tenant": t.get("tenant") or "—",
        "sf": t.get("sf") or 0,
        "type": t.get("suite") or "Inline",
        "status": t.get("status") or "—",
        "pct": "green" if is_executed(t.get("status")) else
               ("amber" if is_loi(t.get("status")) else "gray"),
    } for t in lu_rows]
    pre_leased_sf = sum((t.get("sf") or 0) for t in lu_rows if is_executed(t.get("status")))
    nrsf = metrics.get("rentableSF", 0)
    lease_up = {
        "nrsf": nrsf,
        "preLeasedSf": pre_leased_sf,
        "preLeasedPct": round((pre_leased_sf / nrsf) * 100) if nrsf > 0 else 0,
        "tenants": tenants,
    }

    def fmt_md(iso):
        if not iso:
            return "—"
        try:
            d = datetime.date.fromisoformat(str(iso)[:10])
            return d.strftime("%b %-d") if hasattr(d, "strftime") else str(iso)
        except Exception:
            return str(iso)
    action_items = [{
        "item":     a.get("item") or a.get("details") or "—",
        "owner":    "—",
        "due":      fmt_md(a.get("target")),
        "priority": a.get("priority") or "Med",
    } for a in (raw.get("actions") or [])]
    issues = [{
        "issue":    it.get("issue") or "—",
        "impact":   it.get("impact") or "—",
        "severity": status_color(it.get("impact")),
    } for it in (raw.get("issues") or [])]

    exec_summary = (comments.get("executive_summary") or {}).get("body") or ""

    return {
        "vertical": "tgp", "hasData": True,
        "generatedAt": (last["created_at"].isoformat() if last and last.get("created_at")
                        else (row["created_at"].isoformat() if row.get("created_at") else None)),
        "sourceFile":  last["source_file"] if last else None,
        "data": {
            "psr": {
                "overview": overview, "info": info, "team": team,
                "executiveSummary": exec_summary,
                "budget": budget, "schedule": schedule, "leasingMilestones": leasing_milestones,
                "gmp": gmp, "leaseUp": lease_up,
                "actionItems": action_items, "issues": issues,
            },
            "comments": comments,
        },
    }

@app.route("/api/verticals/snapshot", methods=["GET"])
@login_required
def vd_snapshot():
    v = _vd_vertical()
    if not v:
        return jsonify({"error": "Unknown vertical"}), 400
    try:
        if v == "lighthaven":
            return jsonify(_vd_lighthaven_snapshot())
        if v == "hawthorne":
            return jsonify(_vd_hawthorne_snapshot())
        if v == "tgp":
            return jsonify(_vd_tgp_snapshot())
    except Exception as e:
        app.logger.exception("vd_snapshot failed")
        return jsonify({"error": "Failed to build snapshot", "detail": str(e)}), 500
    return jsonify({"error": "Unsupported vertical"}), 400

# ── PAGE ROUTE ──────────────────────────────────────────────────────
# Renders the wrapped dashboard. Sub-section navigation lives inside
# the page (left pane = verticals), mirroring the project-returns flow.
@app.route("/verticals")
@login_required
def vd_page():
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("verticals", True):
        return redirect(url_for("home"))
    if session.get("is_admin"):
        pa = {**pa, "verticals": True}
    return render_template(
        "verticals.html",
        username=session.get("username"),
        display_name=session.get("display_name", session.get("username")),
        is_admin=session.get("is_admin"),
        page_access=pa,
    )

# ─── FINANCIAL STATEMENTS DASHBOARD ──────────────────────────────────────────
# Sage-Intacct-backed entity-level Balance Sheet · Income Statement ·
# Statement of Cash Flows. Phase 1 ships BS only — IS and SoCF come in
# follow-up PRs. Trial balances are cached in `intacct_tb_cache` so the
# dashboard renders fast; the Refresh-from-Sage button re-pulls.
# (Imports for sage_intacct + frp_mapping live at the top of the file.)


def _can_view_financials() -> bool:
    """Per-user gate for /financials. Admin overrides; otherwise the
    user must hold page_access.financials. Defaults to False — financial
    statements are sensitive. Re-reads from DB so admin grants apply
    without requiring re-login (same pattern as _vd_can_upload)."""
    if session.get("is_admin"):
        return True
    pa = _refresh_page_access_from_db()
    return bool(pa.get("financials", False))


def _fin_cache_get(entity_id: str, period_name: str):
    """Return cached TB rows + fetched_at, or None if no cache exists.
    Includes accounts_ytd if a Phase 2 YTD upload populated it."""
    conn = get_db(); cur = conn.cursor()
    cur.execute(
        "SELECT accounts, accounts_ytd, fetched_at, fetched_by "
        "FROM intacct_tb_cache WHERE entity_id = %s AND period_name = %s",
        (entity_id, period_name)
    )
    row = cur.fetchone(); cur.close(); conn.close()
    if not row:
        return None
    return {
        "accounts":     row["accounts"] or [],
        "accounts_ytd": row.get("accounts_ytd") or [],
        "fetched_at":   row["fetched_at"],
        "fetched_by":   row.get("fetched_by"),
    }


def _fin_cache_put(entity_id: str, period_name: str, accounts: list,
                   accounts_ytd: list = None):
    """Upsert monthly accounts and (optionally) YTD accounts into the
    cache. Passing accounts_ytd=None preserves whatever YTD data is
    already cached; pass an empty list explicitly to clear it."""
    conn = get_db(); cur = conn.cursor()
    if accounts_ytd is None:
        # Don't touch accounts_ytd; preserve whatever's there.
        cur.execute("""
            INSERT INTO intacct_tb_cache (entity_id, period_name, accounts, fetched_at, fetched_by)
            VALUES (%s, %s, %s, NOW(), %s)
            ON CONFLICT (entity_id, period_name) DO UPDATE
              SET accounts   = EXCLUDED.accounts,
                  fetched_at = NOW(),
                  fetched_by = EXCLUDED.fetched_by
        """, (entity_id, period_name, json.dumps(accounts), session.get("user_id")))
    else:
        cur.execute("""
            INSERT INTO intacct_tb_cache
                (entity_id, period_name, accounts, accounts_ytd, fetched_at, fetched_by)
            VALUES (%s, %s, %s, %s, NOW(), %s)
            ON CONFLICT (entity_id, period_name) DO UPDATE
              SET accounts     = EXCLUDED.accounts,
                  accounts_ytd = EXCLUDED.accounts_ytd,
                  fetched_at   = NOW(),
                  fetched_by   = EXCLUDED.fetched_by
        """, (entity_id, period_name, json.dumps(accounts),
              json.dumps(accounts_ytd), session.get("user_id")))
    conn.commit(); cur.close(); conn.close()


@app.route("/financials")
@login_required
def financials_page():
    if not _can_view_financials():
        return redirect(url_for("home"))
    pa = session.get("page_access") or {}
    return render_template(
        "financials.html",
        username=session.get("username"),
        display_name=session.get("display_name", session.get("username")),
        is_admin=session.get("is_admin", False),
        page_access=pa,
        sage_configured=sage_intacct.is_configured(),
    )


@app.route("/api/bva/sage-discovery", methods=["GET"])
@login_required
@admin_required
def api_bva_sage_discovery():
    """Read-only Sage introspection for designing the Budget-vs-Actuals
    template. Hit /api/bva/sage-discovery for the entity list, then
    /api/bva/sage-discovery?entity=<LOCATIONID> for that project's cost
    dimensions, sample values, and commitment-object availability."""
    if not sage_intacct.is_configured():
        return jsonify({"configured": False,
                        "error": "Sage Intacct is not configured on this server."}), 200
    entity = request.args.get("entity")
    try:
        data = sage_intacct.bva_discovery(entity)
        data["configured"] = True
        return jsonify(data)
    except sage_intacct.IntacctConfigurationError as e:
        return jsonify({"configured": False, "error": str(e)}), 200
    except sage_intacct.IntacctAPIError as e:
        return jsonify({"configured": True, "error": str(e)}), 502


# The three dev projects we control accounting on (label → Sage LOCATIONID).
_BVA_ENTITIES = [
    ("GPD",       "16 - GPD"),
    ("Dennison",  "20 - Emp WRRD"),
    ("WRG",       "10 - Emp Angleton"),
]

# Sage has no major-category field on projects, so we bucket by name. Ordered:
# first bucket with a matching keyword (lowercased substring) wins. Unmatched
# projects fall to "Other". Tweak keywords here to re-group.
_BVA_CATEGORIES = [
    ("Plant & Utilities",   ["wwtp", "water plant", "lift station", "liftstation", "wtp"]),
    ("Collector Roads",     ["baethe", "kermier", " rd ", " rd", "road"]),
    ("Detention & Drainage", ["detention", "outfall", "dtn", "drainage", "fdc"]),
    ("Sections",            ["sec.", "sec ", "section"]),
    ("Amenities",           ["rec center", "lake charlie", "sundancer", "amenit"]),
    ("Marketing",           ["m_", "welcome center", "realtor", "marketing", "spring event", "stratgy", "strategy"]),
    ("Soft Costs",          ["land planning", "professional", "field expense", "g&a", "g-a", "operations"]),
    ("MUD / HOA",           ["mud advance", "hoa advance", " mud", " hoa"]),
    ("Site Work",           ["site work", "sitework"]),
]


def _bva_category(project_name: str) -> str:
    """Map a project name to its major cost category (configurable above)."""
    n = " " + (project_name or "").lower() + " "
    for cat, kws in _BVA_CATEGORIES:
        if any(k in n for k in kws):
            return cat
    return "Other"


def _gen_bva_template_xlsx(blocks: list) -> bytes:
    """Budget-vs-Actuals template: one sheet per project, rows = the project's
    real Sage PROJECT × COSTTYPE combos. Budget column is left blank for the
    user; Committed + Actuals are pre-filled from Sage; Variance is a formula."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    ACCENT = "C56028"; HDR = PatternFill("solid", fgColor="F2EFE8")
    SUB = PatternFill("solid", fgColor="FAF6EC")      # project subtotal
    CAT = PatternFill("solid", fgColor="F3E0D4")      # category subtotal
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    rightal = Alignment(horizontal="right")
    def F(b=False, c="1A1A1A", s=10): return Font(name="Calibri", bold=b, color=c, size=s)
    def money(cell, v=None):
        if v is not None:
            cell.value = v
        cell.number_format = '"$"#,##0'

    wb = openpyxl.Workbook(); wb.remove(wb.active)
    used: set = set()
    def sheet_name(nm):
        base = "".join(ch for ch in (nm or "Sheet") if ch not in '[]:*?/\\')[:31] or "Sheet"
        s, i = base, 2
        while s in used:
            s = base[:28] + "_%d" % i; i += 1
        used.add(s); return s

    # Cols: A Category, B Project, C Task, D Cost Type, E Budget, F Committed,
    # G Actuals, H Budget−Committed, I Committed−Actuals, J Budget−Actuals
    headers = ["Category", "Project", "Task", "Cost Type", "Budget", "Committed",
               "Actuals", "Budget − Committed", "Committed − Actuals", "Budget − Actuals"]
    NC = len(headers)
    for blk in blocks:
        ws = wb.create_sheet(sheet_name(blk["label"]))
        ws.cell(1, 1, "%s — Budget vs Actuals" % blk["label"]).font = Font(name="Calibri", bold=True, size=14, color=ACCENT)
        ws.cell(2, 1, "Sage entity: %s · Fill the Budget column; Committed & Actuals pulled from Sage. B−C = uncommitted budget · C−A = open commitment · B−A = budget vs spend." % blk["entity"]).font = F(False, "888888", 9)
        r = 4
        for ci, h in enumerate(headers, 1):
            c = ws.cell(r, ci, h); c.font = F(True, "1A1A1A", 9); c.fill = HDR; c.border = border
            if ci >= 5:
                c.alignment = rightal
        r += 1

        def var_cells(ridx, bold=False):
            for col, fexpr in ((8, "=E%d-F%d"), (9, "=F%d-G%d"), (10, "=E%d-G%d")):
                cc = ws.cell(ridx, col); cc.value = fexpr % (ridx, ridx); money(cc); cc.border = border
                if bold:
                    cc.font = F(True)

        # Group by category → project
        cat_groups: dict = {}; cat_order: list = []
        for row in blk.get("rows", []):
            cat = row.get("category") or "Other"
            if cat not in cat_groups:
                cat_groups[cat] = {"projs": {}, "order": []}; cat_order.append(cat)
            pn = row.get("project_name") or row.get("project") or "(no project)"
            cg = cat_groups[cat]
            if pn not in cg["projs"]:
                cg["projs"][pn] = []; cg["order"].append(pn)
            cg["projs"][pn].append(row)

        cat_subtotal_rows = []
        for cat in cat_order:
            cg = cat_groups[cat]
            proj_subtotal_rows = []
            for pn in cg["order"]:
                prows = sorted(cg["projs"][pn], key=lambda x: x.get("costtype") or "")
                pstart = r
                for row in prows:
                    ws.cell(r, 1, cat).border = border
                    ws.cell(r, 2, pn).border = border
                    ws.cell(r, 3, row.get("task") or "").border = border
                    ws.cell(r, 4, row.get("costtype") or "").border = border
                    bc = ws.cell(r, 5); money(bc); bc.border = border                      # Budget blank
                    cc = ws.cell(r, 6); money(cc, int(round(row.get("committed") or 0))); cc.border = border
                    ac = ws.cell(r, 7); money(ac, int(round(row.get("actual") or 0))); ac.border = border
                    var_cells(r)
                    r += 1
                for ci in range(1, NC + 1):
                    ws.cell(r, ci).fill = SUB; ws.cell(r, ci).border = border
                ws.cell(r, 2, pn + " — Total").font = F(True)
                for col in (5, 6, 7):
                    L = get_column_letter(col)
                    tc = ws.cell(r, col); tc.value = "=SUM(%s%d:%s%d)" % (L, pstart, L, r - 1); money(tc); tc.font = F(True)
                var_cells(r, bold=True)
                proj_subtotal_rows.append(r)
                r += 1
            # Category subtotal = sum of its project subtotals
            for ci in range(1, NC + 1):
                ws.cell(r, ci).fill = CAT; ws.cell(r, ci).border = border
            ws.cell(r, 1, cat + " — TOTAL").font = F(True, ACCENT)
            for col in (5, 6, 7):
                L = get_column_letter(col)
                expr = "+".join("%s%d" % (L, pr) for pr in proj_subtotal_rows) or "0"
                tc = ws.cell(r, col); tc.value = "=" + expr; money(tc); tc.font = F(True, ACCENT)
            var_cells(r, bold=True)
            cat_subtotal_rows.append(r)
            r += 2

        # Grand total = sum of category subtotals
        for ci in range(1, NC + 1):
            ws.cell(r, ci).fill = CAT; ws.cell(r, ci).border = border
        ws.cell(r, 1, "ALL CATEGORIES").font = F(True)
        for col in (5, 6, 7):
            L = get_column_letter(col)
            expr = "+".join("%s%d" % (L, cr) for cr in cat_subtotal_rows) or "0"
            tc = ws.cell(r, col); tc.value = "=" + expr; money(tc); tc.font = F(True)
        var_cells(r, bold=True)

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 26
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 14
        for col in ("E", "F", "G", "H", "I", "J"):
            ws.column_dimensions[col].width = 14

    bio = io.BytesIO(); wb.save(bio)
    return bio.getvalue()


@app.route("/api/bva/template.xlsx", methods=["GET"])
@login_required
@admin_required
def api_bva_template():
    """Generate the BVA budget template (GPD, Dennison, WRG), one tab each,
    rows = each project's Sage PROJECT × COSTTYPE combos with Actuals +
    Committed pre-filled and Budget left blank to fill in."""
    if not sage_intacct.is_configured():
        return jsonify({"error": "Sage Intacct is not configured on this server."}), 400
    blocks = []
    for label, eid in _BVA_ENTITIES:
        try:
            actuals = sage_intacct.bva_actuals(eid)
        except sage_intacct.IntacctAPIError as e:
            return jsonify({"error": "Actuals pull failed for %s: %s" % (label, e)}), 502
        prefixes = {p.split("_", 1)[0].strip() for (p, _ct) in actuals.keys()} or None
        try:
            commits = sage_intacct.bva_commitments(eid, prefixes)
        except Exception:
            commits = {}
        keys = set(actuals.keys()) | set(commits.keys())
        rows = []
        for k in keys:
            proj, ct = k
            a = actuals.get(k, {})
            pname = a.get("project_name") or a.get("project") or proj
            rows.append({
                "project":      a.get("project") or proj,
                "project_name": pname,
                "category":     _bva_category(pname),
                "task":         a.get("task", ""),
                "costtype":     ct,
                "actual":       a.get("actual", 0.0),
                "committed":    commits.get(k, 0.0),
            })
        blocks.append({"label": label, "entity": eid, "rows": rows})
    xlsx = _gen_bva_template_xlsx(blocks)
    fname = "BVA_Template_%s.xlsx" % datetime.datetime.now().strftime("%Y-%m-%d")
    return send_file(io.BytesIO(xlsx),
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=fname)


# ─── BVA budget storage + dashboard data ─────────────────────────────────────
_BVA_KEYSEP = ""   # project<sep>costtype budget key


def _bva_can_view() -> bool:
    """View gate for the BVA page/data: admins always; else page_access.bva."""
    if session.get("is_admin"):
        return True
    return bool(_refresh_page_access_from_db().get("bva", False))


def _bva_load_budgets() -> dict:
    """Stored budgets: {entity_id: {"<project>\\x1f<costtype>": amount}}."""
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT data FROM reports WHERE report_type = 'bva_budget' "
                "ORDER BY uploaded_at DESC LIMIT 1")
    row = cur.fetchone(); cur.close(); conn.close()
    return ((row["data"] if row else {}) or {}).get("budgets", {}) or {}


def _bva_save_budgets(budgets: dict) -> None:
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM reports WHERE report_type = 'bva_budget'")
    cur.execute("INSERT INTO reports (report_type, data, uploaded_by) VALUES (%s, %s, %s)",
                ("bva_budget", json.dumps({"budgets": budgets}), session.get("user_id")))
    conn.commit(); cur.close(); conn.close()


@app.route("/api/bva/data", methods=["GET"])
@login_required
def api_bva_data():
    """Per-project Budget vs Committed vs Actuals rows for GPD/Dennison/WRG,
    merging stored budgets with live Sage actuals + commitments."""
    if not _bva_can_view():
        return jsonify({"error": "forbidden"}), 403
    if not sage_intacct.is_configured():
        return jsonify({"configured": False, "blocks": []}), 200
    budgets = _bva_load_budgets()
    blocks = []
    for label, eid in _BVA_ENTITIES:
        try:
            actuals = sage_intacct.bva_actuals(eid)
        except sage_intacct.IntacctAPIError as e:
            return jsonify({"configured": True, "error": "Sage error for %s: %s" % (label, e)}), 502
        prefixes = {p.split("_", 1)[0].strip() for (p, _ct) in actuals.keys()} or None
        try:
            commits = sage_intacct.bva_commitments(eid, prefixes)
        except Exception:
            commits = {}
        ebud = budgets.get(eid, {})
        rows = []
        seen = set()
        for k in (set(actuals) | set(commits)):
            proj, ct = k
            a = actuals.get(k, {})
            pname = a.get("project_name") or a.get("project") or proj
            bkey = pname + _BVA_KEYSEP + ct
            seen.add(bkey)
            bud = float(ebud.get(bkey, 0) or 0)
            act = round(a.get("actual", 0.0))
            com = round(commits.get(k, 0.0))
            rows.append({"project": pname, "category": _bva_category(pname),
                         "task": a.get("task", ""), "costtype": ct,
                         "budget": round(bud), "committed": com, "actual": act,
                         "variance": round(bud - com - act)})
        # Budgeted lines with no Sage activity yet still show up.
        for bkey, amt in ebud.items():
            if bkey in seen:
                continue
            pname, _, ct = bkey.partition(_BVA_KEYSEP)
            amt = round(float(amt or 0))
            rows.append({"project": pname, "category": _bva_category(pname),
                         "task": "", "costtype": ct,
                         "budget": amt, "committed": 0, "actual": 0, "variance": amt})
        rows.sort(key=lambda r: (r["category"], r["project"], r["costtype"]))
        blocks.append({"label": label, "entity": eid, "rows": rows})
    return jsonify({"configured": True, "blocks": blocks})


@app.route("/api/bva/budget", methods=["POST"])
@login_required
def api_bva_budget_upload():
    """Ingest a filled-in BVA template (the workbook from /api/bva/template.xlsx)
    and store its Budget column, keyed by entity + project + cost type."""
    if not _bva_can_view():
        return jsonify({"error": "forbidden"}), 403
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file uploaded"}), 400
    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
    except Exception as e:
        return jsonify({"error": "Could not read workbook: %s" % e}), 400
    budgets = _bva_load_budgets()
    imported = 0
    for ws in wb.worksheets:
        eid = None
        for lbl, e in _BVA_ENTITIES:
            if ws.title.strip().lower().startswith(lbl.strip().lower()[:28]):
                eid = e; break
        if not eid:
            continue
        # Locate the header row (has both "project" and "budget").
        hdr = None; cols = {}
        for r in range(1, min(ws.max_row, 15) + 1):
            vals = [str(ws.cell(r, c).value or "").strip().lower() for c in range(1, 9)]
            if "project" in vals and "budget" in vals:
                hdr = r; cols = {v: i + 1 for i, v in enumerate(vals) if v}
                break
        if not hdr:
            continue
        pcol, ccol, bcol = cols.get("project"), cols.get("cost type"), cols.get("budget")
        if not (pcol and ccol and bcol):
            continue
        ebud = budgets.setdefault(eid, {})
        for r in range(hdr + 1, ws.max_row + 1):
            proj = str(ws.cell(r, pcol).value or "").strip()
            ct = str(ws.cell(r, ccol).value or "").strip()
            if not proj or not ct or "total" in proj.lower():
                continue
            try:
                amt = float(ws.cell(r, bcol).value)
            except (TypeError, ValueError):
                continue
            ebud[proj + _BVA_KEYSEP + ct] = amt
            imported += 1
    _bva_save_budgets(budgets)
    return jsonify({"ok": True, "imported": imported})


def _gen_bva_actualize_xlsx(blocks: list) -> bytes:
    """Actuals matrix for entering into the pro-forma: one sheet per project,
    rows = (Project ID, Task), columns = months, cells = that month's actual
    spend, plus a row Total."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    ACCENT = "C56028"; HDR = PatternFill("solid", fgColor="F2EFE8")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    rightal = Alignment(horizontal="right")
    def F(b=False, c="1A1A1A", s=10): return Font(name="Calibri", bold=b, color=c, size=s)

    wb = openpyxl.Workbook(); wb.remove(wb.active)
    used: set = set()
    def sheet_name(nm):
        base = "".join(ch for ch in (nm or "Sheet") if ch not in '[]:*?/\\')[:31] or "Sheet"
        s, i = base, 2
        while s in used:
            s = base[:28] + "_%d" % i; i += 1
        used.add(s); return s

    for blk in blocks:
        months = blk.get("months", [])
        ws = wb.create_sheet(sheet_name(blk["label"]))
        ws.cell(1, 1, "%s — Actuals by Project / Task / Month" % blk["label"]).font = Font(name="Calibri", bold=True, size=14, color=ACCENT)
        ws.cell(2, 1, "Sage entity: %s · signed GL actuals by month — paste into the pro-forma" % blk["entity"]).font = F(False, "888888", 9)
        r = 4
        # Category | Project ID | Task | Task Name | <months…> | Total
        headers = ["Category", "Project ID", "Task", "Task Name"] + months + ["Total"]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(r, ci, h); c.font = F(True, "1A1A1A", 9); c.fill = HDR; c.border = border
            if ci > 4:
                c.alignment = rightal
        r += 1
        base = 5  # first month column
        for row in sorted(blk.get("rows", []),
                          key=lambda x: (x.get("category") or "", x.get("project") or "", x.get("task") or "")):
            ws.cell(r, 1, row.get("category", "")).border = border
            ws.cell(r, 2, row.get("project", "")).border = border
            ws.cell(r, 3, row.get("task", "")).border = border
            ws.cell(r, 4, row.get("task_name", "")).border = border
            tot = 0.0
            mvals = row.get("months", {})
            for j, mk in enumerate(months):
                v = mvals.get(mk, 0.0); tot += v
                c = ws.cell(r, base + j)
                if v:
                    c.value = round(v)
                c.number_format = "#,##0"; c.border = border
            tc = ws.cell(r, base + len(months)); tc.value = round(tot)
            tc.number_format = "#,##0"; tc.border = border; tc.font = F(True)
            r += 1
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 24
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 20
        for j in range(len(months) + 1):
            ws.column_dimensions[get_column_letter(base + j)].width = 11

    bio = io.BytesIO(); wb.save(bio)
    return bio.getvalue()


@app.route("/api/bva/actualize.xlsx", methods=["GET"])
@login_required
def api_bva_actualize():
    """Export all actuals by Project ID × Task × month (one tab per project)
    for entering into the development pro-forma."""
    if not _bva_can_view():
        return jsonify({"error": "forbidden"}), 403
    if not sage_intacct.is_configured():
        return jsonify({"error": "Sage Intacct is not configured on this server."}), 400
    blocks = []
    for label, eid in _BVA_ENTITIES:
        try:
            d = sage_intacct.bva_actuals_monthly(eid)
        except sage_intacct.IntacctAPIError as e:
            return jsonify({"error": "Sage error for %s: %s" % (label, e)}), 502
        rows = d["rows"]
        for row in rows:
            row["category"] = _bva_category(row.get("project", ""))
        blocks.append({"label": label, "entity": eid,
                       "rows": rows, "months": d["months"]})
    xlsx = _gen_bva_actualize_xlsx(blocks)
    fname = "BVA_Actuals_by_Month_%s.xlsx" % datetime.datetime.now().strftime("%Y-%m-%d")
    return send_file(io.BytesIO(xlsx),
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=fname)


@app.route("/api/bva/preview", methods=["GET"])
@login_required
@admin_required
def api_bva_preview():
    """Verification dump: per-entity actuals total + breakdown by GL account
    and by project, plus the commitments total — so we can sanity-check the
    numbers (and spot non-cost accounts leaking in) before trusting them."""
    if not sage_intacct.is_configured():
        return jsonify({"error": "Sage Intacct is not configured on this server."}), 400
    sel = request.args.get("entity")
    targets = [(l, e) for (l, e) in _BVA_ENTITIES if (not sel or e == sel)]
    out = []
    for label, eid in targets:
        try:
            audit = sage_intacct.bva_audit(eid)
        except sage_intacct.IntacctAPIError as e:
            return jsonify({"error": "Sage error for %s: %s" % (label, e)}), 502
        prefixes = {p["project"].split("_", 1)[0].strip()
                    for p in audit.get("by_project", []) if p.get("project")} or None
        try:
            commits = sage_intacct.bva_commitments(eid, prefixes)
        except Exception:
            commits = {}
        out.append({
            "label": label, "entity": eid,
            "actuals_total": audit["total"],
            "entries_scanned": audit["entries_scanned"],
            "cost_coded_entries": audit["cost_coded_entries"],
            "dropped_no_tr_type": audit["dropped_no_tr_type"],
            "by_account": audit["by_account"],
            "by_project": audit["by_project"],
            "commitments_total": round(sum(commits.values())) if commits else 0,
            "commitment_lines": len(commits),
        })
    return jsonify({"configured": True, "preview": out})


@app.route("/bva", methods=["GET"])
@app.route("/budget-vs-actuals", methods=["GET"])
@login_required
def bva_page():
    pa = _refresh_page_access_from_db() or {}
    if not session.get("is_admin") and not pa.get("bva", False):
        return redirect(url_for("home"))
    return render_template("bva.html",
                           username=session.get("username"),
                           is_admin=session.get("is_admin", False),
                           page_access=pa)


@app.route("/api/financials/entities", methods=["GET"])
@login_required
def api_financials_entities():
    """Return the entity list. When called with ?period=<name>, also tag
    each entity with has_cache:bool indicating whether intacct_tb_cache
    holds a row for (entity_id, period) — the frontend uses this to dim
    entities the user hasn't uploaded a TB for yet."""
    if not _can_view_financials():
        return jsonify({"error": "forbidden"}), 403
    if not sage_intacct.is_configured():
        return jsonify({"configured": False, "entities": []}), 200
    try:
        entities = sage_intacct.list_entities()
    except sage_intacct.IntacctConfigurationError as e:
        return jsonify({"configured": False, "error": str(e)}), 200
    except sage_intacct.IntacctAPIError as e:
        return jsonify({"configured": True, "entities": [], "error": str(e)}), 502

    period = (request.args.get("period") or "").strip()
    if period and entities:
        try:
            conn = get_db(); cur = conn.cursor()
            cur.execute(
                "SELECT entity_id FROM intacct_tb_cache WHERE period_name = %s",
                (period,),
            )
            cached_ids = {row["entity_id"] for row in cur.fetchall()}
            cur.close(); conn.close()
            for e in entities:
                e["has_cache"] = e["id"] in cached_ids
        except Exception as exc:
            app.logger.warning("entities: has_cache lookup failed: %s", exc)
            # Don't fail the response — just omit the field
    return jsonify({"configured": True, "entities": entities})


@app.route("/api/financials/periods", methods=["GET"])
@login_required
def api_financials_periods():
    if not _can_view_financials():
        return jsonify({"error": "forbidden"}), 403
    if not sage_intacct.is_configured():
        return jsonify({"configured": False, "periods": []}), 200
    try:
        periods = sage_intacct.list_periods(closed_only=True)
        return jsonify({"configured": True, "periods": periods})
    except sage_intacct.IntacctConfigurationError as e:
        return jsonify({"configured": False, "error": str(e)}), 200
    except sage_intacct.IntacctAPIError as e:
        return jsonify({"configured": True, "periods": [], "error": str(e)}), 502


def _fin_render_bs(accounts: list) -> dict:
    """Run the mapping over a TB and shape it for the template."""
    rolled = roll_up_balance_sheet(accounts)
    summary = summarize_bs(rolled)
    # Convert OrderedDicts to plain dicts for jsonify
    structure = []
    for sec, subs in summary["structure"].items():
        sec_obj = {"section": sec, "total": summary["section_totals"][sec], "subsections": []}
        for sub, rows in subs.items():
            sec_obj["subsections"].append({
                "subsection": sub,
                "total":      summary["subsection_totals"][(sec, sub)],
                "line_items": [
                    {"line_item": r["line_item"], "value": r["value"], "is_contra": r["is_contra"]}
                    for r in rows
                ],
            })
        structure.append(sec_obj)
    return {
        "structure":         structure,
        "total_assets":      summary["total_assets"],
        "total_liab_and_eq": summary["total_liab_and_eq"],
        "footing_check":     summary["footing_check"],
    }


@app.route("/api/financials/balance-sheet", methods=["GET"])
@login_required
def api_financials_balance_sheet():
    """Return the rolled-up BS for the cached TB of (entity, period).
    Does NOT call Sage — refresh hits a separate endpoint.

    Routes through tb_parser (same renderer as /balance-sheet-tb) so the
    cached TB-upload data renders correctly for *every* user, not just the
    one whose session set FIN.useTbUpload after uploading. The data was
    always shared in Postgres; this is what makes it visible.

    The legacy _fin_render_bs (frp_mapping) path only computed correctly
    for GLENTRY-aggregated accounts, which we abandoned — keeping the
    route name for backward compatibility but unifying the renderer.
    """
    if not _can_view_financials():
        return jsonify({"error": "forbidden"}), 403
    try:
        import tb_parser
    except ImportError as e:
        return jsonify({"error": f"tb_parser import failed: {e}"}), 500
    entity = (request.args.get("entity") or "").strip()
    period = (request.args.get("period") or "").strip()
    if not entity or not period:
        return jsonify({"error": "entity and period required"}), 400
    cache = _fin_cache_get(entity, period)
    if not cache:
        return jsonify({
            "cached":  False,
            "message": "No data cached for this entity/period. Click Refresh from Sage.",
        }), 200
    v  = tb_parser.compute_bs(cache["accounts"])
    bs = tb_parser.render_bs(v)
    return jsonify({
        "cached":        True,
        "source":        "tb_upload",
        "entity":        entity,
        "period":        period,
        "fetched_at":    cache["fetched_at"].isoformat() if cache["fetched_at"] else None,
        "account_count": len(cache["accounts"]),
        "balance_sheet": bs,
    })


@app.route("/api/financials/refresh", methods=["POST"])
@login_required
def api_financials_refresh():
    """Pull fresh TB from Sage for (entity, period), cache it, return
    the rolled-up BS. This is the slow path (~5-30 sec depending on
    entity size)."""
    if not _can_view_financials():
        return jsonify({"error": "forbidden"}), 403
    entity = (request.args.get("entity") or "").strip()
    period = (request.args.get("period") or "").strip()
    if not entity or not period:
        return jsonify({"error": "entity and period required"}), 400
    if not sage_intacct.is_configured():
        return jsonify({"error": "Sage Intacct credentials not configured"}), 503
    # Wrap everything in try/except so any runtime error returns a
    # JSON payload instead of a Flask HTML error page (the latter
    # is what's causing the "Unexpected token '<', '<html>'" error).
    try:
        return _api_financials_refresh_impl(entity, period)
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        app.logger.error("api_financials_refresh failed: %s\n%s", e, tb)
        return jsonify({
            "error": f"{type(e).__name__}: {e}",
            "traceback": tb.splitlines()[-30:],  # last 30 lines for context
        }), 500


def _api_financials_refresh_impl(entity: str, period: str):
    try:
        accounts = sage_intacct.get_trial_balance(entity, period)
    except sage_intacct.IntacctConfigurationError as e:
        return jsonify({"error": str(e)}), 503
    except sage_intacct.IntacctAPIError as e:
        return jsonify({"error": f"Sage API error: {e}"}), 502
    _fin_cache_put(entity, period, accounts)
    bs = _fin_render_bs(accounts)
    # Always attach a diagnostic block now — we've been blind too many
    # times to "non-zero accounts but 0 line items rendered" cases.
    # Includes the account sample so we can see what classify_bs is
    # seeing (number, name, signed close balance, what line item it
    # mapped to). When 0 entries came back, ALSO includes the sample
    # GLENTRY.LOCATION values so we can fix the filter format.
    from frp_mapping import classify_bs as _classify_bs
    # Full classified account list, sorted by absolute close balance —
    # biggest dollars first, so misclassifications jump out at the top.
    accounts_full = []
    for a in accounts:
        accounts_full.append({
            "no":         a.get("no", ""),
            "name":       a.get("name", ""),
            "close":      a.get("close", 0.0),
            "mapped_to":  _classify_bs(a.get("no", ""), a.get("name", "")) or "(unmapped)",
        })
    accounts_full.sort(key=lambda a: -abs(a.get("close", 0.0)))
    # Per-location breakdown for accounts where it matters most:
    #   - abs(balance) > $1M (big-dollar items),  OR
    #   - more than one LOCATION contributed (potential double-counting).
    # Each entry shows the per-location split, sorted by abs contribution.
    per_loc_raw = getattr(sage_intacct.get_trial_balance,
                          "_last_per_location_sums", {}) or {}
    per_location_balances = []
    for a in accounts_full:
        no = a["no"]
        locs = per_loc_raw.get(no, {})
        if not locs:
            continue
        big = abs(a["close"]) > 1_000_000
        split = len(locs) > 1
        if not (big or split):
            continue
        per_location_balances.append({
            "no":        no,
            "name":      a["name"],
            "close":     a["close"],
            "mapped_to": a["mapped_to"],
            "by_location": sorted(
                [{"loc": L, "sum": S} for L, S in locs.items()],
                key=lambda x: -abs(x["sum"]),
            ),
        })
    # Per-year DR/CR tally + raw-entry samples for the TOP-15 accounts
    # by |close|. Targets the "is this 2x because year-opening JEs
    # replay cumulative balance" theory: if Land shows e.g. DR=$35M in
    # year 1, then DR=$35M again in year 2 (no new acquisitions but
    # same amount appears), that's the opening-JE doubling pattern.
    per_year_raw     = getattr(sage_intacct.get_trial_balance,
                               "_last_per_year_sums", {}) or {}
    samples_raw      = getattr(sage_intacct.get_trial_balance,
                               "_last_account_samples", {}) or {}
    bt_raw           = getattr(sage_intacct.get_trial_balance,
                               "_last_per_batchtitle_sums", {}) or {}
    # Always include these "watch" accounts in the per-account detail
    # regardless of |close|, so AP and other small-but-important
    # accounts stay visible after the commit filter shrinks them.
    PINNED_DETAIL_ACCOUNTS = {"20030"}  # Trade Payables
    pinned_accts = [a for a in accounts_full if a["no"] in PINNED_DETAIL_ACCOUNTS]
    rest_accts   = [a for a in accounts_full if a["no"] not in PINNED_DETAIL_ACCOUNTS]
    detail_pool  = pinned_accts + rest_accts[:max(0, 30 - len(pinned_accts))]
    account_time_detail = []
    for a in detail_pool:
        no = a["no"]
        yrs = per_year_raw.get(no, {})
        if not yrs:
            continue
        # by_year sorted by year string asc
        by_year = sorted(
            [{"year": y, "dr": v["dr"], "cr": v["cr"], "net": v["dr"] - v["cr"], "count": v["count"]}
             for y, v in yrs.items()],
            key=lambda x: x["year"],
        )
        # by_batchtitle sorted by abs(net) desc, top 25 so we see more
        # of the tail. Each row = a JE batch grouping that contributed
        # to this account — the AJE pattern we're hunting will pop out
        # as the biggest contributor.
        bts = bt_raw.get(no, {})
        by_bt = sorted(
            [{"batchtitle": bt, "dr": v["dr"], "cr": v["cr"],
              "net": v["dr"] - v["cr"], "count": v["count"]}
             for bt, v in bts.items()],
            key=lambda x: -abs(x["net"]),
        )[:25]
        account_time_detail.append({
            "no":             no,
            "name":           a["name"],
            "close":          a["close"],
            "mapped_to":      a["mapped_to"],
            "by_year":        by_year,
            "by_batchtitle":  by_bt,
            "samples":        samples_raw.get(no, []),
        })
    # Resolve what sub-locations get queried so we can see the fan-out
    # in the diagnostic (helps when only the parent's name is visible
    # in the dropdown but the actual postings are on children).
    try:
        sub_locs = sage_intacct._entity_location_set(entity)
    except Exception:
        sub_locs = [entity]
    placeholder_filtered = getattr(sage_intacct.get_trial_balance,
                                   "_last_placeholder_filtered", []) or []
    exclude_reason_counts = getattr(sage_intacct.get_trial_balance,
                                    "_last_exclude_reason_counts", {}) or {}
    # Global by-BATCHTITLE rollup across every account. Useful when an
    # account we care about (e.g., AP after the commit filter) drops
    # out of the top-30 list. Whatever AJE pattern is still inflating
    # the BS will show up here as a giant DR or CR batchtitle.
    bt_global: dict = {}
    for acct_no, by_bt in bt_raw.items():
        for bt, v in by_bt.items():
            agg = bt_global.setdefault(bt, {"dr": 0.0, "cr": 0.0, "count": 0})
            agg["dr"]    += v["dr"]
            agg["cr"]    += v["cr"]
            agg["count"] += v["count"]
    by_batchtitle_global = sorted(
        [{"batchtitle": bt, "dr": v["dr"], "cr": v["cr"],
          "net": v["dr"] - v["cr"], "count": v["count"]}
         for bt, v in bt_global.items()],
        key=lambda x: -abs(x["net"]),
    )[:40]
    diagnostic = {
        "filter_value_used":      entity,
        "sub_locations_queried":  sub_locs,
        "account_count":          len(accounts),
        "accounts_full":          accounts_full,
        "per_location_balances":  per_location_balances,
        "account_time_detail":    account_time_detail,
        "by_batchtitle_global":   by_batchtitle_global,
        "placeholder_filtered":   placeholder_filtered,
        "exclude_reason_counts":  exclude_reason_counts,
    }
    if not accounts:
        sample_locs = getattr(sage_intacct.get_trial_balance,
                              "_last_sample_locations", None)
        if sample_locs is not None:
            diagnostic["sample_glentry_locations"] = sample_locs
            diagnostic["note"] = (
                "0 GLENTRY rows matched the LOCATION filter. "
                "Compare `filter_value_used` to `sample_glentry_locations` "
                "to see what format Sage stores in GLENTRY.LOCATION."
            )
    return jsonify({
        "cached":        True,
        "entity":        entity,
        "period":        period,
        "fetched_at":    datetime.datetime.utcnow().isoformat() + "Z",
        "account_count": len(accounts),
        "balance_sheet": bs,
        "diagnostic":    diagnostic,
    })


@app.route("/api/financials/list-reports", methods=["GET"])
@login_required
def api_financials_list_reports():
    """List saved Sage reports so we can find an already-saved TB
    report on this instance. Returns object-probe results — if any
    object returns rows, those are the available reports."""
    if not _can_view_financials():
        return jsonify({"error": "forbidden"}), 403
    if not sage_intacct.is_configured():
        return jsonify({"error": "Sage Intacct credentials not configured"}), 503
    try:
        return jsonify(sage_intacct.list_saved_reports())
    except Exception as e:
        import traceback
        return jsonify({
            "error":     f"{type(e).__name__}: {e}",
            "traceback": traceback.format_exc().splitlines()[-20:],
        }), 500


@app.route("/api/financials/pull-tb-report", methods=["POST"])
@login_required
def api_financials_pull_tb_report():
    """Pull a Trial Balance live from Sage via readReport. Bypasses
    GLENTRY aggregation by invoking the canonical TB report directly."""
    if not _can_view_financials():
        return jsonify({"error": "forbidden"}), 403
    entity = (request.args.get("entity") or "").strip()
    period = (request.args.get("period") or "").strip()
    if not entity or not period:
        return jsonify({"error": "entity and period required"}), 400
    if not sage_intacct.is_configured():
        return jsonify({"error": "Sage Intacct credentials not configured"}), 503
    try:
        result = sage_intacct.pull_tb_via_report(entity, period)
    except sage_intacct.IntacctAPIError as e:
        return jsonify({"error": f"Sage API error: {e}"}), 502
    except Exception as e:
        return jsonify({
            "error":     f"{type(e).__name__}: {e}",
            "ok":        False,
        }), 500
    if not result.get("ok"):
        # Surface the per-report-name attempts so the user (or sysadmin)
        # can see why each candidate failed. Most likely the accountant
        # needs to save a custom report — explain that in the message.
        return jsonify({
            "ok":      False,
            "error":   result.get("message"),
            "attempts": result.get("attempts", []),
        }), 200
    # Success — cache and render.
    accounts = result["accounts"]
    _fin_cache_put(entity, period, accounts)
    # Render via the TB-upload-style code path (frp_builder predicates),
    # since the report rows come pre-aggregated by Sage exactly like an
    # uploaded TB would.
    try:
        import tb_parser
        v  = tb_parser.compute_bs(accounts)
        bs = tb_parser.render_bs(v)
    except Exception as e:
        return jsonify({
            "ok":      False,
            "error":   f"Parsed report but BS render failed: {type(e).__name__}: {e}",
        }), 500
    return jsonify({
        "ok":            True,
        "source":        "readReport",
        "report_name":   result["report_name"],
        "entity":        entity,
        "period":        period,
        "fetched_at":    datetime.datetime.utcnow().isoformat() + "Z",
        "account_count": len(accounts),
        "balance_sheet": bs,
        "attempts":      result["attempts"],
    })


@app.route("/api/financials/upload-tb", methods=["POST"])
@login_required
def api_financials_upload_tb():
    """Upload Sage Intacct HTML-formatted .xls TB exports and render the
    BS directly from them.

    This sidesteps the entire GLENTRY-aggregation rabbit hole: Sage's
    TB export is the canonical, pre-filtered source of truth (matches
    FRP to the penny). User exports Current-Month and Current-YTD TBs
    from the Sage UI, uploads here, we parse + cache + render.

    Form fields:
        monthly_tb: required, the Current-Month TB .xls file
        ytd_tb:     optional (needed for IS/SCF in later phases),
                    Current-YTD TB .xls file
        period:     period name to cache under (e.g. "Month Ended March 2026")
    """
    if not _can_view_financials():
        return jsonify({"error": "forbidden"}), 403
    try:
        import tb_parser
    except ImportError as e:
        return jsonify({"error": f"tb_parser import failed: {e}"}), 500

    monthly_file = request.files.get("monthly_tb")
    if not monthly_file:
        return jsonify({"error": "monthly_tb file is required"}), 400
    period = (request.form.get("period") or "").strip()
    if not period:
        return jsonify({"error": "period is required"}), 400

    try:
        monthly_bytes = monthly_file.read()
        monthly_entities = tb_parser.parse_tb_html(monthly_bytes)
    except Exception as e:
        return jsonify({"error": f"failed to parse monthly TB: {e}"}), 400
    if not monthly_entities:
        return jsonify({"error": "no entities found in monthly TB (expected HTML-formatted .xls)"}), 400

    # YTD file — required for IS and SCF (Phase 2 & 3). BS can render
    # without it, but uploading both at once is the normal workflow.
    ytd_entities = {}
    ytd_file = request.files.get("ytd_tb")
    if ytd_file and ytd_file.filename:
        try:
            ytd_bytes = ytd_file.read()
            ytd_entities = tb_parser.parse_tb_html(ytd_bytes)
        except Exception as e:
            return jsonify({"error": f"failed to parse YTD TB: {e}"}), 400

    # Cache each TB section under both the sage_code (TB header) and the
    # entity-level LOCATIONID the dropdown navigates by. Previously this
    # only mirrored under the user's *currently-selected* entity, so the
    # workaround was to re-upload once per project — now one upload
    # populates every project whose entity appears in the TB.
    try:
        loc_to_entity = sage_intacct.location_to_entity_map()
    except Exception as e:
        app.logger.warning("upload-tb: location_to_entity_map() failed: %s", e)
        loc_to_entity = {}
    # Optional override the form may pass (kept for backwards compat with
    # the old upload UX that scoped to one entity).
    selected_entity = (request.form.get("entity") or "").strip()
    cached_entities = []
    for code, data in monthly_entities.items():
        ytd_accts = None
        ytd_info  = ytd_entities.get(code)
        if ytd_info:
            ytd_accts = ytd_info["accounts"]

        keys_written = []
        # Primary: cache under the TB header's sage_code so direct lookups
        # by Sage code still work.
        _fin_cache_put(code, period, data["accounts"], accounts_ytd=ytd_accts)
        keys_written.append(code)

        # Auto-mirror: resolve sage_code → parent entity_id via Sage's
        # LOCATION graph and cache under that key too. This is what
        # eliminates the "re-upload per project" workflow.
        parent_entity = loc_to_entity.get(code)
        if parent_entity and parent_entity != code:
            _fin_cache_put(parent_entity, period, data["accounts"],
                           accounts_ytd=ytd_accts)
            keys_written.append(parent_entity)

        # Legacy explicit-override mirror (only useful when the dropdown's
        # current selection doesn't naturally match either key above).
        if selected_entity and selected_entity not in keys_written:
            _fin_cache_put(selected_entity, period, data["accounts"],
                           accounts_ytd=ytd_accts)
            keys_written.append(selected_entity)

        cached_entities.append({
            "sage_code":         code,
            "full_name":         data["full_name"],
            "reporting_period":  data["reporting_period"],
            "as_of_date":        data["as_of_date"],
            "account_count":     len(data["accounts"]),
            "ytd_account_count": len(ytd_accts) if ytd_accts else 0,
            "cached_under":      keys_written,
        })
    return jsonify({
        "ok":              True,
        "period":          period,
        "monthly_count":   len(monthly_entities),
        "ytd_count":       len(ytd_entities),
        "cached_entities": cached_entities,
    })


@app.route("/api/financials/download-frp", methods=["GET"])
@login_required
def api_financials_download_frp():
    """Build and return a 3-sheet FRP .xlsx (BS / IS / SCF) for the
    given cached entity+period. Output matches the format the
    accountant's frp_builder.py produces, so users can download the
    same workbook the accounting team distributes."""
    if not _can_view_financials():
        return jsonify({"error": "forbidden"}), 403
    try:
        import frp_excel
    except ImportError as e:
        return jsonify({"error": f"frp_excel import failed: {e}"}), 500
    entity = (request.args.get("entity") or "").strip()
    period = (request.args.get("period") or "").strip()
    entity_name = (request.args.get("entity_name") or "").strip() or entity
    if not entity or not period:
        return jsonify({"error": "entity and period required"}), 400
    cache = _fin_cache_get(entity, period)
    if not cache or not cache["accounts"]:
        return jsonify({"error": "no cached TB for this entity/period — upload first"}), 404

    # Derive a human-readable date label from the period name. The
    # period names look like "Month Ended March 2026" — we convert
    # to "March 31, 2026" to match the FRP workbook format. Falls
    # back to the period name itself if parsing fails.
    date_label = period
    try:
        import re as _re, calendar
        m = _re.search(r"(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{4})",
                       period)
        if m:
            month_name = m.group(1)
            year       = int(m.group(2))
            month_num  = ["January","February","March","April","May","June","July",
                          "August","September","October","November","December"
                          ].index(month_name) + 1
            last_day   = calendar.monthrange(year, month_num)[1]
            date_label = f"{month_name} {last_day}, {year}"
    except Exception:
        pass

    try:
        buf = frp_excel.build_workbook(
            entity_name,
            cache["accounts"],
            cache.get("accounts_ytd") or [],
            date_label,
        )
    except Exception as e:
        import traceback
        app.logger.error("FRP excel build failed: %s\n%s", e, traceback.format_exc())
        return jsonify({"error": f"FRP build failed: {type(e).__name__}: {e}"}), 500

    # Filename sanitized to filesystem-safe characters.
    import re as _re2
    safe_name = _re2.sub(r"[^A-Za-z0-9 ._-]", "_", entity_name).strip()
    filename = f"{safe_name} - {date_label} FRP.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/api/financials/income-statement-tb", methods=["GET"])
@login_required
def api_financials_income_statement_tb():
    """Render IS for a cached TB-uploaded entity/period. Requires both
    Current-Month and YTD accounts (uploaded together via the upload
    endpoint). Returns the same JSON shape as the BS endpoint plus
    each line carries both monthly and YTD values."""
    if not _can_view_financials():
        return jsonify({"error": "forbidden"}), 403
    try:
        import tb_parser
    except ImportError as e:
        return jsonify({"error": f"tb_parser import failed: {e}"}), 500
    entity = (request.args.get("entity") or "").strip()
    period = (request.args.get("period") or "").strip()
    if not entity or not period:
        return jsonify({"error": "entity and period required"}), 400
    cache = _fin_cache_get(entity, period)
    if not cache or not cache["accounts"]:
        return jsonify({
            "cached":  False,
            "message": "No TB upload found for this entity/period. "
                       "Upload both Current-Month and YTD TBs first.",
        }), 200
    if not cache.get("accounts_ytd"):
        return jsonify({
            "cached":  False,
            "message": "Monthly TB cached, but no YTD TB uploaded yet — "
                       "IS needs both. Re-upload with the YTD file too.",
        }), 200
    v  = tb_parser.compute_is(cache["accounts"], cache["accounts_ytd"])
    is_payload = tb_parser.render_is(v)
    fetched_at = cache["fetched_at"]
    return jsonify({
        "cached":          True,
        "source":          "tb_upload",
        "entity":          entity,
        "period":          period,
        "fetched_at":      fetched_at.isoformat() + "Z" if hasattr(fetched_at, "isoformat") else str(fetched_at),
        "income_statement": is_payload,
    })


@app.route("/api/financials/cash-flows-tb", methods=["GET"])
@login_required
def api_financials_cash_flows_tb():
    """Render Statement of Cash Flows for a cached TB-uploaded
    entity/period. Indirect method, mirrors frp_builder.py's SCF.
    Requires both Current-Month and YTD accounts."""
    if not _can_view_financials():
        return jsonify({"error": "forbidden"}), 403
    try:
        import tb_parser
    except ImportError as e:
        return jsonify({"error": f"tb_parser import failed: {e}"}), 500
    entity = (request.args.get("entity") or "").strip()
    period = (request.args.get("period") or "").strip()
    if not entity or not period:
        return jsonify({"error": "entity and period required"}), 400
    cache = _fin_cache_get(entity, period)
    if not cache or not cache["accounts"]:
        return jsonify({
            "cached":  False,
            "message": "No TB upload found for this entity/period.",
        }), 200
    if not cache.get("accounts_ytd"):
        return jsonify({
            "cached":  False,
            "message": "Monthly TB cached, but no YTD TB uploaded yet — "
                       "SCF needs both.",
        }), 200
    v  = tb_parser.compute_scf(cache["accounts"], cache["accounts_ytd"])
    scf = tb_parser.render_scf(v)
    fetched_at = cache["fetched_at"]
    return jsonify({
        "cached":     True,
        "source":     "tb_upload",
        "entity":     entity,
        "period":     period,
        "fetched_at": fetched_at.isoformat() + "Z" if hasattr(fetched_at, "isoformat") else str(fetched_at),
        "cash_flows": scf,
    })


@app.route("/api/financials/balance-sheet-tb", methods=["GET"])
@login_required
def api_financials_balance_sheet_tb():
    """Render the BS for a cached TB-uploaded entity/period using the
    frp_builder.py-derived predicates (different — and more accurate —
    than the frp_mapping.py path used for GLENTRY-based data)."""
    if not _can_view_financials():
        return jsonify({"error": "forbidden"}), 403
    try:
        import tb_parser
    except ImportError as e:
        return jsonify({"error": f"tb_parser import failed: {e}"}), 500
    entity = (request.args.get("entity") or "").strip()
    period = (request.args.get("period") or "").strip()
    if not entity or not period:
        return jsonify({"error": "entity and period required"}), 400
    cache = _fin_cache_get(entity, period)
    if not cache:
        return jsonify({
            "cached":  False,
            "message": "No TB upload found for this entity/period.",
        }), 200
    v = tb_parser.compute_bs(cache["accounts"])
    bs = tb_parser.render_bs(v)
    fetched_at = cache["fetched_at"]
    return jsonify({
        "cached":        True,
        "source":        "tb_upload",
        "entity":        entity,
        "period":        period,
        "fetched_at":    fetched_at.isoformat() + "Z" if hasattr(fetched_at, "isoformat") else str(fetched_at),
        "account_count": len(cache["accounts"]),
        "balance_sheet": bs,
    })


# ── FRP target snapshot used by the cutoff-date calibration ──
# These are the authoritative GPD March 2026 FRP values we're trying
# to reverse-engineer the snapshot date from. Targets are SIGNED close
# balances per account (CR = negative). Picked accounts whose FRP
# values we know directly from the workbook:
#   16000 Land:                    35,570,385
#   20030 Accounts payable:        -2,453,261  (CR balance)
#   13015 Promissory note:          1,100,000
#   30500 Retained earnings (GL):    -972,346  (note: FRP RE includes
#                                                YTD-NI; we use GL
#                                                30500 only as Phase 1
#                                                doesn't roll IS in)
#   20060 Retainage payable:       -1,625,628
#   21015 Bond Payable - ST:       -3,869,000
#   21016 Bond Payable - LT:      -55,240,000
#   22010 Development loan:       -29,558,821
#   16902 MUD Receivable - Bond:  -26,683,664
#
# AP is the strongest signal (it's where the snapshot-date entries land);
# Land is a "shouldn't move" anchor; Retainage / Bond Payables / Dev
# Loan anchor the no-cutoff baseline.
FRP_TARGETS_BY_ENTITY_PERIOD = {
    ("16 - GPD", "Month Ended March 2026"): {
        "16000":  35_570_385.0,
        "20030":  -2_453_261.0,
        "13015":   1_100_000.0,
        "20060":  -1_625_628.0,
        "21015":  -3_869_000.0,
        "21016": -55_240_000.0,
        "22010": -29_558_821.0,
        "16902": -26_683_664.0,
    },
}


@app.route("/api/financials/calibrate", methods=["POST"])
@login_required
def api_financials_calibrate():
    """Reverse-engineer the FRP snapshot date.

    Walks WHENCREATED cutoff dates from latest backward and finds the
    one that minimizes total |actual - FRP_target| across the known
    target accounts. Then recomputes the BS at that cutoff and returns
    it alongside the cutoff date and per-target diff trace.
    """
    if not _can_view_financials():
        return jsonify({"error": "forbidden"}), 403
    entity = (request.args.get("entity") or "").strip()
    period = (request.args.get("period") or "").strip()
    if not entity or not period:
        return jsonify({"error": "entity and period required"}), 400
    targets = FRP_TARGETS_BY_ENTITY_PERIOD.get((entity, period))
    if not targets:
        return jsonify({
            "error": f"no FRP targets configured for ({entity!r}, {period!r}); "
                     "add to FRP_TARGETS_BY_ENTITY_PERIOD in app.py first.",
        }), 400
    # Pull TB (uses cache if available within this process — get_trial_balance
    # has no DB cache lookup, but we just rendered the BS so the per-day
    # data should be stashed on the function from the prior render).
    try:
        accounts = sage_intacct.get_trial_balance(entity, period)
    except sage_intacct.IntacctAPIError as e:
        return jsonify({"error": f"Sage API error: {e}"}), 502
    per_day = getattr(sage_intacct.get_trial_balance,
                      "_last_per_when_created_sums", {}) or {}
    excl_sums = getattr(sage_intacct.get_trial_balance,
                        "_last_exclude_reason_sums", {}) or {}
    # Forward calibration: which WHENCREATED cutoff minimizes loss?
    forward = sage_intacct.calibrate_when_created_cutoff(
        accounts, per_day, targets,
    )
    # Inverse calibration: which subset of exclude-reasons, if RE-
    # INCLUDED, minimizes loss? Useful when the forward calibration
    # gives "no cutoff helps" — that signals our filter is too
    # aggressive and the missing $$ lives in currently-excluded data.
    inverse = sage_intacct.inverse_calibrate_exclude_reasons(
        accounts, excl_sums, targets,
    )
    # Pick the better of the two calibrations.
    use_inverse = inverse["loss_at_best"] < forward["loss_at_best"]
    # Recompute the full BS at the chosen optimum.
    if use_inverse and inverse["best_unexcluded"]:
        # Apply un-exclusion: add back contributions from chosen reasons.
        adj_accounts = []
        adjust_map = {a["no"]: a["close"] for a in accounts}
        for r in inverse["best_unexcluded"]:
            for acct, contrib in excl_sums.get(r, {}).items():
                adjust_map[acct] = adjust_map.get(acct, 0.0) + contrib
        # Build adjusted account list preserving names.
        name_by_no = {a["no"]: a.get("name", "") for a in accounts}
        all_nos = set(adjust_map.keys()) | {a["no"] for a in accounts}
        for no in all_nos:
            adj_accounts.append({
                "no":     no,
                "name":   name_by_no.get(no, ""),
                "open":   0.0,
                "debit":  0.0,
                "credit": 0.0,
                "close":  adjust_map.get(no, 0.0),
            })
        bs = _fin_render_bs(adj_accounts)
    elif forward["best_cutoff"]:
        cutoff_accounts = sage_intacct.get_trial_balance_with_cutoff(
            accounts, per_day, forward["best_cutoff"],
        )
        bs = _fin_render_bs(cutoff_accounts)
    else:
        bs = _fin_render_bs(accounts)
    # Trim forward trace for visibility.
    trace_sorted = sorted(forward["trace"], key=lambda x: x["loss"])[:20]
    return jsonify({
        "entity":              entity,
        "period":              period,
        "best_cutoff":         forward["best_cutoff"],
        "loss_at_best":        forward["loss_at_best"],
        "loss_no_cutoff":      forward["loss_no_cutoff"],
        "balances_at_best":    forward["balances_at_best"],
        "targets":             targets,
        "trace_top20":         trace_sorted,
        "inverse_best":        inverse["best_unexcluded"],
        "inverse_loss":        inverse["loss_at_best"],
        "inverse_baseline":    inverse["loss_baseline"],
        "inverse_balances":    inverse["balances_at_best"],
        "inverse_top_combos":  inverse["top_combos"],
        "used":                "inverse" if use_inverse else ("forward" if forward["best_cutoff"] else "baseline"),
        "balance_sheet":       bs,
    })


@app.route("/api/financials/ping", methods=["GET"])
@login_required
def api_financials_ping():
    """Admin diagnostic — confirms Sage env vars are set + login works."""
    if not session.get("is_admin"):
        return jsonify({"error": "admin only"}), 403
    return jsonify(sage_intacct.ping())


@app.route("/api/financials/diagnose", methods=["GET"])
@login_required
def api_financials_diagnose():
    """Admin diagnostic — probes Sage user permissions across LOCATION,
    REPORTINGPERIOD, and USERINFO with several filter combos so we can
    see exactly which queries succeed/fail/return zero. Useful when
    /api/financials/entities returns an empty list with no error."""
    if not session.get("is_admin"):
        return jsonify({"error": "admin only"}), 403
    return jsonify(sage_intacct.diagnose())


# ─── INVOICE DASHBOARD ───────────────────────────────────────────────────────
# Bi-weekly Stampli HTML invoice analytics. Each upload becomes a
# permanent row in `invoice_periods` keyed by half-month (YYYY-MM-A
# for 1–15, YYYY-MM-B for 16–EOM). The Period Archive rail on the
# /invoice-dashboard page lets users time-travel between snapshots.

_INVD_MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                     "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
_INVD_D_PATTERN  = re.compile(r"const\s+D\s*=\s*(\{[\s\S]*?\});\s*\n?\s*const\s+GL_COLORS", re.MULTILINE)
_INVD_DATE_TUPLE = re.compile(r"(\d{4})[-_]?(\d{2})[-_]?(\d{2})")
_INVD_REPORT_HDR = re.compile(
    r"Updated:\s*([A-Z][a-z]+)\s+(\d{1,2})\s*[–—\-]\s*(?:([A-Z][a-z]+)\s+)?(\d{1,2}),\s*(\d{4})",
)
_INVD_REPORT_DT  = re.compile(r"Report Date:\s*([A-Z][a-z]+)\s+(\d{1,2}),\s*(\d{4})")

def _invd_period_of(d):
    """Half-month period metadata. d is a datetime.date.
    Returns {key, start, end, label, short, year}."""
    import calendar
    first_half = d.day <= 15
    start_day  = 1 if first_half else 16
    end_day    = 15 if first_half else calendar.monthrange(d.year, d.month)[1]
    key   = f"{d.year}-{d.month:02d}-{'A' if first_half else 'B'}"
    mo    = _INVD_MONTH_NAMES[d.month - 1]
    return {
        "key":   key,
        "start": datetime.date(d.year, d.month, start_day),
        "end":   datetime.date(d.year, d.month, end_day),
        "label": f"{mo} {start_day} – {mo} {end_day}, {d.year}",
        "short": f"{mo.upper()} {start_day}–{end_day}",
        "year":  d.year,
    }

def _invd_load_period(period_key=None):
    """Latest row if period_key is None, otherwise the specific period.
    Returns a dict shaped like the JS-side data island expects, or None."""
    conn = get_db(); cur = conn.cursor()
    try:
        if period_key:
            cur.execute(
                "SELECT period_key, period_start, period_end, report_date, "
                "       total_records, total_amount, data, uploaded_at "
                "FROM invoice_periods WHERE period_key = %s", (period_key,))
        else:
            cur.execute(
                "SELECT period_key, period_start, period_end, report_date, "
                "       total_records, total_amount, data, uploaded_at "
                "FROM invoice_periods ORDER BY period_start DESC LIMIT 1")
        row = cur.fetchone()
        if not row:
            return None
        d = dict(row)
        # psycopg2 with RealDictCursor returns JSONB as dict already;
        # defensive if a different driver hands back a string.
        if isinstance(d.get("data"), str):
            d["data"] = json.loads(d["data"])
        # IMPORTANT: leave `period_start` / `period_end` / `report_date`
        # as datetime.date objects. The Jinja template calls `.strftime`
        # on them (see invoice_dashboard.html: `invoice.period_start
        # .strftime(...)`), and Flask's tojson filter / jsonify already
        # serialize dates natively for the JSON paths. Calling
        # .isoformat() here was a hangover from an earlier defensive
        # rewrite and broke the page-render path with AttributeError.
        # Only Decimal → float still needs explicit conversion (psycopg2
        # returns NUMERIC as Decimal which isn't JSON-friendly).
        if d.get("total_amount") is not None:
            d["total_amount"] = float(d["total_amount"])
        return d
    finally:
        cur.close(); conn.close()

def _invd_load_archive_summary():
    """Every period from newest to oldest. Used by the rail.

    Dates are converted to ISO 8601 strings (`YYYY-MM-DD`) before
    return because the rail data flows ONLY through JSON →
    `window.INVOICE_ARCHIVE`, and Flask 3's default JSON provider
    serializes `datetime.date` as RFC-822 (`"Thu, 16 Apr 2026
    00:00:00 GMT"`), which the JS-side `periodOf(dateStr).split("-")`
    can't parse. The resulting NaN cascades through
    `buildPeriodArchive()` and throws at module load — the IIFE
    aborts and *nothing* downstream of it (KPIs, charts, tables,
    rail cards) renders. Converting to ISO here keeps the JS path
    self-consistent.

    `_invd_load_period()` deliberately does NOT do this conversion
    because that path's dates flow through the Jinja template
    (`{{ invoice.period_start.strftime(...) }}`), which requires
    a real `datetime.date` object. Two different serialization
    requirements, two different functions.
    """
    conn = get_db(); cur = conn.cursor()
    try:
        cur.execute(
            "SELECT period_key, period_start, period_end, report_date, "
            "       total_records, total_amount, ember_count, ccdl_count, "
            "       uploaded_at FROM invoice_periods_summary")
        rows = cur.fetchall()
        out = []
        for r in rows:
            d = dict(r)
            for k in ("period_start", "period_end", "report_date"):
                if d.get(k) is not None:
                    d[k] = d[k].isoformat()           # → "YYYY-MM-DD"
            if d.get("uploaded_at") is not None:
                d["uploaded_at"] = d["uploaded_at"].isoformat()
            if d.get("total_amount") is not None:
                d["total_amount"] = float(d["total_amount"])
            out.append(d)
        return out
    finally:
        cur.close(); conn.close()

def _invd_save_upload(*, period_key, period_start, period_end, report_date,
                      data, source_filename=None, source_blob=None,
                      uploaded_by=None):
    """UPSERT one period row. Returns metadata about what landed."""
    records = data.get("records", [])
    total_records = len(records)
    total_amount  = sum(float(r.get("amount", 0) or 0) for r in records)
    conn = get_db(); cur = conn.cursor()
    try:
        cur.execute("SELECT 1 FROM invoice_periods WHERE period_key = %s", (period_key,))
        replaced = cur.fetchone() is not None
        cur.execute("""
            INSERT INTO invoice_periods
                (period_key, period_start, period_end, report_date,
                 total_records, total_amount, data,
                 source_filename, source_blob, uploaded_by, uploaded_at)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s, NOW())
            ON CONFLICT (period_key) DO UPDATE SET
                period_start    = EXCLUDED.period_start,
                period_end      = EXCLUDED.period_end,
                report_date     = EXCLUDED.report_date,
                total_records   = EXCLUDED.total_records,
                total_amount    = EXCLUDED.total_amount,
                data            = EXCLUDED.data,
                source_filename = COALESCE(EXCLUDED.source_filename,
                                           invoice_periods.source_filename),
                source_blob     = COALESCE(EXCLUDED.source_blob,
                                           invoice_periods.source_blob),
                uploaded_by     = COALESCE(EXCLUDED.uploaded_by,
                                           invoice_periods.uploaded_by),
                uploaded_at     = NOW();
        """, (period_key, period_start, period_end, report_date,
              total_records, total_amount, json.dumps(data),
              source_filename, psycopg2.Binary(source_blob) if source_blob else None,
              uploaded_by))
        conn.commit()
    finally:
        cur.close(); conn.close()
    p = _invd_period_of(period_start)
    return {"period_key": period_key, "period_label": p["label"],
            "total_records": total_records, "replaced": replaced}

def _invd_parse_stampli_export(raw, filename=""):
    """Parse a Stampli HTML export. Returns
    {period_key, period_start, period_end, report_date, data}.
    Raises ValueError if the upload doesn't look like a Stampli file."""
    try:
        text = raw.decode("utf-8")
    except UnicodeDecodeError:
        text = raw.decode("utf-8", errors="replace")
    m = _INVD_D_PATTERN.search(text)
    if not m:
        raise ValueError(
            "Could not find `const D = {...}` block in the upload. "
            "Is this really the Stampli Invoice Dashboard HTML export?")
    try:
        D = json.loads(m.group(1))
    except json.JSONDecodeError as e:
        raise ValueError(f"Invoice payload didn't parse as JSON: {e}")
    # Decode HTML entities ONLY for the banner/period detection — the
    # JSON literal extraction above already grabbed the raw D block.
    # Stampli's `Updated: Apr 16 – Apr 30` banner encodes the dash as
    # `&ndash;` in source, which our regex's dash class won't match
    # until we unescape. Without this, period detection silently fails
    # and the parser falls back to the filename date (May 8 → May 1-15)
    # even though the file is the Apr 16-30 report.
    text_decoded = html.unescape(text)
    report_date = _invd_detect_report_date(text_decoded, filename)
    if not report_date:
        raise ValueError(
            "Could not determine the report date. Include YYYY_MM_DD in the "
            "filename or restore the 'Report Date:' header in the source export.")
    start, end = _invd_detect_period_window(text_decoded)
    if not start or not end:
        anchor = datetime.date(report_date.year, report_date.month,
                                max(1, report_date.day - 1))
        p = _invd_period_of(anchor)
        start, end = p["start"], p["end"]
    return {
        "period_key":   _invd_period_of(start)["key"],
        "period_start": start,
        "period_end":   end,
        "report_date":  report_date,
        "data":         D,
    }

def _invd_detect_report_date(text, filename):
    m = _INVD_DATE_TUPLE.search(filename or "")
    if m:
        y, mo, d = map(int, m.groups())
        try: return datetime.date(y, mo, d)
        except ValueError: pass
    m = _INVD_REPORT_DT.search(text)
    if m:
        mo_name, day, year = m.groups()
        try:
            mo = _INVD_MONTH_NAMES.index(mo_name[:3]) + 1
            return datetime.date(int(year), mo, int(day))
        except ValueError: pass
    return None

def _invd_detect_period_window(text):
    m = _INVD_REPORT_HDR.search(text)
    if not m: return (None, None)
    mo1, d1, mo2, d2, year = m.groups()
    if not mo2: mo2 = mo1
    try:
        start = datetime.date(int(year), _INVD_MONTH_NAMES.index(mo1[:3]) + 1, int(d1))
        end   = datetime.date(int(year), _INVD_MONTH_NAMES.index(mo2[:3]) + 1, int(d2))
        return (start, end)
    except ValueError:
        return (None, None)

# ── ROUTES ───────────────────────────────────────────────────────────
@app.route("/invoice-dashboard")
@login_required
def invoice_dashboard():
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("invoice_dashboard", True):
        return redirect(url_for("home"))
    if session.get("is_admin"):
        pa = {**pa, "invoice_dashboard": True}
    period_key = request.args.get("period")
    period   = _invd_load_period(period_key)
    archive  = _invd_load_archive_summary()
    # Data dates footer: "Data From" = the Stampli "Updated" date
    # (period.report_date); "Uploaded" = when this period snapshot was
    # saved to the invoice_periods table.
    data_from   = _fmt_data_date(period.get("report_date")) if period else None
    uploaded_at = _fmt_data_date(period.get("uploaded_at")) if period else None
    return render_template(
        "invoice_dashboard.html",
        invoice=period,           # None when no uploads yet → empty state
        archive=archive,
        data_from=data_from, uploaded_at=uploaded_at,
        page_access=pa,
        is_admin=session.get("is_admin", False),
        username=session.get("username"),
        display_name=session.get("display_name", session.get("username")),
    )

@app.route("/api/invoice-dashboard/period/<period_key>")
@login_required
def api_invoice_period(period_key):
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("invoice_dashboard", True):
        return jsonify({"error": "forbidden"}), 403
    period = _invd_load_period(period_key)
    if not period:
        return jsonify({"error": "not found", "period_key": period_key}), 404
    return jsonify(period)

@app.route("/api/invoice-dashboard/upload", methods=["POST"])
@login_required
def api_invoice_upload():
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("invoice_dashboard", True):
        return jsonify({"error": "forbidden"}), 403
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "no file"}), 400
    raw = f.read()
    try:
        normalized = _invd_parse_stampli_export(raw, filename=f.filename)
        result = _invd_save_upload(
            period_key      = normalized["period_key"],
            period_start    = normalized["period_start"],
            period_end      = normalized["period_end"],
            report_date     = normalized["report_date"],
            data            = normalized["data"],
            source_filename = f.filename,
            source_blob     = raw,
            uploaded_by     = session.get("username") or "?",
        )
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        app.logger.exception("Invoice upload failed")
        return jsonify({"error": f"internal: {e}"}), 500
    return jsonify({
        "ok":            True,
        "period_key":    result["period_key"],
        "period_label":  result["period_label"],
        "replaced":      result["replaced"],
        "total_records": result["total_records"],
    })

# ─── ADMIN: TEAM ACTIVITY DASHBOARD ───────────────────────────────────────────
# Friendly labels for the paths the dashboard surfaces. Anything not in
# this map falls back to the raw path so unknown routes (admin pages,
# nested views) still render a readable row.
_ACTIVITY_PATH_LABELS = {
    "/":               "MPC Underwriting",
    "/home":           "Home",
    "/returns":        "Project Returns",
    "/loans":          "Loans & Debt",
    "/operations":     "Operating Revenues",
    "/ember-capital":  "Ember Capital",
    "/capital":        "Ember Capital",
    "/portfolio":      "Ember Capital",
    "/macro":          "Macro Dashboard",
    "/sales":          "Community Sales",
    "/verticals":      "Vertical Dashboard",
    "/admin/activity": "Team Activity",
    "/admin/users":    "Manage Team",
    "/admin/reports":  "Report Subscriptions",
}

def _activity_label(path):
    if not path:
        return "—"
    if path in _ACTIVITY_PATH_LABELS:
        return _ACTIVITY_PATH_LABELS[path]
    # /capital/report.pdf → "Capital · Report" style fallback for sub-pages
    head = path.split("/")[1] if "/" in path else path
    if head in {"capital", "ember-capital"}:
        return "Ember Capital"
    return path

@app.route("/admin/users")
@login_required
def admin_users_page():
    """Manage Team — full-page replacement for the legacy modal. The
    page hits the existing /api/admin/users endpoints over fetch, so
    no API changes were needed; this is presentation only."""
    pa = session.get("page_access") or {"mpc_underwriting": True, "returns": True, "loans": True, "operations": True}
    if not session.get("is_admin"):
        return render_template("admin_users.html", forbidden=True, is_admin=False, page_access=pa), 403
    return render_template(
        "admin_users.html",
        forbidden=False,
        username=session.get("username"),
        display_name=session.get("display_name", session.get("username")),
        is_admin=True,
        page_access=pa,
    )

@app.route("/admin/reports")
@login_required
def admin_reports_page():
    """Report Subscriptions — full-page replacement for the legacy
    modal. Per-user × per-report dropdown matrix, autosaves on change."""
    pa = session.get("page_access") or {"mpc_underwriting": True, "returns": True, "loans": True, "operations": True}
    if not session.get("is_admin"):
        return render_template("admin_reports.html", forbidden=True, is_admin=False, page_access=pa), 403
    return render_template(
        "admin_reports.html",
        forbidden=False,
        username=session.get("username"),
        display_name=session.get("display_name", session.get("username")),
        is_admin=True,
        page_access=pa,
    )

@app.route("/admin/activity")
@login_required
def admin_activity():
    """Team Activity dashboard. Per-user roll-up of logins and page
    views over the last 30 days, with a 14-day daily sparkline. Admin
    only (we serve a 403-style page rather than redirecting so a
    non-admin landing here gets a clear message)."""
    if not session.get("is_admin"):
        pa = session.get("page_access") or {"mpc_underwriting": True, "returns": True, "loans": True, "operations": True}
        return render_template("admin_activity.html", forbidden=True, is_admin=False, page_access=pa), 403

    conn = get_db()
    cur = conn.cursor()

    # ── KPI band ──
    # Active users (7d): distinct user_ids with any event in last 7 days.
    cur.execute("""
        SELECT COUNT(DISTINCT user_id) AS n
        FROM activity_log
        WHERE ts >= NOW() - INTERVAL '7 days' AND user_id IS NOT NULL
    """)
    active_7d = (cur.fetchone() or {}).get("n") or 0

    # Total logins (7d): how many sign-in events landed in the last week.
    cur.execute("""
        SELECT COUNT(*) AS n
        FROM activity_log
        WHERE event_type = 'login' AND ts >= NOW() - INTERVAL '7 days'
    """)
    logins_7d = (cur.fetchone() or {}).get("n") or 0

    # Most-visited page (7d).
    cur.execute("""
        SELECT path, COUNT(*) AS n
        FROM activity_log
        WHERE event_type = 'page_view' AND ts >= NOW() - INTERVAL '7 days'
              AND path IS NOT NULL
        GROUP BY path
        ORDER BY n DESC
        LIMIT 1
    """)
    top_row = cur.fetchone()
    top_page_label = _activity_label(top_row["path"]) if top_row else "—"
    top_page_count = top_row["n"] if top_row else 0

    # Total page views (7d) for the avg-views-per-active-user KPI.
    cur.execute("""
        SELECT COUNT(*) AS n
        FROM activity_log
        WHERE event_type = 'page_view' AND ts >= NOW() - INTERVAL '7 days'
    """)
    views_7d = (cur.fetchone() or {}).get("n") or 0
    avg_views_per_user = round(views_7d / active_7d, 1) if active_7d else 0

    # ── Per-user roll-up ──
    # Pull every active user, plus anyone who logged in in the last 30
    # days even if their account was later removed (LEFT JOIN so the row
    # survives a deleted user; username comes from the denormalized
    # column on activity_log).
    cur.execute("""
        SELECT u.id AS user_id, u.username, u.is_admin,
               COALESCE(u.first_name, '') AS first_name,
               COALESCE(u.last_name, '')  AS last_name
        FROM users u
        ORDER BY u.username
    """)
    user_rows = cur.fetchall()

    # Aggregate activity counts per user in one round-trip.
    cur.execute("""
        SELECT user_id,
               MAX(ts) AS last_seen,
               COUNT(*) FILTER (WHERE event_type = 'login'
                                AND ts >= NOW() - INTERVAL '7 days')  AS logins_7d,
               COUNT(*) FILTER (WHERE event_type = 'login'
                                AND ts >= NOW() - INTERVAL '30 days') AS logins_30d,
               COUNT(*) FILTER (WHERE event_type = 'page_view'
                                AND ts >= NOW() - INTERVAL '7 days')  AS views_7d,
               COUNT(*) FILTER (WHERE event_type = 'page_view'
                                AND ts >= NOW() - INTERVAL '30 days') AS views_30d
        FROM activity_log
        WHERE user_id IS NOT NULL
        GROUP BY user_id
    """)
    by_user = {r["user_id"]: r for r in cur.fetchall()}

    # 14-day daily page-view counts per user for the row sparkline.
    cur.execute("""
        SELECT user_id,
               DATE_TRUNC('day', ts)::date AS d,
               COUNT(*) AS n
        FROM activity_log
        WHERE event_type = 'page_view'
          AND ts >= (NOW() - INTERVAL '14 days')::date
          AND user_id IS NOT NULL
        GROUP BY user_id, d
    """)
    spark_rows = cur.fetchall()
    today = datetime.date.today()
    days = [today - datetime.timedelta(days=13 - i) for i in range(14)]
    spark_by_user = {}
    for r in spark_rows:
        d = r["d"]
        if hasattr(d, "date"):
            d = d.date()
        spark_by_user.setdefault(r["user_id"], {})[d] = int(r["n"])

    # Top 3 visited paths per user (last 30d). Single query, partition
    # in memory so we only hit Postgres once.
    cur.execute("""
        SELECT user_id, path, COUNT(*) AS n
        FROM activity_log
        WHERE event_type = 'page_view'
          AND ts >= NOW() - INTERVAL '30 days'
          AND user_id IS NOT NULL
          AND path IS NOT NULL
        GROUP BY user_id, path
    """)
    top_paths = {}
    for r in cur.fetchall():
        top_paths.setdefault(r["user_id"], []).append((r["path"], r["n"]))
    for uid in top_paths:
        top_paths[uid].sort(key=lambda t: -t[1])

    cur.close(); conn.close()

    # Build the rows the template renders.
    rows = []
    for u in user_rows:
        agg = by_user.get(u["user_id"]) or {}
        last_seen = agg.get("last_seen")
        if last_seen:
            # Send the raw UTC ISO timestamp. The template wraps it in a
            # <time datetime="..."> element and JS rewrites the visible
            # text to the viewer's local timezone — Postgres TIMESTAMPTZ
            # already returns tz-aware datetimes so .isoformat() includes
            # the offset.
            last_seen_iso   = last_seen.isoformat() if hasattr(last_seen, "isoformat") else None
            last_seen_label = last_seen.strftime("%b %-d, %H:%M UTC") if hasattr(last_seen, "strftime") else str(last_seen)
        else:
            last_seen_iso   = None
            last_seen_label = "Never"
        spark_map = spark_by_user.get(u["user_id"], {})
        spark_values = [spark_map.get(d, 0) for d in days]
        spark_svg = _ops_sparkline(spark_values, "#F25929", bold=True) if any(spark_values) else ""
        top3 = top_paths.get(u["user_id"], [])[:3]
        top3_labels = [{"label": _activity_label(p), "n": n} for p, n in top3]
        display = (f"{u['first_name']} {u['last_name']}".strip()) or u["username"]
        rows.append({
            "user_id":       u["user_id"],
            "username":      u["username"],
            "display":       display,
            "is_admin":      u["is_admin"],
            "last_seen":     last_seen_label,
            "last_seen_iso": last_seen_iso,
            "logins_7d":     int(agg.get("logins_7d") or 0),
            "logins_30d":  int(agg.get("logins_30d") or 0),
            "views_7d":    int(agg.get("views_7d") or 0),
            "views_30d":   int(agg.get("views_30d") or 0),
            "top_pages":   top3_labels,
            "spark_svg":   spark_svg,
            "active":      bool(spark_values and any(spark_values)),
        })
    # Sort: most-active in last 7d first, then by name. Inactive users
    # bubble to the bottom so the admin sees who's lapsing at a glance.
    rows.sort(key=lambda r: (-r["views_7d"], r["display"].lower()))

    kpis = [
        {"label": "Active Users",       "val": str(active_7d),
         "sub":   "with any activity · last 7d"},
        {"label": "Total Logins",       "val": str(logins_7d),
         "sub":   "sign-ins · last 7d"},
        {"label": "Most-Visited Page",  "val": top_page_label,
         "sub":   f"{top_page_count} views · last 7d"},
        {"label": "Avg Views / User",   "val": str(avg_views_per_user),
         "sub":   "page views per active user · last 7d"},
    ]

    pa = session.get("page_access") or {"mpc_underwriting": True, "returns": True, "loans": True, "operations": True}
    # Send generated_at as a UTC ISO string; the template formats it for
    # the viewer's locale via the shared <time datetime="..."> JS.
    generated_at_iso = datetime.datetime.now(datetime.timezone.utc).isoformat()
    return render_template(
        "admin_activity.html",
        forbidden=False,
        kpis=kpis,
        rows=rows,
        username=session.get("username"),
        display_name=session.get("display_name", session.get("username")),
        is_admin=True,
        page_access=pa,
        generated_at_iso=generated_at_iso,
    )

# ─── ADMIN API ────────────────────────────────────────────────────────────────
@app.route("/api/admin/users", methods=["GET"])
@login_required
@admin_required
def list_users():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, username, email, is_admin, page_access, created_at,
               report_opt_in, report_format, report_subscriptions,
               first_name, last_name
        FROM users ORDER BY id
    """)
    rows = cur.fetchall()
    cur.close(); conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/admin/users", methods=["POST"])
@login_required
@admin_required
def create_user():
    data = request.json or {}
    username = data.get("username", "").strip()
    password = data.get("password", "")
    email = data.get("email", "").strip() or None
    is_admin = data.get("is_admin", False)
    page_access = data.get("page_access", {
        "mpc_underwriting": True, "returns": True, "loans": True,
        "operations": True, "macro": True, "sales": True,
        "portfolio": True, "reports": True, "verticals": True,
        "verticals_comment": True, "invoice_dashboard": True,
    })
    if not username or not password:
        return jsonify({"error": "Username and password required"}), 400
    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute(
            "INSERT INTO users (username, password_hash, email, is_admin, page_access) VALUES (%s, %s, %s, %s, %s) RETURNING id",
            (username, generate_password_hash(password), email, is_admin, json.dumps(page_access))
        )
        uid = cur.fetchone()["id"]
        conn.commit()
    except psycopg2.errors.UniqueViolation:
        conn.rollback()
        cur.close(); conn.close()
        return jsonify({"error": "Username already exists"}), 409
    cur.close(); conn.close()
    return jsonify({"id": uid, "username": username})

@app.route("/api/admin/users/<int:uid>", methods=["DELETE"])
@login_required
@admin_required
def delete_user(uid):
    if uid == session["user_id"]:
        return jsonify({"error": "Cannot delete yourself"}), 400
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM users WHERE id = %s", (uid,))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/admin/users/<int:uid>/password", methods=["PUT"])
@login_required
@admin_required
def reset_password(uid):
    data = request.json or {}
    password = data.get("password", "")
    if not password:
        return jsonify({"error": "Password required"}), 400
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE users SET password_hash = %s WHERE id = %s",
                (generate_password_hash(password), uid))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/account", methods=["GET"])
@login_required
def get_account():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT username, email, report_opt_in, report_format, first_name, last_name FROM users WHERE id = %s", (session["user_id"],))
    row = cur.fetchone()
    cur.close(); conn.close()
    return jsonify({
        "username": row["username"],
        "email": row["email"] or "",
        "report_opt_in": bool(row["report_opt_in"]),
        "report_format": row["report_format"] or "pdf",
        "first_name": row["first_name"] or "",
        "last_name": row["last_name"] or ""
    })

@app.route("/api/admin/users/<int:uid>/email", methods=["PUT"])
@login_required
@admin_required
def set_user_email(uid):
    data = request.json or {}
    email = data.get("email", "").strip() or None
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE users SET email = %s WHERE id = %s", (email, uid))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/account/password", methods=["PUT"])
@login_required
def change_own_password():
    data = request.json or {}
    current_pw = data.get("current_password", "")
    new_pw = data.get("new_password", "")
    if not current_pw or not new_pw:
        return jsonify({"error": "All fields are required"}), 400
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT password_hash FROM users WHERE id = %s", (session["user_id"],))
    user = cur.fetchone()
    if not user or not check_password_hash(user["password_hash"], current_pw):
        cur.close(); conn.close()
        return jsonify({"error": "Current password is incorrect"}), 400
    cur.execute("UPDATE users SET password_hash = %s WHERE id = %s",
                (generate_password_hash(new_pw), session["user_id"]))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/admin/users/<int:uid>/admin", methods=["PUT"])
@login_required
@admin_required
def set_user_admin(uid):
    """Promote/demote a user to admin. Guards: an admin can't remove their own
    admin access, and the last remaining admin can't be demoted (so the team
    never locks itself out of the admin tools)."""
    data = request.json or {}
    make_admin = bool(data.get("is_admin", False))
    if uid == session.get("user_id") and not make_admin:
        return jsonify({"error": "You can't remove your own admin access"}), 400
    conn = get_db()
    cur = conn.cursor()
    if not make_admin:
        cur.execute("SELECT COUNT(*) AS n FROM users WHERE is_admin = TRUE")
        if (cur.fetchone()["n"] or 0) <= 1:
            cur.close(); conn.close()
            return jsonify({"error": "At least one admin must remain"}), 400
    cur.execute("UPDATE users SET is_admin = %s WHERE id = %s", (make_admin, uid))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True, "is_admin": make_admin})


@app.route("/api/admin/users/<int:uid>/access", methods=["PUT"])
@login_required
@admin_required
def update_page_access(uid):
    data = request.json or {}
    page_access = data.get("page_access", {})
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE users SET page_access = %s WHERE id = %s",
                (json.dumps(page_access), uid))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})


# Reports each user can subscribe to. Storage shape on `users`:
#   report_subscriptions = {report_key: 'pdf' | 'excel'}
# Missing key = not subscribed. Order here drives the column order in
# the admin Reports management table.
_REPORT_SUBSCRIPTION_KEYS = ("returns", "ember_capital", "operations", "loans")
_REPORT_FORMATS           = ("pdf", "excel")


@app.route("/api/admin/users/<int:uid>/reports", methods=["PUT"])
@login_required
@admin_required
def update_user_report_subscriptions(uid):
    """Set the per-report subscription map for a user.

    Body:
        {"subscriptions": {"returns": "pdf", "ember_capital": "excel", ...}}

    Unknown keys and unknown formats are silently dropped. Pass the full
    map every time — server replaces the column wholesale.
    """
    data = request.json or {}
    incoming = data.get("subscriptions") or {}
    clean = {}
    for k, v in incoming.items():
        if k not in _REPORT_SUBSCRIPTION_KEYS:
            continue
        if v in _REPORT_FORMATS:
            clean[k] = v
    # Mirror to the legacy single-flag columns so the existing
    # _send_monthly_emails path keeps working until we cut it over.
    legacy_opt_in = bool(clean)
    legacy_format = next(iter(clean.values()), "pdf") if clean else "pdf"
    conn = get_db(); cur = conn.cursor()
    cur.execute(
        "UPDATE users SET report_subscriptions = %s, "
        "                 report_opt_in = %s, "
        "                 report_format = %s "
        "WHERE id = %s",
        (json.dumps(clean), legacy_opt_in, legacy_format, uid),
    )
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True, "subscriptions": clean})

@app.route("/api/account/report-settings", methods=["PUT"])
@login_required
def update_report_settings():
    data = request.json or {}
    opt_in = bool(data.get("report_opt_in", False))
    fmt = data.get("report_format", "pdf")
    if fmt not in ("pdf", "excel"):
        fmt = "pdf"
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE users SET report_opt_in = %s, report_format = %s WHERE id = %s",
                (opt_in, fmt, session["user_id"]))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/account/name", methods=["PUT"])
@login_required
def update_own_name():
    data = request.json or {}
    first_name = (data.get("first_name") or "").strip() or None
    last_name = (data.get("last_name") or "").strip() or None
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE users SET first_name = %s, last_name = %s WHERE id = %s",
                (first_name, last_name, session["user_id"]))
    conn.commit(); cur.close(); conn.close()
    # Update session display name immediately
    fn = first_name or ""
    ln = last_name or ""
    session["display_name"] = f"{fn} {ln}".strip() or session.get("username")
    return jsonify({"ok": True})

@app.route("/api/admin/users/<int:uid>/name", methods=["PUT"])
@login_required
@admin_required
def set_user_name(uid):
    data = request.json or {}
    first_name = (data.get("first_name") or "").strip() or None
    last_name = (data.get("last_name") or "").strip() or None
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE users SET first_name = %s, last_name = %s WHERE id = %s",
                (first_name, last_name, uid))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/admin/send-reports-now", methods=["POST"])
@login_required
@admin_required
def send_reports_now():
    """Run the monthly send job synchronously and return the per-recipient
    diagnostics so the admin UI can show who got skipped and why.

    Runs in-process (not a thread) so the response carries the result —
    the job is bounded by the recipient count and SendGrid latency, and
    knowing what happened beats fire-and-forget for debugging."""
    try:
        count = _send_monthly_emails(force=True)
    except Exception as e:
        print(f"[Reports] Send failed: {type(e).__name__}: {e}", flush=True)
        return jsonify({"ok": False, "error": f"{type(e).__name__}: {e}"}), 500
    return jsonify({
        "ok":    True,
        "count": count,
        "diag":  getattr(_send_monthly_emails, "last_diag", []),
    })

@app.route("/api/admin/users/<int:uid>/send-reports", methods=["POST"])
@login_required
@admin_required
def send_reports_to_user(uid):
    """Ad-hoc per-user report send. Powers the Send button on each row
    of /admin/reports. Returns the same {ok, count, diag} shape as the
    bulk endpoint so the client can reuse the diag-rendering logic.
    The diag will contain exactly one entry — either status 'sent' or
    'skipped' with a reason ('no email on file', 'no subscriptions',
    'no accessible reports …'). Period claim is bypassed (irrelevant
    for ad-hoc); subscription gate at SQL level is bypassed (we want a
    real diag entry rather than a silent zero-row result)."""
    try:
        count = _send_monthly_emails(force=True, recipient_ids=[uid])
    except Exception as e:
        print(f"[Reports] Per-user send failed (uid={uid}): {type(e).__name__}: {e}", flush=True)
        return jsonify({"ok": False, "error": f"{type(e).__name__}: {e}"}), 500
    return jsonify({
        "ok":    True,
        "count": count,
        "diag":  getattr(_send_monthly_emails, "last_diag", []),
    })

# ─── DEFAULT INPUTS TEMPLATE ──────────────────────────────────────────────────
def default_inputs(name="New Project"):
    # Lot size defaults match Excel Cost Inputs rows 72-87 exactly
    # Columns: FF, on, yield/ac, pace lots/mo, home_price, wsd/ff, paving/ff, landscaping/lot, urd/lot, lots_per_streetlight, fence/ff
    lot_size_defaults = [
        {"front_footage":25,  "on":0, "yield_per_ac":8.25, "pace":5,    "home_price":200000,    "wsd_per_ff":290, "paving_per_ff":220, "dev_start_month":1, "landscaping_per_lot":2000, "urd_per_lot":35, "lots_per_streetlight":4, "fence_cost_per_ff":94},
        {"front_footage":30,  "on":0, "yield_per_ac":5.54, "pace":5,    "home_price":360000,    "wsd_per_ff":290, "paving_per_ff":220, "dev_start_month":1, "landscaping_per_lot":2000, "urd_per_lot":35, "lots_per_streetlight":4, "fence_cost_per_ff":94},
        {"front_footage":35,  "on":0, "yield_per_ac":8.25, "pace":6,    "home_price":275000,    "wsd_per_ff":290, "paving_per_ff":220, "dev_start_month":1, "landscaping_per_lot":2000, "urd_per_lot":35, "lots_per_streetlight":4, "fence_cost_per_ff":94},
        {"front_footage":40,  "on":1, "yield_per_ac":5.5,  "pace":7,    "home_price":330168,    "wsd_per_ff":290, "paving_per_ff":220, "dev_start_month":1, "landscaping_per_lot":2000, "urd_per_lot":35, "lots_per_streetlight":4, "fence_cost_per_ff":94},
        {"front_footage":45,  "on":1, "yield_per_ac":5.0,  "pace":6,    "home_price":380000,    "wsd_per_ff":290, "paving_per_ff":220, "dev_start_month":1, "landscaping_per_lot":2000, "urd_per_lot":35, "lots_per_streetlight":4, "fence_cost_per_ff":94},
        {"front_footage":50,  "on":1, "yield_per_ac":4.5,  "pace":5,    "home_price":430000,    "wsd_per_ff":290, "paving_per_ff":220, "dev_start_month":1, "landscaping_per_lot":2000, "urd_per_lot":35, "lots_per_streetlight":4, "fence_cost_per_ff":94},
        {"front_footage":55,  "on":0, "yield_per_ac":4.0,  "pace":5,    "home_price":500000,    "wsd_per_ff":290, "paving_per_ff":220, "dev_start_month":1, "landscaping_per_lot":2000, "urd_per_lot":35, "lots_per_streetlight":4, "fence_cost_per_ff":94},
        {"front_footage":60,  "on":1, "yield_per_ac":3.5,  "pace":2,    "home_price":580000,    "wsd_per_ff":290, "paving_per_ff":220, "dev_start_month":1, "landscaping_per_lot":2000, "urd_per_lot":35, "lots_per_streetlight":4, "fence_cost_per_ff":94},
        {"front_footage":65,  "on":0, "yield_per_ac":3.0,  "pace":2,    "home_price":615000,    "wsd_per_ff":290, "paving_per_ff":220, "dev_start_month":1, "landscaping_per_lot":2000, "urd_per_lot":35, "lots_per_streetlight":3, "fence_cost_per_ff":94},
        {"front_footage":70,  "on":0, "yield_per_ac":2.5,  "pace":1,    "home_price":675000,    "wsd_per_ff":290, "paving_per_ff":220, "dev_start_month":1, "landscaping_per_lot":2000, "urd_per_lot":35, "lots_per_streetlight":3, "fence_cost_per_ff":94},
        {"front_footage":75,  "on":0, "yield_per_ac":2.0,  "pace":1,    "home_price":720000,    "wsd_per_ff":290, "paving_per_ff":220, "dev_start_month":1, "landscaping_per_lot":2000, "urd_per_lot":35, "lots_per_streetlight":3, "fence_cost_per_ff":94},
        {"front_footage":80,  "on":1, "yield_per_ac":1.5,  "pace":0.75, "home_price":750000,    "wsd_per_ff":290, "paving_per_ff":220, "dev_start_month":1, "landscaping_per_lot":2000, "urd_per_lot":35, "lots_per_streetlight":3, "fence_cost_per_ff":94},
        {"front_footage":85,  "on":0, "yield_per_ac":5.5,  "pace":0.75, "home_price":325000,    "wsd_per_ff":290, "paving_per_ff":220, "dev_start_month":1, "landscaping_per_lot":2000, "urd_per_lot":35, "lots_per_streetlight":2, "fence_cost_per_ff":94},
        {"front_footage":90,  "on":0, "yield_per_ac":5.5,  "pace":0.75, "home_price":360000,    "wsd_per_ff":290, "paving_per_ff":220, "dev_start_month":1, "landscaping_per_lot":2000, "urd_per_lot":35, "lots_per_streetlight":2, "fence_cost_per_ff":94},
        {"front_footage":95,  "on":0, "yield_per_ac":1.15, "pace":0.75, "home_price":385000,    "wsd_per_ff":290, "paving_per_ff":220, "dev_start_month":1, "landscaping_per_lot":2000, "urd_per_lot":35, "lots_per_streetlight":2, "fence_cost_per_ff":94},
        {"front_footage":100, "on":0, "yield_per_ac":1.0,  "pace":0.75, "home_price":410000,    "wsd_per_ff":290, "paving_per_ff":220, "dev_start_month":1, "landscaping_per_lot":2000, "urd_per_lot":35, "lots_per_streetlight":2, "fence_cost_per_ff":94},
    ]
    return {
        "project_name": name,
        "address": "",
        "gross_acreage": 0,
        "land_escalator": 0.05,
        "purchase_price_per_acre": 0,
        "closing_costs_pct": 0.045,
        "closing_date": "",
        "default_other_pct": 0.17,
        "sectional_other_pct": 0.25,       # Excel B6 = 0.17
        "landscaping_other_pct": 0.12,
        "contingency": 0.05,
        "site_work_pct": 0.01,
        "fenced_pct": 0.25,
        "cost_per_mailbox": 200,
        "cost_per_streetlight": 1700,
        "default_start_month": 1,
        "det_storage_rate": 1.1,            # Excel B31 = 1.1
        "det_depth": 9,                     # Excel B33 = 9
        "det_num_projects": 6,              # Excel B34 = 6
        "det_cost_per_cy": 10.0,            # Excel A37 = $10/CY
        "parks_pct": 0.03,                  # Excel B51 = 3%
        "drill_site_acres": 0,
        "commercial_pod_acres": 0,
        "residential_pod_acres": 0,
        "plants": [{"type":"None","notes":""} for _ in range(8)],
        "amenities": [{"type":"None","acres":0,"notes":""} for _ in range(6)],
        "other_netouts": [{"desc":"","acres":0,"notes":""} for _ in range(6)],
        "roads": [{"type":"","lf":0,"width":0,"road_setback":0,"landscaping_setback":0,"notes":""} for _ in range(6)],
        "takedowns": [{"period":0,"pct":0.5},{"period":36,"pct":0.5},{"period":0,"pct":0.0}],
        "plant_costs": [{"base_cost":0,"other_pct":0.17,"start_month":1,"ph2_base_cost":0,"ph2_other_pct":0.17,"ph2_start_month":37} for _ in range(8)],
        "amenity_costs": [{"base_cost":0,"other_pct":0.17,"start_month":1} for _ in range(6)],
        "det_costs": [{"other_pct":0.17,"landscaping_per_foot":2} for _ in range(6)],
        "other_costs": [{"base_cost":0,"other_pct":0.17,"start_month":1,"duration":1} for _ in range(6)],
        "road_costs": [{"other_pct":0.17,"start_month":1,"landscaping_per_sf":2,"light_spacing":0} for _ in range(6)],
        "lot_sizes": lot_size_defaults,
        "timing_method": "50/25/25",        # Excel B2 = 50/25/25
        "bem_period": 9,                    # Excel B3 = 9
        "bem_pct": 0.18,                    # Excel B4 = 18%
        "brokerage_fees": 0.03,             # Excel B5 = 3%
        "lot_closing_costs": 0.015,         # Excel B6 = 1.5%
        "take1_pct": 0.50,
        "take2_pct": 0.25,
        "take3_pct": 0.25,
        "price_per_ff": {str(yr): 1800 for yr in range(11)},
        "res_pod_acreage": 0,
        "res_pod_count": 1,
        "res_pods": [{"price_per_acre":120000,"closing_costs_pct":0.045,"implied_lots_per_acre":3.5,"impact_fee_per_lot":10000,"sale_period":12} for _ in range(6)],
        "comm_pod_acreage": 0,
        "comm_pod_count": 6,
        "comm_pods": [{"price_per_sf":8,"closing_costs_pct":0.045,"sale_period":12+i*24,"av_per_acre":1200000,"av_delay_months":18} for i in range(6)],
        "mud_bond": {"toggle":1,"amount":0,"reimbursement_pct":0.85,"first_bond_period":48,"bond_interval":12,"pct_to_dev":0.85,"receivables_fee":0.025,"debt_ratio":0.12},
        "wcid_bond": {"toggle":1,"amount":0,"reimbursement_pct":0.85,"first_bond_period":48,"bond_interval":12,"pct_to_dev":0.85,"receivables_fee":0.025,"debt_ratio":0.042},
        "marketing_pct": 0.02,
        "prof_svc_pct": 0.015,              # Excel B95 = 1.5%
        "dmf_pct": 0.025,                   # Excel B99 = 2.5%
        "personnel_monthly": 50000,         # Excel C103 = 50,000
        "marketing_personnel_monthly": 15000, # Excel C104 = 15,000
        "legal_monthly": 10000,             # Excel C108 = 10,000
        "mud_monthly": 35000,               # Excel C112 = 35,000
        "mud_pct": 0.2,                     # Excel D112 = 20% (what % of project MUD runs)
        "insurance_monthly": 10000,         # Excel C116 = 10,000
        "bookkeeping_monthly": 10000,       # Excel C120 = 10,000
    }

# ─── CHANGE LOG HELPERS ───────────────────────────────────────────────────────
_CHANGE_LOG_FIELDS = [
    ("purchase_price_per_acre", "Purchase Price / Acre"),
    ("gross_acreage",           "Gross Acreage"),
    ("land_escalator",          "Land Escalator"),
    ("contingency",             "Contingency"),
    ("closing_costs_pct",       "Closing Costs %"),
    ("timing_method",           "Revenue Timing Method"),
    ("dmf_pct",                 "DMF Rate"),
    ("personnel_monthly",       "Personnel Cost / Mo"),
]
_CHANGE_LOG_LOT_FIELDS = [
    ("home_price",   "Home Price"),
    ("pace",         "Pace (lots/mo)"),
    ("yield_per_ac", "Yield / Acre"),
    ("pct_mix",      "Mix %"),
]

def _compare_inputs(old_inp, new_inp):
    """Return list of {field, label, old, new} dicts for tracked changes."""
    changes = []
    for key, label in _CHANGE_LOG_FIELDS:
        ov, nv = old_inp.get(key), new_inp.get(key)
        if ov != nv and not (ov is None and nv is None):
            changes.append({"field": key, "label": label,
                            "old": ov, "new": nv})
    # Per-lot changes for active lots
    old_lots = old_inp.get("lot_sizes", [])
    new_lots = new_inp.get("lot_sizes", [])
    for i, new_lot in enumerate(new_lots):
        if not new_lot.get("on"):
            continue
        old_lot = old_lots[i] if i < len(old_lots) else {}
        ff = new_lot.get("ff", (i+1)*5+20)
        for key, label in _CHANGE_LOG_LOT_FIELDS:
            ov, nv = old_lot.get(key), new_lot.get(key)
            if ov != nv and not (ov is None and nv is None):
                changes.append({"field": f"lot_sizes[{i}].{key}",
                                "label": f"{label} ({ff}' lot)",
                                "old": ov, "new": nv})
    return changes


def _apply_sensitivity_override(inp, field, value):
    """Deep-copy inp, apply sensitivity override, return modified copy."""
    import copy
    inp2 = copy.deepcopy(inp)
    if field == "price_per_ff_base":
        # Scale all per-FF price years proportionally to the new base (year-0) value
        ppff = inp2.get("price_per_ff", {})
        ref = float(ppff.get("0", ppff.get(0, 1800)) or 1800)
        scale = float(value) / ref if ref else 1.0
        inp2["price_per_ff"] = {k: float(v or 0) * scale for k, v in ppff.items()}
    elif field == "lot_sizes.dev_cost_per_lot":
        # Scale wsd_per_ff and paving_per_ff proportionally across all active lots
        active = [r for r in inp2.get("lot_sizes", []) if r.get("on")]
        if active:
            costs = [(r.get("wsd_per_ff", 0) + r.get("paving_per_ff", 0)) * r.get("ff", 0)
                     for r in active]
            ref_avg = sum(costs) / len(costs) if costs else 0
            scale = float(value) / ref_avg if ref_avg else 1.0
            for row in inp2.get("lot_sizes", []):
                if row.get("on"):
                    row["wsd_per_ff"] = row.get("wsd_per_ff", 0) * scale
                    row["paving_per_ff"] = row.get("paving_per_ff", 0) * scale
    elif field.startswith("lot_sizes."):
        sub = field[len("lot_sizes."):]
        for row in inp2.get("lot_sizes", []):
            if row.get("on"):
                row[sub] = value
    else:
        inp2[field] = value
    return inp2


# ─────────────────────────────────────────────────────────────────────────────
# Ember Capital — redesigned page (Concept C two-column cockpit)
#
# Pulls live data from four sources:
#   • reports[returns]                       — project list, IRR/EM, yearly distributions
#   • reports[ember_capital_commitments]     — investor groups (raw $)
#   • reports[ember_capital_settings]        — LP/Promote recycle % per project
#   • reports[ember_capital_asset_classes]   — asset class assigned per project name
#   • projects table (status='Active')       — pipeline rows from MPC Underwriting
#
# Asset class is a piece of metadata layered onto each project independent of
# the source. Storage is a JSON blob keyed by project name (so Active and
# Pipeline rows with the same name share an assignment).
# ─────────────────────────────────────────────────────────────────────────────

# Canonical asset-class palette — order drives the donut wedge order.
# Colors are pulled directly from this list into the chip, donut, and
# Returns rail (no CSS-side hex duplication), so adding a class only
# requires editing this one place.
_CAP_ASSET_CLASSES = [
    {"id": "mpc-hub",        "label": "MPC Hub",        "color": "#F25929"},
    {"id": "mpc-spoke",      "label": "MPC Spoke",      "color": "#b058df"},
    {"id": "mpc-commercial", "label": "MPC Commercial", "color": "#C8A96E"},
    {"id": "condos",         "label": "Condos",         "color": "#3D87C7"},
    {"id": "mf-btr",         "label": "MF / BTR",       "color": "#5E9E8C"},
    {"id": "land",           "label": "Land",           "color": "#E07A3E"},
]
_CAP_VALID_CLASSES = {c["id"] for c in _CAP_ASSET_CLASSES}


def _capital_slug(name: str) -> str:
    """Stable slug for a project name. Used as the row id in the active +
    returns views (commitments use their own ids)."""
    import re
    s = (name or "").lower().strip()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-") or "project"


def _capital_load_asset_classes() -> dict:
    """{project_name: asset_class_id} from the reports JSONB blob."""
    conn = get_db(); cur = conn.cursor()
    cur.execute(
        "SELECT data FROM reports "
        "WHERE report_type = 'ember_capital_asset_classes' "
        "ORDER BY uploaded_at DESC LIMIT 1"
    )
    row = cur.fetchone(); cur.close(); conn.close()
    d = (row["data"] if row else None) or {}
    return d.get("by_project_name", {}) if isinstance(d, dict) else {}


def _capital_save_asset_class(project_name: str, class_id: str) -> None:
    """Upsert {project_name: class_id} into the asset-classes blob."""
    by = _capital_load_asset_classes()
    by[project_name] = class_id
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM reports WHERE report_type = 'ember_capital_asset_classes'")
    cur.execute(
        "INSERT INTO reports (report_type, data, uploaded_by) VALUES (%s, %s, %s)",
        ("ember_capital_asset_classes",
         json.dumps({"by_project_name": by}),
         session.get("user_id")),
    )
    conn.commit(); cur.close(); conn.close()


def _capital_commitments_totals(groups: list[dict]) -> dict:
    """Roll up the four commit/allocated columns + derived totals."""
    mpc       = sum((g.get("mpc")                or 0) for g in groups)
    mpc_a     = sum((g.get("mpc_allocated")      or 0) for g in groups)
    vert      = sum((g.get("vertical")           or 0) for g in groups)
    vert_a    = sum((g.get("vertical_allocated") or 0) for g in groups)
    return {
        "mpc": mpc, "mpc_allocated": mpc_a,
        "vertical": vert, "vertical_allocated": vert_a,
        "total_committed": mpc + vert,
        "total_allocated": mpc_a + vert_a,
        "available":       (mpc + vert) - (mpc_a + vert_a),
    }


# ---------------------------------------------------------------------------
# Investor portfolios (cap tables × project returns)
# ---------------------------------------------------------------------------
# The Investors view scales each project's LP cashflow streams by an
# investment vehicle's ownership % (from the cap tables) to derive that
# vehicle's portfolio and returns. Cap-table positions are stored in the
# reports table under report_type='ember_capital_captable'; the returns
# streams come from the same 'returns' blob the rest of /capital uses.
#
# Unit note: returns arrays are in $000s; cap-table contributions are in
# actual dollars. Everything below is normalized to ACTUAL DOLLARS (returns
# streams are multiplied by 1000) so the Investors tab is internally
# consistent and reads in real money.

def _capital_load_captable() -> dict:
    """{positions: [...], uploaded_at} from reports[ember_capital_captable].

    Position shape: {project, vehicle, contribution, pct, equity_class}.
    Returns an empty list when nothing has been uploaded yet.
    """
    conn = get_db(); cur = conn.cursor()
    cur.execute(
        "SELECT data, uploaded_at FROM reports "
        "WHERE report_type = 'ember_capital_captable' "
        "ORDER BY uploaded_at DESC LIMIT 1"
    )
    row = cur.fetchone(); cur.close(); conn.close()
    if not row:
        return {"positions": [], "uploaded_at": None}
    d = (row["data"] or {})
    return {
        "positions":   d.get("positions", []) or [],
        "uploaded_at": row["uploaded_at"].isoformat() if row.get("uploaded_at") else None,
    }


def _simple_irr(cashflows: list[float], lo: float = -0.9999, hi: float = 10.0) -> float | None:
    """Annual IRR for a per-year net-cashflow series via bisection.

    cashflows[t] is the net flow in year t (contributions negative,
    distributions positive). Returns the rate as a decimal (0.18 = 18%),
    or None when the series has no sign change (IRR undefined).
    """
    if not cashflows or all(abs(c) < 1e-9 for c in cashflows):
        return None
    has_pos = any(c > 0 for c in cashflows)
    has_neg = any(c < 0 for c in cashflows)
    if not (has_pos and has_neg):
        return None

    def npv(rate: float) -> float:
        return sum(cf / ((1.0 + rate) ** t) for t, cf in enumerate(cashflows))

    f_lo, f_hi = npv(lo), npv(hi)
    if f_lo == 0:
        return lo
    if f_hi == 0:
        return hi
    if (f_lo > 0) == (f_hi > 0):
        # No bracketed root in [lo, hi] — IRR outside our search range.
        return None
    for _ in range(200):
        mid = (lo + hi) / 2.0
        f_mid = npv(mid)
        if abs(f_mid) < 1e-6:
            return mid
        if (f_mid > 0) == (f_lo > 0):
            lo, f_lo = mid, f_mid
        else:
            hi, f_hi = mid, f_mid
    return (lo + hi) / 2.0


# Legal-entity suffixes and Ember project codes stripped when collapsing
# vehicle-name variants into one rolled-up investor (e.g. "CCDL Ventures GPD,
# LLC", "CCDL Ventures WRRD 16663 39 LLC", "CCDL Ventures, LLC" -> one row).
_VEHICLE_LEGAL_TOKENS = {
    "llc", "l.l.c", "l.l.c.", "inc", "incorporated", "corp", "corporation",
    "co", "company", "lp", "l.p", "l.p.", "llp", "ltd", "limited",
    "partners", "partnership", "holdings", "holding", "group", "the",
}
_VEHICLE_PROJECT_TOKENS = {"gpd", "hld", "wrg", "wrrd"}


def _canon_vehicle_key(name: str) -> str:
    """Normalized grouping key for an investment vehicle. Merges punctuation,
    legal-suffix, and project-code variants so the same entity rolls up into a
    single investor row regardless of how each cap table spelled it."""
    cleaned = re.sub(r"[.,\-/&]", " ", (name or "").lower())
    keep = [
        t for t in cleaned.split()
        if t and t not in _VEHICLE_LEGAL_TOKENS
        and t not in _VEHICLE_PROJECT_TOKENS and not t.isdigit()
    ]
    return " ".join(keep)


def _clean_project_label(name: str) -> str:
    """Drop a trailing equity-class parenthetical (e.g. 'LightHaven (Preferred
    Equity)' -> 'LightHaven') so a position's Class column isn't echoed in the
    project name and common/preferred rows don't look duplicated."""
    out = re.sub(r"\s*\([^)]*equity[^)]*\)\s*$", "", name or "", flags=re.I).strip()
    return out or (name or "")


def _build_investor_view(raw_projects: list, years_str: list, years_int: list,
                         src_months: list | None = None) -> dict:
    """Aggregate cap-table positions into per-vehicle portfolios scaled by
    each project's LP returns streams. Grouping unit = investment vehicle.

    Distributions are split into "to date" (actually received through today)
    and "forecast" (everything after today). When the returns workbook ships
    a Monthly Cashflows block (`src_months`), the split is month-accurate;
    otherwise it falls back to whole past years plus a pro-rated current year.

    Returns {investors: [...], years, projects_matched, projects_unmatched,
    uploaded_at, total_*}. All money values in actual dollars.
    """
    from captable_parser import normalize_project, resolve_returns_project

    blob = _capital_load_captable()
    positions = blob.get("positions", []) or []
    # Legacy compatibility: cap tables uploaded before LightHaven was split into
    # two standalone returns deals stored both blocks under one "LightHaven"
    # project. Remap by equity class so older stored blobs wire correctly
    # without forcing a re-upload.
    for pos in positions:
        if (pos.get("project") or "").strip() == "LightHaven":
            pos["project"] = ("LightHaven Preferred"
                              if pos.get("equity_class") == "preferred"
                              else "LightHaven Common")
    n_years = len(years_str)

    # ── To-date boundary ──────────────────────────────────────────────
    today      = datetime.date.today()
    cur_year   = today.year
    today_iso  = today.isoformat()
    months_iso = [str(m)[:10] for m in (src_months or [])]
    year_frac  = today.timetuple().tm_yday / 366.0  # current-year pro-rate

    def _to_date_sum(yearly: list, monthly: list) -> float:
        """Portion of a stream (same units as inputs) received on/before
        today. Month-accurate when a monthly array aligned to src_months is
        present; else past years in full + current year pro-rated."""
        if months_iso and monthly and len(monthly) == len(months_iso):
            return sum(float(monthly[i] or 0)
                       for i, iso in enumerate(months_iso) if iso <= today_iso)
        s = 0.0
        for i, y in enumerate(years_int):
            if i >= len(yearly):
                break
            v = float(yearly[i] or 0)
            if y < cur_year:
                s += v
            elif y == cur_year:
                s += v * year_frac
        return s

    # ── Forecast axis (cashflow columns) ──────────────────────────────
    # Rolling near-term window + long-term roll-up: the NEXT 4 quarters
    # (starting at the quarter today falls in) get their own columns; each
    # subsequent year is a single annual column. As a quarter ends it falls
    # into "to date" and the window advances to the next 4. ("To Date" is its
    # own column upstream.)
    cur_q = (today.month - 1) // 3 + 1
    last_year = max(years_int) if years_int else cur_year
    _wins: list[tuple[int, int]] = []          # next 4 quarters as (year, q)
    _wy, _wq = cur_year, cur_q
    for _ in range(4):
        _wins.append((_wy, _wq))
        _wq += 1
        if _wq > 4:
            _wq = 1; _wy += 1
    window_index = {(y, q): i for i, (y, q) in enumerate(_wins)}
    annual_years = list(range(cur_year + 1, last_year + 1))
    quarter_labels = (["Q%d %d" % (q, y) for (y, q) in _wins]
                      + [str(y) for y in annual_years])
    n_q = len(quarter_labels)

    def _col_index(year: int, q: int) -> int | None:
        """Forecast-axis column for a (year, quarter): one of the next-4-quarter
        columns, else the matching annual column; None if before the window."""
        j = window_index.get((year, q))
        if j is not None:
            return j
        if annual_years and year >= cur_year + 1:
            return 4 + (min(year, last_year) - (cur_year + 1))
        return None

    def _future_quarter_buckets(yearly: list, monthly: list, pct: float,
                                forecast_total: float) -> list[float]:
        """Distribute a position's forecast (actual $) across the forecast
        axis (next-4-quarter columns + annual columns). Uses the monthly block
        when it aligns to src_months; else splits each future year's value
        across its future quarters and routes them to the right column.
        Buckets scale to sum to `forecast_total`."""
        raw = [0.0] * n_q
        if months_iso and monthly and len(monthly) == len(months_iso):
            for i, iso in enumerate(months_iso):
                if iso > today_iso:
                    y = int(iso[:4]); q = (int(iso[5:7]) - 1) // 3 + 1
                    j = _col_index(y, q)
                    if j is not None:
                        raw[j] += float(monthly[i] or 0) * pct * 1000.0
        else:
            for i, y in enumerate(years_int):
                if y < cur_year or i >= len(yearly):
                    continue
                yv = float(yearly[i] or 0) * pct * 1000.0
                fq = [q for q in range(1, 5) if not (y == cur_year and q < cur_q)]
                if not fq:
                    continue
                if y == cur_year:
                    yv *= max(0.0, 1.0 - year_frac)
                per = yv / len(fq)
                for q in fq:
                    j = _col_index(y, q)
                    if j is not None:
                        raw[j] += per
        tot = sum(raw)
        if tot > 1e-9:
            scale = forecast_total / tot
            return [r * scale for r in raw]
        # No shape info — spread evenly across the axis.
        if n_q and forecast_total:
            return [forecast_total / n_q] * n_q
        return raw

    # ── Per-project returns lookup (keyed by normalized returns name) ──
    # Each entry carries the streams we scale: distributions, preferred
    # return, contributions (for IRR timing), plus project IRR/EM.
    proj_lookup: dict[str, dict] = {}
    returns_names: list[str] = []
    for p in raw_projects:
        if p.get("active") is False:
            continue
        name = (p.get("name") or "").strip()
        if not name:
            continue
        returns_names.append(name)
        by_label = {m.get("label"): m for m in (p.get("metrics") or [])}

        def _y(label: str) -> list[float]:
            v = (by_label.get(label) or {}).get("yearly") or []
            return [float(x or 0) for x in (list(v) + [0] * max(0, n_years - len(v)))[:n_years]]

        def _m(label: str) -> list[float]:
            """Monthly array for the metric (parallel to top-level src['months'])."""
            v = (by_label.get(label) or {}).get("monthly") or []
            return [float(x or 0) for x in v]

        def _t(label: str, default: float = 0.0) -> float:
            v = (by_label.get(label) or {}).get("total")
            try:
                return float(v) if v is not None else default
            except (TypeError, ValueError):
                return default

        irr_dec = _t("LP IRR", 0.0)
        irr_pct = irr_dec * 100 if abs(irr_dec) <= 1.5 else irr_dec
        proj_lookup[normalize_project(name)] = {
            "name":     name,
            "dist_y":   _y("Total LP Distributions"),    # $000s, yearly (timing)
            "contrib_y": _y("Total LP Contributions"),   # $000s, stored negative
            "dist_m":   _m("Total LP Distributions"),    # $000s, monthly (timing)
            # Authoritative lifetime total from the model's own total column
            # (E). The yearly/monthly grids only drive the to-date vs forecast
            # timing split; magnitude comes from this so a vehicle's EM ties to
            # the project EM (which is also computed off the total column).
            "dist_total": _t("Total LP Distributions"),  # $000s
            "irr":      round(irr_pct, 1),
            "em":       round(_t("LP Equity Multiple", 0.0), 2),
        }

    # Resolve each cap-table project to a returns project once.
    captable_projects = sorted({pos["project"] for pos in positions})
    resolved: dict[str, str | None] = {}
    for cp in captable_projects:
        match = resolve_returns_project(cp, returns_names)
        resolved[cp] = normalize_project(match) if match else None

    projects_matched   = sorted(cp for cp in captable_projects if resolved.get(cp))
    projects_unmatched = sorted(cp for cp in captable_projects if not resolved.get(cp))

    # ── Aggregate by vehicle ──
    vehicles: dict[str, dict] = {}
    for pos in positions:
        cp = pos["project"]
        norm = resolved.get(cp)
        if not norm or norm not in proj_lookup:
            continue  # match-by-name only: ignore positions with no return
        pj = proj_lookup[norm]
        pct = float(pos.get("pct") or 0.0)
        equity_class = pos.get("equity_class", "common")
        contribution = float(pos.get("contribution") or 0.0)  # actual $

        # Each position uses its matched project's Total LP Distributions. The
        # yearly/monthly arrays carry only the *timing shape*; `base_total` is
        # the authoritative lifetime magnitude from the model's total column.
        # (LightHaven preferred vs common are now distinct returns projects, so
        # there's no longer any per-class stream subtraction — equity_class is
        # carried through only as a display label.)
        base       = pj["dist_y"]
        base_m     = pj["dist_m"]
        base_total = pj["dist_total"]

        # Magnitude from the total column; timing from the yearly grid. Rescale
        # the yearly stream so it sums to the authoritative total (preserving
        # the year-by-year shape for IRR), then convert $000s -> actual $.
        shape_sum      = sum(base)
        mag_scale      = (base_total / shape_sum) if abs(shape_sum) > 1e-9 else 0.0
        dist_stream    = [v * pct * 1000.0 * mag_scale for v in base]
        contrib_stream = [v * pct * 1000.0 for v in pj["contrib_y"]]  # negative
        total_dist_pos = base_total * pct * 1000.0
        # To-date fraction from the timing streams, applied to the true total.
        td_frac        = (_to_date_sum(base, base_m) / shape_sum) if abs(shape_sum) > 1e-9 else 0.0
        td_frac        = min(max(td_frac, 0.0), 1.0)
        to_date_pos    = total_dist_pos * td_frac
        forecast_pos   = total_dist_pos - to_date_pos
        q_buckets      = _future_quarter_buckets(base, base_m, pct, forecast_pos)

        vname = pos["vehicle"]
        vkey = _canon_vehicle_key(vname) or vname.strip().lower()
        v = vehicles.setdefault(vkey, {
            "name": vname,
            "variants": set(),
            "committed": 0.0,
            "dist_total": 0.0,
            "dist_yearly": [0.0] * n_years,
            "contrib_yearly": [0.0] * n_years,
            "dist_to_date": 0.0,
            "q_forecast": [0.0] * n_q,
            "positions": [],
        })
        v["variants"].add(vname.strip())
        v["committed"] += contribution
        v["dist_total"] += total_dist_pos
        v["dist_to_date"] += to_date_pos
        for i in range(n_years):
            v["dist_yearly"][i]    += dist_stream[i]
            v["contrib_yearly"][i] += contrib_stream[i]
        for j in range(n_q):
            v["q_forecast"][j] += q_buckets[j]
        v["positions"].append({
            "project":      _clean_project_label(pj["name"]),
            "equity_class": equity_class,
            "pct":          round(pct * 100, 4),
            "committed":    int(round(contribution)),
            "distributions": int(round(total_dist_pos)),
            "dist_to_date":  int(round(to_date_pos)),
            "dist_forecast": int(round(forecast_pos)),
            "q_forecast":   [int(round(x)) for x in q_buckets],
            "irr":          pj["irr"],
            "em":           pj["em"],
        })

    # ── Finalize per-vehicle metrics ──
    investors = []
    for vkey, v in vehicles.items():
        # Canonical display = the cleanest spelled variant (fewest characters,
        # alphabetical tie-break) so a rolled-up entity shows one tidy name.
        if v["variants"]:
            v["name"] = min(v["variants"], key=lambda s: (len(s), s))
        total_dist = v["dist_total"]   # authoritative (model total column)
        to_date    = max(0.0, min(v["dist_to_date"], total_dist))
        forecast   = total_dist - to_date
        committed  = v["committed"]
        # EM / profit anchored on the cap-table committed dollars (the
        # ownership basis). Falls back to returns-derived contributions
        # when the cap table has no contribution figure.
        basis = committed if committed > 0 else abs(sum(v["contrib_yearly"]))
        em = round(total_dist / basis, 2) if basis else 0.0
        profit = total_dist - basis
        # Investment-weighted IRR: blend each position's project LP IRR by the
        # dollars invested in that position (committed).
        irr_w = sum(p["committed"] for p in v["positions"])
        irr_pct = (round(sum((p["irr"] or 0) * p["committed"]
                             for p in v["positions"]) / irr_w, 1)
                   if irr_w > 0 else None)
        investors.append({
            "id":            _capital_slug(v["name"]),
            "name":          v["name"],
            "committed":     int(round(committed)),
            "distributions": int(round(total_dist)),
            "distributions_to_date":  int(round(to_date)),
            "distributions_forecast": int(round(forecast)),
            "profit":        int(round(profit)),
            "em":            em,
            "irr":           irr_pct,
            "yearly":        [int(round(x)) for x in v["dist_yearly"]],
            "q_forecast":    [int(round(x)) for x in v["q_forecast"]],
            "positions":     sorted(v["positions"], key=lambda r: r["project"]),
        })

    investors.sort(key=lambda r: r["committed"], reverse=True)

    # ── Match coverage (both directions) ──────────────────────────────
    # Returns side: which active returns projects do the cap tables cover,
    # and which have no cap-table positions at all (so they never appear).
    matched_norms = {n for n in resolved.values() if n}
    returns_matched   = sorted(proj_lookup[n]["name"] for n in matched_norms
                               if n in proj_lookup)
    returns_unmatched = sorted(pj["name"] for n, pj in proj_lookup.items()
                               if n not in matched_norms)
    # Portfolio quarterly forecast = sum of investors' rows.
    total_q_forecast = [0] * n_q
    for inv in investors:
        for j, val in enumerate(inv["q_forecast"]):
            total_q_forecast[j] += val

    # Portfolio investment-weighted IRR (every position's project IRR weighted
    # by dollars invested).
    _all_pos = [p for inv in investors for p in inv["positions"]]
    _w = sum(p["committed"] for p in _all_pos)
    total_irr = (round(sum((p["irr"] or 0) * p["committed"] for p in _all_pos) / _w, 1)
                 if _w > 0 else None)

    return {
        "investors":          investors,
        "positions":          positions,   # flat raw rows (pct as fraction)
        "years":              years_str,
        "quarters":           quarter_labels,
        "total_q_forecast":   total_q_forecast,
        "projects_matched":   projects_matched,       # cap-table side
        "projects_unmatched": projects_unmatched,     # cap-table → no return
        "returns_matched":    returns_matched,        # returns w/ cap-table coverage
        "returns_unmatched":  returns_unmatched,      # active returns, no cap table
        "returns_total":      len(returns_names),
        "uploaded_at":        blob.get("uploaded_at"),
        "total_committed":    int(round(sum(i["committed"] for i in investors))),
        "total_distributions": int(round(sum(i["distributions"] for i in investors))),
        "total_distributions_to_date":  int(round(sum(i["distributions_to_date"] for i in investors))),
        "total_distributions_forecast": int(round(sum(i["distributions_forecast"] for i in investors))),
        "total_irr":          total_irr,   # investment-weighted across positions
        "count":              len(investors),
    }


def _build_capital_view_context() -> dict:
    """Assemble the dict the redesigned /capital template consumes.

    Defensive on every section: returns a usable shape even when there's no
    returns upload, no commitments, no pipeline. The template's "no data"
    fallbacks render gracefully against empty lists.
    """
    today = datetime.datetime.now()
    current_year = today.year
    months_elapsed = today.month - 1 + (today.day / 31.0)  # rough YTD fraction
    ytd_fraction = months_elapsed / 12.0
    months_left_this_year = 12 - months_elapsed
    ytd_label = f"Jan – {today.strftime('%b')} {current_year}"

    asset_class_map = _capital_load_asset_classes()
    def _class_for(name: str) -> str:
        return asset_class_map.get(name) or "mpc-hub"

    # ── Returns blob: drives Active list + Returns view + recycle chart ──
    conn = get_db(); cur = conn.cursor()
    cur.execute(
        "SELECT data, uploaded_at FROM reports "
        "WHERE report_type = 'returns' ORDER BY uploaded_at DESC LIMIT 1"
    )
    rrow = cur.fetchone()
    src = (rrow["data"] if rrow else {}) or {}
    years_int = list(src.get("years", []) or [])
    years_str = [str(y) for y in years_int]
    cur_year_str = str(current_year)
    # Monthly date headers from the new "Monthly Cashflows" block on
    # the Consolidated Project Returns tab. ISO YYYY-MM-DD strings,
    # one per monthly column. Empty list when the workbook predates
    # the monthly block — to_date falls back to year-pro-rate then.
    src_months = list(src.get("months", []) or [])

    # Pull project metrics into a clean per-project dict.
    raw_projects = src.get("projects", []) or []
    active = []
    returns_by_project = {}
    portfolio_lp_dist_yearly = [0.0] * len(years_str)
    portfolio_promote_yearly = [0.0] * len(years_str)
    # Two distinct distribution roll-ups, both fed from per-project
    # monthly arrays (with year-pro-rate fallback when the workbook
    # predates the monthly block):
    #   - portfolio_distributed_ltd: every distribution from project
    #     inception through today. Drives the per-project "To Date"
    #     column on the Active Projects table (column = sum of this).
    #   - portfolio_distributed_ytd_current: current calendar year
    #     only, Jan 1 → today. Drives the bottom-right "Distributed
    #     YTD" KPI on page 2 of the report.
    portfolio_distributed_ltd          = 0.0
    portfolio_distributed_ytd_current  = 0.0
    portfolio_to_be_distributed_18mo   = 0.0  # next 18 months after today
    # Pre-compute reusable monthly-array index sets. Walk src_months
    # once instead of per-project.
    #   cur_year_indices  — current year, on/before today (YTD bucket)
    #   next_18mo_indices — first 18 months strictly after today
    today_iso = today.date().isoformat()
    cur_year_indices = [
        mi for mi, iso in enumerate(src_months)
        if str(iso)[:4] == cur_year_str and str(iso) <= today_iso
    ]
    next_18mo_indices = [
        mi for mi, iso in enumerate(src_months)
        if str(iso) > today_iso
    ][:18]
    eq_weighted_irr_num = 0.0
    eq_weighted_irr_den = 0.0
    total_lp_profit = 0.0
    total_promote   = 0.0
    total_equity    = 0.0
    for p in raw_projects:
        if p.get("active") is False:
            continue
        name = (p.get("name") or "").strip()
        if not name:
            continue
        by_label = {m.get("label"): m for m in (p.get("metrics") or [])}
        def _t(label, default=0.0):
            v = (by_label.get(label) or {}).get("total")
            try: return float(v) if v is not None else default
            except (TypeError, ValueError): return default
        def _y(label):
            v = (by_label.get(label) or {}).get("yearly") or []
            n = len(years_str)
            return list((list(v) + [0] * max(0, n - len(v)))[:n])
        def _m(label):
            """Monthly array for the metric (parallel to top-level src['months'])."""
            return list((by_label.get(label) or {}).get("monthly") or [])

        irr_dec  = _t("LP IRR", 0.0)
        irr_pct  = irr_dec * 100 if abs(irr_dec) <= 1.5 else irr_dec
        em_val   = _t("LP Equity Multiple", 0.0)
        profit   = _t("Total LP Profit", 0.0)
        promote  = _t("Promote", 0.0)
        contrib  = abs(_t("Total LP Contributions", 0.0))   # stored negative; equity is positive
        dist_y   = _y("Total LP Distributions")
        dist_m   = _m("Total LP Distributions")
        prom_y   = _y("Promote")
        prom_m   = _m("Promote")

        # "To Date" (per-project column) — life-to-date LP distributions
        # received through the most recently completed month. Sum of
        # every monthly cell whose ISO date is on/before today when
        # the workbook ships a Monthly Cashflows block; falls back to
        # prior years summed + current year × ytd_fraction otherwise.
        # Some workbook uploads have buggy monthly cells (rows shifted,
        # spurious values). When sum(monthly) doesn't tie to the yearly
        # total within tolerance, fall back to yearly. We compute this
        # check per-metric per-project so e.g. Mid Main's broken Promote
        # row falls back while Hawthorne's clean rows keep using
        # exact monthly precision.
        def _monthly_reliable(metric):
            mlist = (metric or {}).get("monthly") or []
            if not mlist or not src_months:
                return False
            try: total = float(metric.get("total") or 0)
            except (TypeError, ValueError): return False
            try: sm = sum(float(v or 0) for v in mlist)
            except (TypeError, ValueError): return False
            tol = max(0.5, abs(total) * 0.001)  # $0.5K or 0.1% of total
            return abs(sm - total) <= tol

        dist_metric = by_label.get("Total LP Distributions") or {}
        prom_metric = by_label.get("Promote") or {}
        dist_reliable = _monthly_reliable(dist_metric)
        prom_reliable = _monthly_reliable(prom_metric)
        idx_cur = years_int.index(current_year) if current_year in years_int else -1

        # ── To Date (life-to-date through today) ─────────────────────
        if dist_reliable:
            cutoff = sum(1 for d in src_months if str(d) <= today_iso)
            to_date = sum(dist_m[:cutoff])
        elif idx_cur >= 0:
            # Fallback: prior years summed; current year treated as
            # all-future (conservative — broken monthly means we don't
            # know what's been paid yet, so don't double-count).
            to_date = sum(dist_y[:idx_cur])
        else:
            to_date = 0.0

        # ── YTD (current calendar year, Jan 1 → today) ──────────────
        if dist_reliable:
            ytd_current = sum(
                float(dist_m[mi] or 0)
                for mi in cur_year_indices
                if mi < len(dist_m)
            )
        else:
            # No monthly precision → conservative 0 for the current
            # year (we don't know how much of the yearly amount has
            # already been paid out).
            ytd_current = 0.0

        # ── To Be Distributed — LP next 18mo ────────────────────────
        if dist_reliable:
            next18_lp = sum(
                float(dist_m[mi] or 0)
                for mi in next_18mo_indices
                if mi < len(dist_m)
            )
        else:
            # Yearly fallback. The current year is treated as
            # all-future (matches the conservative YTD=0 above), and
            # subsequent years are pro-rated by how many of their
            # months fall inside the 18-month window.
            next18_lp = 0.0
            if idx_cur >= 0:
                if idx_cur < len(dist_y):
                    next18_lp += float(dist_y[idx_cur])
                for j in range(idx_cur + 1, len(years_int)):
                    months_to_start = (years_int[j] - current_year) * 12 - months_elapsed
                    if months_to_start >= 18:
                        break
                    fraction = min(12.0, 18.0 - months_to_start) / 12.0
                    next18_lp += float(dist_y[j] if j < len(dist_y) else 0) * fraction

        # ── To Be Distributed — Promote next 18mo ───────────────────
        if prom_reliable:
            next18_pr = sum(
                float(prom_m[mi] or 0)
                for mi in next_18mo_indices
                if mi < len(prom_m)
            )
        else:
            next18_pr = 0.0
            if idx_cur >= 0:
                if idx_cur < len(prom_y):
                    next18_pr += float(prom_y[idx_cur])
                for j in range(idx_cur + 1, len(years_int)):
                    months_to_start = (years_int[j] - current_year) * 12 - months_elapsed
                    if months_to_start >= 18:
                        break
                    fraction = min(12.0, 18.0 - months_to_start) / 12.0
                    next18_pr += float(prom_y[j] if j < len(prom_y) else 0) * fraction

        next_18mo_total = next18_lp + next18_pr

        slug = _capital_slug(name)
        active.append({
            "id":          slug,
            "name":        name,
            "asset_class": _class_for(name),
            "equity":      int(round(contrib)),
            "irr":         round(irr_pct, 1),
            "em":          round(em_val, 2),
            "to_date":     int(round(to_date)),
            "profit":      int(round(profit)),
        })
        returns_by_project[slug] = {
            "name":        name,
            "asset_class": _class_for(name),
            "yearly":      [int(round(v)) for v in dist_y],
        }

        # Portfolio aggregates
        for i in range(len(years_str)):
            portfolio_lp_dist_yearly[i] += dist_y[i] if i < len(dist_y) else 0
            portfolio_promote_yearly[i] += prom_y[i] if i < len(prom_y) else 0
        # Three roll-ups, all summed across active projects:
        #   life-to-date  → Active Projects column
        #   current-year-YTD → Distributed YTD KPI
        #   next-18mo forecast → To Be Distributed KPI
        portfolio_distributed_ltd         += to_date
        portfolio_distributed_ytd_current += ytd_current
        portfolio_to_be_distributed_18mo  += next_18mo_total
        if contrib:
            eq_weighted_irr_num += irr_pct * contrib
            eq_weighted_irr_den += contrib
        total_lp_profit += profit
        total_promote   += promote
        total_equity    += contrib

    forecasted_lp_irr = round(eq_weighted_irr_num / eq_weighted_irr_den, 1) if eq_weighted_irr_den else 0.0

    # ── Recycle settings (LP/Promote % per project) → annual recycle bars ──
    cur.execute(
        "SELECT data FROM reports WHERE report_type = 'ember_capital_settings' "
        "ORDER BY uploaded_at DESC LIMIT 1"
    )
    srow = cur.fetchone()
    settings = (srow["data"] if srow else {}) or {}
    recycle_map = settings.get("recycle", {}) or {}
    # Default 100/100 if a project hasn't been touched (matches the Returns
    # page's new default).
    def _rec_for(name):
        v = recycle_map.get(name) or {}
        return float(v.get("lp", 100)) / 100.0, float(v.get("prom", 100)) / 100.0

    # ── Annual Capital Recycling & Equity Investment ───────────────────
    # Per the v3 chart redesign (CHANGELOG-v3.md): two stacked bars per
    # year over a fixed 2021-2035 window. Left bar = Contributions
    # (active solid + pipeline dashed top); right bar = Distributions
    # (active LP solid base + active Promote stack + pipeline dashed top).
    # All values in $K. Active rows roll up the live returns blob;
    # pipeline rows roll up the manual_pipeline contributions /
    # distributions arrays.
    recycle_year_start = 2021
    recycle_year_end   = 2035
    recycle_years = list(range(recycle_year_start, recycle_year_end + 1))
    # Pre-load manual pipeline once for the recycle aggregation
    manual_for_recycle = list((manual_blob_for_recycle := _capital_load_manual_pipeline()).get("projects") or [])

    # Pre-compute which monthly cells are realized (ended on or before
    # today) vs forecast (in the future). Used to split actDLP for the
    # current year — past years are 100% realized, future years are
    # 100% forecast, and the bar for the CURRENT year is split month-
    # by-month so the report draws a clean line at "today".
    today_iso = today.date().isoformat()
    monthly_idx_by_year: dict[int, dict] = {}
    if src_months:
        for _idx, _iso in enumerate(src_months):
            try: _yr = int(str(_iso)[:4])
            except (TypeError, ValueError): continue
            slot = monthly_idx_by_year.setdefault(_yr, {"realized": [], "forecast": []})
            if str(_iso) <= today_iso:
                slot["realized"].append(_idx)
            else:
                slot["forecast"].append(_idx)

    recycle_rows = []
    for yr in recycle_years:
        actC = actDLP_real = actDLP_fore = actDProm = pipC = 0.0
        # Active side — sum across the live returns projects, indexed by
        # the project's native year array (typically 2023-2036).
        if yr in years_int:
            i_yr = years_int.index(yr)
            for p in raw_projects:
                if p.get("active") is False:
                    continue
                by_label = {m.get("label"): m for m in (p.get("metrics") or [])}
                contrib_y = (by_label.get("Total LP Contributions") or {}).get("yearly") or []
                dist_y    = (by_label.get("Total LP Distributions") or {}).get("yearly") or []
                dist_m    = (by_label.get("Total LP Distributions") or {}).get("monthly") or []
                prom_y    = (by_label.get("Promote") or {}).get("yearly") or []
                actC     += abs(float(contrib_y[i_yr])) if i_yr < len(contrib_y) else 0.0
                actDProm += float(prom_y[i_yr])        if i_yr < len(prom_y)    else 0.0
                # Split the LP distribution into realized (≤ today) vs
                # forecast (> today). Prefer monthly precision when the
                # workbook ships a Monthly Cashflows block; otherwise
                # fall back to a year-based split (past = realized, this
                # year = ytd_fraction split, future = forecast).
                year_total = float(dist_y[i_yr]) if i_yr < len(dist_y) else 0.0
                slot = monthly_idx_by_year.get(yr) if dist_m else None
                if slot:
                    for mi in slot["realized"]:
                        if mi < len(dist_m): actDLP_real += float(dist_m[mi] or 0)
                    for mi in slot["forecast"]:
                        if mi < len(dist_m): actDLP_fore += float(dist_m[mi] or 0)
                elif yr < current_year:
                    actDLP_real += year_total
                elif yr > current_year:
                    actDLP_fore += year_total
                else:  # current year, no monthly data — pro-rate
                    actDLP_real += year_total * ytd_fraction
                    actDLP_fore += year_total * (1.0 - ytd_fraction)

        # Pipeline side — only contributions; per partner feedback
        # we no longer render the dashed-teal "Distributions · Pipeline"
        # series. Pipeline distributions still flow through Capital's
        # KPI roll-up (to_be_distributed etc.) — just not on this chart.
        for mp in manual_for_recycle:
            ystart = mp.get("years_start", _CAP_MANUAL_PIPELINE_YEAR_START)
            idx = yr - ystart
            mc = mp.get("contributions_yearly") or []
            if 0 <= idx < len(mc):
                pipC += abs(float(mc[idx])) / 1000.0

        recycle_rows.append({
            "year":           yr,
            "actC":           int(round(actC)),
            # Split LP distributions into two stack segments. actDLP is
            # kept for any downstream consumer that just wants the
            # combined total without caring about realization status.
            "actDLP":         int(round(actDLP_real + actDLP_fore)),
            "actDLP_real":    int(round(actDLP_real)),
            "actDLP_fore":    int(round(actDLP_fore)),
            "actDProm":       int(round(actDProm)),
            "pipC":           int(round(pipC)),
            # Pipeline distributions deliberately omitted from the chart
            # (legacy `pipD: 0` for backwards compatibility).
            "pipD":           0,
        })

    # Y-axis ceiling: round to next 10K above the tallest bar so the
    # chart auto-scales when numbers grow. Default 50K matches the
    # design canvas reference numbers.
    max_bar = 0
    for r in recycle_rows:
        contrib_total = r["actC"] + r["pipC"]
        # actDLP already = real + fore, no double-counting needed.
        dist_total    = r["actDLP"] + r["actDProm"]
        max_bar = max(max_bar, contrib_total, dist_total)
    y_axis_max = max(50_000, int((max_bar // 10_000 + 1) * 10_000)) if max_bar else 50_000

    # ── Commitments ────────────────────────────────────────────────────
    cur.execute(
        "SELECT data FROM reports WHERE report_type = 'ember_capital_commitments' "
        "ORDER BY uploaded_at DESC LIMIT 1"
    )
    crow = cur.fetchone()
    cdata = (crow["data"] if crow else {}) or {}
    commit_groups = []
    for g in (cdata.get("groups") or []):
        commit_groups.append({
            "id":   _capital_slug(g.get("name", "")),
            "name": g.get("name", ""),
            "mpc":                int(g.get("mpc")                or 0),
            "mpc_allocated":      int(g.get("mpc_allocated")      or 0),
            "vertical":           int(g.get("vertical")           or 0),
            "vertical_allocated": int(g.get("vertical_allocated") or 0),
        })
    commit_totals = _capital_commitments_totals(commit_groups)
    total_committed_k         = round(commit_totals["total_committed"] / 1000)
    unallocated_commitments_k = round(commit_totals["available"]       / 1000)

    # ── Pipeline: MPC underwriting projects with status='Active' ───────
    hidden_pipeline = _capital_load_pipeline_visibility()
    cur.execute("""
        SELECT id, name, address, updated_at,
               outputs, inputs,
               COALESCE(p.status, 'Active') AS status, archived
        FROM projects p
        WHERE COALESCE(archived, FALSE) = FALSE
          AND COALESCE(p.status, 'Active') = 'Active'
        ORDER BY updated_at DESC
    """)
    pipeline_rows = cur.fetchall() or []
    pipeline = []

    # Build a name→LP-equity map from the returns blob so MPC pipeline
    # rows can show forecasted equity (sum of |LP contributions|, in $K)
    # alongside land price. Manual rows compute their own from the
    # contributions_yearly array. Falls back to land_cost_k when no
    # match is found in the returns blob.
    equity_by_name = {}
    for _rp in raw_projects:
        _name = (_rp.get("name") or "").strip()
        if not _name:
            continue
        _by_label = {m.get("label"): m for m in (_rp.get("metrics") or [])}
        _v = (_by_label.get("Total LP Contributions") or {}).get("total")
        try: equity_by_name[_name] = int(round(abs(float(_v or 0))))  # already $K in returns blob
        except (TypeError, ValueError): pass

    for r in pipeline_rows:
        out = r.get("outputs") or {}
        inp = r.get("inputs")  or {}
        # Best-effort field mapping. `total_land_cost` is calc.py output;
        # if missing, fall back to gross-revenue × an estimate so the row
        # at least renders.
        land_cost = (out.get("total_land_cost")
                     or inp.get("raw_land_cost")
                     or 0)
        try: land_cost_k = int(round(float(land_cost) / 1000))
        except (TypeError, ValueError): land_cost_k = 0
        try: irr_pct = float(out.get("unlevered_irr") or 0) * 100
        except (TypeError, ValueError): irr_pct = 0.0
        try: gm = float(out.get("gross_margin_pct") or 0)  # decimal
        except (TypeError, ValueError): gm = 0.0
        try: dur_months = int(round(float(out.get("project_length_years") or 0) * 12))
        except (TypeError, ValueError): dur_months = 0
        # No EM in our outputs schema — rough proxy: 1 + gross_margin.
        em_val = round(1 + gm, 2) if gm else 1.0
        pid = f"proj_{r['id']}"
        proj_name = r.get("name") or f"Project {r['id']}"
        # Forecasted equity = LP capital commitment forecast for this
        # project. Pulled from the returns blob when we have a match by
        # name; otherwise fall back to land_cost_k so the column never
        # shows blank.
        forecasted_equity_k = equity_by_name.get(proj_name, land_cost_k)
        pipeline.append({
            "id":           pid,
            "name":         proj_name,
            "address":      r.get("address") or "",
            "asset_class":  _class_for(proj_name),
            "land_price":   land_cost_k,
            "forecasted_equity": forecasted_equity_k,
            "duration":     dur_months,
            "gross_margin": gm,
            "irr":          round(irr_pct, 1),
            "em":           em_val,
            "updated":      (r["updated_at"].strftime("%Y-%m-%d")
                             if r.get("updated_at") else ""),
            "show_in_report": pid not in hidden_pipeline,
        })
    cur.close(); conn.close()

    # Manual pipeline entries — speculative deals the user typed in directly.
    # These coexist with the live MPC underwriting rows in the same Pipeline
    # table; we tag them with a `source: 'manual'` flag so the UI can show
    # an Edit/Delete affordance and skip the asset-class chip dropdown
    # (it's edited via the modal instead).
    # Reuse the manual pipeline blob already loaded for the recycle chart.
    for mp in manual_for_recycle:
        contribs = mp.get("contributions_yearly") or []
        distribs = mp.get("distributions_yearly") or []
        # Total upfront contribution = |sum(contributions)| in $K
        contrib_total = sum(abs(c) for c in contribs)
        land_k = int(round(contrib_total / 1000))
        # Duration = gap from FIRST contribution year to LAST distribution
        # year, in months. We use the year DIFFERENCE (not inclusive
        # count): contribute in 2026 + last distribution in 2029 = 3y
        # life, not 4y. Trailing-only contributions don't extend the
        # span; a project's life ends when capital stops coming back. If
        # there are no distributions yet, duration is 0 (capital still
        # at work, no full life span to report). Same-year contribute +
        # distribute floors at 12mo so we don't render "0mo".
        first_contrib_idx = next(
            (i for i, c in enumerate(contribs) if abs(c) > 0),
            None,
        )
        last_dist_idx = next(
            (i for i in range(len(distribs) - 1, -1, -1) if abs(distribs[i]) > 0),
            None,
        )
        if first_contrib_idx is None or last_dist_idx is None or last_dist_idx < first_contrib_idx:
            dur_months = 0
        else:
            dur_months = max(12, (last_dist_idx - first_contrib_idx) * 12)
        mpid = f"manual_{mp.get('id', '')}"
        pipeline.append({
            "id":           mpid,
            "name":         mp.get("name", ""),
            "address":      mp.get("location", ""),
            "asset_class":  _class_for(mp.get("name") or "") or mp.get("asset_class", "mpc-hub"),
            "land_price":   land_k,
            # Manual rows store LP contributions directly via the modal
            # / Excel import — that IS the forecasted equity.
            "forecasted_equity": land_k,
            "duration":     dur_months,
            "gross_margin": 0.0,
            "irr":          float(mp.get("irr") or 0),
            "em":           float(mp.get("em")  or 0),
            "updated":      (mp.get("updated_at") or "")[:10],
            "source":       "manual",
            "manual_id":    mp.get("id", ""),
            "show_in_report": mpid not in hidden_pipeline,
            # Pre-populate the year inputs in the edit modal
            "_form": {
                "name":                 mp.get("name", ""),
                "location":             mp.get("location", ""),
                "asset_class":          mp.get("asset_class", "mpc-hub"),
                "irr":                  mp.get("irr") or 0,
                "em":                   mp.get("em")  or 0,
                "contributions_yearly": contribs,
                "distributions_yearly": distribs,
            },
        })

    pipeline_land_total = sum(p["land_price"] for p in pipeline)
    weighted_irr = (sum(p["irr"] * p["land_price"] for p in pipeline) / pipeline_land_total
                    if pipeline_land_total else 0.0)

    # ── Distributed YTD + To Be Distributed (next 18mo) ─────────────────
    # KPI is current-year YTD only — Jan 1 of the current year through
    # today. Computed exactly from the monthly distribution arrays when
    # the workbook ships them, year-pro-rated otherwise. The Active
    # Projects table's per-project To Date column is a SEPARATE
    # life-to-date sum (see portfolio_distributed_ltd above) — keep
    # those two figures distinct, they answer different questions.
    idx_cur = years_int.index(current_year) if current_year in years_int else -1
    distributed_ytd_k = int(round(portfolio_distributed_ytd_current))

    # 18 months — sum of LP + Promote distributions scheduled for the
    # 18 months strictly AFTER today, computed at month granularity
    # when the workbook ships a Monthly Cashflows block. No more
    # year-pro-rate guessing; "next 18 months" is exactly that.
    if src_months and portfolio_to_be_distributed_18mo > 0:
        to_be_distributed_k = int(round(portfolio_to_be_distributed_18mo))
    else:
        # Legacy fallback: year-pro-rate when the workbook predates
        # the monthly block. Remaining-of-current-year + full next
        # years × 12mo until the 18-month budget is exhausted.
        to_be_distributed_k = 0
        if idx_cur >= 0:
            remaining = 18.0
            for j in range(idx_cur, len(years_int)):
                if remaining <= 0:
                    break
                ld = portfolio_lp_dist_yearly[j]
                pd = portfolio_promote_yearly[j]
                month_window = months_left_this_year if j == idx_cur else 12
                use = min(remaining, month_window)
                to_be_distributed_k += int(round((ld + pd) * (use / 12.0)))
                remaining -= use

    quarter = (today.month - 1) // 3 + 1
    pipeline_year_range = list(range(_CAP_MANUAL_PIPELINE_YEAR_START,
                                     _CAP_MANUAL_PIPELINE_YEAR_END + 1))
    return {
        "as_of":         f"Q{quarter} {current_year}",
        "asset_classes": _CAP_ASSET_CLASSES,
        "active":        active,
        "pipeline":      pipeline,
        "pipeline_year_range": pipeline_year_range,
        "returns": {
            "years":        years_str,
            "current_year": cur_year_str,
            "by_project":   returns_by_project,
        },
        "recycle": {
            "years":        recycle_years,
            "rows":         recycle_rows,
            "y_axis_max":   y_axis_max,        # $K ceiling, auto-scaled
            "current_year": current_year,
            # Palette (active series solid, pipeline series dashed). The
            # SVG inlines these so it renders correctly inside print PDFs
            # where CSS variables don't always resolve.
            "color_contrib":      "#F25929",   # active contributions (orange)
            # LP distributions are split into two stack segments —
            # realized (past) renders in the deeper teal, forecasted
            # (future) in the lighter sage. Same hue family so they read
            # as one logical series with a "this is what's been paid"
            # demarcation visible at a glance.
            "color_dist_lp":           "#5E9E8C",   # legacy alias = forecast color
            "color_dist_lp_realized":  "#1F7A4D",   # realized — deep "good" green
            "color_dist_lp_forecast":  "#5E9E8C",   # forecast — established teal
            "color_dist_promote":      "#C8A96E",   # active promote (gold)
            "color_pipeline_contrib":  "#F25929",   # dashed orange
            "color_pipeline_dist":     "#5E9E8C",   # legacy — no longer rendered
        },
        "commitments": {"groups": commit_groups, "totals": commit_totals},
        "investors": _build_investor_view(raw_projects, years_str, years_int, src_months),
        "kpis": {
            "active_count":            len(active),
            "total_equity":            int(round(total_equity)),
            "total_committed":         total_committed_k,
            "forecasted_lp_profit":    int(round(total_lp_profit)),
            "forecasted_lp_irr":       forecasted_lp_irr,
            "forecasted_promote":      int(round(total_promote)),
            "distributed_ytd":         distributed_ytd_k,
            "to_be_distributed":       to_be_distributed_k,
            "pipeline_count":          len(pipeline),
            "pipeline_land":           pipeline_land_total,
            "weighted_irr":            round(weighted_irr, 1),
            "unallocated_commitments": unallocated_commitments_k,
            "ytd_label":               ytd_label,
        },
    }


@app.route("/portfolio")
@app.route("/ember-capital")
@app.route("/capital")
@login_required
def portfolio_page():
    # Pull the *current* permissions from the DB (not the stale login-time
    # session copy) so newly-granted, default-False perms like captable_edit
    # take effect without forcing a log-out/in.
    pa = _refresh_page_access_from_db() or {
        "mpc_underwriting": True, "returns": True, "loans": True,
        "operations": True, "portfolio": True, "reports": True,
    }
    # Mirror the redirect-on-no-access pattern used by every other
    # dashboard (/returns, /loans, /operations, /macro, /sales,
    # /verticals). Sidebar already hides the link when access is
    # revoked, but URL access was open until this guard landed.
    if not session.get("is_admin") and not pa.get("portfolio", True):
        return redirect(url_for("home"))
    pa.setdefault("portfolio", True)
    pa.setdefault("reports",   True)

    capital = _build_capital_view_context()
    # Capital reads from the same returns blob as /returns, so its
    # "Data From" / "Uploaded" should match. Pull both from the latest
    # reports[returns] row for the dashboard footer.
    conn = get_db(); cur = conn.cursor()
    cur.execute(
        "SELECT data->>'data_from' AS data_from, uploaded_at "
        "FROM reports WHERE report_type = 'returns' "
        "ORDER BY uploaded_at DESC LIMIT 1"
    )
    rrow = cur.fetchone(); cur.close(); conn.close()
    data_from   = _fmt_data_date(rrow["data_from"]   if rrow else None)
    uploaded_at = _fmt_data_date(rrow["uploaded_at"] if rrow else None)
    return render_template(
        "capital.html",
        username=session.get("username"),
        is_admin=session.get("is_admin", False),
        page_access=pa,
        capital=capital,
        data_from=data_from, uploaded_at=uploaded_at,
    )


def _capital_report_context() -> dict:
    """Build the dict consumed by capital_report.html. Reuses the live
    /capital page context and layers on a few report-specific keys
    (period label, generated date), then filters the pipeline list to
    just rows the user has flagged visible in the report and re-rolls
    the pipeline KPIs against that subset.
    """
    now = datetime.datetime.now()
    capital = _build_capital_view_context()
    capital["report_period_label"] = now.strftime("%B %Y").upper()
    capital["report_period_short"] = now.strftime("%b %Y").upper()
    capital["generated_date"]      = now.strftime("%Y-%m-%d")

    # Pipeline visibility filter — drop hidden rows from the report,
    # then re-derive the pipeline-side KPIs (count, land total,
    # weighted IRR) so the cockpit numbers match the displayed table.
    visible = [p for p in capital.get("pipeline", []) if p.get("show_in_report", True)]
    capital["pipeline"] = visible
    land_tot = sum(p["land_price"] for p in visible)
    capital["kpis"]["pipeline_count"] = len(visible)
    capital["kpis"]["pipeline_land"]  = land_tot
    capital["kpis"]["weighted_irr"]   = round(
        sum(p["irr"] * p["land_price"] for p in visible) / land_tot, 1
    ) if land_tot else 0.0
    return capital


@app.route("/api/ember-capital/report", methods=["GET"])
@app.route("/capital/report", methods=["GET"])
@login_required
def ember_capital_report_html():
    """HTML preview of the 3-page Capital Report. Useful for debugging
    layout without round-tripping through WeasyPrint. Use /report.pdf
    for the actual download."""
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("reports", True):
        return jsonify({"error": "Reports access disabled by admin"}), 403
    capital = _capital_report_context()
    return render_template("capital_report.html", capital=capital)


@app.route("/api/ember-capital/report.pdf", methods=["GET"])
@app.route("/capital/report.pdf", methods=["GET"])
@login_required
def ember_capital_report_pdf():
    """Stream the 3-page Capital Report as a real PDF (WeasyPrint).

    Falls back to the legacy fpdf2 2-page report if WeasyPrint can't
    load (missing libpango/cairo, ImportError, etc.) so the button
    never returns nothing.
    """
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("reports", True):
        return jsonify({"error": "Reports access disabled by admin"}), 403

    capital = _capital_report_context()
    html = render_template("capital_report.html", capital=capital)

    try:
        from weasyprint import HTML
        pdf_bytes = HTML(
            string=html,
            base_url=request.host_url,
            url_fetcher=_weasyprint_local_fetcher,
        ).write_pdf()
    except (ImportError, OSError) as e:
        app.logger.warning(
            "WeasyPrint unavailable for capital report (%s: %s); falling back to fpdf2",
            type(e).__name__, e,
        )
        # Last-resort fallback — the older 2-page fpdf2 report
        payload = _build_ember_capital_payload()
        pdf_bytes = bytes(_gen_pdf_ember_capital(payload))

    as_attachment = request.args.get("download") in ("1", "true", "yes")
    fname = f"Ember_Capital_Report_{datetime.datetime.now().strftime('%Y-%m')}.pdf"
    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=as_attachment,
        download_name=fname,
    )


@app.route("/api/projects/<project_id>/asset-class", methods=["PATCH"])
@login_required
def update_project_asset_class(project_id):
    """Save the asset class assignment for a project (active or pipeline).

    `project_id` is the slug used in the template — either a returns slug
    (`grand-prairie-east-ccc`) or a pipeline `proj_<int>`. Both resolve back
    to the project's display name, which is the storage key.
    """
    body = request.get_json(silent=True) or {}
    new_cls = body.get("asset_class")
    if new_cls not in _CAP_VALID_CLASSES:
        return jsonify({"error": "invalid asset_class"}), 400

    # Resolve project_id → project name. Pipeline ids are `proj_<int>`;
    # returns slugs we look up in the latest returns blob.
    project_name = None
    if project_id.startswith("proj_"):
        try:
            pid = int(project_id.split("_", 1)[1])
            conn = get_db(); cur = conn.cursor()
            cur.execute("SELECT name FROM projects WHERE id = %s", (pid,))
            row = cur.fetchone()
            cur.close(); conn.close()
            if row:
                project_name = row.get("name")
        except (ValueError, IndexError):
            pass
    else:
        # Returns slug — find the matching name from the returns blob
        conn = get_db(); cur = conn.cursor()
        cur.execute(
            "SELECT data FROM reports WHERE report_type = 'returns' "
            "ORDER BY uploaded_at DESC LIMIT 1"
        )
        rrow = cur.fetchone()
        cur.close(); conn.close()
        if rrow and rrow.get("data"):
            for p in (rrow["data"].get("projects") or []):
                if _capital_slug(p.get("name", "")) == project_id:
                    project_name = p.get("name")
                    break

    if not project_name:
        return jsonify({"error": "project not found"}), 404

    _capital_save_asset_class(project_name, new_cls)
    return jsonify({"id": project_id, "name": project_name, "asset_class": new_cls})


# ─────────────────────────────────────────────────────────────────────────────
# Manual pipeline projects — for forecast/placeholder deals that aren't yet
# in MPC Underwriting. Stored as a single reports[ember_capital_pipeline_manual]
# blob with shape {"projects": [{id, name, location, asset_class, irr, em,
# years_start, contributions_yearly, distributions_yearly, created_at,
# updated_at}]}. Dollar amounts are raw $ (not $K) — the data shaper rolls
# them into $K before they hit the Pipeline table to match MPC rows.
# ─────────────────────────────────────────────────────────────────────────────
_CAP_MANUAL_PIPELINE_YEAR_START = 2024
_CAP_MANUAL_PIPELINE_YEAR_END   = 2045  # inclusive — matches the years axis used by /api/ember-capital
_CAP_MANUAL_PIPELINE_N_YEARS    = _CAP_MANUAL_PIPELINE_YEAR_END - _CAP_MANUAL_PIPELINE_YEAR_START + 1


def _capital_load_manual_pipeline() -> dict:
    conn = get_db(); cur = conn.cursor()
    cur.execute(
        "SELECT data FROM reports WHERE report_type = 'ember_capital_pipeline_manual' "
        "ORDER BY uploaded_at DESC LIMIT 1"
    )
    row = cur.fetchone(); cur.close(); conn.close()
    d = (row["data"] if row else None) or {}
    return d if isinstance(d, dict) and "projects" in d else {"projects": []}


def _capital_save_manual_pipeline(data: dict) -> None:
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM reports WHERE report_type = 'ember_capital_pipeline_manual'")
    cur.execute(
        "INSERT INTO reports (report_type, data, uploaded_by) VALUES (%s, %s, %s)",
        ("ember_capital_pipeline_manual", json.dumps(data), session.get("user_id")),
    )
    conn.commit(); cur.close(); conn.close()


def _clean_manual_pipeline_project(body: dict) -> dict | None:
    """Validate + coerce a manual-pipeline payload. Returns None on bad input.

    Year arrays are padded/truncated to N_YEARS so the data shape stays
    consistent regardless of which years the form posted.
    """
    if not isinstance(body, dict):
        return None
    name = (body.get("name") or "").strip()
    if not name:
        return None

    asset_class = body.get("asset_class") or "mpc-hub"
    if asset_class not in _CAP_VALID_CLASSES:
        asset_class = "mpc-hub"

    def _f(v):
        try: return float(v or 0)
        except (TypeError, ValueError): return 0.0

    def _arr(v):
        out = []
        for x in (v or []):
            try: out.append(float(x or 0))
            except (TypeError, ValueError): out.append(0.0)
        # pad / truncate to N_YEARS
        if len(out) < _CAP_MANUAL_PIPELINE_N_YEARS:
            out += [0.0] * (_CAP_MANUAL_PIPELINE_N_YEARS - len(out))
        return out[:_CAP_MANUAL_PIPELINE_N_YEARS]

    return {
        "name":                 name,
        "location":             (body.get("location") or "").strip(),
        "asset_class":          asset_class,
        "irr":                  round(_f(body.get("irr")), 1),
        "em":                   round(_f(body.get("em")),  2),
        "years_start":          _CAP_MANUAL_PIPELINE_YEAR_START,
        "contributions_yearly": _arr(body.get("contributions_yearly")),
        "distributions_yearly": _arr(body.get("distributions_yearly")),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Pipeline visibility — per-project toggle for the Capital Report. Hides a
# row from the report's pipeline table (and rolled-up KPIs) without removing
# it from the live /capital page. Storage is a single JSONB blob with a list
# of hidden project ids; default is "show". Project ids are the same shape
# the pipeline table renders: `proj_<int>` for MPC projects, `manual_<id>`
# for manually-added ones.
# ─────────────────────────────────────────────────────────────────────────────

def _capital_load_pipeline_visibility() -> set[str]:
    conn = get_db(); cur = conn.cursor()
    cur.execute(
        "SELECT data FROM reports WHERE report_type = 'ember_capital_pipeline_visibility' "
        "ORDER BY uploaded_at DESC LIMIT 1"
    )
    row = cur.fetchone(); cur.close(); conn.close()
    d = (row["data"] if row else None) or {}
    return set((d or {}).get("hidden") or [])


def _capital_save_pipeline_visibility(hidden: set[str]) -> None:
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM reports WHERE report_type = 'ember_capital_pipeline_visibility'")
    cur.execute(
        "INSERT INTO reports (report_type, data, uploaded_by) VALUES (%s, %s, %s)",
        ("ember_capital_pipeline_visibility",
         json.dumps({"hidden": sorted(hidden)}),
         session.get("user_id")),
    )
    conn.commit(); cur.close(); conn.close()


@app.route("/api/ember-capital/pipeline-visibility/<project_id>", methods=["POST"])
@login_required
def ember_capital_pipeline_visibility(project_id):
    """Toggle whether a pipeline project shows in the Capital Report.

    Body: {"show": true|false}. Returns the updated hidden-id list so
    the client can verify state.
    """
    body = request.get_json(silent=True) or {}
    show = bool(body.get("show", True))
    hidden = _capital_load_pipeline_visibility()
    if show:
        hidden.discard(project_id)
    else:
        hidden.add(project_id)
    _capital_save_pipeline_visibility(hidden)
    return jsonify({"id": project_id, "show": show, "hidden": sorted(hidden)})


@app.route("/api/ember-capital/pipeline", methods=["GET", "POST"])
@login_required
def ember_capital_manual_pipeline():
    """List or create manual pipeline projects."""
    if request.method == "GET":
        return jsonify(_capital_load_manual_pipeline())

    body = request.get_json(silent=True) or {}
    proj = _clean_manual_pipeline_project(body)
    if not proj:
        return jsonify({"error": "invalid project"}), 400

    data = _capital_load_manual_pipeline()
    projects = list(data.get("projects") or [])
    now = datetime.datetime.now().isoformat(timespec="seconds")
    proj["id"]         = f"{_capital_slug(proj['name'])}-{int(datetime.datetime.now().timestamp())}"
    proj["created_at"] = now
    proj["updated_at"] = now
    projects.append(proj)
    _capital_save_manual_pipeline({"projects": projects})

    # Layer in the asset class assignment so the donut + chip stay in sync
    # with the Active/Returns side of the page.
    _capital_save_asset_class(proj["name"], proj["asset_class"])

    return jsonify({"ok": True, "project": proj})


@app.route("/api/ember-capital/pipeline/<pid>", methods=["PUT", "DELETE"])
@login_required
def ember_capital_manual_pipeline_one(pid):
    data = _capital_load_manual_pipeline()
    projects = list(data.get("projects") or [])

    if request.method == "DELETE":
        new_list = [p for p in projects if p.get("id") != pid]
        if len(new_list) == len(projects):
            return jsonify({"error": "not found"}), 404
        _capital_save_manual_pipeline({"projects": new_list})
        return jsonify({"ok": True})

    body = request.get_json(silent=True) or {}
    cleaned = _clean_manual_pipeline_project(body)
    if not cleaned:
        return jsonify({"error": "invalid project"}), 400

    found = False
    for i, p in enumerate(projects):
        if p.get("id") == pid:
            cleaned["id"]         = pid
            cleaned["created_at"] = p.get("created_at") or datetime.datetime.now().isoformat(timespec="seconds")
            cleaned["updated_at"] = datetime.datetime.now().isoformat(timespec="seconds")
            projects[i] = cleaned
            found = True
            break
    if not found:
        return jsonify({"error": "not found"}), 404
    _capital_save_manual_pipeline({"projects": projects})
    _capital_save_asset_class(cleaned["name"], cleaned["asset_class"])
    return jsonify({"ok": True, "project": cleaned})


# ─────────────────────────────────────────────────────────────────────────────
# Pipeline Excel import — parse a project's underwriting "Returns" tab and
# pull out the LP-level summary (IRR, equity multiple, yearly contributions
# and distributions) so it can prefill the Add Pipeline modal or replace the
# numeric fields of an existing pipeline row.
#
# Field map (per the underwriting workbook layout — see _design_reference for
# Mllican Returns Tab.xlsx):
#   D26  Total LP Contributions ($)         used as a sanity-check total
#   D25  Total LP Distributions ($)         used as a sanity-check total
#   D28  LP IRR (decimal, e.g. 0.1362)
#   D29  LP Equity Multiple (e.g. 2.55)
#   F19→ month index 0..N
#   F20→ year per column (2026, 2027, ...)
#   F25→ monthly LP distributions
#   F26→ monthly LP contributions
# ─────────────────────────────────────────────────────────────────────────────
def _parse_returns_excel_for_pipeline(file_bytes: bytes) -> dict:
    """Read an underwriting workbook's Returns tab and return a dict shaped
    for the pipeline form / API:

        {"irr": 13.62, "em": 2.55,
         "contributions_yearly": [...N_YEARS floats...],
         "distributions_yearly": [...N_YEARS floats...],
         "totals": {"contrib": 33618766, "dist": 85663163},
         "year_range": [2026, 2041]}

    On failure returns {"error": "..."}.
    """
    import openpyxl as _opx
    import io as _io
    from collections import defaultdict as _dd

    try:
        wb = _opx.load_workbook(_io.BytesIO(file_bytes), data_only=True, read_only=False)
    except Exception as e:
        return {"error": f"Could not open workbook: {type(e).__name__}: {e}"}

    # Find the right sheet — the one whose D19 cell contains "Analysis Period".
    # The underwriting workbook sometimes ships with extra tabs, so don't
    # blindly take the first one.
    ws = None
    for name in wb.sheetnames:
        try:
            label = wb[name].cell(row=19, column=4).value
            if isinstance(label, str) and "analysis period" in label.lower():
                ws = wb[name]
                break
        except Exception:
            continue
    if ws is None:
        # Fall back to a sheet named "Returns" if it exists, else first sheet.
        for name in wb.sheetnames:
            if name.strip().lower() == "returns":
                ws = wb[name]
                break
        if ws is None:
            ws = wb[wb.sheetnames[0]]

    def _f(v):
        if v is None: return 0.0
        try: return float(v)
        except (TypeError, ValueError): return 0.0

    # ── Headline metrics ──
    total_contrib = _f(ws.cell(row=26, column=4).value)
    total_dist    = _f(ws.cell(row=25, column=4).value)
    irr_raw       = ws.cell(row=28, column=4).value     # decimal
    em_raw        = ws.cell(row=29, column=4).value

    try:
        irr_pct = round(float(irr_raw) * 100, 2) if irr_raw not in (None, "", "N/A") else 0.0
    except (TypeError, ValueError):
        irr_pct = 0.0
    try:
        em = round(float(em_raw), 2) if em_raw not in (None, "", "N/A") else 0.0
    except (TypeError, ValueError):
        em = 0.0

    # ── Yearly buckets ──
    by_year_contrib: dict = _dd(float)
    by_year_dist:    dict = _dd(float)
    seen_years = set()

    # Walk columns F (=6) onward until we run out of data on row 20 (year row).
    last_col = ws.max_column or 6
    for c in range(6, last_col + 1):
        yv = ws.cell(row=20, column=c).value
        if yv is None:
            # Stop at the first fully-empty year cell so we don't pick up
            # trailing Excel calc tail columns. But check if there's any
            # cashflow value here too — if not, break.
            if (ws.cell(row=25, column=c).value in (None, 0) and
                ws.cell(row=26, column=c).value in (None, 0)):
                break
            continue
        try:
            yi = int(yv)
        except (TypeError, ValueError):
            continue
        seen_years.add(yi)
        by_year_contrib[yi] += _f(ws.cell(row=26, column=c).value)
        by_year_dist[yi]    += _f(ws.cell(row=25, column=c).value)

    if not seen_years:
        return {"error": "No yearly cashflow rows found — verify this is a Returns tab."}

    # Build N_YEARS-long arrays anchored at the pipeline window.
    contributions = [0.0] * _CAP_MANUAL_PIPELINE_N_YEARS
    distributions = [0.0] * _CAP_MANUAL_PIPELINE_N_YEARS
    for y, v in by_year_contrib.items():
        idx = y - _CAP_MANUAL_PIPELINE_YEAR_START
        if 0 <= idx < _CAP_MANUAL_PIPELINE_N_YEARS:
            contributions[idx] = round(v, 2)
    for y, v in by_year_dist.items():
        idx = y - _CAP_MANUAL_PIPELINE_YEAR_START
        if 0 <= idx < _CAP_MANUAL_PIPELINE_N_YEARS:
            distributions[idx] = round(v, 2)

    out_of_range = sorted([y for y in seen_years if not (
        _CAP_MANUAL_PIPELINE_YEAR_START <= y <= _CAP_MANUAL_PIPELINE_YEAR_END
    )])

    return {
        "irr": irr_pct,
        "em":  em,
        "contributions_yearly": contributions,
        "distributions_yearly": distributions,
        "totals": {"contrib": round(total_contrib, 2), "dist": round(total_dist, 2)},
        "year_range": [min(seen_years), max(seen_years)] if seen_years else [None, None],
        "out_of_range_years": out_of_range,
    }


@app.route("/api/ember-capital/pipeline/parse-excel", methods=["POST"])
@login_required
def ember_capital_pipeline_parse_excel():
    """Parse an uploaded underwriting workbook's Returns tab and return the
    extracted LP fields for the pipeline form. Does NOT save anything — the
    caller (Add modal or per-row "Update from Excel") decides whether/how
    to persist via POST/PUT to /api/ember-capital/pipeline.
    """
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file uploaded"}), 400
    try:
        data = file.read()
    except Exception as e:
        return jsonify({"error": f"Could not read upload: {e}"}), 400
    parsed = _parse_returns_excel_for_pipeline(data)
    if parsed.get("error"):
        return jsonify(parsed), 400
    return jsonify(parsed)


@app.route("/api/ember-capital/excel", methods=["GET"])
@login_required
def ember_capital_excel():
    """Excel export — same workbook the existing PDF endpoint draws from."""
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("reports", True):
        return jsonify({"error": "Reports access disabled by admin"}), 403
    payload = _build_ember_capital_payload()
    xlsx_bytes = _gen_excel_ember_capital(payload)
    return send_file(
        io.BytesIO(xlsx_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"Ember_Capital_{datetime.datetime.now().strftime('%Y-%m')}.xlsx",
    )


@app.route("/api/ember-capital", methods=["GET"])
@login_required
def ember_capital_data():
    """
    Return the shape the Ember Capital dashboard needs:
      {
        "years":        [2024, 2025, ..., 2045],
        "uploaded_at":  "YYYY-MM-DD" or None,
        "projects": [
          {
            "name": "...",
            "lp_irr":       0.22,     # decimal
            "lp_em":        1.84,     # multiplier
            "lp_profit":    12345.67, # total ($K)
            "lp_distributions_total":  ...,
            "lp_contributions_total":  ...,    # negative
            "promote_total":           ...,
            "lp_distributions_yearly": [...],  # per year ($K)
            "lp_contributions_yearly": [...],  # per year ($K), negative
            "lp_profit_yearly":        [...],  # per year ($K)
            "promote_yearly":          [...],  # per year ($K)
          },
          ...
        ]
      }
    Values are pulled from the latest 'returns' report uploaded via the
    Ember Dashboard Excel.
    """
    if not session.get("is_admin"):
        return jsonify({"error": "Access denied"}), 403
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "SELECT data, uploaded_at FROM reports "
        "WHERE report_type = 'returns' ORDER BY uploaded_at DESC LIMIT 1"
    )
    row = cur.fetchone()
    cur.close(); conn.close()

    if not row or not row["data"]:
        return jsonify({"years": [], "projects": [], "uploaded_at": None})

    src = row["data"] or {}
    years = src.get("years", []) or []
    uploaded_at = row["uploaded_at"].isoformat() if row["uploaded_at"] else None

    projects = []
    for p in src.get("projects", []) or []:
        by_label = {m.get("label"): m for m in (p.get("metrics") or [])}

        def _m(label):
            return by_label.get(label) or {}

        def _yearly(label):
            m = _m(label)
            y = m.get("yearly") or []
            # pad/truncate to match years length
            n = len(years)
            if len(y) < n:
                y = list(y) + [0] * (n - len(y))
            return y[:n]

        def _total(label):
            m = _m(label)
            t = m.get("total")
            try:
                return float(t) if t is not None else 0.0
            except (TypeError, ValueError):
                return 0.0

        projects.append({
            "name": p.get("name", ""),
            "lp_irr":                  _total("LP IRR"),
            "lp_em":                   _total("LP Equity Multiple"),
            "lp_profit":               _total("Total LP Profit"),
            "lp_distributions_total":  _total("Total LP Distributions"),
            "lp_contributions_total":  _total("Total LP Contributions"),
            "promote_total":           _total("Promote"),
            "lp_distributions_yearly": _yearly("Total LP Distributions"),
            "lp_contributions_yearly": _yearly("Total LP Contributions"),
            "lp_profit_yearly":        _yearly("Total LP Profit"),
            "promote_yearly":          _yearly("Promote"),
        })

    # Pull saved recycle assumptions + commitments so the UI can hydrate from DB
    cur2 = get_db().cursor()
    cur2.execute(
        "SELECT data FROM reports WHERE report_type = 'ember_capital_settings' "
        "ORDER BY uploaded_at DESC LIMIT 1"
    )
    srow = cur2.fetchone()
    settings = (srow["data"] or {}) if srow else {}

    cur2.execute(
        "SELECT data FROM reports WHERE report_type = 'ember_capital_commitments' "
        "ORDER BY uploaded_at DESC LIMIT 1"
    )
    crow = cur2.fetchone()
    commitments = (crow["data"] or {}) if crow else {"groups": []}
    cur2.close()

    return jsonify({
        "years":       years,
        "uploaded_at": uploaded_at,
        "projects":    projects,
        "settings":    settings,       # {recycle: {ProjName: {lp, prom}}}
        "commitments": commitments,    # {groups: [{name, mpc, vertical}]}
    })


@app.route("/api/ember-capital/settings", methods=["POST"])
@login_required
def ember_capital_save_settings():
    """Save the recycle-% assumptions. Admin-only; shared across users."""
    if not session.get("is_admin"):
        return jsonify({"error": "Access denied"}), 403
    body = request.get_json(silent=True) or {}
    recycle = body.get("recycle") or {}
    # Validate shape: {name: {lp: 0..100, prom: 0..100}}
    clean = {}
    for name, v in recycle.items():
        if not isinstance(v, dict):
            continue
        try:
            lp  = max(0, min(100, float(v.get("lp",  0))))
            pr  = max(0, min(100, float(v.get("prom", 0))))
        except (TypeError, ValueError):
            continue
        clean[str(name)] = {"lp": lp, "prom": pr}

    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM reports WHERE report_type = 'ember_capital_settings'")
    cur.execute(
        "INSERT INTO reports (report_type, data, uploaded_by) VALUES (%s, %s, %s)",
        ("ember_capital_settings", json.dumps({"recycle": clean}), session["user_id"])
    )
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})


@app.route("/api/ember-capital/commitments", methods=["GET", "POST"])
@login_required
def ember_capital_commitments():
    """GET or replace the commitments object.

    Group shape:
        {name, mpc, vertical, mpc_allocated, vertical_allocated}

        - `mpc` / `vertical` = total dollars committed by the group to
          each asset class.
        - `mpc_allocated` / `vertical_allocated` = dollars Ember has
          already drawn against those commitments. Difference is what
          remains available for new deals.

    Older saved rows without the *_allocated keys read back as 0.
    """
    if request.method == "GET":
        conn = get_db(); cur = conn.cursor()
        cur.execute(
            "SELECT data, uploaded_at FROM reports "
            "WHERE report_type = 'ember_capital_commitments' "
            "ORDER BY uploaded_at DESC LIMIT 1"
        )
        row = cur.fetchone()
        cur.close(); conn.close()
        if not row:
            return jsonify({"groups": [], "uploaded_at": None})
        d = row["data"] or {"groups": []}
        d["uploaded_at"] = row["uploaded_at"].isoformat() if row["uploaded_at"] else None
        return jsonify(d)

    # POST — admin-only
    if not session.get("is_admin"):
        return jsonify({"error": "Access denied"}), 403
    body = request.get_json(silent=True) or {}
    groups_in = body.get("groups") or []

    def _f(v):
        try: return float(v or 0)
        except (TypeError, ValueError): return 0.0

    clean_groups = []
    for g in groups_in:
        if not isinstance(g, dict):
            continue
        name = str(g.get("name", "")).strip()
        if not name:
            continue
        clean_groups.append({
            "name":               name,
            "mpc":                _f(g.get("mpc")),
            "vertical":           _f(g.get("vertical")),
            "mpc_allocated":      _f(g.get("mpc_allocated")),
            "vertical_allocated": _f(g.get("vertical_allocated")),
        })

    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM reports WHERE report_type = 'ember_capital_commitments'")
    cur.execute(
        "INSERT INTO reports (report_type, data, uploaded_by) VALUES (%s, %s, %s)",
        ("ember_capital_commitments", json.dumps({"groups": clean_groups}), session["user_id"])
    )
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True, "groups": clean_groups})


def _capital_returns_projects_years() -> tuple[list, list, list, list]:
    """(raw_projects, years_str, years_int, src_months) from the latest
    returns blob.

    Shared by the Investors API so it can recompute portfolios without
    rebuilding the entire /capital context. Returns empty lists when no
    returns upload exists."""
    conn = get_db(); cur = conn.cursor()
    cur.execute(
        "SELECT data FROM reports WHERE report_type = 'returns' "
        "ORDER BY uploaded_at DESC LIMIT 1"
    )
    row = cur.fetchone(); cur.close(); conn.close()
    src = (row["data"] if row else {}) or {}
    years_int = list(src.get("years", []) or [])
    years_str = [str(y) for y in years_int]
    raw_projects = src.get("projects", []) or []
    src_months = list(src.get("months", []) or [])
    return raw_projects, years_str, years_int, src_months


def _capital_save_captable(positions: list[dict]) -> None:
    """Replace the stored cap-table positions blob (latest-snapshot pattern)."""
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM reports WHERE report_type = 'ember_capital_captable'")
    cur.execute(
        "INSERT INTO reports (report_type, data, uploaded_by) VALUES (%s, %s, %s)",
        ("ember_capital_captable", json.dumps({"positions": positions}),
         session.get("user_id")),
    )
    conn.commit(); cur.close(); conn.close()


def _clean_captable_positions(rows: list) -> list[dict]:
    """Validate/normalize incoming position rows.

    Accepts ownership as a fraction (0..1) or a percentage (>1.5 -> /100).
    Drops rows missing a project or vehicle. equity_class is coerced to
    'preferred' or 'common'."""
    clean = []
    for r in rows or []:
        if not isinstance(r, dict):
            continue
        project = str(r.get("project", "")).strip()
        vehicle = str(r.get("vehicle", "")).strip()
        if not project or not vehicle:
            continue
        try:
            pct = float(r.get("pct") or 0)
        except (TypeError, ValueError):
            pct = 0.0
        if pct > 1.5:           # sent as a percentage (e.g. 20.78)
            pct = pct / 100.0
        try:
            contribution = float(r.get("contribution") or 0)
        except (TypeError, ValueError):
            contribution = 0.0
        equity_class = "preferred" if str(r.get("equity_class", "")).strip().lower() == "preferred" else "common"
        clean.append({
            "project":      project,
            "vehicle":      vehicle,
            "contribution": round(contribution, 2),
            "pct":          pct,
            "equity_class": equity_class,
        })
    return clean


@app.route("/api/ember-capital/captable", methods=["GET", "POST"])
@login_required
def ember_capital_captable():
    """Investor cap-table positions + derived portfolios.

    GET  — returns {positions, investors, can_edit, uploaded_at}. Viewing
           is gated by the page's `portfolio` access (same as the rest of
           Ember Capital); `can_edit` tells the UI whether to show edit /
           upload affordances.
    POST — replace the positions. Two content types:
             * multipart/form-data with `file` = the cap-table .xlsx
               (parsed server-side via captable_parser), or
             * application/json {positions:[...]} for manual grid edits.
           Both require captable_edit permission (admins always pass).
    """
    if request.method == "GET":
        pa = session.get("page_access") or {}
        if not session.get("is_admin") and not pa.get("portfolio", True):
            return jsonify({"error": "Access denied"}), 403
        blob = _capital_load_captable()
        raw_projects, years_str, years_int, src_months = _capital_returns_projects_years()
        view = _build_investor_view(raw_projects, years_str, years_int, src_months)
        return jsonify({
            "positions":   blob.get("positions", []),
            "uploaded_at": blob.get("uploaded_at"),
            "investors":   view.get("investors", []),
            "years":       view.get("years", []),
            "projects_matched":   view.get("projects_matched", []),
            "projects_unmatched": view.get("projects_unmatched", []),
            "can_edit":    _capital_can_edit_captable(),
        })

    # POST — permission-gated write
    if not _capital_can_edit_captable():
        return jsonify({"error": "Access denied"}), 403

    f = request.files.get("file")
    if f:
        try:
            from captable_parser import parse_captable_workbook
            positions = parse_captable_workbook(io.BytesIO(f.read()))
        except Exception as e:
            return jsonify({"error": f"Failed to parse cap-table file: {e}"}), 400
    else:
        body = request.get_json(silent=True) or {}
        positions = _clean_captable_positions(body.get("positions"))

    _capital_save_captable(positions)
    raw_projects, years_str, years_int, src_months = _capital_returns_projects_years()
    view = _build_investor_view(raw_projects, years_str, years_int, src_months)
    return jsonify({
        "ok":          True,
        "positions":   positions,
        "investors":   view.get("investors", []),
        "years":       view.get("years", []),
        "projects_matched":   view.get("projects_matched", []),
        "projects_unmatched": view.get("projects_unmatched", []),
    })


def _investor_export_view(selected_id: str | None = None) -> dict:
    """Build the investor view for export, optionally narrowed to one vehicle.
    Returns the view dict (with `investors` filtered when an id is given)."""
    raw_projects, years_str, years_int, src_months = _capital_returns_projects_years()
    view = _build_investor_view(raw_projects, years_str, years_int, src_months)
    if selected_id:
        view = dict(view)
        view["investors"] = [i for i in view.get("investors", [])
                             if i["id"] == selected_id]
    return view


def _investor_export_gate():
    """Same view gate as the investor portfolios table (portfolio access);
    admins always pass. Returns an error response tuple or None."""
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("portfolio", True):
        return jsonify({"error": "Access denied"}), 403
    return None


@app.route("/api/ember-capital/investors.xlsx", methods=["GET"])
@login_required
def ember_capital_investors_xlsx():
    """Excel export of investor positions + quarterly schedule. `?id=<vehicle>`
    exports a single investor (for handing to a visiting investor); omit it to
    export every investor, one sheet each."""
    gate = _investor_export_gate()
    if gate:
        return gate
    sel = request.args.get("id")
    view = _investor_export_view(sel)
    if not view.get("investors"):
        return jsonify({"error": "No investor positions to export"}), 404
    xlsx = _gen_excel_investor_positions(view)
    base = view["investors"][0]["name"] if sel else "All_Investors"
    fname = "Positions_%s_%s.xlsx" % (_capital_slug(base),
                                      datetime.datetime.now().strftime("%Y-%m-%d"))
    return send_file(
        io.BytesIO(xlsx),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=fname)


@app.route("/api/ember-capital/investors.pdf", methods=["GET"])
@login_required
def ember_capital_investors_pdf():
    """PDF position statement for one investor (`?id=`) or all investors."""
    gate = _investor_export_gate()
    if gate:
        return gate
    sel = request.args.get("id")
    view = _investor_export_view(sel)
    if not view.get("investors"):
        return jsonify({"error": "No investor positions to export"}), 404
    html = render_template("investor_positions.html", view=view,
                           generated=datetime.date.today().isoformat())
    try:
        from weasyprint import HTML
        pdf_bytes = HTML(string=html, base_url=request.host_url,
                         url_fetcher=_weasyprint_local_fetcher).write_pdf()
    except (ImportError, OSError) as e:
        app.logger.warning("WeasyPrint unavailable for investor positions "
                           "(%s: %s); returning HTML", type(e).__name__, e)
        return Response(html, mimetype="text/html")
    base = view["investors"][0]["name"] if sel else "All_Investors"
    fname = "Positions_%s_%s.pdf" % (_capital_slug(base),
                                     datetime.datetime.now().strftime("%Y-%m-%d"))
    return send_file(io.BytesIO(pdf_bytes), mimetype="application/pdf",
                     as_attachment=True, download_name=fname)


def _gen_excel_investor_positions(view: dict) -> bytes:
    """One worksheet per investor: summary KPIs + the per-project positions
    table with the quarterly forecast schedule and a total row."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    investors = view.get("investors", []) or []
    quarters  = view.get("quarters", []) or []
    uploaded  = (view.get("uploaded_at") or "")[:10] or datetime.date.today().isoformat()

    ACCENT = "C56028"; HDR_FILL = PatternFill("solid", fgColor="F2EFE8")
    TOT_FILL = PatternFill("solid", fgColor="FAF6EC")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    right = Alignment(horizontal="right")

    def font(bold=False, color="1A1A1A", size=10):
        return Font(name="Calibri", size=size, bold=bold, color=color)
    def money(cell, v):
        cell.value = v if isinstance(v, (int, float)) else None
        cell.number_format = '"$"#,##0'
    def pct(cell, v):
        if v is not None:
            cell.value = v; cell.number_format = '0.0%'
    def emx(cell, v):
        if v is not None:
            cell.value = v; cell.number_format = '0.00"x"'

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    used: set[str] = set()
    def sheet_name(nm):
        base = "".join(c for c in (nm or "Investor") if c not in '[]:*?/\\')[:31] or "Investor"
        s, i = base, 2
        while s in used:
            s = (base[:28] + "_%d" % i); i += 1
        used.add(s); return s

    for inv in investors:
        ws = wb.create_sheet(sheet_name(inv["name"]))
        ws.cell(1, 1, inv["name"] + " — Position Statement").font = font(True, ACCENT, 14)
        ws.cell(2, 1, "As of %s · values in actual USD" % uploaded).font = font(False, "888888", 9)

        # Summary block (label / value rows)
        irr_dec = (inv["irr"] / 100.0) if inv.get("irr") is not None else None
        summary = [
            ("Total Investment",       inv["committed"],              "money"),
            ("Distributions To Date",  inv["distributions_to_date"],  "money"),
            ("Forecast Distributions", inv["distributions_forecast"], "money"),
            ("Profit",                 inv["profit"],                 "money"),
            ("Equity Multiple",        inv["em"],                     "em"),
            ("IRR",                    irr_dec,                       "pct"),
        ]
        r = 4
        for lbl, val, kind in summary:
            ws.cell(r, 1, lbl).font = font(True, "666666", 9)
            c = ws.cell(r, 2)
            if kind == "money": money(c, val)
            elif kind == "em":  emx(c, val)
            else:               pct(c, val)
            r += 1

        # Positions table
        r += 1
        headers = (["Project", "Class", "Ownership", "Investment", "Dist. To Date"]
                   + list(quarters) + ["Forecast", "Proj IRR", "Proj EM"])
        for ci, h in enumerate(headers, start=1):
            cell = ws.cell(r, ci, h)
            cell.font = font(True, "1A1A1A", 9)
            cell.fill = HDR_FILL; cell.border = border
            if ci > 2:
                cell.alignment = right
        hdr_row = r
        r += 1
        nq = len(quarters)
        for p in sorted(inv.get("positions", []), key=lambda x: x["project"]):
            ws.cell(r, 1, p["project"]).border = border
            ws.cell(r, 2, "Preferred" if p["equity_class"] == "preferred" else "Common").border = border
            pc = ws.cell(r, 3); pct(pc, (p["pct"] or 0) / 100.0); pc.border = border
            mc = ws.cell(r, 4); money(mc, p["committed"]); mc.border = border
            dc = ws.cell(r, 5); money(dc, p["dist_to_date"]); dc.border = border
            for j in range(nq):
                qv = p["q_forecast"][j] if j < len(p["q_forecast"]) else 0
                qc = ws.cell(r, 6 + j); money(qc, qv); qc.border = border
            fc = ws.cell(r, 6 + nq); money(fc, p["dist_forecast"]); fc.border = border
            ic = ws.cell(r, 7 + nq); pct(ic, (p["irr"] or 0) / 100.0); ic.border = border
            ec = ws.cell(r, 8 + nq); emx(ec, p["em"]); ec.border = border
            r += 1
        # Total row
        tc = ws.cell(r, 1, "Total"); tc.font = font(True)
        for ci in range(1, 9 + nq):
            ws.cell(r, ci).fill = TOT_FILL; ws.cell(r, ci).border = border
        ws.cell(r, 2).font = font(True)
        m = ws.cell(r, 4); money(m, inv["committed"]); m.font = font(True)
        m = ws.cell(r, 5); money(m, inv["distributions_to_date"]); m.font = font(True)
        for j in range(nq):
            qv = inv["q_forecast"][j] if j < len(inv["q_forecast"]) else 0
            m = ws.cell(r, 6 + j); money(m, qv); m.font = font(True)
        m = ws.cell(r, 6 + nq); money(m, inv["distributions_forecast"]); m.font = font(True)
        ti = ws.cell(r, 7 + nq); pct(ti, (inv["irr"] / 100.0) if inv.get("irr") is not None else None); ti.font = font(True)
        te = ws.cell(r, 8 + nq); emx(te, inv["em"]); te.font = font(True)

        # Column widths
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 11
        for ci in range(3, 9 + nq):
            ws.column_dimensions[get_column_letter(ci)].width = 13

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


@app.route("/api/ember-capital/pdf", methods=["GET"])
@login_required
def ember_capital_pdf():
    """Stream the branded 2-page Ember Capital executive report as a PDF.
    Opens inline in the browser — user can print or save from there.
    ?download=1 forces a download instead of inline view."""
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("reports", True):
        return jsonify({"error": "Reports access disabled by admin"}), 403
    try:
        payload = _build_ember_capital_payload()
        pdf_bytes = bytes(_gen_pdf_ember_capital(payload))
    except Exception as e:
        return jsonify({"error": f"PDF generation failed: {e}"}), 500

    as_attachment = request.args.get("download") in ("1", "true", "yes")
    fname = f"Ember_Capital_Executive_Report_{datetime.datetime.now().strftime('%Y-%m')}.pdf"
    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=as_attachment,
        download_name=fname,
    )


# ---------------------------------------------------------------------------
# Branded executive PDFs for Returns / Loans / Operations
# Same look-and-feel as the Ember Capital PDF — full header/footer, branded
# tables, etc. Mount behind /api/<report>/pdf with ?download=1 for attachment.
# ---------------------------------------------------------------------------
_EXEC_REPORT_CONFIG = {
    "returns":    {"label": "Active_Project_Returns",         "access": "returns"},
    "loans":      {"label": "Loan_Capacities_and_Debt_Schedules", "access": "loans"},
    "operations": {"label": "Ember_Operating_Revenues",       "access": "operations"},
}


def _send_exec_report_pdf(report_type):
    cfg = _EXEC_REPORT_CONFIG.get(report_type)
    if not cfg:
        return jsonify({"error": "Unknown report type"}), 404

    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get(cfg["access"], True):
        return jsonify({"error": "Access denied"}), 403

    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "SELECT data FROM reports WHERE report_type = %s ORDER BY uploaded_at DESC LIMIT 1",
        (report_type,),
    )
    row = cur.fetchone()
    cur.close(); conn.close()

    if not row or not row["data"]:
        return jsonify({"error": "No data uploaded yet for this report."}), 404

    try:
        pdf_bytes = bytes(_gen_pdf_report(report_type, row["data"]))
    except Exception as e:
        return jsonify({"error": f"PDF generation failed: {e}"}), 500

    as_attachment = request.args.get("download") in ("1", "true", "yes")
    stamp = datetime.datetime.now().strftime("%Y-%m")
    fname = f"{cfg['label']}_{stamp}.pdf"
    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=as_attachment,
        download_name=fname,
    )


def _load_project_metadata():
    """Load the optional project metadata blob (image filename, location,
    role per project name) from the `reports` table. The blob is uploaded
    via the admin "Project Images" panel (see issue #project-images) and
    stored as a row with `report_type='project_metadata'`. Returns an
    empty dict if the row hasn't been created yet, so the PDF renders
    without images cleanly.

    Expected shape:
        {"projects": {
            "Grand Prairie East (CCC)": {
                "image_filename": "grand-prairie-east.jpg",   # under /static/img/projects/
                "location": "Grand Prairie, TX",
                "role": "Land · Common Equity"
            },
            ...
        }}
    """
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute(
            "SELECT data FROM reports WHERE report_type='project_metadata' "
            "ORDER BY uploaded_at DESC LIMIT 1"
        )
        row = cur.fetchone()
        cur.close(); conn.close()
    except Exception:
        return {}
    if not row or not row.get("data"):
        return {}
    blob = row["data"] or {}
    raw = blob.get("projects") or {}
    out = {}
    for name, meta in raw.items():
        if not isinstance(meta, dict):
            continue
        entry = {
            "location": meta.get("location"),
            "role":     meta.get("role"),
        }
        # New (persistent) path: image bytes are stored as base64 in the
        # JSON blob under image_data + image_mime. Survives Railway
        # redeploys because Railway's filesystem is ephemeral and any
        # files written under /static/img/projects/ get wiped when the
        # container restarts.
        img_data = meta.get("image_data")
        img_mime = meta.get("image_mime") or "image/jpeg"
        if img_data:
            entry["hero_image_url"] = f"data:{img_mime};base64,{img_data}"
        else:
            # Legacy path (kept for back-compat with older uploads): a
            # filename pointing at /static/img/projects/. WeasyPrint's
            # _weasyprint_local_fetcher serves this from disk if present.
            img_filename = meta.get("image_filename")
            if img_filename:
                entry["hero_image_url"] = f"/static/img/projects/{img_filename}"
            elif meta.get("hero_image_url"):
                entry["hero_image_url"] = meta["hero_image_url"]
        out[name] = entry
    return out


# ─── Project Library (admin-only image + metadata uploader) ──────────────────

import re as _re

_ALLOWED_IMG_EXTS = {".jpg", ".jpeg", ".png", ".webp"}
_PROJECTS_IMG_DIR = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "static", "img", "projects"
)


def _slugify_project(name: str) -> str:
    """Lowercase, hyphenate, strip non-alphanumerics. Stable per project name
    so re-uploads overwrite the same file in place."""
    s = (name or "").lower()
    s = _re.sub(r"[^\w\s-]", "", s, flags=_re.UNICODE)
    s = _re.sub(r"[\s_]+", "-", s).strip("-")
    s = _re.sub(r"-+", "-", s)
    return s or "project"


def _save_project_metadata_blob(blob: dict):
    """Upsert the singleton project_metadata row in the `reports` table."""
    import json as _json
    conn = get_db(); cur = conn.cursor()
    cur.execute(
        "INSERT INTO reports (report_type, data, uploaded_at) VALUES "
        "(%s, %s::jsonb, NOW())",
        ("project_metadata", _json.dumps(blob)),
    )
    conn.commit()
    cur.close(); conn.close()


def _read_project_metadata_blob() -> dict:
    """Latest project_metadata blob (raw, with image_filename keys)."""
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute(
            "SELECT data FROM reports WHERE report_type='project_metadata' "
            "ORDER BY uploaded_at DESC LIMIT 1"
        )
        row = cur.fetchone()
        cur.close(); conn.close()
    except Exception:
        return {"projects": {}}
    if not row or not row.get("data"):
        return {"projects": {}}
    blob = row["data"] or {}
    if not isinstance(blob.get("projects"), dict):
        blob = {"projects": {}}
    return blob


def _active_project_names() -> list[str]:
    """Names of projects in the most recent returns upload, in order."""
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute(
            "SELECT data FROM reports WHERE report_type='returns' "
            "ORDER BY uploaded_at DESC LIMIT 1"
        )
        row = cur.fetchone()
        cur.close(); conn.close()
    except Exception:
        return []
    if not row or not row.get("data"):
        return []
    return [
        (p.get("name") or "").strip()
        for p in (row["data"].get("projects") or [])
        if (p.get("name") or "").strip() and p.get("active") is not False
    ]


@app.route("/api/diagnostics/pdf", methods=["GET"])
@login_required
def diagnostics_pdf():
    """Read-out of which PDF backends are working on the live host.
    Admin-only. Every check is wrapped in its own try/except so the
    endpoint can never 500 on us; it always returns useful JSON.
    """
    if not session.get("is_admin"):
        return jsonify({"error": "Access denied"}), 403

    out = {}

    def _safe(name, fn):
        try:
            fn(out)
        except Exception as e:
            out[f"{name}_error"] = f"{type(e).__name__}: {e}"

    def _platform(o):
        import sys, platform
        o["python_version"] = sys.version.split()[0]
        o["platform"] = platform.platform()

    def _commit(o):
        o["git_commit"] = (
            os.environ.get("RAILWAY_GIT_COMMIT_SHA")
            or os.environ.get("RENDER_GIT_COMMIT")
            or os.environ.get("HEROKU_SLUG_COMMIT")
            or os.environ.get("SOURCE_VERSION")
            or "unknown"
        )

    def _weasyprint(o):
        try:
            import weasyprint as wp
            o["weasyprint_importable"] = True
            o["weasyprint_version"] = getattr(wp, "__version__", "?")
        except Exception as e:
            o["weasyprint_importable"] = False
            o["weasyprint_import_error"] = f"{type(e).__name__}: {e}"
            return
        try:
            wp.HTML(string="<p>ping</p>").write_pdf()
            o["weasyprint_render_ok"] = True
        except Exception as e:
            o["weasyprint_render_ok"] = False
            o["weasyprint_render_error"] = f"{type(e).__name__}: {e}"

    def _pillow(o):
        try:
            import PIL
            o["pillow_importable"] = True
            o["pillow_version"] = getattr(PIL, "__version__", "?")
        except Exception as e:
            o["pillow_importable"] = False
            o["pillow_import_error"] = f"{type(e).__name__}: {e}"

    def _templates(o):
        from flask import current_app
        for tpl in ("returns_report.html", "returns.html"):
            try:
                current_app.jinja_env.get_template(tpl)
                o[f"template_{tpl}"] = "found"
            except Exception as e:
                o[f"template_{tpl}"] = f"missing ({e})"

    def _uploads(o):
        conn = get_db(); cur = conn.cursor()
        cur.execute(
            "SELECT report_type, uploaded_at FROM reports "
            "WHERE report_type IN ('returns','project_metadata') "
            "ORDER BY uploaded_at DESC"
        )
        rows = cur.fetchall(); cur.close(); conn.close()
        seen, latest = set(), {}
        for r in rows or []:
            try:
                t  = r["report_type"]  if isinstance(r, dict) else r[0]
                ts = r["uploaded_at"]  if isinstance(r, dict) else r[1]
            except Exception:
                continue
            if t not in seen:
                seen.add(t)
                latest[t] = ts.isoformat() if hasattr(ts, "isoformat") else str(ts)
        o["latest_uploads"] = latest

    _safe("platform", _platform)
    _safe("commit", _commit)
    _safe("weasyprint", _weasyprint)
    _safe("pillow", _pillow)
    _safe("templates", _templates)
    _safe("uploads", _uploads)

    return jsonify(out)


@app.route("/api/admin/project-meta", methods=["GET"])
@login_required
def admin_project_meta_list():
    """Return the project list (from latest returns) joined with any
    saved metadata (image / location / role). Admin-only."""
    if not session.get("is_admin"):
        return jsonify({"error": "Access denied"}), 403
    def _entry_to_dict(name, meta, *, orphaned=False):
        meta = meta or {}
        img_data = meta.get("image_data")
        img_mime = meta.get("image_mime") or "image/jpeg"
        if img_data:
            url = f"data:{img_mime};base64,{img_data}"
        elif meta.get("image_filename"):
            url = f"/static/img/projects/{meta['image_filename']}"
        else:
            url = None
        d = {
            "name": name,
            "location": meta.get("location") or "",
            "role": meta.get("role") or "",
            "image_url": url,
        }
        if orphaned:
            d["orphaned"] = True
        return d

    names = _active_project_names()
    blob = _read_project_metadata_blob()
    raw = blob.get("projects") or {}
    projects = [_entry_to_dict(n, raw.get(n) or raw.get(n.strip())) for n in names]
    # Surface saved entries that no longer appear in the latest returns
    # upload (admin can still see + remove them).
    seen = {p["name"] for p in projects}
    for n, meta in raw.items():
        if n not in seen:
            projects.append(_entry_to_dict(n, meta, orphaned=True))
    return jsonify({"projects": projects})


@app.route("/api/admin/project-meta/<path:name>", methods=["POST"])
@login_required
def admin_project_meta_save(name):
    """Save (multipart) location/role and optionally a new hero image for
    the named project. Image, if provided, is resized to ≤1200px wide
    and saved as JPEG under /static/img/projects/<slug>.jpg."""
    if not session.get("is_admin"):
        return jsonify({"error": "Access denied"}), 403

    name = (name or "").strip()
    if not name:
        return jsonify({"error": "Missing project name"}), 400

    location = (request.form.get("location") or "").strip()
    role = (request.form.get("role") or "").strip()

    blob = _read_project_metadata_blob()
    projects = blob.setdefault("projects", {})
    entry = projects.get(name) or {}
    entry["location"] = location
    entry["role"] = role

    # Handle optional image — resize and base64-encode into the JSON
    # blob. We do NOT write to /static/img/projects/ because Railway's
    # container filesystem is ephemeral; any file written there is
    # wiped on the next deploy. Storing the bytes in Postgres survives
    # restarts and ships with the metadata as a single atomic record.
    file = request.files.get("image")
    if file and file.filename:
        ext = os.path.splitext(file.filename)[1].lower()
        if ext not in _ALLOWED_IMG_EXTS:
            return jsonify({
                "error": f"Unsupported image type {ext}. Allowed: "
                         + ", ".join(sorted(_ALLOWED_IMG_EXTS))
            }), 400
        try:
            from PIL import Image
        except ImportError:
            return jsonify({"error": "Server missing Pillow library"}), 500
        try:
            img = Image.open(file.stream)
            img.load()
        except Exception as e:
            return jsonify({"error": f"Could not read image: {e}"}), 400
        # Resize to max 1200px wide preserving aspect ratio. The PDF
        # template uses object-fit: cover on a 16:7 box so the user
        # doesn't have to crop themselves.
        if img.mode in ("RGBA", "LA", "P"):
            img = img.convert("RGB")
        max_w = 1200
        if img.width > max_w:
            ratio = max_w / float(img.width)
            new_h = max(1, int(img.height * ratio))
            img = img.resize((max_w, new_h), Image.LANCZOS)

        # Encode to JPEG bytes, then base64 for storage.
        import io as _io, base64 as _base64
        buf = _io.BytesIO()
        img.save(buf, "JPEG", quality=85, optimize=True, progressive=True)
        b64 = _base64.b64encode(buf.getvalue()).decode("ascii")
        entry["image_data"] = b64
        entry["image_mime"] = "image/jpeg"
        # Drop any legacy on-disk filename so the data: URI takes precedence.
        entry.pop("image_filename", None)

        # Best-effort cache to disk too — if the volume sticks around it
        # speeds up subsequent reads via the url_fetcher. Not required.
        try:
            os.makedirs(_PROJECTS_IMG_DIR, exist_ok=True)
            slug = _slugify_project(name)
            with open(os.path.join(_PROJECTS_IMG_DIR, f"{slug}.jpg"), "wb") as fh:
                fh.write(buf.getvalue())
        except Exception:
            pass

    projects[name] = entry
    _save_project_metadata_blob(blob)

    img_mime = entry.get("image_mime")
    img_data = entry.get("image_data")
    return jsonify({
        "ok": True,
        "name": name,
        "location": entry.get("location") or "",
        "role": entry.get("role") or "",
        "image_url": (f"data:{img_mime};base64,{img_data}" if img_data else None),
    })


@app.route("/api/admin/project-meta/<path:name>", methods=["DELETE"])
@login_required
def admin_project_meta_delete(name):
    """Remove a project's metadata entry. The image file on disk is kept
    (cheap to leave; expensive to recover after a mis-click)."""
    if not session.get("is_admin"):
        return jsonify({"error": "Access denied"}), 403
    name = (name or "").strip()
    blob = _read_project_metadata_blob()
    projects = blob.setdefault("projects", {})
    if name in projects:
        del projects[name]
        _save_project_metadata_blob(blob)
    return jsonify({"ok": True})


_STATIC_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static")


def _weasyprint_local_fetcher(url, timeout=10, ssl_context=None):
    """Custom WeasyPrint URL fetcher.

    /static/... URLs (font files, project hero images, the Ember logo)
    are read directly from disk instead of going back through HTTP. This
    avoids the deadlock case: WeasyPrint is rendering inside a gunicorn
    worker, makes an HTTP request to /static/foo.ttf, the request lands
    on the same worker pool, and either deadlocks with workers=1 or
    waits for a slow recursive request. Reading from disk is also ~100x
    faster than the round-trip.

    Anything else (data:, file://, or external https://) falls through
    to WeasyPrint's default fetcher.
    """
    import urllib.parse
    import mimetypes

    try:
        parsed = urllib.parse.urlparse(url)
        path = parsed.path or ""
        if path.startswith("/static/"):
            rel = path[len("/static/"):]
            disk_path = os.path.normpath(os.path.join(_STATIC_DIR, rel))
            # Refuse to escape the static directory.
            if not disk_path.startswith(_STATIC_DIR):
                raise FileNotFoundError(f"refused: {url}")
            if os.path.isfile(disk_path):
                with open(disk_path, "rb") as f:
                    data = f.read()
                mime, _ = mimetypes.guess_type(disk_path)
                return {
                    "string": data,
                    "mime_type": mime or "application/octet-stream",
                    "redirected_url": url,
                    "filename": os.path.basename(disk_path),
                }
            # File missing — raise so WeasyPrint logs a warning and falls
            # back to the next font in the CSS font-family chain. We do
            # NOT fall through to default_url_fetcher for /static/ URLs
            # because that would HTTP-fetch ourselves (the deadlock).
            raise FileNotFoundError(disk_path)
    except FileNotFoundError:
        raise
    except Exception:
        # Don't let a fetcher bug crash the whole render.
        pass

    # Non-/static URL: defer to WeasyPrint's default. Import lazily so
    # the fetcher itself never triggers a WeasyPrint import error on
    # Windows / dev machines where it isn't installed.
    from weasyprint.urls import default_url_fetcher
    return default_url_fetcher(url, timeout=timeout, ssl_context=ssl_context)


def _render_returns_report_pdf(raw_data, uploaded_at, tone="institutional",
                               *, preview_html=False):
    """Build the new design-handoff Project Returns PDF (or HTML preview).

    Pipeline:
        raw `reports.data`
          -> report.normalize()  (canonical shape, layered with project_metadata)
          -> report.build_context()  (KPIs, charts, formatters)
          -> Jinja render templates/returns_report.html
          -> WeasyPrint -> PDF bytes (with a local url_fetcher so /static/...
             URLs read from disk, never HTTP-fetch ourselves)
    """
    from report import normalize, build_context

    if isinstance(uploaded_at, str):
        try:
            uploaded_at = datetime.datetime.fromisoformat(
                uploaded_at.replace("Z", "+00:00")
            )
        except ValueError:
            uploaded_at = datetime.datetime.now()
    if uploaded_at is None:
        uploaded_at = datetime.datetime.now()

    project_meta = _load_project_metadata()
    data = normalize(raw_data, project_meta=project_meta)
    ctx = build_context(data, run_date=uploaded_at, tone=tone)

    html = render_template("returns_report.html", **ctx)
    if preview_html:
        return html

    from weasyprint import HTML  # imported lazily so a missing system lib
                                  # only breaks PDF rendering, not the app
    return HTML(
        string=html,
        base_url=request.host_url,
        url_fetcher=_weasyprint_local_fetcher,
    ).write_pdf()


@app.route("/api/returns/pdf", methods=["GET"])
@app.route("/returns/pdf", methods=["GET"])
@login_required
def returns_pdf():
    """Server-rendered Project Returns PDF.

    Visual ground truth: design_handoff_returns_pdf/_design_reference/
    Returns Report.html (Tone 1 = institutional, Tone 2 = editorial).

    Query params:
        download=1   -> Content-Disposition: attachment (otherwise inline)
        tone=editorial   -> dark cover + italic display variants
        preview=1   -> return raw HTML instead of PDF (debug)
    """
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("returns", True):
        return jsonify({"error": "Access denied"}), 403
    # Separate report-export gate (admin can disable downloads while
    # leaving page access intact).
    if not session.get("is_admin") and not pa.get("reports", True):
        return jsonify({"error": "Reports access disabled by admin"}), 403

    conn = get_db(); cur = conn.cursor()
    cur.execute(
        "SELECT data, uploaded_at FROM reports WHERE report_type = 'returns' "
        "ORDER BY uploaded_at DESC LIMIT 1"
    )
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row or not row.get("data"):
        return jsonify({"error": "No returns data has been uploaded yet."}), 404

    tone = request.args.get("tone", "institutional")
    if tone not in ("institutional", "editorial"):
        tone = "institutional"
    preview = request.args.get("preview") == "1"

    # WeasyPrint can fail two ways:
    #   1) ImportError      — pip didn't install (or its native deps for cffi)
    #   2) OSError / others — installed but Pango/Cairo .so files missing at
    #                          runtime (Railway Nixpacks needs nixpacks.toml,
    #                          not Aptfile). The error from cffi is usually
    #                          OSError("cannot load library 'libpango-1.0-0'").
    # Either way: fall back to the existing fpdf2 executive PDF so the
    # button always returns *something* readable. Diagnostic logged.
    try:
        result = _render_returns_report_pdf(
            row["data"], row["uploaded_at"], tone=tone, preview_html=preview
        )
    except (ImportError, OSError) as e:
        app.logger.warning("WeasyPrint unavailable (%s: %s), falling back to fpdf2",
                           type(e).__name__, e)
        if preview:
            # In preview mode we still want HTML — no PDF library needed for that.
            try:
                return _render_returns_report_pdf(
                    row["data"], row["uploaded_at"], tone=tone, preview_html=True
                )
            except Exception as ee:
                return f"<pre>HTML render failed: {ee}</pre>", 500
        return _send_exec_report_pdf("returns")
    except Exception as e:
        # Genuinely unexpected — log full trace, but still try the fpdf2
        # fallback so the user gets a working PDF rather than a 500.
        app.logger.exception("Returns PDF render failed: %s", e)
        if preview:
            return f"<pre>Render error: {e}</pre>", 500
        try:
            return _send_exec_report_pdf("returns")
        except Exception:
            return jsonify({"error": f"PDF render failed: {e}"}), 500

    if preview:
        return result  # raw HTML

    as_attachment = request.args.get("download") in ("1", "true", "yes")
    stamp = (row["uploaded_at"] or datetime.datetime.now()).strftime("%Y-%m-%d")
    fname = f"EmberAcquisitions_Returns_{stamp}.pdf"
    disp = "attachment" if as_attachment else "inline"
    return Response(
        result,
        mimetype="application/pdf",
        headers={
            "Content-Disposition": f'{disp}; filename="{fname}"',
            "Cache-Control": "no-store",
        },
    )


@app.route("/api/loans/pdf", methods=["GET"])
@login_required
def loans_pdf():
    """Branded executive PDF for Loan Capacities & Debt Schedules."""
    return _send_exec_report_pdf("loans")


@app.route("/api/operations/pdf", methods=["GET"])
@login_required
def operations_pdf():
    """Branded executive PDF for Ember Operating Revenues."""
    return _send_exec_report_pdf("operations")


@app.route("/api/portfolio", methods=["GET"])
@login_required
def portfolio():
    conn = get_db()
    cur = conn.cursor()
    include_archived = request.args.get("include_archived") == "true"
    where = "" if include_archived else "WHERE p.archived = FALSE"
    cur.execute(f"""
        SELECT p.id, p.name, p.address, p.outputs, p.scenarios, COALESCE(p.status, 'Active') as status
        FROM projects p
        {where}
        ORDER BY p.name
    """)
    rows = cur.fetchall()
    cur.close(); conn.close()
    result = []
    for r in rows:
        scens = list(r["scenarios"] or [])
        o = (scens[0]["outputs"] if scens else None) or r["outputs"] or {}
        result.append({"id": r["id"], "name": r["name"], "address": r["address"],
                       "status": r["status"], "outputs": o})
    return jsonify(result)

@app.route("/api/projects/<int:pid>/export_excel", methods=["GET"])
@login_required
def export_excel(pid):
    try:
        from excel_export import export_excel as _export
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT * FROM projects WHERE id=%s", (pid,))
        proj = cur.fetchone()
        cur.close(); conn.close()
        if not proj:
            return jsonify({"error": "not found"}), 404
        inputs = proj["inputs"] or {}
        excel_bytes = _export(inputs)
        name = (proj.get("name") or "project").replace(" ", "_")
        return send_file(
            io.BytesIO(excel_bytes),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"{name}_Underwriting.xlsx"
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/projects/<int:pid>/backup", methods=["GET"])
@login_required
def backup_project(pid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM projects WHERE id=%s", (pid,))
    proj = dict(cur.fetchone() or {})
    cur.close(); conn.close()
    if not proj:
        return jsonify({"error": "not found"}), 404
    for k, v in proj.items():
        if hasattr(v, 'isoformat'):
            proj[k] = v.isoformat()
    name = (proj.get("name") or "project").replace(" ", "_")
    backup_data = json.dumps(proj, indent=2)
    return send_file(
        io.BytesIO(backup_data.encode()),
        mimetype="application/json",
        as_attachment=True,
        download_name=f"{name}_backup.json"
    )

@app.route("/api/projects/restore", methods=["POST"])
@login_required
def restore_project():
    data = request.json or {}
    inputs  = data.get("inputs", {})
    outputs = data.get("outputs", {})
    name    = data.get("name", "Restored Project")
    address = data.get("address", "")
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO projects (name, address, created_by, inputs, outputs) VALUES (%s,%s,%s,%s,%s) RETURNING id",
        (name, address, session["user_id"], json.dumps(inputs), json.dumps(outputs))
    )
    new_id = cur.fetchone()["id"]
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True, "id": new_id})

@app.route("/api/projects/import_excel", methods=["POST"])
@login_required
def import_excel_project():
    """Upload an Ember underwriting Excel and create a new project from it."""
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file provided"}), 400
    try:
        from excel_import import import_excel
        file_bytes = f.read()
        inputs = import_excel(file_bytes)
    except Exception as e:
        return jsonify({"error": f"Failed to parse Excel: {e}"}), 400
    try:
        outputs = calculate(inputs)
    except Exception:
        outputs = {}
    name = inputs.get("project_name", "Imported Project")
    address = inputs.get("address", "")
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO projects (name, address, created_by, inputs, outputs) VALUES (%s, %s, %s, %s, %s) RETURNING id",
        (name, address, session["user_id"], json.dumps(inputs), json.dumps(outputs))
    )
    pid = cur.fetchone()["id"]
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True, "id": pid, "name": name})

@app.route("/api/parse_excel", methods=["POST"])
@login_required
def parse_excel():
    """Parse an Ember underwriting Excel and return the inputs dict (no project created)."""
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file provided"}), 400
    try:
        from excel_import import import_excel
        file_bytes = f.read()
        inputs = import_excel(file_bytes)
    except Exception as e:
        return jsonify({"error": f"Failed to parse Excel: {e}"}), 400
    return jsonify({"ok": True, "inputs": inputs})

# ─── DASHBOARD REPORTS ────────────────────────────────────────────────────────
@app.route("/api/upload-dashboard", methods=["POST"])
@login_required
@admin_required
def upload_dashboard():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file provided"}), 400
    try:
        file_bytes = f.read()
        data = parse_dashboard(file_bytes)
    except Exception as e:
        return jsonify({"error": f"Failed to parse file: {e}"}), 400
    conn = get_db()
    cur = conn.cursor()
    # Upsert returns
    cur.execute("DELETE FROM reports WHERE report_type = 'returns'")
    cur.execute(
        "INSERT INTO reports (report_type, data, uploaded_by) VALUES (%s, %s, %s)",
        ("returns", json.dumps(data.get("returns", {})), session["user_id"])
    )
    # Upsert loans
    cur.execute("DELETE FROM reports WHERE report_type = 'loans'")
    cur.execute(
        "INSERT INTO reports (report_type, data, uploaded_by) VALUES (%s, %s, %s)",
        ("loans", json.dumps(data.get("loans", {})), session["user_id"])
    )
    # Upsert operations
    cur.execute("DELETE FROM reports WHERE report_type = 'operations'")
    if data.get("operations"):
        cur.execute(
            "INSERT INTO reports (report_type, data, uploaded_by) VALUES (%s, %s, %s)",
            ("operations", json.dumps(data["operations"]), session["user_id"])
        )
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/export-returns-excel")
@login_required
def export_returns_excel():
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("returns", True):
        return jsonify({"error": "Access denied"}), 403
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT data, uploaded_at FROM reports WHERE report_type = 'returns' ORDER BY uploaded_at DESC LIMIT 1")
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row or not row["data"]:
        return jsonify({"error": "No data available"}), 404

    data = row["data"]
    uploaded_at = row["uploaded_at"].strftime("%B %d, %Y") if row["uploaded_at"] else ""

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    LABEL_MAP = {"LP IRR": "Net Cashflow", "LP Equity Multiple": "Cumulative Net Cashflow"}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Project Returns"

    # Light-background palette — legible on white Excel
    PROJ_FILL   = PatternFill("solid", fgColor="F2EFE8")   # warm tan for project header
    SUMM_FILL   = PatternFill("solid", fgColor="E8F0EE")   # light teal for summary header
    HEADER_FILL = PatternFill("solid", fgColor="F7F6F3")   # near-white for column headers
    thin = Side(style="thin", color="CCCCCC")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    TEXT       = "1A1A1A"   # near-black for data
    HDR_TEXT   = "555555"   # medium grey for column header labels
    PROJ_TEXT  = "6B4E1E"   # dark brown for project title
    SUMM_TEXT  = "2D6B5A"   # dark teal for summary title
    ACCENT     = "7A5C1E"   # dark gold for highlighted metric rows (IRR/EM)

    def _f(bold=False, color=TEXT, size=9):
        return Font(name="Calibri", size=size, bold=bold, color=color)

    def _set_num(cell, val):
        """Write val as #,##0; blank if zero/None."""
        if isinstance(val, (int, float)) and val != 0:
            cell.value = val
            cell.number_format = "#,##0"
        else:
            cell.value = None

    years = data.get("years", [])
    all_idxs = list(range(len(years)))   # include every year column
    num_cols = 2 + len(years)            # label + Total + one per year

    r = 1
    ws.cell(row=r, column=1, value="Consolidated Ember Project Returns").font = Font(name="Calibri", bold=True, size=14, color=PROJ_TEXT)
    r += 1
    ws.cell(row=r, column=1, value=f"Last updated: {uploaded_at}  |  ($ in 000s)").font = _f(color="888888")
    r += 2

    # ── Project Returns Summary Table ──
    SUMMARY_HDR_FILL = PatternFill("solid", fgColor="EDE8DF")
    summary_cols = ["Project", "LP IRR", "Equity Multiple", "Total LP Profit", "Promote"]
    for ci, h in enumerate(summary_cols, 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = _f(bold=True, color=HDR_TEXT)
        c.fill = SUMMARY_HDR_FILL
        c.border = cell_border
        c.alignment = Alignment(horizontal="left" if ci == 1 else "center")
    r += 1

    for proj in data.get("projects", []):
        metrics_by_label = {m["label"]: m for m in proj.get("metrics", [])}
        irr_val  = metrics_by_label.get("LP IRR", {}).get("total", None)
        em_val   = metrics_by_label.get("LP Equity Multiple", {}).get("total", None)
        pft_val  = metrics_by_label.get("Total LP Profit", {}).get("total", None)
        prom_val = metrics_by_label.get("Promote", {}).get("total", None)

        # Project name
        nc = ws.cell(row=r, column=1, value=proj["name"])
        nc.font = _f(bold=True, color=PROJ_TEXT)
        nc.border = cell_border

        # LP IRR — display as percentage
        ic = ws.cell(row=r, column=2)
        ic.font = _f(bold=True, color=ACCENT)
        ic.alignment = Alignment(horizontal="right")
        ic.border = cell_border
        if isinstance(irr_val, (int, float)) and irr_val:
            ic.value = irr_val
            ic.number_format = "0.0%"

        # Equity Multiple — display as multiplier
        ec = ws.cell(row=r, column=3)
        ec.font = _f(bold=True, color=ACCENT)
        ec.alignment = Alignment(horizontal="right")
        ec.border = cell_border
        if isinstance(em_val, (int, float)) and em_val:
            ec.value = em_val
            ec.number_format = '0.00"x"'

        # Total LP Profit
        pc = ws.cell(row=r, column=4)
        pc.font = _f()
        pc.alignment = Alignment(horizontal="right")
        pc.border = cell_border
        _set_num(pc, pft_val)

        # Promote
        prc = ws.cell(row=r, column=5)
        prc.font = _f()
        prc.alignment = Alignment(horizontal="right")
        prc.border = cell_border
        _set_num(prc, prom_val)

        r += 1

    r += 1  # blank spacer before detail sections

    def write_section_header(r, title, fill, color):
        c = ws.cell(row=r, column=1, value=title)
        c.font = Font(name="Calibri", bold=True, size=10, color=color)
        c.fill = fill
        c.border = cell_border
        for ci in range(2, num_cols + 1):
            cell = ws.cell(row=r, column=ci)
            cell.fill = fill
            cell.border = cell_border
        return r + 1

    def write_col_headers(r, col_labels):
        ws.cell(row=r, column=1, value="Metric").font = _f(bold=True, color=HDR_TEXT)
        ws.cell(row=r, column=1).fill = HEADER_FILL
        ws.cell(row=r, column=1).border = cell_border
        ws.cell(row=r, column=2, value="Total").font = _f(bold=True, color=HDR_TEXT)
        ws.cell(row=r, column=2).fill = HEADER_FILL
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2).border = cell_border
        for ci, lbl in enumerate(col_labels, 3):
            c = ws.cell(row=r, column=ci, value=lbl)
            c.font = _f(bold=True, color=HDR_TEXT)
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center")
            c.border = cell_border
        return r + 1

    def write_project(r, proj):
        metrics = proj.get("metrics", [])
        r = write_section_header(r, proj["name"], PROJ_FILL, PROJ_TEXT)
        r = write_col_headers(r, years)

        for m in metrics:
            label = m["label"]
            display = LABEL_MAP.get(label, label)
            is_accent = label in ("LP IRR", "LP Equity Multiple")
            txt_color = ACCENT if is_accent else TEXT

            # Total: for renamed rows use sum of yearly (Net Cashflow) or last non-zero (Cum. Net CF)
            if label == "LP IRR":
                total = sum(v for v in m.get("yearly", []) if isinstance(v, (int, float)))
            elif label == "LP Equity Multiple":
                yvals = [v for v in m.get("yearly", []) if isinstance(v, (int, float)) and v != 0]
                total = yvals[-1] if yvals else 0
            else:
                total = m.get("total", 0)

            lc = ws.cell(row=r, column=1, value=display)
            lc.font = _f(bold=is_accent, color=txt_color)
            lc.border = cell_border

            tc = ws.cell(row=r, column=2)
            tc.font = _f(bold=is_accent, color=txt_color)
            tc.alignment = Alignment(horizontal="right")
            tc.border = cell_border
            _set_num(tc, total)

            for ci, i in enumerate(all_idxs, 3):
                yc = ws.cell(row=r, column=ci)
                val = m["yearly"][i] if i < len(m.get("yearly", [])) else 0
                yc.font = _f(color=txt_color)
                yc.alignment = Alignment(horizontal="right")
                yc.border = cell_border
                _set_num(yc, val)
            r += 1
        return r + 1

    for proj in data.get("projects", []):
        r = write_project(r, proj)

    # Portfolio Summary
    summary = data.get("summary", [])
    if summary:
        r = write_section_header(r, "Portfolio Summary", SUMM_FILL, SUMM_TEXT)
        r = write_col_headers(r, years)
        for s in summary:
            lc = ws.cell(row=r, column=1, value=s["label"])
            lc.font = _f()
            lc.border = cell_border
            tc = ws.cell(row=r, column=2)
            tc.font = _f()
            tc.alignment = Alignment(horizontal="right")
            tc.border = cell_border
            _set_num(tc, s.get("total", 0))
            for ci, i in enumerate(all_idxs, 3):
                yc = ws.cell(row=r, column=ci)
                val = s["yearly"][i] if i < len(s.get("yearly", [])) else 0
                yc.font = _f()
                yc.alignment = Alignment(horizontal="right")
                yc.border = cell_border
                _set_num(yc, val)
            r += 1

    # Column widths — B=Total/IRR, C=EquityMult/yr1, D onwards = years
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 14   # "Equity Multiple" header needs a touch more
    ws.column_dimensions["D"].width = 14   # "Total LP Profit"
    ws.column_dimensions["E"].width = 13   # "Promote"
    for ci in range(6, 3 + len(years)):
        ws.column_dimensions[get_column_letter(ci)].width = 11

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    from flask import send_file
    return send_file(output, as_attachment=True,
                     download_name="Ember_Project_Returns.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def _enrich_returns_payload(raw):
    """Augment the raw returns report with per-project + portfolio aggregates
    needed for the redesigned two-pane workspace.

    Returns a dict shaped for the new returns template:
      {
        years: [...],
        portfolio: {
          active_count, total_lp_profit_k, weighted_irr, weighted_em,
          total_distributions_k, total_contributions_k,
        },
        projects: [
          {
            name, vintage, irr, em, profit, distributions, contributions,
            net_yearly, cumulative_yearly, metrics: [...]   # raw, for the table
          }, ...
        ],
        summary: [...]   # untouched portfolio summary (for fallback)
      }
    Values that come from the report stay in $K (matching the rest of the
    platform and the 'Cashflow detail · $ in 000s' caption).
    """
    if not raw:
        return None
    years = raw.get("years") or []
    raw_projects = raw.get("projects") or []

    def _f(by_label, label):
        v = (by_label.get(label) or {}).get("total")
        try: return float(v) if v is not None else 0.0
        except (TypeError, ValueError): return 0.0

    def _yearly(by_label, label, n):
        m = by_label.get(label) or {}
        y = list(m.get("yearly") or [])
        # pad/truncate so all series have the same length as years
        if len(y) < n: y = y + [0.0] * (n - len(y))
        try:
            y = [float(v) if v is not None else 0.0 for v in y]
        except (TypeError, ValueError):
            y = [0.0] * n
        return y[:n]

    enriched_projects = []
    for p in raw_projects:
        by_label = {m.get("label"): m for m in (p.get("metrics") or [])}
        nyrs = len(years) or len((p.get("metrics") or [{}])[0].get("yearly") or [])
        contrib_total = abs(_f(by_label, "Total LP Contributions"))
        distrib_total = _f(by_label, "Total LP Distributions")
        irr = _f(by_label, "LP IRR")
        em  = _f(by_label, "LP Equity Multiple")
        profit = _f(by_label, "Total LP Profit")

        # Vintage = first year with a non-zero LP contribution. Fall back to
        # the first year with any non-zero metric.
        vintage = None
        contrib_yearly = _yearly(by_label, "Total LP Contributions", nyrs)
        for i, v in enumerate(contrib_yearly):
            if v and abs(v) > 0:
                vintage = years[i] if i < len(years) else None
                break
        if vintage is None:
            for i in range(nyrs):
                hit = False
                for m in (p.get("metrics") or []):
                    yr = m.get("yearly") or []
                    if i < len(yr):
                        try:
                            if yr[i] and abs(float(yr[i])) > 0:
                                hit = True; break
                        except (TypeError, ValueError):
                            pass
                if hit:
                    vintage = years[i] if i < len(years) else None
                    break

        # Net cashflow series: prefer the Net Cashflow row if present;
        # otherwise compute as Distributions + Contributions (negative).
        net_y = _yearly(by_label, "Net Cashflow", nyrs)
        if not any(abs(v) > 0 for v in net_y):
            d = _yearly(by_label, "Total LP Distributions", nyrs)
            c = _yearly(by_label, "Total LP Contributions", nyrs)
            net_y = [d[i] + c[i] for i in range(nyrs)]

        # Cumulative series: prefer the report's row, else running sum of net.
        cum_y = _yearly(by_label, "Cumulative Net Cashflow", nyrs)
        if not any(abs(v) > 0 for v in cum_y):
            running = 0.0
            cum_y = []
            for v in net_y:
                running += v
                cum_y.append(running)

        enriched_projects.append({
            "name":          p.get("name", ""),
            "vintage":       vintage,
            "irr":           irr,
            "em":            em,
            "profit":        profit,
            "distributions": distrib_total,
            "contributions": contrib_total,
            "net_yearly":    net_y,
            "cumulative_yearly": cum_y,
            "metrics":       p.get("metrics") or [],
        })

    # Portfolio aggregates. Weight IRR & EM by abs(LP contributions): bigger
    # checks pull the average. Fall back to a simple average if nothing has
    # any contribution data.
    total_profit_k = sum(p["profit"] for p in enriched_projects)
    total_dist_k   = sum(p["distributions"] for p in enriched_projects)
    total_contr_k  = sum(p["contributions"] for p in enriched_projects)
    irr_w = sum(p["irr"] * p["contributions"] for p in enriched_projects if p["contributions"] > 0)
    em_w  = sum(p["em"]  * p["contributions"] for p in enriched_projects if p["contributions"] > 0)
    weight = sum(p["contributions"] for p in enriched_projects if p["contributions"] > 0)
    weighted_irr = (irr_w / weight) if weight > 0 else 0.0
    weighted_em  = (em_w  / weight) if weight > 0 else 0.0

    return {
        "years": years,
        "portfolio": {
            "active_count":            len(enriched_projects),
            "total_lp_profit_k":       total_profit_k,
            "weighted_irr":            weighted_irr,
            "weighted_em":             weighted_em,
            "total_distributions_k":   total_dist_k,
            "total_contributions_k":   total_contr_k,
        },
        "projects": enriched_projects,
        "summary":  raw.get("summary") or [],
    }


@app.route("/returns")
@login_required
def returns_report():
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("returns", True):
        return redirect(url_for("home"))
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT data, uploaded_at FROM reports WHERE report_type = 'returns' ORDER BY uploaded_at DESC LIMIT 1")
    row = cur.fetchone()
    cur.close(); conn.close()
    data = row["data"] if row else None
    # Data dates footer: "Data From" comes from the parser (cell E2 on
    # Consolidated Project Returns); "Uploaded" is the row's upload time.
    data_from   = _fmt_data_date(data.get("data_from") if data else None)
    uploaded_at = _fmt_data_date(row["uploaded_at"]) if row else None
    enriched = _enrich_returns_payload(data)
    pa = session.get("page_access") or {"mpc_underwriting": True, "returns": True, "loans": True, "operations": True}
    if session.get("is_admin"):
        pa = {"mpc_underwriting": True, "returns": True, "loans": True, "operations": True}
    return render_template("returns.html",
        data=data, enriched=enriched,
        data_from=data_from, uploaded_at=uploaded_at,
        username=session.get("username"),
        is_admin=session.get("is_admin"), page_access=pa)

@app.route("/loans")
@login_required
def loans_report():
    """Loans & Debt page — Concept B Hybrid Stacks redesign.

    Reads the latest reports[loans] row, runs it through the legacy
    adapter (`_loans_translate_legacy`) so existing parser-uploaded
    Excel data still works, then enriches each loan with formatted
    strings + health classes. Empty state renders when no data.
    """
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("loans", True):
        return redirect(url_for("home"))
    pa.setdefault("loans",   True)
    pa.setdefault("reports", True)

    conn = get_db(); cur = conn.cursor()
    cur.execute(
        "SELECT data, uploaded_at FROM reports "
        "WHERE report_type = 'loans' ORDER BY uploaded_at DESC LIMIT 1"
    )
    row = cur.fetchone(); cur.close(); conn.close()
    raw_data    = row["data"]        if row else None
    uploaded_at = row["uploaded_at"] if row else None

    # Debug helper for diagnosing data-shape issues. Hit /loans?debug=1
    # to get a JSON dump of (a) the raw DB row, (b) the enriched view
    # context's per-loan capacityHealth fields. Admin-only.
    if request.args.get("debug") == "1" and session.get("is_admin"):
        view = _build_loans_view_context(raw_data, uploaded_at) or {}
        diag = {
            "uploaded_at": uploaded_at.isoformat() if hasattr(uploaded_at, "isoformat") else str(uploaded_at),
            "raw_top_keys": list(raw_data.keys()) if isinstance(raw_data, dict) else None,
            "raw_mpc_first_row": (raw_data.get("mpc_loans") or {}).get("rows", [{}])[:1] if isinstance(raw_data, dict) else None,
            "raw_vert_first_row": (raw_data.get("vertical_loans") or {}).get("rows", [{}])[:1] if isinstance(raw_data, dict) else None,
            "enriched_mpc": [
                {
                    "community":          l.get("community"),
                    "capacityHealth_raw": l.get("capacityHealth"),
                    "capacityHealth_fmt": l.get("capacityHealth_fmt"),
                    "cap_cls":            l.get("cap_cls"),
                    "irHealth_raw":       l.get("irHealth"),
                    "irHealth_fmt":       l.get("irHealth_fmt"),
                    "ir_cls":             l.get("ir_cls"),
                } for l in (view.get("mpc", {}).get("loans") or [])
            ],
            "enriched_vert": [
                {
                    "community":          l.get("community"),
                    "capacityHealth_raw": l.get("capacityHealth"),
                    "capacityHealth_fmt": l.get("capacityHealth_fmt"),
                    "cap_cls":            l.get("cap_cls"),
                } for l in (view.get("vert", {}).get("loans") or [])
            ],
        }
        return jsonify(diag)

    loans_ctx = _build_loans_view_context(raw_data, uploaded_at)
    # Data dates footer:
    #   "Data From"  — Loan Capacities & DS!U3 (loan capacities cutoff)
    #   "Debt From"  — Debt!D1 (debt-schedules detail cutoff; pulled
    #                  into the loans tab via formulas, so its own
    #                  date is often different from U3)
    #   "Uploaded"   — when this row landed in the reports table
    data_from        = _fmt_data_date(raw_data.get("data_from")      if raw_data else None)
    debt_from        = _fmt_data_date(raw_data.get("debt_data_from") if raw_data else None)
    uploaded_at_fmt  = _fmt_data_date(uploaded_at)
    extras           = [("Debt From", debt_from)] if debt_from else []
    return render_template(
        "loans.html",
        loans=loans_ctx,
        data_from=data_from,
        data_dates_extra=extras,
        uploaded_at=uploaded_at_fmt,
        is_admin=session.get("is_admin", False),
        page_access=pa,
    )


@app.route("/loans/pdf")
@login_required
def loans_pdf_view():
    """5-page landscape PDF of the Loans & Debt report.

    ?preview=1   → return raw HTML (debug)
    ?download=1  → force attachment Content-Disposition
    Falls back to the legacy fpdf2 generator when WeasyPrint can't load.
    """
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("loans", True):
        return jsonify({"error": "Access denied"}), 403
    if not session.get("is_admin") and not pa.get("reports", True):
        return jsonify({"error": "Reports access disabled by admin"}), 403

    conn = get_db(); cur = conn.cursor()
    cur.execute(
        "SELECT data, uploaded_at FROM reports "
        "WHERE report_type = 'loans' ORDER BY uploaded_at DESC LIMIT 1"
    )
    row = cur.fetchone(); cur.close(); conn.close()
    if not row or not row.get("data"):
        return jsonify({"error": "No loans data uploaded"}), 404

    view_ctx = _build_loans_view_context(row["data"], row.get("uploaded_at"))
    if view_ctx is None:
        return jsonify({"error": "Loans data is empty"}), 404
    rpt_ctx = _build_loans_report_context(view_ctx, run_date=row.get("uploaded_at"))
    html = render_template("loans_report.html", **rpt_ctx)

    if request.args.get("preview") == "1":
        return html

    try:
        from weasyprint import HTML
        pdf_bytes = HTML(
            string=html,
            base_url=request.host_url,
            url_fetcher=_weasyprint_local_fetcher,
        ).write_pdf()
    except (ImportError, OSError) as e:
        app.logger.warning(
            "WeasyPrint unavailable for loans report (%s: %s); falling back to fpdf2",
            type(e).__name__, e,
        )
        return _send_exec_report_pdf("loans")

    as_attachment = request.args.get("download") in ("1", "true", "yes")
    fname = f"Ember_Loans_{datetime.datetime.now().strftime('%Y-%m')}.pdf"
    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=as_attachment,
        download_name=fname,
    )


@app.route("/api/export-loans-excel")
@login_required
def export_loans_excel_view():
    """Excel export — drives the existing `_gen_excel_loans` workbook."""
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("loans", True):
        return jsonify({"error": "Access denied"}), 403
    if not session.get("is_admin") and not pa.get("reports", True):
        return jsonify({"error": "Reports access disabled by admin"}), 403
    conn = get_db(); cur = conn.cursor()
    cur.execute(
        "SELECT data, uploaded_at FROM reports "
        "WHERE report_type = 'loans' ORDER BY uploaded_at DESC LIMIT 1"
    )
    row = cur.fetchone(); cur.close(); conn.close()
    if not row or not row.get("data"):
        return jsonify({"error": "No loans data uploaded"}), 404
    xlsx_bytes = _gen_excel_loans(row["data"])
    return send_file(
        io.BytesIO(xlsx_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="Ember_Loans_Capacities.xlsx",
    )

@app.route("/api/export-operations-excel")
@login_required
def export_operations_excel():
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("operations", True):
        return jsonify({"error": "Access denied"}), 403
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT data, uploaded_at FROM reports WHERE report_type = 'operations' ORDER BY uploaded_at DESC LIMIT 1")
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row or not row["data"]:
        return jsonify({"error": "No data available"}), 404

    data = row["data"]
    uploaded_at = row["uploaded_at"].strftime("%B %d, %Y") if row["uploaded_at"] else ""

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Operating Revenues"

    WHITE = "FFFFFF"
    HEADER_FILL = PatternFill("solid", fgColor="1E2535")
    TOTALS_FILL = PatternFill("solid", fgColor="161B24")
    thin = Side(style="thin", color="2E3750")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Font palette mirrors the user's hand-tweaked workbook
    # (Downloads/Ember_Operating_Revenues (2).xlsx):
    #   - Title + section headers: default (black), bold, sized.
    #   - Column headers + totals rows: white text on the dark navy fills
    #     so they read; previously muted-gray / black-on-dark was illegible.
    #   - KPI labels + data cells: unchanged (default black).
    def _hdr_font(bold=False):
        return Font(name="Calibri", size=9, bold=bold, color=WHITE)

    def _val_font(bold=False):
        return Font(name="Calibri", size=9, bold=bold)

    def _totals_font(bold=True):
        return Font(name="Calibri", size=9, bold=bold, color=WHITE)

    def _section_font(size=11):
        return Font(name="Calibri", size=size, bold=True)

    def write_section(r, title):
        c = ws.cell(row=r, column=1, value=title)
        c.font = _section_font(size=11)
        return r + 1

    def write_table(r, col_headers, data_rows, totals):
        # Header row
        for ci, h in enumerate(col_headers, 1):
            c = ws.cell(row=r, column=ci, value=h)
            c.font = _hdr_font(bold=True)
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center" if ci > 1 else "left")
            c.border = cell_border
        r += 1
        # Data rows
        for dr in data_rows:
            for ci, v in enumerate(dr, 1):
                cell = ws.cell(row=r, column=ci, value=v if v != 0 else None)
                cell.font = _val_font()
                cell.border = cell_border
                cell.alignment = Alignment(horizontal="left" if ci == 1 else "right")
                if ci > 1 and isinstance(v, (int, float)) and v:
                    cell.number_format = "#,##0"
            r += 1
        # Totals row — white text on the navy TOTALS_FILL so it reads.
        ws.cell(row=r, column=1, value="Total").font = _totals_font(bold=True)
        ws.cell(row=r, column=1).border = cell_border
        ws.cell(row=r, column=1).fill = TOTALS_FILL
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left")
        for ci, v in enumerate(totals, 2):
            cell = ws.cell(row=r, column=ci, value=v if v else None)
            cell.font = _totals_font(bold=True)
            cell.fill = TOTALS_FILL
            cell.border = cell_border
            cell.alignment = Alignment(horizontal="right")
            if isinstance(v, (int, float)):
                cell.number_format = "#,##0"
        return r + 2

    r = 1
    # Title — default-color bold (gold removed to match the tweaked palette).
    ws.cell(row=r, column=1, value="Ember Operating Revenues").font = Font(name="Calibri", bold=True, size=14)
    r += 1
    ws.cell(row=r, column=1, value=f"Last updated: {uploaded_at}").font = Font(name="Calibri", size=9, color="8B95A8")
    r += 2

    # KPIs
    r = write_section(r, "KPI Summary")
    for kpi in data.get("kpis", []):
        ws.cell(row=r, column=1, value=kpi["label"]).font = _val_font()
        vc = ws.cell(row=r, column=2, value=kpi["value"])
        vc.font = _val_font(bold=True)
        vc.number_format = "#,##0"
        vc.alignment = Alignment(horizontal="right")
        r += 1
    r += 1

    # Annual Forecast
    yr = data.get("yearly_rollup", {})
    if yr.get("years"):
        r = write_section(r, "Annual Revenue Forecast (Next 5 Years)")
        headers = ["Revenue Source"] + [str(y) for y in yr["years"]]
        rows = [[row["label"]] + row["values"] for row in yr.get("rows", [])]
        r = write_table(r, headers, rows, yr.get("totals", []))

    # Monthly Revenue — by-project roll-up + detailed Project × Category
    mo = data.get("monthly", {})
    if mo.get("dates"):
        dates    = mo["dates"]
        mo_rows  = mo.get("rows", []) or []
        mo_total = mo.get("totals", []) or []
        month_hdrs = [f"{d[5:7]}/{d[2:4]}" for d in dates]

        # Roll-up: sum across the 6 fee categories per project per month.
        # Preserves the first-seen project order from the detail rows so
        # this section and the detail below read in the same order.
        proj_order: list[str] = []
        proj_totals: dict[str, list[float]] = {}
        for row in mo_rows:
            p = row.get("project")
            vals = row.get("values") or []
            if not p:
                continue
            if p not in proj_totals:
                proj_order.append(p)
                proj_totals[p] = [0.0] * len(dates)
            for i, v in enumerate(vals[: len(dates)]):
                try:
                    proj_totals[p][i] += float(v or 0)
                except (TypeError, ValueError):
                    pass

        if proj_order:
            r = write_section(r, "Monthly Revenue by Project (Roll-Up)")
            headers = ["Project"] + month_hdrs
            rolled  = [[p] + proj_totals[p] for p in proj_order]
            r = write_table(r, headers, rolled, mo_total)

        # Detail: Project × Category, unchanged.
        r = write_section(r, "Monthly Fee Revenue")
        headers = ["Project / Category"] + month_hdrs
        rows = [[f"{row['project']} — {row['category']}"] + row["values"] for row in mo_rows]
        r = write_table(r, headers, rows, mo_total)

    # Next 12 Months
    n12 = data.get("next_12_months", {})
    if n12.get("dates"):
        r = write_section(r, "Next 12 Months")
        dates = n12["dates"]
        headers = ["Revenue Source"] + [f"{d[5:7]}/{d[2:4]}" for d in dates]
        rows = [[row["label"]] + row["values"] for row in n12.get("rows", [])]
        r = write_table(r, headers, rows, n12.get("totals", []))

    # Next 12 Quarters
    qr = data.get("quarterly_rollup", {})
    if qr.get("quarters"):
        r = write_section(r, "Next 12 Quarters")
        headers = ["Revenue Source"] + qr["quarters"]
        rows = [[row["label"]] + row["values"] for row in qr.get("rows", [])]
        r = write_table(r, headers, rows, qr.get("totals", []))

    ws.column_dimensions["A"].width = 36
    for ci in range(2, 50):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(ci)].width = 11

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    from flask import send_file
    return send_file(output, as_attachment=True,
                     download_name="Ember_Operating_Revenues.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ─────────────────────────────────────────────────────────────────────────────
# Operating Cashflows — page + PDF report (Concept C · Pivot Grid)
#
# Reads the latest reports[operations] blob (per-project × per-category × 18
# months, anchored to the current period) and shapes it for both the live
# page and the WeasyPrint PDF. Math mirrors the JS prototype in the design
# canvas (operations-data.jsx) so the live page matches the design ref.
# ─────────────────────────────────────────────────────────────────────────────

_OPS_CATS = [
    "Development Fees",
    "Project Personnel",
    "Bookkeeping",
    "Receivables & Bond Fees",
    "EB Fees (Lots)",
    "EB Fees (Pods & Commercial)",
]
_OPS_CAT_COLORS = {
    "Development Fees":             "#F25929",
    "Project Personnel":            "#08233B",
    "Bookkeeping":                  "#5B9BD5",
    "Receivables & Bond Fees":      "#7E5BA6",
    "EB Fees (Lots)":               "#1F7A4D",
    "EB Fees (Pods & Commercial)":  "#C8A96E",
}
_OPS_MONTHS_SHORT = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]


def _ops_normalize(raw, anchor):
    """Canonical shape: {anchor, projects, monthly{project: {category: [18 floats]}}}"""
    projects = raw.get("projects") or []
    monthly  = raw.get("monthly")  or {}
    for p in projects:
        monthly.setdefault(p, {})
        for c in _OPS_CATS:
            arr = monthly[p].get(c) or []
            if len(arr) < 18:
                arr = list(arr) + [0.0] * (18 - len(arr))
            elif len(arr) > 18:
                arr = arr[:18]
            monthly[p][c] = [float(v or 0) for v in arr]
    out = {"anchor": anchor, "projects": list(projects), "monthly": monthly}
    history = raw.get("history")
    if isinstance(history, dict):
        out["history"] = {str(k): float(v or 0) for k, v in history.items()}
    return out


def _ops_month_dates(anchor):
    """18 months: anchor-3 .. anchor+14 (anchor at index 3)."""
    out = []
    for off in range(-3, 15):
        m_idx = anchor.month - 1 + off
        y = anchor.year + (m_idx // 12)
        m = (m_idx % 12) + 1
        out.append({
            "short": _OPS_MONTHS_SHORT[m - 1],
            "yy":    str(y)[-2:],
            "year":  y,
            "iso":   f"{y}-{m:02d}-01",
        })
    return out


def _ops_year_month(s):
    """Best-effort 'YYYY-MM' extractor for the legacy parser's date strings.

    The Excel parser writes one of: ISO date ('2024-01-01'), full ISO
    timestamp, or a Python str() of a date. Pull out year+month from
    any of those without crashing.
    """
    if s is None:
        return None
    s = str(s)
    # Try ISO date first (handles '2024-01-01' and '2024-01-01T00:00:00' forms)
    try:
        d = datetime.date.fromisoformat(s[:10])
        return f"{d.year:04d}-{d.month:02d}"
    except (ValueError, TypeError):
        pass
    try:
        d = datetime.datetime.fromisoformat(s[:19])
        return f"{d.year:04d}-{d.month:02d}"
    except (ValueError, TypeError):
        pass
    # Last-resort: first 7 chars look like YYYY-MM?
    if len(s) >= 7 and s[4:5] == '-':
        return s[:7]
    return None


def _ops_translate_legacy(raw):
    """Translate the legacy uploaded-Excel shape to the new format the
    redesigned /operations page expects.

    Legacy shape (from report_parser._parse_operations):
        {
          "kpis": [...],
          "yearly_rollup": {...},
          "monthly": {
              "dates": ["2024-01-01", "2024-02-01", ...],   # full lifetime
              "rows":  [{"project": "...", "category": "...",
                         "values": [...]}, ...],
              "totals": [...],
          },
          ...
        }

    New shape (consumed by _build_operations_view_context):
        {
          "anchor": "YYYY-MM-01",
          "projects": ["..."],
          "monthly": {
              "<project>": { "<category>": [18 floats], ... },
              ...
          },
        }

    We anchor on today's month (1st) and slice an 18-month window
    (anchor-3 .. anchor+14). Months that fall outside the legacy
    series are filled with 0.
    """
    if not isinstance(raw, dict):
        return None
    legacy_monthly = raw.get("monthly") or {}
    rows  = legacy_monthly.get("rows")  or []
    dates = legacy_monthly.get("dates") or []
    if not rows or not dates:
        return None

    # Build {project: {category: {YYYY-MM: value}}} so the window slice
    # below is a series of dict lookups regardless of how the legacy
    # date series was sized.
    by_pc = {}
    project_order = []
    for row in rows:
        p = row.get("project")
        c = row.get("category")
        v = row.get("values") or []
        if not p or not c:
            continue
        if p not in by_pc:
            project_order.append(p)
            by_pc[p] = {}
        if c not in by_pc[p]:
            by_pc[p][c] = {}
        for i, val in enumerate(v):
            if i >= len(dates):
                break
            ym = _ops_year_month(dates[i])
            if not ym:
                continue
            by_pc[p][c][ym] = float(val or 0)

    if not project_order:
        return None

    # Anchor on today's month; slice 18 months (-3..+14).
    anchor = datetime.date.today().replace(day=1)
    monthly_out = {}
    for p in project_order:
        monthly_out[p] = {}
        for c in _OPS_CATS:
            month_map = by_pc[p].get(c, {})
            values = []
            for off in range(-3, 15):
                m_idx = anchor.month - 1 + off
                yr = anchor.year + (m_idx // 12)
                mo = (m_idx % 12) + 1
                values.append(month_map.get(f"{yr:04d}-{mo:02d}", 0.0))
            monthly_out[p][c] = values

    # Full-history monthly grand totals (sum across projects/cats per
    # YYYY-MM). The chart window only carries 3 months of past data, so
    # KPIs needing a wider lookback (Trailing 12 Months, Realized YTD
    # past April, FY actuals) read from this dict instead.
    history = {}
    for cats in by_pc.values():
        for ym_map in cats.values():
            for ym, v in ym_map.items():
                history[ym] = history.get(ym, 0.0) + v

    return {
        "anchor":   anchor.isoformat(),
        "projects": project_order,
        "monthly":  monthly_out,
        "history":  history,
    }


def _ops_fmtM(v):
    return f"${v / 1_000_000:.2f}M"


def _ops_kpis(monthly, projects, anchor, history=None):
    grand = [0.0] * 18
    for p in projects:
        for c in _OPS_CATS:
            for i, v in enumerate(monthly[p][c]):
                grand[i] += v
    fy_year = anchor.year

    def _shift(year, month, off):
        idx = (month - 1) + off
        return year + (idx // 12), (idx % 12) + 1

    def _hist_sum(start_y, start_m, n):
        if not history:
            return None
        total = 0.0
        for k in range(n):
            y, m = _shift(start_y, start_m, k)
            total += history.get(f"{y:04d}-{m:02d}", 0.0)
        return total

    # Realized YTD: Jan..anchor of fy_year. The window only holds 3 past
    # months, so once we're past April the history dict is the only way
    # to get an accurate Jan-onward total.
    ytd_hist = _hist_sum(fy_year, 1, anchor.month)
    if ytd_hist is not None:
        realized_ytd = ytd_hist
    else:
        realized_ytd = (
            sum(grand[3 - anchor.month + 1 : 4]) if anchor.month <= 4
            else sum(grand[max(0, 4 - anchor.month) : 4])
        )

    # FY total: realized YTD + remaining months of fy_year from forecast window.
    fy_total = realized_ytd + sum(grand[4 : 4 + (12 - anchor.month)])

    # Trailing 12: 12 months ending at anchor (inclusive).
    t_y, t_m = _shift(fy_year, anchor.month, -11)
    trailing12_hist = _hist_sum(t_y, t_m, 12)
    if trailing12_hist is not None:
        trailing12 = trailing12_hist
        trailing12_sub = (
            f"{_OPS_MONTHS_SHORT[t_m - 1]} {t_y} – "
            f"{_OPS_MONTHS_SHORT[anchor.month - 1]} {fy_year}"
        )
    else:
        trailing12 = sum(grand[0:4])
        trailing12_sub = "(window proxy — 4 past months)"

    # Next 12: month after anchor through 12 months later.
    next12 = sum(grand[4 : 16])
    n_start_y, n_start_m = _shift(fy_year, anchor.month, 1)
    n_end_y,   n_end_m   = _shift(fy_year, anchor.month, 12)
    next12_sub = (
        f"{_OPS_MONTHS_SHORT[n_start_m - 1]} {n_start_y} – "
        f"{_OPS_MONTHS_SHORT[n_end_m - 1]} {n_end_y}"
    )

    return [
        {"label": f"FY {fy_year} Forecast", "val": _ops_fmtM(fy_total),
         "sub": "across all projects"},
        {"label": "Realized YTD", "val": _ops_fmtM(realized_ytd),
         "sub": f"Jan – {_OPS_MONTHS_SHORT[anchor.month - 1]} {fy_year}"},
        {"label": "Trailing 12 Months", "val": _ops_fmtM(trailing12),
         "sub": trailing12_sub},
        {"label": "Expected Next 12 Months", "val": _ops_fmtM(next12),
         "sub": next12_sub},
    ]


def _ops_sparkline(values, color, bold=False, now_offset=None):
    """Server-rendered SVG sparkline (88×22)."""
    w, h, pad = 88, 22, 2
    if not values: return ""
    vmax, vmin = max(values), min(values)
    rng = (vmax - vmin) or 1
    step = (w - pad * 2) / max(1, len(values) - 1)
    parts = []
    for i, v in enumerate(values):
        x = pad + i * step
        y = pad + (1 - (v - vmin) / rng) * (h - pad * 2)
        parts.append(f"{'M' if i == 0 else 'L'} {x:.1f} {y:.1f}")
    d = " ".join(parts)
    dot = ""
    if now_offset is not None and 0 <= now_offset < len(values):
        nx = pad + now_offset * step
        ny = pad + (1 - (values[now_offset] - vmin) / rng) * (h - pad * 2)
        dot = f'<circle cx="{nx:.1f}" cy="{ny:.1f}" r="1.8" fill="{color}"/>'
    return (
        f'<svg width="{w}" height="{h}" viewBox="0 0 {w} {h}">'
        f'<path d="{d}" fill="none" stroke="{color}" '
        f'stroke-width="{1.6 if bold else 1.2}" opacity="{1 if bold else 0.85}"/>'
        f'{dot}</svg>'
    )


def _build_operations_view_context(raw_data, uploaded_at):
    """Build the dict the new operations.html template consumes.

    Accepts two input shapes — auto-detects which one is on disk:
      • New shape: {anchor, projects, monthly{project: {category: [18 vals]}}}
      • Legacy shape (from the existing Excel uploader / report_parser):
        {monthly: {dates, rows, totals}, kpis, yearly_rollup, ...}
    Legacy data is reshaped in-memory through `_ops_translate_legacy`
    so the user can keep uploading the same Excel without a parser
    rewrite.
    """
    if not raw_data:
        return None
    # Detect legacy: a dict-shaped `monthly` with `rows` + `dates` keys.
    legacy_monthly = raw_data.get("monthly")
    if (isinstance(legacy_monthly, dict)
        and "rows"  in legacy_monthly
        and "dates" in legacy_monthly
        and "anchor" not in raw_data):
        translated = _ops_translate_legacy(raw_data)
        if translated is None:
            return None
        raw_data = translated
    anchor = raw_data.get("anchor")
    if isinstance(anchor, str):
        try: anchor = datetime.date.fromisoformat(anchor[:10])
        except Exception: anchor = None
    if anchor is None:
        if uploaded_at:
            anchor = uploaded_at.date().replace(day=1)
        else:
            anchor = datetime.date.today().replace(day=1)

    data = _ops_normalize(raw_data, anchor)
    projects = data["projects"]
    monthly  = data["monthly"]
    if not projects:
        return None
    month_dates = _ops_month_dates(anchor)
    now_idx = 3

    by_cat = {c: [0.0] * 18 for c in _OPS_CATS}
    for p in projects:
        for c in _OPS_CATS:
            for i, v in enumerate(monthly[p][c]):
                by_cat[c][i] += v

    proj_totals = {p: [0.0] * 18 for p in projects}
    for p in projects:
        for c in _OPS_CATS:
            for i, v in enumerate(monthly[p][c]):
                proj_totals[p][i] += v

    grand = [0.0] * 18
    for c in _OPS_CATS:
        for i, v in enumerate(by_cat[c]):
            grand[i] += v

    kpis = _ops_kpis(monthly, projects, anchor, history=data.get("history"))

    default_from, default_to = 0, 17
    window_months = month_dates[default_from : default_to + 1]
    window_now    = now_idx - default_from
    window_grand  = grand[default_from : default_to + 1]

    pivot_rows_default = []
    for p in projects:
        series = proj_totals[p][default_from : default_to + 1]
        total_k = round(sum(series) / 1000)
        spark = _ops_sparkline(series, "#F25929", bold=False, now_offset=window_now)
        pivot_rows_default.append({
            "label":     p,
            "series":    series,
            "total_k":   total_k,
            "spark_svg": spark,
        })
    total_spark = _ops_sparkline(window_grand, "#08233B", bold=True, now_offset=window_now)

    client_json = json.dumps({
        "cats":       _OPS_CATS,
        "projects":   projects,
        "cat_colors": _OPS_CAT_COLORS,
        "month_dates": month_dates,
        "now_idx":    now_idx,
        "monthly_totals_by_cat":  by_cat,
        "project_monthly_totals": proj_totals,
        "monthly_grand_total":    grand,
        "default_from_idx":       default_from,
        "default_to_idx":         default_to,
    }, separators=(",", ":"))

    period_label = f"{_OPS_MONTHS_SHORT[anchor.month - 1]} {anchor.year}"
    return {
        "cats":         _OPS_CATS,
        "kpis":         kpis,
        "month_dates":  month_dates,
        "default_from_idx":     default_from,
        "default_to_idx":       default_to,
        "default_window_count": default_to - default_from + 1,
        "now_idx":      now_idx,
        "window_months":      window_months,
        "window_now_idx":     window_now,
        "window_grand_total": window_grand,
        "window_total_k":     round(sum(window_grand) / 1000),
        "pivot_rows_default": pivot_rows_default,
        "total_spark_svg":    total_spark,
        "period_label":       period_label,
        "period_short":       f"FY {anchor.year} · Q{(anchor.month - 1) // 3 + 1}",
        "updated_at_short":   datetime.datetime.now().strftime("%b %d %Y"),
        "client_json":        client_json,
        # Internal — used by the report builder, not the live template
        "_anchor":   anchor,
        "_monthly":  monthly,
        "_projects": projects,
        "_by_cat":   by_cat,
        "_grand":    grand,
        "_month_dates": month_dates,
    }


def _ops_pie_donut_svg(slices, total, size=170, stroke_w=26):
    import math as _math
    r = size / 2 - 18
    cx = cy = size / 2
    C = 2 * _math.pi * r
    if total <= 0:
        total = 1
    parts = [
        f'<svg width="{size}" height="{size}" viewBox="0 0 {size} {size}" style="flex:0 0 auto">',
        f'<circle cx="{cx}" cy="{cy}" r="{r}" fill="none" stroke="rgba(8,35,59,0.10)" stroke-width="{stroke_w}"/>',
    ]
    acc = 0.0
    for s in slices:
        seg = (s["value"] / total) * C
        parts.append(
            f'<circle cx="{cx}" cy="{cy}" r="{r}" fill="none" '
            f'stroke="{s["color"]}" stroke-width="{stroke_w}" '
            f'stroke-dasharray="{seg:.2f} {(C - seg):.2f}" '
            f'stroke-dashoffset="{-acc:.2f}" '
            f'transform="rotate(-90 {cx} {cy})"/>'
        )
        acc += seg
    parts.append("</svg>")
    return "".join(parts)


def _build_operations_report_context(view_ctx, run_date=None):
    """Pre-renders the Category × Month pivot, donut SVG, KPIs for the PDF."""
    anchor   = view_ctx["_anchor"]
    by_cat   = view_ctx["_by_cat"]
    projects = view_ctx["_projects"]
    months   = view_ctx["_month_dates"]
    now_idx  = view_ctx["now_idx"]

    month_totals = [sum(by_cat[c][i] for c in _OPS_CATS) for i in range(18)]
    cat_rows = [{
        "name":    c,
        "color":   _OPS_CAT_COLORS[c],
        "row_k":   [int(round(v / 1000)) for v in by_cat[c]],
        "total_k": int(round(sum(by_cat[c]) / 1000)),
    } for c in _OPS_CATS]
    grand_total_k  = int(round(sum(month_totals) / 1000))
    month_totals_k = [int(round(v / 1000)) for v in month_totals]

    next12_slices = [
        {"name": c, "color": _OPS_CAT_COLORS[c], "value": sum(by_cat[c][4:16])}
        for c in _OPS_CATS
    ]
    next12_total = sum(s["value"] for s in next12_slices)
    pie_svg = _ops_pie_donut_svg(next12_slices, next12_total)
    pie_legend = sorted(
        [{"name": s["name"], "color": s["color"],
          "value_k": int(round(s["value"] / 1000)),
          "pct": int(round(s["value"] / next12_total * 100)) if next12_total else 0}
         for s in next12_slices],
        key=lambda x: x["value_k"], reverse=True,
    )
    next12_label = "{} {} — {} {}".format(
        _OPS_MONTHS_SHORT[anchor.month % 12], anchor.year,
        _OPS_MONTHS_SHORT[(anchor.month - 1) % 12], anchor.year + 1,
    ).upper()

    if next12_total > 0:
        top_cat = max(next12_slices, key=lambda s: s["value"])
        recurring_total = sum(
            sum(by_cat[c][4:16])
            for c in ("Development Fees", "Project Personnel", "Bookkeeping")
        )
        mix_note = (
            f"<strong>{top_cat['name']}</strong> is the largest forecast revenue "
            f"source over the next twelve months at "
            f"${top_cat['value'] / 1_000_000:.2f}M "
            f"({int(round(top_cat['value'] / next12_total * 100))}% of total). "
            f"Recurring fees — Development, Personnel, Bookkeeping — provide a "
            f"stable base of roughly ${recurring_total / 1_000_000:.1f}M annually. "
            "Receivables &amp; Bond Fees lump on quarter-end months."
        )
    else:
        mix_note = "No forward forecast yet."

    generated = (run_date or datetime.datetime.now()).strftime("%Y-%m-%d")
    period_label = (
        f"{['January','February','March','April','May','June','July','August','September','October','November','December'][anchor.month - 1]}"
        f" {anchor.year}"
    )

    return {
        "rpt": {
            "period_label":   period_label,
            "generated_iso":  generated,
            "project_count":  len(projects),
            "cat_count":      len(_OPS_CATS),
            "months":         months,
            "now_idx":        now_idx,
            "kpis":           view_ctx["kpis"],
            "cat_rows":       cat_rows,
            "month_totals_k": month_totals_k,
            "grand_total_k":  grand_total_k,
            "pie_svg":        pie_svg,
            "pie_legend":     pie_legend,
            "next12_label":   next12_label,
            "mix_note":       mix_note,
        }
    }


# Jinja filter — used by templates/_partials/_operations_pivot.html
@app.template_filter("int_comma")
def _jinja_int_comma(v):
    if v is None or v == 0:
        return "0"
    try:
        return f"{int(round(float(v))):,}"
    except (TypeError, ValueError):
        return "0"


@app.route("/operations")
@login_required
def operations_report():
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("operations", True):
        return redirect(url_for("home"))
    pa.setdefault("operations", True)
    pa.setdefault("reports",    True)

    conn = get_db(); cur = conn.cursor()
    cur.execute(
        "SELECT data, uploaded_at FROM reports "
        "WHERE report_type = 'operations' ORDER BY uploaded_at DESC LIMIT 1"
    )
    row = cur.fetchone(); cur.close(); conn.close()

    raw_data    = row["data"]        if row else None
    uploaded_at = row["uploaded_at"] if row else None
    ops_ctx = _build_operations_view_context(raw_data, uploaded_at)
    # Data dates footer: "Data From" from cell D1 on the Operations tab.
    data_from       = _fmt_data_date(raw_data.get("data_from") if raw_data else None)
    uploaded_at_fmt = _fmt_data_date(uploaded_at)
    return render_template(
        "operations.html",
        ops=ops_ctx,
        data_from=data_from, uploaded_at=uploaded_at_fmt,
        is_admin=session.get("is_admin", False),
        page_access=pa,
    )


@app.route("/operations/pdf")
@login_required
def operations_pdf_view():
    """3-page landscape PDF version of the Operating Cashflows page.

    ?preview=1   → return raw HTML (debug)
    ?download=1  → force attachment Content-Disposition
    Falls back to the legacy fpdf2 export when WeasyPrint can't load.
    """
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("operations", True):
        return jsonify({"error": "Access denied"}), 403
    if not session.get("is_admin") and not pa.get("reports", True):
        return jsonify({"error": "Reports access disabled by admin"}), 403

    conn = get_db(); cur = conn.cursor()
    cur.execute(
        "SELECT data, uploaded_at FROM reports "
        "WHERE report_type = 'operations' ORDER BY uploaded_at DESC LIMIT 1"
    )
    row = cur.fetchone(); cur.close(); conn.close()
    if not row or not row.get("data"):
        return jsonify({"error": "No operations forecast uploaded"}), 404

    view_ctx = _build_operations_view_context(row["data"], row.get("uploaded_at"))
    if view_ctx is None:
        return jsonify({"error": "Operations data is empty"}), 404
    rpt_ctx = _build_operations_report_context(view_ctx, run_date=row.get("uploaded_at"))
    html = render_template("operations_report.html", **rpt_ctx)

    if request.args.get("preview") == "1":
        return html

    try:
        from weasyprint import HTML
        pdf_bytes = HTML(
            string=html,
            base_url=request.host_url,
            url_fetcher=_weasyprint_local_fetcher,
        ).write_pdf()
    except (ImportError, OSError) as e:
        app.logger.warning(
            "WeasyPrint unavailable for operations report (%s: %s); falling back to fpdf2",
            type(e).__name__, e,
        )
        return _send_exec_report_pdf("operations")

    as_attachment = request.args.get("download") in ("1", "true", "yes")
    fname = f"Ember_Operations_{datetime.datetime.now().strftime('%Y-%m')}.pdf"
    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=as_attachment,
        download_name=fname,
    )


@app.route("/operations/excel")
@login_required
def operations_excel_view():
    """Alias to the existing /api/export-operations-excel workbook so the
    cockpit's "Export Excel" button has a clean URL alongside /operations/pdf."""
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("operations", True):
        return jsonify({"error": "Access denied"}), 403
    if not session.get("is_admin") and not pa.get("reports", True):
        return jsonify({"error": "Reports access disabled by admin"}), 403
    return export_operations_excel()


# ─────────────────────────────────────────────────────────────────────────────
# Loans & Debt — page + PDF report (Concept B · Hybrid Stacks)
#
# Pulls the latest reports[loans] blob and reshapes it for both the
# redesigned /loans page and the 5-page WeasyPrint executive report.
# Health-class thresholds (term / IR / capacity) drive the dial colors,
# stripe colors, and pill colors in lockstep.
# ─────────────────────────────────────────────────────────────────────────────

# ---------------------------------------------------------------------------
# Format helpers
# ---------------------------------------------------------------------------
def _loans_fmt_money_short(v):
    if v is None:
        return "—"
    try: v = float(v)
    except (TypeError, ValueError): return "—"
    if abs(v) >= 1_000_000:
        return f"${v / 1_000_000:.1f}M"
    if abs(v) >= 1_000:
        return f"${round(v / 1_000)}K"
    return f"${round(v)}"


def _loans_fmt_money_m(v, places=2):
    if v is None:
        return "—"
    try: return f"${float(v) / 1_000_000:.{places}f}M"
    except (TypeError, ValueError): return "—"


def _loans_fmt_pct(v, places=2):
    if v is None:
        return "—"
    try: return f"{float(v) * 100:.{places}f}%"
    except (TypeError, ValueError): return "—"


def _loans_fmt_pct1(v):
    return _loans_fmt_pct(v, 1)


def _loans_fmt_date_short(iso):
    """2027-03-15 → 'Mar 27'."""
    if not iso:
        return "—"
    try:
        d = datetime.date.fromisoformat(str(iso)[:10])
    except Exception:
        return str(iso)
    return d.strftime("%b %y")


# ---------------------------------------------------------------------------
# Health-class thresholds (shared with the design ref + the PDF)
# ---------------------------------------------------------------------------
def _loans_term_class(months_remaining):
    if months_remaining is None: return "na"
    try: m = float(months_remaining)
    except (TypeError, ValueError): return "na"
    if m < 12: return "bad"
    if m < 18: return "warn"
    return "good"


def _loans_ir_class(ir_health):
    if ir_health is None: return "na"
    try: ih = float(ir_health)
    except (TypeError, ValueError): return "na"
    if ih < 0.5:  return "bad"
    if ih < 0.75: return "warn"
    return "good"


def _loans_cap_class(cap_health):
    if cap_health is None: return "na"
    try: ch = float(cap_health)
    except (TypeError, ValueError): return "na"
    if ch < 1.0: return "bad"
    if ch < 1.1: return "warn"
    return "good"


# ---------------------------------------------------------------------------
# Legacy adapter — translates the report_parser._parse_loans shape
# ({mpc_loans: {headers, rows, totals}, vertical_loans: {...},
#   debt_schedules: [...]}) into the canonical shape the new templates
# expect ({mpcLoans: [...], mpcTotals: {...}, verticalLoans: [...],
# verticalTotals: {...}, portfolio: {...}, debtSchedules: [...]}).
# ---------------------------------------------------------------------------
_LOANS_HEADER_TO_KEY = {
    "Community":            "community",
    "Lender":               "lender",
    "Collateral":           "collateral",
    "Recourse":             "recourse",
    "Loan Origination":     "origination",
    "Loan Term Date":       "termDate",
    "Months Remaining":     "monthsRemaining",
    "Rem. Interest Reserve":"remIR",
    "Monthly Interest Burn":"monthlyBurn",
    "Remaining Mos. of IR": "remMosIR",
    "IR Health":            "irHealth",
    "Index + Spread":       "indexSpread",
    "Today's Rate":         "todayRate",
    "Extensions Remaining": "extensionsRem",
    "Extension Cost":       "extensionCost",
    "Loan Amount":          "loanAmount",
    "Drawn":                "drawn",
    "Balance":              "balance",
    "Utilization":          "utilization",
    "Remaining":            "remaining",
    "Forecasted Thru Term": "forecastedThruTerm",
    "Capacity Health":      "capacityHealth",
}


def _loans_translate_legacy(raw):
    """Map the report_parser legacy shape onto the new design's contract.

    Legacy:
      {
        "mpc_loans":      {"headers": [...], "rows": [...], "totals": {...}},
        "vertical_loans": {"headers": [...], "rows": [...], "totals": {...}, "footnote": "..."},
        "debt_schedules": [...]
      }
    New shape (matches `server/sample_data.json` from the handoff):
      {
        "anchor": "YYYY-MM-01",
        "mpcLoans": [...], "mpcTotals": {...},
        "verticalLoans": [...], "verticalTotals": {...},
        "portfolio": {...},
        "debtSchedules": [...],
        "months": [12 ISO YYYY-MM strings],
      }
    """
    if not isinstance(raw, dict):
        return None
    if "mpcLoans" in raw or "mpc_loans" not in raw:
        # Either already in new shape or unrecognized — return as-is.
        return raw

    def _coerce_health(v):
        """Robustly coerce an IR/Capacity Health cell value to a float ratio.

        The legacy parser stores these as strings via _str(), so we get
        e.g. "1.26", "0.85", or sometimes percentage-formatted text like
        "126%" / "85%", a label like "Healthy", an empty cell ("" / None
        / "—"), or a leftover formula string starting with "=" if the
        workbook never cached its values. Normalize:

          numeric         → use as-is
          str "126%"      → strip "%", divide by 100 → 1.26
          str "1.26"      → 1.26
          str "=A1/B1"    → None (uncached formula; show as missing)
          str label       → return label as-is (e.g. "Healthy")
          empty / None    → None
        """
        if v is None: return None
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip()
        if s in ("", "—", "-"):
            return None
        if s.startswith("="):  # uncached formula
            return None
        is_pct = s.endswith("%")
        cleaned = s.replace("%", "").replace("$", "").replace(",", "").strip()
        try:
            n = float(cleaned)
        except (TypeError, ValueError):
            return s  # opaque label — keep for display
        return n / 100.0 if is_pct else n

    def _row_to_obj(row):
        """Map a header-keyed row dict to camelCase keys."""
        out = {}
        for k, v in (row or {}).items():
            new_key = _LOANS_HEADER_TO_KEY.get(k)
            if new_key:
                if new_key in ("irHealth", "capacityHealth"):
                    v = _coerce_health(v)
                out[new_key] = v
        return out

    mpc_loans  = [_row_to_obj(r) for r in (raw.get("mpc_loans",      {}) or {}).get("rows") or []]
    vert_loans = [_row_to_obj(r) for r in (raw.get("vertical_loans", {}) or {}).get("rows") or []]
    mpc_tot    = _row_to_obj((raw.get("mpc_loans",      {}) or {}).get("totals") or {})
    vert_tot   = _row_to_obj((raw.get("vertical_loans", {}) or {}).get("totals") or {})

    # Roll up portfolio KPIs from the loan lists. Legacy data didn't
    # carry these, so we derive: outstanding = sum(balance), committed =
    # sum(loanAmount), weighted-avg-rate = balance-weighted, term-at-risk
    # count = how many have <12 months left.
    def _f(v, default=0.0):
        """Coerce a possibly-string field to float; non-numeric → default."""
        if v is None: return default
        try: return float(v)
        except (TypeError, ValueError): return default

    all_loans = mpc_loans + vert_loans
    total_outstanding = sum(_f(l.get("balance"))    for l in all_loans)
    total_committed   = sum(_f(l.get("loanAmount")) for l in all_loans)
    total_remaining   = sum(_f(l.get("remaining"))  for l in all_loans)
    weighted_rate_num = sum(_f(l.get("todayRate")) * _f(l.get("balance")) for l in all_loans)
    weighted_rate     = (weighted_rate_num / total_outstanding) if total_outstanding else 0.0
    monthly_burn      = sum(_f(l.get("monthlyBurn")) for l in mpc_loans)
    # Risk counts: use a sentinel that DOESN'T trigger the threshold so
    # missing/non-numeric values don't false-positive.
    term_expiring = sum(1 for l in all_loans if _f(l.get("monthsRemaining"),  default=999) < 12)
    ir_at_risk    = sum(1 for l in mpc_loans if _f(l.get("irHealth"),         default=999) < 0.5)
    cap_at_risk   = sum(1 for l in all_loans if _f(l.get("capacityHealth"),   default=999) < 1.0)

    # Translate debt-schedule shape: legacy parser uses snake_case keys
    # (cumulative_revenues / cumulative_payments) but the renderers
    # expect camelCase. Without this fix the cumulative-coverage row
    # and the page-5 coverage curve silently render empty.
    raw_scheds = raw.get("debt_schedules") or []
    schedules = []
    for s in raw_scheds:
        s2 = dict(s)
        if "cumulative_revenues" in s2 and "cumulativeRevenues" not in s2:
            s2["cumulativeRevenues"] = s2["cumulative_revenues"]
        if "cumulative_payments" in s2 and "cumulativePayments" not in s2:
            s2["cumulativePayments"] = s2["cumulative_payments"]
        schedules.append(s2)

    # Use the *schedule's* months (parsed from row 57 of the spreadsheet)
    # as the canonical month axis when present — that's what the
    # cumulative arrays line up to. Falls back to a generated next-12-
    # month list only when no schedule was uploaded.
    today = datetime.date.today().replace(day=1)
    months = []
    if schedules and schedules[0].get("months"):
        for d_iso in schedules[0]["months"]:
            # _date_iso gave us full ISO strings; normalize to YYYY-MM.
            s = str(d_iso)[:7]
            if len(s) == 7 and s[4] == "-":
                months.append(s)
    if not months:
        y, m = today.year, today.month
        for _ in range(12):
            m += 1
            if m > 12:
                m = 1; y += 1
            months.append(f"{y:04d}-{m:02d}")

    return {
        "anchor": today.isoformat(),
        "today":  datetime.date.today().isoformat(),
        "months": months,
        "portfolio": {
            "totalOutstanding":    int(round(total_outstanding)),
            "totalCommitted":      int(round(total_committed)),
            "totalRemaining":      int(round(total_remaining)),
            "weightedAvgRate":     round(weighted_rate, 4),
            "monthlyInterestBurn": int(round(monthly_burn)),
            "termExpiringCount":   int(term_expiring),
            "irAtRiskCount":       int(ir_at_risk),
            "capacityAtRiskCount": int(cap_at_risk),
        },
        "mpcLoans":      mpc_loans,
        "mpcTotals":     mpc_tot,
        "verticalLoans": vert_loans,
        "verticalTotals": vert_tot,
        "debtSchedules": schedules,
    }


# ---------------------------------------------------------------------------
# Per-loan + totals enrichment (formatted strings + health classes)
# ---------------------------------------------------------------------------
def _loans_enrich_loan(raw, kind):
    l = dict(raw)
    is_vert = kind == "vert"
    l["loanAmount_fmt"]    = _loans_fmt_money_short(l.get("loanAmount"))
    l["drawn_fmt"]         = _loans_fmt_money_short(l.get("drawn"))
    l["balance_fmt"]       = _loans_fmt_money_short(l.get("balance"))
    l["remaining_fmt"]     = _loans_fmt_money_short(l.get("remaining"))
    l["forecastedThruTerm_fmt"] = _loans_fmt_money_short(l.get("forecastedThruTerm"))
    l["extensionCost_fmt"] = _loans_fmt_money_short(l.get("extensionCost")) if l.get("extensionCost") else "—"
    l["utilization_fmt"]   = _loans_fmt_pct1(l.get("utilization"))
    try:    l["utilization_pct"] = round(float(l.get("utilization") or 0) * 100, 1)
    except (TypeError, ValueError): l["utilization_pct"] = 0
    l["todayRate_fmt"]     = _loans_fmt_pct(l.get("todayRate"), 2)
    try:    l["indexSpread_fmt"] = f"{float(l.get('indexSpread') or 0) * 100:.2f}%"
    except (TypeError, ValueError): l["indexSpread_fmt"] = "0.00%"
    l["termDate_short"]    = _loans_fmt_date_short(l.get("termDate"))
    l["origination_short"] = _loans_fmt_date_short(l.get("origination"))
    # IR + monthly-burn fields: ALWAYS compute from raw data, regardless
    # of whether the loan is MPC or vertical. The report's per-template
    # rendering may still suppress these for vertical lines (e.g. via
    # the {% if kind == 'mpc' %} branches in _loans_report_card.html),
    # but the page card surfaces whatever is in the spreadsheet so
    # zeros / partial fills are visible to the partner.
    ih = l.get("irHealth")
    if ih is None:
        l["irHealth_fmt"] = "—"
    elif isinstance(ih, (int, float)):
        l["irHealth_fmt"] = _loans_fmt_pct1(ih)
    else:
        s = str(ih).strip()
        try:
            l["irHealth_fmt"] = _loans_fmt_pct1(float(s))
        except (TypeError, ValueError):
            l["irHealth_fmt"] = s if s else "—"

    l["remIR_fmt"]       = _loans_fmt_money_short(l.get("remIR"))
    l["monthlyBurn_fmt"] = _loans_fmt_money_short(l.get("monthlyBurn"))
    try:
        rm = l.get("remMosIR")
        l["remMosIR_fmt"] = f"{float(rm):.1f}" if rm is not None else "—"
    except (TypeError, ValueError):
        l["remMosIR_fmt"] = "—"

    # Pre-format capacityHealth and monthsRemaining so the templates
    # don't crash on None (legacy parser sometimes stores `Capacity
    # Health` as a non-numeric string label, which our translator
    # coerces to None — `'%.2f'|format(None)` would TypeError).
    ch = l.get("capacityHealth")
    try:
        l["capacityHealth_fmt"] = f"{float(ch):.2f}×" if ch is not None else "—"
    except (TypeError, ValueError):
        # Fall back to the raw string label so the user sees the
        # spreadsheet's text rather than an unhelpful em-dash.
        s = str(ch).strip() if ch is not None else ""
        l["capacityHealth_fmt"] = s if s else "—"
    mr = l.get("monthsRemaining")
    try:
        l["monthsRemaining_fmt"] = f"{int(round(float(mr)))}" if mr is not None else "—"
    except (TypeError, ValueError):
        l["monthsRemaining_fmt"] = "—"

    l["term_cls"] = _loans_term_class(l.get("monthsRemaining"))
    l["ir_cls"]   = _loans_ir_class(l.get("irHealth")) if not is_vert else "na"
    l["cap_cls"]  = _loans_cap_class(l.get("capacityHealth"))
    return l


def _loans_enrich_totals(raw, kind):
    t = dict(raw or {})
    t["loanAmount_fmt"]    = _loans_fmt_money_short(t.get("loanAmount"))
    t["drawn_fmt"]         = _loans_fmt_money_short(t.get("drawn"))
    t["balance_fmt"]       = _loans_fmt_money_short(t.get("balance"))
    t["remaining_fmt"]     = _loans_fmt_money_short(t.get("remaining"))
    t["forecastedThruTerm_fmt"] = _loans_fmt_money_short(t.get("forecastedThruTerm"))
    t["utilization_fmt"]   = _loans_fmt_pct1(t.get("utilization"))
    t["todayRate_fmt"]     = _loans_fmt_pct(t.get("todayRate"), 2)
    t["remIR_fmt"]         = _loans_fmt_money_short(t.get("remIR")) if kind == "mpc" else "—"
    t["monthlyBurn_fmt"]   = _loans_fmt_money_short(t.get("monthlyBurn")) if kind == "mpc" else "—"
    return t


def _loans_normalize(raw, anchor):
    raw = _loans_translate_legacy(raw)
    if not raw:
        return None
    mpc_loans  = [_loans_enrich_loan(l, "mpc")  for l in (raw.get("mpcLoans")      or [])]
    vert_loans = [_loans_enrich_loan(l, "vert") for l in (raw.get("verticalLoans") or [])]
    return {
        "anchor":  anchor,
        "today":   raw.get("today"),
        "months":  raw.get("months") or [],
        "mpc":  {"loans": mpc_loans,  "totals": _loans_enrich_totals(raw.get("mpcTotals")      or {}, "mpc")},
        "vert": {"loans": vert_loans, "totals": _loans_enrich_totals(raw.get("verticalTotals") or {}, "vert")},
        "portfolio":     raw.get("portfolio") or {},
        "debt_schedules": raw.get("debtSchedules") or [],
    }


def _loans_kpis(data):
    p = data["portfolio"]
    return [
        {"label": "Outstanding", "val": _loans_fmt_money_m(p.get("totalOutstanding")),
         "sub": f"of {_loans_fmt_money_m(p.get('totalCommitted'))} capacity", "tone": ""},
        {"label": "Remaining Capacity", "val": _loans_fmt_money_m(p.get("totalRemaining")),
         "sub": "across all facilities", "tone": ""},
        {"label": "Wtd Avg Rate", "val": _loans_fmt_pct(p.get("weightedAvgRate"), 2),
         "sub": "on outstanding balance", "tone": ""},
        {"label": "Monthly IR Burn", "val": _loans_fmt_money_m(p.get("monthlyInterestBurn")),
         "sub": "MPC interest reserves", "tone": ""},
        {"label": "Term < 12 mo", "val": str(p.get("termExpiringCount") or 0),
         "sub": "requires action",
         "tone": "bad" if (p.get("termExpiringCount") or 0) > 0 else ""},
        {"label": "IR Runway < 50%", "val": str(p.get("irAtRiskCount") or 0),
         "sub": "below break-even",
         "tone": "warn" if (p.get("irAtRiskCount") or 0) > 0 else ""},
    ]


def _build_loans_view_context(raw_data, uploaded_at):
    """Build the dict the redesigned loans.html template consumes."""
    if not raw_data:
        return None
    # Anchor on `anchor` if the new shape carries one, else today's
    # 1st-of-month. The legacy adapter sets it itself.
    anchor = None
    if isinstance(raw_data, dict):
        a = raw_data.get("anchor")
        if isinstance(a, str):
            try: anchor = datetime.date.fromisoformat(a[:10])
            except Exception: anchor = None
    if anchor is None:
        if uploaded_at:
            anchor = uploaded_at.date().replace(day=1) if hasattr(uploaded_at, "date") else datetime.date.today().replace(day=1)
        else:
            anchor = datetime.date.today().replace(day=1)

    data = _loans_normalize(raw_data, anchor)
    if data is None or (not data["mpc"]["loans"] and not data["vert"]["loans"]):
        return None

    today_iso = data.get("today")
    try:
        updated = datetime.datetime.fromisoformat(today_iso) if today_iso else datetime.datetime.now()
    except Exception:
        updated = datetime.datetime.now()

    ctx = {
        "kpis":            _loans_kpis(data),
        "mpc":             data["mpc"],
        "vert":            data["vert"],
        "portfolio":       data["portfolio"],
        "debt_schedules":  data["debt_schedules"],
        "period_label":    anchor.strftime("%B %Y"),
        "period_short":    anchor.strftime("%b %Y").upper(),
        "updated_at_short": updated.strftime("%Y-%m-%d"),
        # Internal — used by the report builder
        "_anchor":   anchor,
        "_today":    data.get("today"),
        "_months":   data.get("months"),
    }
    return ctx


# ---------------------------------------------------------------------------
# PDF Report — server-rendered SVG widgets (capacity stack, maturity wall,
# coverage curve), per-project schedule rows, KPI band, and the 5-page
# layout in templates/loans_report.html.
# ---------------------------------------------------------------------------
_LOANS_REPORT_COLORS = {
    "ink":     "#08233B",
    "ink_2":   "#13344E",
    "muted":   "#5b6b7b",
    "subtle":  "#97a3b0",
    "line":    "rgba(8,35,59,0.10)",
    "line_2":  "rgba(8,35,59,0.20)",
    "warm":    "#F8F4EE",
    "accent":  "#F25929",
    "accent_soft": "rgba(242,89,41,0.10)",
    "good":    "#1F7A4D",
    "warn":    "#C9871F",
    "bad":     "#C0311A",
}
_LOANS_MONTHS_SHORT = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]


def _loans_month_idx(iso, start_year):
    parts = str(iso).split("-")
    y, m = int(parts[0]), int(parts[1])
    return (y - start_year) * 12 + (m - 1)


def _loans_short_month_label(iso):
    parts = str(iso).split("-")
    y, m = int(parts[0]), int(parts[1])
    return f"{_LOANS_MONTHS_SHORT[m - 1]} {str(y)[2:]}"


def _loans_render_capacity_stack(loans):
    """HTML rows — uses display:table classes (.capstack-row) defined in
    templates/loans_report.html so WeasyPrint paginates reliably.
    """
    if not loans:
        return ""
    max_loan = max((l.get("loanAmount") or 0) for l in loans) or 1
    rows = []
    for l in loans:
        drawn_pct = ((l.get("balance") or 0) / max_loan) * 100
        rem_pct   = ((l.get("remaining") or 0) / max_loan) * 100
        irh = l.get("irHealth")
        ircls = _loans_ir_class(irh) if irh is not None else None
        ir_dot = (
            f'<span class="dot" style="background:var(--{ircls})"></span>'
            if ircls and ircls != "na" else ""
        )
        rows.append(f"""
<div class="capstack-row">
  <span class="nm">
    <span class="name">{l.get('community','')}</span>
    <span class="ld">{l.get('lender','')}</span>
  </span>
  <span class="bar-cell">
    <div class="bar">
      <div class="drawn" style="width:{drawn_pct:.2f}%"></div>
      <div class="rem"   style="left:{drawn_pct:.2f}%;width:{rem_pct:.2f}%"></div>
      <div class="lbl">{_loans_fmt_money_short(l.get('balance'))}</div>
    </div>
  </span>
  <span class="util">{ir_dot}{_loans_fmt_pct1(l.get('utilization'))}</span>
</div>""")
    legend = """
<div class="capstack-legend">
  <div class="it"><span class="sw" style="background:var(--accent)"></span>Drawn</div>
  <div class="it"><span class="sw" style="background:var(--accent-soft);border:1px solid var(--accent)"></span>Remaining</div>
  <div class="it right">● IR Health</div>
</div>"""
    return "".join(rows) + legend


def _loans_render_maturity_wall(mpc_loans, vert_loans, today, start_year=2023, end_year=2029):
    """Origination → term timeline with today line and term-tone bars.

    Sized to the new design ref: 900×variable, 30px row height, 18px bar
    height, 120/18px paddings. The SVG fills its container via
    `width:100%; height:auto` in the report stylesheet.
    """
    all_loans = (
        [{**l, "kind": "MPC"} for l in mpc_loans] +
        [{**l, "kind": "VRT"} for l in vert_loans]
    )
    if not all_loans:
        return ""
    today_idx = _loans_month_idx(today.isoformat(), start_year)
    total = (end_year - start_year) * 12 or 1
    W, ROW = 900, 30
    PADL, PADR, PADT = 120, 18, 22
    H = PADT + len(all_loans) * ROW + 18
    inner_w = W - PADL - PADR
    C = _LOANS_REPORT_COLORS

    def x(m): return PADL + (m / total) * inner_w

    parts = [f'<svg viewBox="0 0 {W} {H}" preserveAspectRatio="xMidYMid meet">']
    # Year ticks
    for i in range(end_year - start_year + 1):
        xx = x(i * 12)
        dash = ' stroke-dasharray="2 3"' if i != 0 else ""
        parts.append(f'<line x1="{xx}" y1="{PADT-3}" x2="{xx}" y2="{H-14}" stroke="{C["line"]}"{dash}/>')
        parts.append(
            f'<text x="{xx}" y="{PADT-8}" text-anchor="middle" font-size="8" '
            f'fill="{C["subtle"]}" font-family="JetBrains Mono" font-weight="700">'
            f'{start_year + i}</text>'
        )
    # Today line
    parts.append(
        f'<line x1="{x(today_idx)}" y1="{PADT-3}" x2="{x(today_idx)}" y2="{H-14}" '
        f'stroke="{C["accent"]}" stroke-width="1.5"/>'
    )
    parts.append(
        f'<text x="{x(today_idx)}" y="{PADT-12}" text-anchor="middle" font-size="7.5" '
        f'fill="{C["accent"]}" font-weight="700" font-family="Plus Jakarta Sans" '
        f'letter-spacing="1.4" style="text-transform:uppercase">Today</text>'
    )
    # Loan rows
    for i, l in enumerate(all_loans):
        y = PADT + i * ROW + 6
        try:
            x0 = x(_loans_month_idx(l.get("origination") or today.isoformat(), start_year))
            x1 = x(_loans_month_idx(l.get("termDate")    or today.isoformat(), start_year))
        except Exception:
            continue
        bar_h = 18
        cls = _loans_term_class(l.get("monthsRemaining"))
        col = {"bad": C["bad"], "warn": C["warn"]}.get(cls, C["ink_2"])
        parts.append(
            f'<text x="{PADL-8}" y="{y+10}" text-anchor="end" font-size="9" '
            f'fill="{C["ink"]}" font-family="Plus Jakarta Sans" font-weight="600">'
            f'{l.get("community","")}</text>'
        )
        parts.append(
            f'<text x="{PADL-8}" y="{y+21}" text-anchor="end" font-size="7.5" '
            f'fill="{C["subtle"]}" font-family="JetBrains Mono">'
            f'{l.get("lender","")} · {l["kind"]}</text>'
        )
        past_end = min(x(today_idx), x1)
        parts.append(
            f'<rect x="{x0}" y="{y}" width="{past_end-x0}" height="{bar_h}" '
            f'fill="{C["ink_2"]}" fill-opacity="0.12" stroke="{C["line_2"]}"/>'
        )
        if x1 > x(today_idx):
            fx = max(x0, x(today_idx))
            opacity = 0.65 if cls == "bad" else 0.35
            parts.append(
                f'<rect x="{fx}" y="{y}" width="{x1-fx}" height="{bar_h}" '
                f'fill="{col}" fill-opacity="{opacity}" stroke="{col}"/>'
            )
        parts.append(f'<circle cx="{x1}" cy="{y+bar_h/2}" r="2.6" fill="{col}"/>')
        parts.append(
            f'<text x="{x1+5}" y="{y+bar_h/2+3}" font-size="8" '
            f'fill="{col}" font-weight="700" font-family="JetBrains Mono">'
            f'{_loans_fmt_date_short(l.get("termDate"))}</text>'
        )
        parts.append(
            f'<text x="{x0+6}" y="{y+bar_h/2+3}" font-size="8" '
            f'fill="{C["ink"]}" font-family="JetBrains Mono" font-weight="700">'
            f'{_loans_fmt_money_short(l.get("balance"))}</text>'
        )
    parts.append(f'<g transform="translate({PADL},{H-6})">')
    for i, (cls, lbl) in enumerate([("bad", "&lt;12 mo"), ("warn", "12–18 mo"), ("ink_2", "&gt;18 mo")]):
        ox = i * 94
        opacity = "0.65" if cls == "bad" else "0.35"
        parts.append(
            f'<rect x="{ox}" y="-5" width="9" height="6" fill="{C[cls]}" fill-opacity="{opacity}"/>'
            f'<text x="{ox+13}" y="0" font-size="6.5" fill="{C["muted"]}" '
            f'font-family="Plus Jakarta Sans" letter-spacing="1" '
            f'style="text-transform:uppercase">{lbl}</text>'
        )
    parts.append("</g></svg>")
    return "".join(parts)


def _loans_render_coverage_curve(scheds, months):
    if not scheds or not months:
        return ""
    n = len(months)
    cum_p = [sum((s.get("cumulativePayments") or [0]*n)[i] for s in scheds) for i in range(n)]
    cum_r = [sum((s.get("cumulativeRevenues") or [0]*n)[i] for s in scheds) for i in range(n)]
    if max(cum_p + cum_r) == 0:
        return ""
    W, H = 900, 160
    PADL, PADR, PADT, PADB = 46, 70, 12, 24
    iw, ih = W - PADL - PADR, H - PADT - PADB
    max_v = max(max(cum_p), max(cum_r)) * 1.05 or 1.0
    C = _LOANS_REPORT_COLORS

    def xp(i):  return PADL + (i / max(1, n - 1)) * iw
    def yp(v):  return PADT + ih - (v / max_v) * ih
    def path(vals):
        return " ".join(f"{'M' if i == 0 else 'L'} {xp(i):.1f} {yp(v):.1f}" for i, v in enumerate(vals))
    def area(vals):
        return path(vals) + f" L {xp(n-1):.1f} {yp(0):.1f} L {xp(0):.1f} {yp(0):.1f} Z"

    parts = [f'<svg viewBox="0 0 {W} {H}" preserveAspectRatio="xMidYMid meet">']
    for p in (0, 0.25, 0.5, 0.75, 1):
        v, y = max_v * p, yp(max_v * p)
        op = "" if p == 0 else ' stroke-opacity="0.5"'
        parts.append(f'<line x1="{PADL}" y1="{y}" x2="{PADL+iw}" y2="{y}" stroke="{C["line"]}"{op}/>')
        lbl = f"{v/1e6:.1f}M" if v >= 1e6 else f"{round(v/1000)}K"
        parts.append(
            f'<text x="{PADL-5}" y="{y+2.5}" text-anchor="end" font-size="7" '
            f'fill="{C["subtle"]}" font-family="JetBrains Mono">{lbl}</text>'
        )
    parts.append(f'<path d="{area(cum_r)}" fill="{C["good"]}" fill-opacity="0.12"/>')
    parts.append(f'<path d="{path(cum_r)}" fill="none" stroke="{C["good"]}" stroke-width="1.8"/>')
    parts.append(f'<path d="{path(cum_p)}" fill="none" stroke="{C["accent"]}" stroke-width="1.8"/>')
    last = n - 1
    parts.append(f'<circle cx="{xp(last)}" cy="{yp(cum_r[last])}" r="3" fill="{C["good"]}"/>')
    parts.append(f'<circle cx="{xp(last)}" cy="{yp(cum_p[last])}" r="3" fill="{C["accent"]}"/>')
    parts.append(
        f'<text x="{xp(last)+5}" y="{yp(cum_r[last])+2.5}" font-size="8" '
        f'fill="{C["good"]}" font-weight="700" font-family="JetBrains Mono">'
        f'{_loans_fmt_money_m(cum_r[last])}</text>'
    )
    parts.append(
        f'<text x="{xp(last)+5}" y="{yp(cum_p[last])+2.5}" font-size="8" '
        f'fill="{C["accent"]}" font-weight="700" font-family="JetBrains Mono">'
        f'{_loans_fmt_money_m(cum_p[last])}</text>'
    )
    for i, m in enumerate(months):
        col = C["accent"] if i == 0 else C["subtle"]
        weight = ' font-weight="700"' if i == 0 else ''
        parts.append(
            f'<text x="{xp(i)}" y="{H-PADB+10}" text-anchor="middle" font-size="7" '
            f'fill="{col}" font-family="JetBrains Mono"{weight}>'
            f'{_loans_short_month_label(m)}</text>'
        )
    parts.append(f'<g transform="translate({PADL},{PADT-2})">')
    parts.append(
        f'<rect x="0" y="-5" width="8" height="2.5" fill="{C["good"]}"/>'
        f'<text x="12" y="-1.5" font-size="7" fill="{C["good"]}" '
        f'font-weight="600" letter-spacing="1" style="text-transform:uppercase">Cumulative Revenue</text>'
    )
    parts.append(
        f'<rect x="140" y="-5" width="8" height="2.5" fill="{C["accent"]}"/>'
        f'<text x="152" y="-1.5" font-size="7" fill="{C["accent"]}" '
        f'font-weight="600" letter-spacing="1" style="text-transform:uppercase">Cumulative Payments</text>'
    )
    parts.append("</g></svg>")
    return "".join(parts)


def _loans_coverage_cls(ratio):
    if ratio is None:    return "na"
    if ratio >= 1.5:     return "good"
    if ratio >= 1.0:     return "warn"
    return "bad"


def _loans_build_schedule(sch, months):
    n = len(months)
    monthly_pmts = []
    for m in months:
        amt = sum(p.get("amount", 0) for p in (sch.get("payments") or []) if str(p.get("date","")).startswith(m))
        monthly_pmts.append(amt)
    total_pmts = sum(monthly_pmts)
    total_rev  = sum(r.get("total", 0) for r in (sch.get("revenues") or []))
    monthly_rev = [
        sum((r.get("monthly") or [0]*n)[i] for r in (sch.get("revenues") or []))
        for i in range(n)
    ]
    coverage = (total_rev / total_pmts) if total_pmts else None
    cum_p = sch.get("cumulativePayments") or [0] * n
    cum_r = sch.get("cumulativeRevenues") or [0] * n
    cum_ratios = [
        {
            "ratio": (cum_r[i] / cum_p[i]) if cum_p[i] else None,
            "cls":   _loans_coverage_cls((cum_r[i] / cum_p[i]) if cum_p[i] else None),
        } for i in range(n)
    ]
    return {
        "project":      sch.get("project", ""),
        "month_labels": [_loans_short_month_label(m) for m in months],
        "revenues": [{
            "type":        r.get("type", ""),
            "pct_fmt":     f"{round((r.get('pct') or 0) * 100)}% of mix",
            "monthly_fmt": [_loans_fmt_money_short(v) if v else "—" for v in (r.get("monthly") or [0]*n)],
            "total_fmt":   _loans_fmt_money_short(r.get("total")),
        } for r in (sch.get("revenues") or [])],
        "monthly_rev_fmt": [_loans_fmt_money_short(v) for v in monthly_rev],
        "total_rev_fmt":   _loans_fmt_money_m(total_rev, 2),
        "monthly_pmts_fmt":[_loans_fmt_money_short(v) if v else "—" for v in monthly_pmts],
        "total_pmts_fmt":  _loans_fmt_money_m(total_pmts, 2),
        "cum_ratios": [{
            "ratio_fmt": f"{r['ratio']:.2f}×" if r["ratio"] is not None else "—",
            "cls": r["cls"],
        } for r in cum_ratios],
        "coverage_fmt":  f"{coverage:.2f}×" if coverage is not None else "—",
        "coverage_cls":  _loans_coverage_cls(coverage),
    }


def _build_loans_report_context(view_ctx, run_date=None):
    """Build the dict consumed by templates/loans_report.html."""
    anchor   = view_ctx["_anchor"]
    today    = view_ctx.get("_today")
    months   = view_ctx.get("_months") or []
    try:
        today_d = datetime.date.fromisoformat(str(today)[:10]) if today else anchor
    except Exception:
        today_d = anchor

    mpc = view_ctx["mpc"]
    vert = view_ctx["vert"]
    portfolio = view_ctx["portfolio"]
    scheds = view_ctx["debt_schedules"]

    # Capacity stack expects raw loan dicts (not enriched). Strip the
    # _fmt fields by keeping only the numeric ones.
    cap_stack = _loans_render_capacity_stack(list(mpc["loans"]) + list(vert["loans"]))
    mat_wall  = _loans_render_maturity_wall(mpc["loans"], vert["loans"], today_d)
    cov_curve = _loans_render_coverage_curve(scheds, months)
    schedules = [_loans_build_schedule(s, months) for s in scheds]

    p = portfolio
    overview_kpis = [
        {"label": "Outstanding", "val": _loans_fmt_money_m(p.get("totalOutstanding")),
         "sub": f"of {_loans_fmt_money_m(p.get('totalCommitted'))} capacity · "
                f"{_loans_fmt_pct1((p.get('totalOutstanding') or 0) / (p.get('totalCommitted') or 1))} drawn",
         "tone": ""},
        {"label": "Remaining Capacity", "val": _loans_fmt_money_m(p.get("totalRemaining")),
         "sub": "across all facilities", "tone": ""},
        {"label": "Wtd Avg Rate", "val": _loans_fmt_pct(p.get("weightedAvgRate"), 2),
         "sub": "on outstanding balance", "tone": ""},
        {"label": "Monthly IR Burn", "val": _loans_fmt_money_m(p.get("monthlyInterestBurn")),
         "sub": "MPC interest reserves", "tone": ""},
        {"label": "Term < 12 mo", "val": str(p.get("termExpiringCount") or 0),
         "sub": "requires action",
         "tone": "bad" if (p.get("termExpiringCount") or 0) > 0 else ""},
        {"label": "IR Runway < 50%", "val": str(p.get("irAtRiskCount") or 0),
         "sub": "below break-even",
         "tone": "warn" if (p.get("irAtRiskCount") or 0) > 0 else ""},
    ]
    run = run_date or datetime.datetime.now()

    return {
        "report": {
            "period_label":  anchor.strftime("%B %Y"),
            "generated":     run.strftime("%Y-%m-%d"),
            "active_facilities": len(mpc["loans"]) + len(vert["loans"]),
            "active_facilities_label": (
                f"{len(mpc['loans']) + len(vert['loans'])} — "
                f"{len(mpc['loans'])} MPC + {len(vert['loans'])} Vert"
            ),
            "outstanding_fmt": _loans_fmt_money_m(p.get("totalOutstanding")),
            "committed_fmt":   _loans_fmt_money_m(p.get("totalCommitted")),
            "kpis":            overview_kpis,
            "mpc":             mpc,
            "vert":            vert,
            "cap_stack_svg":   cap_stack,
            "mat_wall_svg":    mat_wall,
            "cov_curve_svg":   cov_curve,
            "schedules":       schedules,
            "months":          months,
        }
    }


# ─── REPORT GENERATORS ────────────────────────────────────────────────────────

def _gen_excel_loans(data):
    """Generate loans Excel workbook bytes from report data."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Loan Capacities"

    NAVY = "1A3A5C"
    GOLD = "C8A96E"
    HDR_FILL = PatternFill("solid", fgColor="1E2535")
    ALT_FILL = PatternFill("solid", fgColor="F7F8FA")
    thin = Side(style="thin", color="D0D5DD")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hf(bold=False, size=9): return Font(name="Calibri", size=size, bold=bold, color="8B95A8")
    def vf(bold=False, color="1A1A1A"): return Font(name="Calibri", size=9, bold=bold, color=color)
    def gf(size=12): return Font(name="Calibri", size=size, bold=True, color=GOLD)

    r = 1
    ws.cell(row=r, column=1, value="Loan Capacities & Debt Schedules").font = gf()
    r += 2

    def write_table(title, headers, rows, totals=None):
        nonlocal r
        ws.cell(row=r, column=1, value=title).font = Font(name="Calibri", size=11, bold=True, color=NAVY)
        r += 1
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=r, column=ci, value=h)
            c.font = hf(bold=True)
            c.fill = HDR_FILL
            c.border = bdr
            c.alignment = Alignment(horizontal="center" if ci > 1 else "left")
        r += 1
        for ri, row_data in enumerate(rows):
            fill = ALT_FILL if ri % 2 == 0 else PatternFill()
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=r, column=ci, value=val)
                c.font = vf()
                c.fill = fill
                c.border = bdr
                c.alignment = Alignment(horizontal="left" if ci == 1 else "right")
                if ci > 1 and isinstance(val, (int, float)):
                    c.number_format = "#,##0"
            r += 1
        if totals:
            tot_fill = PatternFill("solid", fgColor="E8EEF5")
            ws.cell(row=r, column=1, value="Total").font = vf(bold=True, color=NAVY)
            ws.cell(row=r, column=1).fill = tot_fill
            ws.cell(row=r, column=1).border = bdr
            for ci, val in enumerate(totals, 2):
                c = ws.cell(row=r, column=ci, value=val if val else None)
                c.font = vf(bold=True, color=NAVY)
                c.fill = tot_fill
                c.border = bdr
                c.alignment = Alignment(horizontal="right")
                if isinstance(val, (int, float)):
                    c.number_format = "#,##0"
            r += 1
        r += 1

    # MPC Loans table
    mpc = data.get("mpc_loans", {})
    if mpc.get("headers") and mpc.get("rows"):
        totals_row = [mpc["totals"].get(h, "") for h in mpc["headers"][1:]] if mpc.get("totals") else None
        write_table("MPC Loan Capacities", mpc["headers"],
                    [[row_d.get(h, "") for h in mpc["headers"]] for row_d in mpc["rows"]], totals_row)

    # Vertical Loans table
    vl = data.get("vertical_loans", {})
    if vl.get("headers") and vl.get("rows"):
        totals_row = [vl["totals"].get(h, "") for h in vl["headers"][1:]] if vl.get("totals") else None
        write_table("Vertical Loan Capacities", vl["headers"],
                    [[row_d.get(h, "") for h in vl["headers"]] for row_d in vl["rows"]], totals_row)

    # Debt Schedules — one mini-table per project
    for sched in data.get("debt_schedules", []):
        proj_name = sched.get("project", "Project")
        months = sched.get("months", [])
        if not months:
            continue
        headers = [""] + [str(m) for m in months]
        rows_data = []
        # Revenue category rows (each has a monthly array)
        for rev in sched.get("revenues", []):
            rows_data.append([rev.get("type", "")] + list(rev.get("monthly", [])))
        # Total revenues monthly
        total_rev = sched.get("total_revenues", {})
        if total_rev:
            rows_data.append(["Total Revenues"] + list(total_rev.get("monthly", [])))
        # Cumulative revenues and payments are already flat monthly arrays
        rows_data.append(["Cumulative Revenues"] + list(sched.get("cumulative_revenues", [])))
        rows_data.append(["Cumulative Payments"] + list(sched.get("cumulative_payments", [])))
        write_table(f"Debt Schedule — {proj_name}", headers, rows_data)

    ws.column_dimensions["A"].width = 30
    for ci in range(2, 30):
        ws.column_dimensions[get_column_letter(ci)].width = 12

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


def _build_ember_capital_payload():
    """Assemble the Ember Capital data bundle for reports.
    Same shape as /api/ember-capital, plus resolved commitments + settings.
    Used by both the Excel and PDF generators so they render from the same
    source of truth the UI shows."""
    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        "SELECT data, uploaded_at FROM reports "
        "WHERE report_type = 'returns' ORDER BY uploaded_at DESC LIMIT 1"
    )
    returns_row = cur.fetchone()
    src = (returns_row["data"] if returns_row else {}) or {}
    years = src.get("years", []) or []

    projects = []
    for p in src.get("projects", []) or []:
        by_label = {m.get("label"): m for m in (p.get("metrics") or [])}

        def _m(lbl):
            return by_label.get(lbl) or {}

        def _yearly(lbl):
            m = _m(lbl); y = m.get("yearly") or []
            n = len(years)
            return (list(y) + [0] * max(0, n - len(y)))[:n]

        def _total(lbl):
            t = _m(lbl).get("total")
            try: return float(t) if t is not None else 0.0
            except (TypeError, ValueError): return 0.0

        projects.append({
            "name":                    p.get("name", ""),
            "lp_irr":                  _total("LP IRR"),
            "lp_em":                   _total("LP Equity Multiple"),
            "lp_profit":               _total("Total LP Profit"),
            "lp_contributions_total":  _total("Total LP Contributions"),
            "lp_distributions_total":  _total("Total LP Distributions"),
            "promote_total":           _total("Promote"),
            "lp_distributions_yearly": _yearly("Total LP Distributions"),
            "promote_yearly":          _yearly("Promote"),
        })

    cur.execute(
        "SELECT data FROM reports WHERE report_type = 'ember_capital_settings' "
        "ORDER BY uploaded_at DESC LIMIT 1"
    )
    srow = cur.fetchone()
    settings = (srow["data"] or {}) if srow else {}
    recycle_map = settings.get("recycle", {}) or {}

    cur.execute(
        "SELECT data FROM reports WHERE report_type = 'ember_capital_commitments' "
        "ORDER BY uploaded_at DESC LIMIT 1"
    )
    crow = cur.fetchone()
    commitments = (crow["data"] or {"groups": []}) if crow else {"groups": []}

    cur.close(); conn.close()

    return {
        "years":       years,
        "projects":    projects,
        "recycle":     recycle_map,
        "commitments": commitments,
        "uploaded_at": (returns_row["uploaded_at"].isoformat() if returns_row and returns_row["uploaded_at"] else None),
    }


def _gen_excel_ember_capital(data):
    """Excel export of the Ember Capital executive report."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ember Capital"

    GOLD = "C8A96E"; LP_COL = "2D8B76"; PR_COL = "A4832A"; LV_COL = "C56028"
    HDR_FILL = PatternFill("solid", fgColor="F2EFE8")
    TOT_FILL = PatternFill("solid", fgColor="FAF6EC")
    thin = Side(style="thin", color="CCCCCC")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _f(bold=False, color="1A1A1A", size=10):
        return Font(name="Calibri", size=size, bold=bold, color=color)
    def _money(cell, v):
        if isinstance(v, (int, float)) and v:
            cell.value = v; cell.number_format = '"$"#,##0'
        else:
            cell.value = None
    def _numK(cell, v):
        if isinstance(v, (int, float)) and v:
            cell.value = v; cell.number_format = '#,##0'
        else:
            cell.value = None
    def _pct(cell, v):
        if isinstance(v, (int, float)):
            cell.value = v; cell.number_format = '0.0%'
    def _em(cell, v):
        if isinstance(v, (int, float)):
            cell.value = v; cell.number_format = '0.00"x"'

    r = 1
    ws.cell(row=r, column=1, value="Ember Capital — Executive Report").font = Font(name="Calibri", bold=True, size=14, color=GOLD); r += 1
    ws.cell(row=r, column=1, value=f"Generated {datetime.datetime.now().strftime('%B %d, %Y')}  |  Portfolio values in $000s").font = _f(color="888888", size=9); r += 2

    projects = data.get("projects", []) or []
    recycle  = data.get("recycle",  {}) or {}

    # --- Portfolio summary / KPIs -----------------------------------------
    ws.cell(row=r, column=1, value="Portfolio Summary").font = _f(bold=True, color=GOLD, size=12); r += 1
    kpis = []
    lp_profit_tot = sum((p.get("lp_profit") or 0) for p in projects)
    prom_tot      = sum((p.get("promote_total") or 0) for p in projects)
    # Weighted LP IRR (by LP contributions)
    eq, eq_irr = 0.0, 0.0
    for p in projects:
        w = abs(p.get("lp_contributions_total") or 0)
        if w and p.get("lp_irr") is not None:
            eq += w; eq_irr += (p.get("lp_irr") or 0) * w
    w_irr = (eq_irr / eq) if eq else None

    # Commitments totals
    groups = (data.get("commitments") or {}).get("groups") or []
    mpc_tot       = sum((g.get("mpc")                or 0) for g in groups)
    vert_tot      = sum((g.get("vertical")           or 0) for g in groups)
    mpc_alloc_tot = sum((g.get("mpc_allocated")      or 0) for g in groups)
    vert_alloc_tot= sum((g.get("vertical_allocated") or 0) for g in groups)
    alloc_tot     = mpc_alloc_tot + vert_alloc_tot
    avail_tot     = (mpc_tot + vert_tot) - alloc_tot

    kpis = [
        ("Active Projects",        len(projects),       "count"),
        ("Total LP Profit",        lp_profit_tot * 1000,"money"),
        ("Total Promote",          prom_tot * 1000,     "money"),
        ("Weighted LP IRR",        w_irr,               "pct"),
        ("MPC Committed",          mpc_tot,             "money"),
        ("Vertical Committed",     vert_tot,            "money"),
        ("Total Committed Capital",mpc_tot + vert_tot,  "money"),
        ("Capital Allocated",      alloc_tot,           "money"),
        ("Available to Deploy",    avail_tot,           "money"),
    ]
    for label, val, kind in kpis:
        ws.cell(row=r, column=1, value=label).font = _f()
        vc = ws.cell(row=r, column=2, value=None); vc.font = _f(bold=True); vc.alignment = Alignment(horizontal="right")
        if kind == "money":   _money(vc, val)
        elif kind == "pct":   _pct(vc, val)
        elif kind == "count": vc.value = val; vc.number_format = '#,##0'
        r += 1
    r += 1

    # --- Projects table ---------------------------------------------------
    ws.cell(row=r, column=1, value="Projects").font = _f(bold=True, color=GOLD, size=12); r += 1
    hdrs = ["Project", "LP IRR", "LP EM", "LP Equity", "Distributions", "LP Recycle %",
            "Promote", "Promote Recycle %", "LP Recycled", "Promote Recycled",
            "LP Leaving", "Promote Leaving"]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=r, column=ci, value=h); c.font = _f(bold=True, color="555555", size=9)
        c.fill = HDR_FILL; c.border = cell_border
        c.alignment = Alignment(horizontal="left" if ci == 1 else "right")
    r += 1

    tots = {"eq":0, "dist":0, "prom":0, "lpR":0, "prR":0, "lpL":0, "prL":0}
    for p in projects:
        rec = recycle.get(p["name"], {"lp": 0, "prom": 0})
        rLp, rPr = (rec.get("lp") or 0)/100.0, (rec.get("prom") or 0)/100.0
        eq   = abs(p.get("lp_contributions_total") or 0)
        dist = p.get("lp_distributions_total") or 0
        prom = p.get("promote_total") or 0
        lpR, lpL  = dist * rLp, dist * (1 - rLp)
        prR, prL  = prom * rPr, prom * (1 - rPr)
        tots["eq"]  += eq;   tots["dist"] += dist; tots["prom"] += prom
        tots["lpR"] += lpR;  tots["prR"]  += prR;  tots["lpL"]  += lpL;  tots["prL"] += prL

        row_vals = [p["name"], p.get("lp_irr"), p.get("lp_em"), eq, dist,
                    (rec.get("lp") or 0)/100.0, prom, (rec.get("prom") or 0)/100.0,
                    lpR, prR, lpL, prL]
        for ci, v in enumerate(row_vals, 1):
            c = ws.cell(row=r, column=ci); c.border = cell_border; c.font = _f(size=9)
            c.alignment = Alignment(horizontal="left" if ci == 1 else "right")
            if ci == 1: c.value = v
            elif ci == 2: _pct(c, v)
            elif ci == 3: _em(c, v)
            elif ci in (6, 8): _pct(c, v)
            else: _numK(c, v)
        r += 1

    # Totals row
    tot_row_vals = ["Totals", None, None, tots["eq"], tots["dist"], None, tots["prom"], None,
                    tots["lpR"], tots["prR"], tots["lpL"], tots["prL"]]
    for ci, v in enumerate(tot_row_vals, 1):
        c = ws.cell(row=r, column=ci); c.border = cell_border; c.fill = TOT_FILL
        c.font = _f(bold=True, color="7A5C1E", size=9)
        c.alignment = Alignment(horizontal="left" if ci == 1 else "right")
        if ci == 1: c.value = v
        elif ci in (4, 5, 7, 9, 10, 11, 12): _numK(c, v)
    r += 2

    # --- Yearly breakdown -------------------------------------------------
    years = data.get("years") or []
    if years:
        ws.cell(row=r, column=1, value="Yearly Cashflow Breakdown").font = _f(bold=True, color=GOLD, size=12); r += 1
        y_hdrs = ["Year", "LP Distributions", "LP Recycled", "LP Leaving",
                  "Promote", "Promote Recycled", "Promote Leaving", "Total Recycled"]
        for ci, h in enumerate(y_hdrs, 1):
            c = ws.cell(row=r, column=ci, value=h); c.font = _f(bold=True, color="555555", size=9)
            c.fill = HDR_FILL; c.border = cell_border
            c.alignment = Alignment(horizontal="left" if ci == 1 else "right")
        r += 1

        # Aggregate yearly
        n = len(years)
        agg = {k: [0.0]*n for k in ("lpD", "prD", "lpR", "lpL", "prR", "prL")}
        for p in projects:
            rec = recycle.get(p["name"], {"lp": 0, "prom": 0})
            rLp, rPr = (rec.get("lp") or 0)/100.0, (rec.get("prom") or 0)/100.0
            lpY = p.get("lp_distributions_yearly") or []
            prY = p.get("promote_yearly") or []
            for i in range(n):
                ld = lpY[i] if i < len(lpY) else 0
                pd = prY[i] if i < len(prY) else 0
                agg["lpD"][i] += ld; agg["prD"][i] += pd
                agg["lpR"][i] += ld*rLp;   agg["lpL"][i] += ld*(1-rLp)
                agg["prR"][i] += pd*rPr;   agg["prL"][i] += pd*(1-rPr)

        tots_y = {k: 0.0 for k in agg}
        for i, yr in enumerate(years):
            if not (agg["lpD"][i] or agg["prD"][i]):
                continue
            tot_rec = agg["lpR"][i] + agg["prR"][i]
            row_vals = [yr, agg["lpD"][i], agg["lpR"][i], agg["lpL"][i],
                        agg["prD"][i], agg["prR"][i], agg["prL"][i], tot_rec]
            for ci, v in enumerate(row_vals, 1):
                c = ws.cell(row=r, column=ci); c.border = cell_border; c.font = _f(size=9)
                c.alignment = Alignment(horizontal="left" if ci == 1 else "right")
                if ci == 1: c.value = v
                else: _numK(c, v)
            for k, i2 in zip(("lpD","lpR","lpL","prD","prR","prL"), range(6)):
                tots_y[k] += agg[k][i]
            r += 1
        # Yearly totals
        row_tot = ["Total", tots_y["lpD"], tots_y["lpR"], tots_y["lpL"],
                   tots_y["prD"], tots_y["prR"], tots_y["prL"],
                   tots_y["lpR"] + tots_y["prR"]]
        for ci, v in enumerate(row_tot, 1):
            c = ws.cell(row=r, column=ci); c.border = cell_border; c.fill = TOT_FILL
            c.font = _f(bold=True, color="7A5C1E", size=9)
            c.alignment = Alignment(horizontal="left" if ci == 1 else "right")
            if ci == 1: c.value = v
            else: _numK(c, v)
        r += 2

    # --- Commitments ------------------------------------------------------
    if groups:
        ws.cell(row=r, column=1, value="Capital Commitments").font = _f(bold=True, color=GOLD, size=12); r += 1
        c_hdrs = ["Group", "MPC Commitment", "MPC Allocated",
                  "Vertical Commitment", "Vertical Allocated",
                  "Total Committed", "Total Allocated", "Remaining"]
        for ci, h in enumerate(c_hdrs, 1):
            c = ws.cell(row=r, column=ci, value=h); c.font = _f(bold=True, color="555555", size=9)
            c.fill = HDR_FILL; c.border = cell_border
            c.alignment = Alignment(horizontal="left" if ci == 1 else "right")
        r += 1
        for g in groups:
            mpc_v   = g.get("mpc") or 0
            vrt_v   = g.get("vertical") or 0
            mpc_a   = g.get("mpc_allocated") or 0
            vrt_a   = g.get("vertical_allocated") or 0
            tot_c   = mpc_v + vrt_v
            tot_a   = mpc_a + vrt_a
            vals = [g.get("name",""), mpc_v, mpc_a, vrt_v, vrt_a,
                    tot_c, tot_a, tot_c - tot_a]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=r, column=ci); c.border = cell_border; c.font = _f(size=9)
                c.alignment = Alignment(horizontal="left" if ci == 1 else "right")
                if ci == 1: c.value = v
                else: _money(c, v)
            r += 1
        row_tot = ["Total", mpc_tot, mpc_alloc_tot, vert_tot, vert_alloc_tot,
                   mpc_tot + vert_tot, alloc_tot, avail_tot]
        for ci, v in enumerate(row_tot, 1):
            c = ws.cell(row=r, column=ci); c.border = cell_border; c.fill = TOT_FILL
            c.font = _f(bold=True, color="7A5C1E", size=9)
            c.alignment = Alignment(horizontal="left" if ci == 1 else "right")
            if ci == 1: c.value = v
            else: _money(c, v)
        r += 2

    # Column widths
    ws.column_dimensions["A"].width = 32
    for ci in range(2, 14):
        ws.column_dimensions[get_column_letter(ci)].width = 15

    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out.read()


def _gen_pdf_ember_capital(data):
    """Branded 2-page executive PDF for the Ember Capital dashboard.

    Layout — A4 Landscape, 2 pages max:
      Page 1  Branded header, Portfolio-at-a-glance KPI grid, Capital Commitments
      Page 2  Project Returns table + Annual Capital Recycling visual

    Uses Ember brand palette from the 1.25.24 logo guide:
      Ember Blue #13344E / Ember Orange #F25929, warm paper #FAF7F2.
    """
    from fpdf import FPDF

    # ---- Brand assets ----------------------------------------------------
    # Official Ember artwork. Two variants available:
    #   ember_logo_white.png  = orange mark + white wordmark, tight-cropped
    #                           (used on dark Ember-blue header bars)
    #   ember_mark.png        = the orange mark alone (fallback)
    # Falls back gracefully if assets are missing so the PDF still renders.
    _STATIC = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static")
    _LOGO_WHITE_PATH = os.path.join(_STATIC, "ember_logo_white.png")
    if not os.path.exists(_LOGO_WHITE_PATH):
        _LOGO_WHITE_PATH = None
    _MARK_PATH = os.path.join(_STATIC, "ember_mark.png")
    if not os.path.exists(_MARK_PATH):
        _MARK_PATH = None

    # ---- Brand palette ---------------------------------------------------
    BLUE      = (19, 52, 78)       # #13344E — primary
    BLUE_DK   = (13, 43, 68)       # #0D2B44
    BLUE_XDK  = (8, 35, 59)        # #08233B
    ORANGE    = (242, 89, 41)      # #F25929 — accent
    ORANGE_DK = (217, 68, 20)      # #D94414
    ORANGE_T  = (254, 238, 231)    # tint for subtle fills
    PAPER     = (250, 247, 242)    # #FAF7F2 — warm canvas
    ROW_ALT   = (248, 245, 239)    # slightly darker paper tint for alt rows
    G700 = (88, 89, 91); G500 = (147, 149, 152); G300 = (209, 211, 212)
    G200 = (229, 230, 231); G100 = (241, 242, 243)
    WHITE = (255, 255, 255)
    BLUE_SOFT = (210, 220, 230)    # header subtitle on dark

    # ---- Data resolution -------------------------------------------------
    projects = data.get("projects", []) or []
    recycle  = data.get("recycle", {}) or {}
    groups   = (data.get("commitments") or {}).get("groups") or []
    years    = data.get("years", []) or []

    lp_profit_tot = sum((p.get("lp_profit") or 0) for p in projects)
    prom_tot      = sum((p.get("promote_total") or 0) for p in projects)
    eq_w, irr_w = 0.0, 0.0
    for p in projects:
        w = abs(p.get("lp_contributions_total") or 0)
        if w and p.get("lp_irr") is not None:
            eq_w += w; irr_w += (p.get("lp_irr") or 0) * w
    w_irr = (irr_w / eq_w) if eq_w else 0.0
    mpc_tot       = sum((g.get("mpc")                or 0) for g in groups)
    vert_tot      = sum((g.get("vertical")           or 0) for g in groups)
    mpc_alloc_tot = sum((g.get("mpc_allocated")      or 0) for g in groups)
    vert_alloc_tot= sum((g.get("vertical_allocated") or 0) for g in groups)
    alloc_tot     = mpc_alloc_tot + vert_alloc_tot
    avail_tot     = (mpc_tot + vert_tot) - alloc_tot

    # Aggregate annual recycling data across all projects
    n_yr = len(years)
    agg = {k: [0.0]*n_yr for k in ("lpD", "prD", "lpR", "lpL", "prR", "prL")}
    for p in projects:
        rec  = recycle.get(p.get("name",""), {"lp": 0, "prom": 0})
        rLp, rPr = (rec.get("lp") or 0)/100.0, (rec.get("prom") or 0)/100.0
        lpY = p.get("lp_distributions_yearly") or []
        prY = p.get("promote_yearly") or []
        for i in range(n_yr):
            ld = lpY[i] if i < len(lpY) else 0
            pd = prY[i] if i < len(prY) else 0
            agg["lpD"][i] += ld; agg["prD"][i] += pd
            agg["lpR"][i] += ld*rLp;  agg["lpL"][i] += ld*(1-rLp)
            agg["prR"][i] += pd*rPr;  agg["prL"][i] += pd*(1-rPr)

    today_label = datetime.datetime.now().strftime("%B %Y")

    # ---- Text safety: Helvetica is Latin-1 only; normalize user strings --
    _UNI_MAP = {"\u2014": "-", "\u2013": "-", "\u2018": "'", "\u2019": "'",
                "\u201C": '"', "\u201D": '"', "\u2026": "...", "\u2022": "*"}
    def safe(s):
        if s is None: return ""
        t = str(s)
        for k, v in _UNI_MAP.items():
            t = t.replace(k, v)
        try:
            t.encode("latin-1")
            return t
        except UnicodeEncodeError:
            return t.encode("latin-1", "replace").decode("latin-1")

    # ---- Formatters ------------------------------------------------------
    def fmt_money_k(v):
        """Values from returns data are already in $000s."""
        if v is None or v == 0: return "--"
        av = abs(v)
        if av >= 1_000:   return f"${v/1000:,.1f}M"
        elif av >= 1:     return f"${v:,.0f}K"
        else:             return f"${v*1000:,.0f}"
    def fmt_money_raw(v):
        """Raw dollar values (commitments)."""
        if v is None or v == 0: return "--"
        av = abs(v)
        if av >= 1_000_000:   return f"${v/1_000_000:,.2f}M"
        elif av >= 1_000:     return f"${v/1000:,.0f}K"
        else:                 return f"${v:,.0f}"
    def fmt_pct(v):
        if v is None: return "--"
        return f"{v*100:,.1f}%"
    def fmt_em(v):
        if not v: return "--"
        return f"{v:,.2f}x"

    # ---- PDF class -------------------------------------------------------
    class PDF(FPDF):
        def __init__(self):
            super().__init__(orientation="L", unit="mm", format="A4")
            self.set_auto_page_break(auto=False)
            self.set_margins(10, 10, 10)

        # helpers
        def _fill(self, rgb):  self.set_fill_color(*rgb)
        def _draw(self, rgb):  self.set_draw_color(*rgb)
        def _text(self, rgb):  self.set_text_color(*rgb)

        def tracked_caps(self, x, y, w, h, text, rgb, size=7, bold=True,
                         spacing=0.35, align="L"):
            """Render text as tracked uppercase caps — for eyebrows/labels."""
            self.set_xy(x, y)
            self.set_font("Helvetica", "B" if bold else "", size)
            self._text(rgb)
            # Approximate letter spacing by inserting thin spaces between chars
            txt = " ".join(list(text.upper()))
            # Character spacing — a more precise approach
            try:
                # fpdf2 supports set_char_spacing in newer versions
                self.set_char_spacing(spacing)
                self.cell(w, h, text.upper(), ln=False, align=align)
                self.set_char_spacing(0)
            except Exception:
                self.cell(w, h, txt, ln=False, align=align)

        def header_bar_tall(self):
            """Full-width branded Ember-blue header (page 1). 22mm tall."""
            self._fill(BLUE); self.rect(0, 0, 297, 22, style="F")
            # Subtle darker lower band for depth
            self._fill(BLUE_DK); self.rect(0, 18, 297, 4, style="F")
            # Orange accent stripe below
            self._fill(ORANGE); self.rect(0, 22, 297, 1.2, style="F")

            # Official Ember logo lockup (orange mark + white wordmark in
            # brand display font). Falls back to mark + Helvetica wordmark
            # if the lockup asset is missing.
            _LOGO_OK = False
            if _LOGO_WHITE_PATH:
                try:
                    # Cropped asset is ~7.78:1; h=6.5mm → w≈50.6mm
                    self.image(_LOGO_WHITE_PATH, x=12, y=7.5, h=6.5)
                    _LOGO_OK = True
                except Exception:
                    _LOGO_OK = False

            if not _LOGO_OK:
                # Fallback: mark image + typed wordmark
                if _MARK_PATH:
                    try:
                        self.image(_MARK_PATH, x=11, y=4, h=15)
                    except Exception:
                        pass
                self.set_xy(29, 6.5)
                self.set_font("Helvetica", "B", 16)
                self._text(WHITE)
                try:
                    self.set_char_spacing(1.4)
                    self.cell(80, 8, "EMBER", ln=False)
                    self.set_char_spacing(0)
                except Exception:
                    self.cell(80, 8, "E M B E R", ln=False)

            # Tagline under the lockup
            self.set_xy(12, 15.5)
            self.set_font("Helvetica", "", 6.5)
            self._text(BLUE_SOFT)
            try:
                self.set_char_spacing(0.9)
                self.cell(80, 4, "FINANCE & ANALYTICS", ln=False)
                self.set_char_spacing(0)
            except Exception:
                self.cell(80, 4, "F I N A N C E  &  A N A L Y T I C S", ln=False)

            # Right side: EXECUTIVE REPORT + date
            self.set_xy(297-12-80, 6.5)
            self.set_font("Helvetica", "B", 8)
            self._text(WHITE)
            try:
                self.set_char_spacing(1.2)
                self.cell(80, 5, "EXECUTIVE REPORT", ln=False, align="R")
                self.set_char_spacing(0)
            except Exception:
                self.cell(80, 5, "EXECUTIVE REPORT", ln=False, align="R")
            self.set_xy(297-12-80, 13)
            self.set_font("Helvetica", "", 7.5)
            self._text(BLUE_SOFT)
            self.cell(80, 4, f"As of {today_label}", ln=False, align="R")

            self._text(BLUE)

        def header_bar_thin(self):
            """Slim ember-blue bar for page 2. 13mm tall."""
            self._fill(BLUE); self.rect(0, 0, 297, 13, style="F")
            self._fill(ORANGE); self.rect(0, 13, 297, 0.9, style="F")

            # Official Ember logo lockup (orange mark + white wordmark).
            # h=5mm × 7.78 ratio ≈ 39mm wide — compact slim-header scale.
            _LOGO_OK = False
            if _LOGO_WHITE_PATH:
                try:
                    self.image(_LOGO_WHITE_PATH, x=12, y=4, h=5)
                    _LOGO_OK = True
                except Exception:
                    _LOGO_OK = False

            if not _LOGO_OK:
                # Fallback path
                if _MARK_PATH:
                    try:
                        self.image(_MARK_PATH, x=11, y=1.8, h=9.5)
                    except Exception:
                        pass
                self.set_xy(23, 4)
                self.set_font("Helvetica", "B", 10)
                self._text(WHITE)
                try:
                    self.set_char_spacing(1.0)
                    self.cell(80, 5, "EMBER", ln=False)
                    self.set_char_spacing(0)
                except Exception:
                    self.cell(80, 5, "EMBER", ln=False)

            # Executive report tag to the right of the lockup
            self.set_xy(54, 4.3)
            self.set_font("Helvetica", "", 8)
            self._text(BLUE_SOFT)
            self.cell(100, 5, "- EXECUTIVE REPORT", ln=False)

            self.set_xy(297-12-60, 4)
            self.set_font("Helvetica", "", 7.5)
            self._text(BLUE_SOFT)
            self.cell(60, 5, f"As of {today_label}", ln=False, align="R")

            self._text(BLUE)

        def section_heading(self, x, y, eyebrow, title):
            # Orange eyebrow with tracked caps
            self.set_xy(x, y)
            self.set_font("Helvetica", "B", 7)
            self._text(ORANGE)
            try:
                self.set_char_spacing(1.0)
                self.cell(120, 4, eyebrow.upper(), ln=False)
                self.set_char_spacing(0)
            except Exception:
                self.cell(120, 4, eyebrow.upper(), ln=False)
            # Title
            self.set_xy(x, y+4.5)
            self.set_font("Helvetica", "B", 15)
            self._text(BLUE)
            self.cell(200, 8, title, ln=False)
            # Accent underline rule
            self._draw(ORANGE); self.set_line_width(0.7)
            self.line(x, y+13, x+14, y+13)

        def kpi_card(self, x, y, w, h, label, value, caption="", accent=ORANGE):
            # Background
            self._fill(WHITE); self._draw(G300)
            self.set_line_width(0.3)
            self.rect(x, y, w, h, style="DF")
            # Left accent line
            self._fill(accent)
            self.rect(x, y, 1.8, h, style="F")
            # Label (eyebrow)
            self.set_xy(x+6, y+5)
            self.set_font("Helvetica", "B", 6.5)
            self._text(accent)
            try:
                self.set_char_spacing(0.9)
                self.cell(w-8, 3.5, label.upper(), ln=False)
                self.set_char_spacing(0)
            except Exception:
                self.cell(w-8, 3.5, label.upper(), ln=False)
            # Big value
            self.set_xy(x+6, y+10)
            self.set_font("Helvetica", "B", 20)
            self._text(BLUE)
            self.cell(w-8, 11, str(value), ln=False)
            # Caption
            if caption:
                self.set_xy(x+6, y+h-6)
                self.set_font("Helvetica", "", 7)
                self._text(G500)
                self.cell(w-8, 4, caption, ln=False)

        def footer_bar(self, page_num, total):
            y = 199
            self._draw(G300); self.set_line_width(0.2)
            self.line(12, y, 297-12, y)
            self.set_xy(12, y+2)
            self.set_font("Helvetica", "B", 6.5)
            self._text(G500)
            try:
                self.set_char_spacing(0.9)
                self.cell(0, 4, "EMBER CAPITAL   |   FINANCE & ANALYTICS", ln=False)
                self.set_char_spacing(0)
            except Exception:
                self.cell(0, 4, "EMBER CAPITAL | FINANCE & ANALYTICS", ln=False)
            self.set_xy(297-12-40, y+2)
            self.set_font("Helvetica", "B", 7)
            self._text(BLUE)
            self.cell(40, 4, f"PAGE {page_num} OF {total}", ln=False, align="R")
            # Confidential watermark
            self.set_xy(12, y+6)
            self.set_font("Helvetica", "", 6.5)
            self._text(G500)
            self.cell(0, 3, "Confidential - for internal Ember stakeholders only.", ln=False)

    pdf = PDF()
    # ========================= PAGE 1 =========================
    pdf.add_page()
    pdf.header_bar_tall()

    # Section: Portfolio at a glance
    pdf.section_heading(12, 29, "Ember Capital", "Portfolio at a Glance")

    # Subtitle line
    pdf.set_xy(12, 45)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(*G700)
    pdf.cell(200, 4,
             f"Capital position across {len(projects)} active project(s). "
             f"Returns shown net to LP.",
             ln=False)

    # --- KPI row 1: four hero metrics ---
    y1 = 53
    cw1 = (277 - 3*4) / 4
    kpis1 = [
        ("Active Projects",    f"{len(projects):,}",              "",                 ORANGE),
        ("Total LP Profit",    fmt_money_k(lp_profit_tot),        "Net of recycling", BLUE),
        ("Weighted LP IRR",    fmt_pct(w_irr),                    "Equity-weighted",  ORANGE),
        ("Total Promote",      fmt_money_k(prom_tot),             "GP economics",     BLUE),
    ]
    for i, (lbl, val, cap, acc) in enumerate(kpis1):
        x = 10 + i*(cw1+4)
        pdf.kpi_card(x, y1, cw1, 30, lbl, val, cap, acc)

    # --- KPI row 2: commitments ---
    y2 = 88
    cw2 = (277 - 4*4) / 5  # five tiles now (added Allocated + Available)
    kpis2 = [
        ("MPC Committed",          fmt_money_raw(mpc_tot),               "Master-planned",       BLUE),
        ("Vertical Committed",     fmt_money_raw(vert_tot),              "Vertical product",     BLUE),
        ("Total Committed Capital",fmt_money_raw(mpc_tot+vert_tot),      "Across asset classes", ORANGE),
        ("Capital Allocated",      fmt_money_raw(alloc_tot),             "Drawn against commits", BLUE),
        ("Available to Deploy",    fmt_money_raw(avail_tot),             "Unallocated remaining", ORANGE),
    ]
    for i, (lbl, val, cap, acc) in enumerate(kpis2):
        x = 10 + i*(cw2+4)
        pdf.kpi_card(x, y2, cw2, 26, lbl, val, cap, acc)

    # --- Capital Commitments section ---
    pdf.section_heading(12, 122, "Investor Groups", "Capital Commitments")

    pdf.set_xy(12, 138)
    pdf.set_font("Helvetica", "", 8.5)
    pdf.set_text_color(*G700)
    if groups:
        pdf.cell(0, 4,
                 f"{len(groups)} investor group(s) committed {fmt_money_raw(mpc_tot+vert_tot)} "
                 f"({fmt_money_raw(mpc_tot)} MPC + {fmt_money_raw(vert_tot)} Vertical) - "
                 f"{fmt_money_raw(alloc_tot)} allocated, {fmt_money_raw(avail_tot)} remaining.",
                 ln=False)
    else:
        pdf.cell(0, 4, "No investor commitments recorded yet.", ln=False)

    # Commitments table — wider with allocated columns. 8 cols summing to 277.
    y = 146
    col_w = [54, 30, 30, 30, 30, 33, 35, 35]  # Group / MPC C / MPC A / V C / V A / Tot C / Tot A / Rem
    x0 = 10
    # Header
    pdf.set_xy(x0, y)
    pdf.set_fill_color(*BLUE)
    pdf.set_text_color(*WHITE)
    pdf.set_font("Helvetica", "B", 7)
    hdrs = ["GROUP", "MPC COMMIT", "MPC ALLOC", "VERT COMMIT", "VERT ALLOC",
            "TOT COMMIT", "TOT ALLOC", "REMAINING"]
    for i, h in enumerate(hdrs):
        try: pdf.set_char_spacing(0.6)
        except Exception: pass
        pdf.cell(col_w[i], 7, h, border=0, fill=True,
                 align="L" if i == 0 else "R")
    try: pdf.set_char_spacing(0)
    except Exception: pass
    pdf.ln()
    y += 7
    # Body
    pdf.set_font("Helvetica", "", 8.5)
    shown_groups = groups[:10]  # keep page tight
    for i, g in enumerate(shown_groups):
        mpc_v = g.get("mpc") or 0; vrt_v = g.get("vertical") or 0
        mpc_a = g.get("mpc_allocated") or 0; vrt_a = g.get("vertical_allocated") or 0
        tot_c = mpc_v + vrt_v
        tot_a = mpc_a + vrt_a
        rem   = tot_c - tot_a
        pdf.set_xy(x0, y)
        pdf.set_fill_color(*(PAPER if i % 2 == 0 else WHITE))
        pdf.set_text_color(*BLUE_XDK)
        vals = [safe(g.get("name","")),
                fmt_money_raw(mpc_v), fmt_money_raw(mpc_a),
                fmt_money_raw(vrt_v), fmt_money_raw(vrt_a),
                fmt_money_raw(tot_c), fmt_money_raw(tot_a),
                fmt_money_raw(rem)]
        for j, v in enumerate(vals):
            pdf.cell(col_w[j], 6, v, border=0, fill=True,
                     align="L" if j == 0 else "R")
        pdf.ln()
        y += 6
    if len(groups) > len(shown_groups):
        pdf.set_xy(x0, y)
        pdf.set_font("Helvetica", "I", 7.5)
        pdf.set_text_color(*G500)
        pdf.cell(0, 5,
                 f"+ {len(groups) - len(shown_groups)} additional group(s) not shown.",
                 ln=False)
        pdf.ln()
        y += 5
    # Totals row
    if groups:
        pdf.set_xy(x0, y)
        pdf.set_fill_color(*BLUE)
        pdf.set_text_color(*WHITE)
        pdf.set_font("Helvetica", "B", 8.5)
        vals = ["TOTAL",
                fmt_money_raw(mpc_tot), fmt_money_raw(mpc_alloc_tot),
                fmt_money_raw(vert_tot), fmt_money_raw(vert_alloc_tot),
                fmt_money_raw(mpc_tot + vert_tot),
                fmt_money_raw(alloc_tot),
                fmt_money_raw(avail_tot)]
        for j, v in enumerate(vals):
            pdf.cell(col_w[j], 7.5, v, border=0, fill=True,
                     align="L" if j == 0 else "R")
        pdf.ln()

    pdf.footer_bar(1, 2)

    # ========================= PAGE 2 =========================
    pdf.add_page()
    pdf.header_bar_thin()

    pdf.section_heading(12, 19, "Project Detail", "Returns & Capital Recycling")

    pdf.set_xy(12, 35)
    pdf.set_font("Helvetica", "", 8.5)
    pdf.set_text_color(*G700)
    pdf.cell(0, 4,
             "Per-project returns with LP and Promote recycling splits. "
             "Values in $000s unless noted.", ln=False)

    # --- Projects table -------------------------------------------------
    # 10 columns tuned for landscape
    hdrs = ["PROJECT", "LP IRR", "LP EM", "EQUITY", "DIST", "PROMOTE",
            "LP REC %", "LP RECYCLED", "PROM REC %", "PROM RECYCLED"]
    widths = [52, 17, 17, 26, 26, 26, 22, 30, 22, 30]  # sum = 268
    y = 42
    x0 = 10
    # Header
    pdf.set_xy(x0, y)
    pdf.set_fill_color(*BLUE)
    pdf.set_text_color(*WHITE)
    pdf.set_font("Helvetica", "B", 7)
    try: pdf.set_char_spacing(0.7)
    except Exception: pass
    for i, h in enumerate(hdrs):
        pdf.cell(widths[i], 7, h, border=0, fill=True,
                 align="L" if i == 0 else "R")
    try: pdf.set_char_spacing(0)
    except Exception: pass
    pdf.ln()
    y += 7
    # Body rows
    pdf.set_font("Helvetica", "", 8)
    tots = {"eq":0, "dist":0, "prom":0, "lpR":0, "prR":0}
    # Cap at 12 projects to keep to one page
    shown_projects = projects[:12]
    for i, p in enumerate(shown_projects):
        rec = recycle.get(p.get("name",""), {"lp": 0, "prom": 0})
        rLp, rPr = (rec.get("lp") or 0)/100.0, (rec.get("prom") or 0)/100.0
        eq   = abs(p.get("lp_contributions_total") or 0)
        dist = p.get("lp_distributions_total") or 0
        prom = p.get("promote_total") or 0
        lpR  = dist * rLp
        prR  = prom * rPr
        tots["eq"]  += eq;  tots["dist"] += dist; tots["prom"] += prom
        tots["lpR"] += lpR; tots["prR"]  += prR

        pdf.set_xy(x0, y)
        pdf.set_fill_color(*(ROW_ALT if i % 2 == 0 else WHITE))
        pdf.set_text_color(*BLUE_XDK)
        name = safe(p.get("name","") or "")
        if len(name) > 34: name = name[:32] + ".."
        vals = [
            name,
            fmt_pct(p.get("lp_irr")),
            fmt_em(p.get("lp_em")),
            fmt_money_k(eq),
            fmt_money_k(dist),
            fmt_money_k(prom),
            f"{(rec.get('lp') or 0):.0f}%",
            fmt_money_k(lpR),
            f"{(rec.get('prom') or 0):.0f}%",
            fmt_money_k(prR),
        ]
        for j, v in enumerate(vals):
            pdf.cell(widths[j], 5.5, v, border=0, fill=True,
                     align="L" if j == 0 else "R")
        pdf.ln()
        y += 5.5
    if len(projects) > len(shown_projects):
        pdf.set_xy(x0, y)
        pdf.set_font("Helvetica", "I", 7.5)
        pdf.set_text_color(*G500)
        pdf.cell(0, 4,
                 f"+ {len(projects) - len(shown_projects)} additional project(s) not shown.",
                 ln=False)
        pdf.ln()
        y += 4
    # Totals row
    pdf.set_xy(x0, y)
    pdf.set_fill_color(*ORANGE_T); pdf.set_text_color(*ORANGE_DK)
    pdf.set_font("Helvetica", "B", 8)
    tot_vals = ["PORTFOLIO TOTAL", "", "",
                fmt_money_k(tots["eq"]), fmt_money_k(tots["dist"]),
                fmt_money_k(tots["prom"]), "",
                fmt_money_k(tots["lpR"]), "",
                fmt_money_k(tots["prR"])]
    for j, v in enumerate(tot_vals):
        pdf.cell(widths[j], 7, v, border=0, fill=True,
                 align="L" if j == 0 else "R")
    pdf.ln()
    y += 7 + 4

    # --- Annual Capital Recycling visual --------------------------------
    # Footer bar begins at y=199; leave ~5mm safety margin above it.
    FOOTER_TOP = 199
    CHART_BOTTOM = FOOTER_TOP - 5  # 194
    if years and any(agg["lpR"][i] + agg["prR"][i] for i in range(n_yr)):
        pdf.section_heading(12, y, "Capital Flow", "Annual Capital Recycling")
        y += 15

        # Summary line — placed above the chart so it can never collide
        # with the page footer regardless of how many years we render.
        total_recycled_all = sum(agg["lpR"]) + sum(agg["prR"])
        pdf.set_xy(12, y)
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_text_color(*ORANGE_DK)
        try: pdf.set_char_spacing(0.7)
        except Exception: pass
        pdf.cell(0, 4,
                 f"PORTFOLIO RECYCLING   {fmt_money_k(total_recycled_all)}   "
                 f"LP {fmt_money_k(sum(agg['lpR']))}   +   "
                 f"PROMOTE {fmt_money_k(sum(agg['prR']))}",
                 ln=False)
        try: pdf.set_char_spacing(0)
        except Exception: pass
        y += 6

        # Legend
        pdf.set_xy(12, y)
        pdf.set_fill_color(*ORANGE)
        pdf.rect(12, y+1, 3, 3, style="F")
        pdf.set_xy(16.5, y)
        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(*G700)
        pdf.cell(28, 5, "LP Recycled", ln=False)
        pdf.set_fill_color(*BLUE)
        pdf.rect(44, y+1, 3, 3, style="F")
        pdf.set_xy(48.5, y)
        pdf.cell(34, 5, "Promote Recycled", ln=False)
        pdf.set_fill_color(*G300)
        pdf.rect(83, y+1, 3, 3, style="F")
        pdf.set_xy(87.5, y)
        pdf.cell(34, 5, "Capital Leaving", ln=False)
        y += 7

        # Compute scale
        year_totals = [agg["lpD"][i] + agg["prD"][i] for i in range(n_yr)]
        max_total = max(year_totals) if year_totals else 0
        if max_total > 0:
            # Filter to years with activity
            active_years = [(i, years[i]) for i in range(n_yr)
                            if agg["lpD"][i] or agg["prD"][i]]
            # Cap at 8 years for space
            active_years = active_years[:8]
            n_rows = len(active_years)
            # Adaptive row height: fill available space without overflow
            avail = CHART_BOTTOM - y
            if n_rows > 0:
                row_h = max(5.5, min(8.0, avail / n_rows))
            else:
                row_h = 8.0
            # Bar area:
            label_w = 18
            bar_x = 30
            bar_right = 297 - 12 - 72  # leave room for right-side numbers
            bar_w_max = bar_right - bar_x
            numbers_x = bar_right + 4

            for i, yr in active_years:
                total_yr = agg["lpD"][i] + agg["prD"][i]
                lpR = agg["lpR"][i]; prR = agg["prR"][i]
                lpL = agg["lpL"][i]; prL = agg["prL"][i]
                rec_total = lpR + prR
                pct = (rec_total / total_yr * 100) if total_yr else 0

                # Year label
                pdf.set_xy(12, y+1.2)
                pdf.set_font("Helvetica", "B", 9)
                pdf.set_text_color(*BLUE)
                pdf.cell(label_w, 4, str(yr), ln=False)

                # Proportional widths
                scale = bar_w_max / max_total
                w_lpR = lpR * scale
                w_prR = prR * scale
                w_leav = (lpL + prL) * scale
                bar_y = y + 1.0
                bar_h = min(4.5, row_h - 2.0)
                # Background track
                pdf.set_fill_color(*G100)
                pdf.rect(bar_x, bar_y, bar_w_max, bar_h, style="F")
                # Segments
                cx = bar_x
                if w_lpR > 0:
                    pdf.set_fill_color(*ORANGE)
                    pdf.rect(cx, bar_y, w_lpR, bar_h, style="F")
                    cx += w_lpR
                if w_prR > 0:
                    pdf.set_fill_color(*BLUE)
                    pdf.rect(cx, bar_y, w_prR, bar_h, style="F")
                    cx += w_prR
                if w_leav > 0:
                    pdf.set_fill_color(*G300)
                    pdf.rect(cx, bar_y, w_leav, bar_h, style="F")

                # Right-side totals
                pdf.set_xy(numbers_x, y+0.2)
                pdf.set_font("Helvetica", "B", 8.5)
                pdf.set_text_color(*BLUE)
                pdf.cell(28, 4, fmt_money_k(rec_total), ln=False, align="R")
                pdf.set_xy(numbers_x+30, y+0.2)
                pdf.set_font("Helvetica", "", 7.5)
                pdf.set_text_color(*G700)
                pdf.cell(28, 4, f"{pct:.0f}% recycled", ln=False, align="L")
                # Second line: raw distribution total — only show if we
                # have enough row height to fit it comfortably.
                if row_h >= 7.0:
                    pdf.set_xy(numbers_x, y+4.2)
                    pdf.set_font("Helvetica", "", 7)
                    pdf.set_text_color(*G500)
                    pdf.cell(58, 3.5,
                             f"Total distributions {fmt_money_k(total_yr)}",
                             ln=False)
                y += row_h

    pdf.footer_bar(2, 2)

    return pdf.output()


def _gen_pdf_report(report_type, data):
    """Branded executive PDF for returns / loans / operations reports.

    Matches the Ember Capital dashboard PDF styling: official Ember logo
    lockup on an Ember-blue header bar with orange accent stripe,
    tracked-caps eyebrows, orange section-underline rules, and a
    confidential footer bar with page numbering.

    Palette (Ember 1.25.24 logo guide):
        Ember Blue #13344E, Ember Orange #F25929, warm paper #FAF7F2.
    """
    from fpdf import FPDF

    # Ember Capital uses its own bespoke 2-page layout.
    if report_type == "ember_capital":
        return _gen_pdf_ember_capital(data)

    # ---- Brand assets (shared with Ember Capital PDF) ---------------------
    _STATIC = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static")
    _LOGO_WHITE_PATH = os.path.join(_STATIC, "ember_logo_white.png")
    if not os.path.exists(_LOGO_WHITE_PATH):
        _LOGO_WHITE_PATH = None
    _MARK_PATH = os.path.join(_STATIC, "ember_mark.png")
    if not os.path.exists(_MARK_PATH):
        _MARK_PATH = None

    # ---- Brand palette ---------------------------------------------------
    BLUE      = (19, 52, 78)       # #13344E — primary
    BLUE_DK   = (13, 43, 68)       # #0D2B44
    ORANGE    = (242, 89, 41)      # #F25929 — accent
    PAPER     = (250, 247, 242)    # #FAF7F2 — warm canvas
    ROW_ALT   = (243, 238, 229)    # slightly deeper paper tint for alt rows
    G700 = (88, 89, 91); G500 = (147, 149, 152); G300 = (209, 211, 212)
    WHITE = (255, 255, 255)
    BLUE_SOFT = (210, 220, 230)

    titles = {
        "returns":    "Active Project Returns",
        "loans":      "Loan Capacities & Debt Schedules",
        "operations": "Ember Operating Revenues",
    }
    eyebrows = {
        "returns":    "Ember Capital",
        "loans":      "Ember Capital",
        "operations": "Ember Operations",
    }
    right_tags = {
        "returns":    "PROJECT RETURNS",
        "loans":      "LOAN CAPACITIES",
        "operations": "OPERATING REVENUES",
    }
    page_title  = titles.get(report_type, "Ember Report")
    eyebrow_txt = eyebrows.get(report_type, "Ember")
    right_tag   = right_tags.get(report_type, "EXECUTIVE REPORT")
    today_label = datetime.datetime.now().strftime("%B %Y")

    # ---- Text safety: Helvetica is Latin-1 only ---------------------------
    _UNI_MAP = {"\u2014": "-", "\u2013": "-", "\u2018": "'", "\u2019": "'",
                "\u201C": '"', "\u201D": '"', "\u2026": "...", "\u2022": "*"}
    def safe(s):
        if s is None: return ""
        t = str(s)
        for k, v in _UNI_MAP.items():
            t = t.replace(k, v)
        try:
            t.encode("latin-1"); return t
        except UnicodeEncodeError:
            return t.encode("latin-1", "replace").decode("latin-1")

    def fmt_cell(v):
        if v is None or v == "":
            return ""
        if isinstance(v, bool):
            return "Yes" if v else "No"
        if isinstance(v, float):
            av = abs(v)
            if av >= 1_000_000:   return f"{v/1_000_000:,.2f}M"
            if av >= 1_000:       return f"{v:,.0f}"
            if 0 < av < 1:        return f"{v:.1%}"
            return f"{v:,.2f}"
        if isinstance(v, int):
            return f"{v:,}"
        return safe(str(v))

    class PDF(FPDF):
        def __init__(self):
            super().__init__(orientation="L", unit="mm", format="A4")
            self.set_auto_page_break(auto=True, margin=20)
            self.set_margins(12, 12, 12)
            self.alias_nb_pages("{nb}")

        # ---- palette helpers
        def _fill(self, rgb): self.set_fill_color(*rgb)
        def _draw(self, rgb): self.set_draw_color(*rgb)
        def _text(self, rgb): self.set_text_color(*rgb)

        # ---- standard header/footer (drive auto page-break)
        def header(self):
            if self.page_no() == 1:
                self._header_tall()
                self.set_xy(12, 29)
                self._section_block(eyebrow_txt, page_title)
                self.set_y(self.get_y() + 2)
            else:
                self._header_thin()
                self.set_xy(12, 18)

        def footer(self):
            self._footer_bar()

        # ---- branded header (page 1, 22mm)
        def _header_tall(self):
            self._fill(BLUE);    self.rect(0, 0, 297, 22, style="F")
            self._fill(BLUE_DK); self.rect(0, 18, 297, 4, style="F")
            self._fill(ORANGE);  self.rect(0, 22, 297, 1.2, style="F")

            _LOGO_OK = False
            if _LOGO_WHITE_PATH:
                try:
                    self.image(_LOGO_WHITE_PATH, x=12, y=7.5, h=6.5)
                    _LOGO_OK = True
                except Exception:
                    _LOGO_OK = False
            if not _LOGO_OK:
                if _MARK_PATH:
                    try: self.image(_MARK_PATH, x=11, y=4, h=15)
                    except Exception: pass
                self.set_xy(29, 6.5)
                self.set_font("Helvetica", "B", 16); self._text(WHITE)
                try:
                    self.set_char_spacing(1.4)
                    self.cell(80, 8, "EMBER", ln=False)
                    self.set_char_spacing(0)
                except Exception:
                    self.cell(80, 8, "EMBER", ln=False)

            # Tagline
            self.set_xy(12, 15.5)
            self.set_font("Helvetica", "", 6.5); self._text(BLUE_SOFT)
            try:
                self.set_char_spacing(0.9)
                self.cell(80, 4, "FINANCE & ANALYTICS", ln=False)
                self.set_char_spacing(0)
            except Exception:
                self.cell(80, 4, "FINANCE & ANALYTICS", ln=False)

            # Right: section tag + as-of
            self.set_xy(297-12-80, 6.5)
            self.set_font("Helvetica", "B", 8); self._text(WHITE)
            try:
                self.set_char_spacing(1.2)
                self.cell(80, 5, right_tag, ln=False, align="R")
                self.set_char_spacing(0)
            except Exception:
                self.cell(80, 5, right_tag, ln=False, align="R")
            self.set_xy(297-12-80, 13)
            self.set_font("Helvetica", "", 7.5); self._text(BLUE_SOFT)
            self.cell(80, 4, f"As of {today_label}", ln=False, align="R")

            self._text(BLUE)

        # ---- slim header (page 2+, 13mm)
        def _header_thin(self):
            self._fill(BLUE);   self.rect(0, 0, 297, 13, style="F")
            self._fill(ORANGE); self.rect(0, 13, 297, 0.9, style="F")

            _LOGO_OK = False
            if _LOGO_WHITE_PATH:
                try:
                    self.image(_LOGO_WHITE_PATH, x=12, y=4, h=5)
                    _LOGO_OK = True
                except Exception:
                    _LOGO_OK = False
            if not _LOGO_OK:
                if _MARK_PATH:
                    try: self.image(_MARK_PATH, x=11, y=1.8, h=9.5)
                    except Exception: pass
                self.set_xy(23, 4)
                self.set_font("Helvetica", "B", 10); self._text(WHITE)
                self.cell(80, 5, "EMBER", ln=False)

            # Report title next to lockup
            self.set_xy(54, 4.3)
            self.set_font("Helvetica", "", 8); self._text(BLUE_SOFT)
            self.cell(140, 5, f"- {page_title.upper()}", ln=False)

            self.set_xy(297-12-60, 4)
            self.set_font("Helvetica", "", 7.5); self._text(BLUE_SOFT)
            self.cell(60, 5, f"As of {today_label}", ln=False, align="R")

            self._text(BLUE)

        # ---- footer bar
        def _footer_bar(self):
            y = 199
            self._draw(G300); self.set_line_width(0.2)
            self.line(12, y, 297-12, y)
            self.set_xy(12, y+2)
            self.set_font("Helvetica", "B", 6.5); self._text(G500)
            try:
                self.set_char_spacing(0.9)
                self.cell(0, 4, "EMBER CAPITAL   |   FINANCE & ANALYTICS", ln=False)
                self.set_char_spacing(0)
            except Exception:
                self.cell(0, 4, "EMBER CAPITAL | FINANCE & ANALYTICS", ln=False)
            # Page number — avoid char_spacing so {nb} alias substitutes cleanly
            self.set_xy(297-12-40, y+2)
            self.set_font("Helvetica", "B", 7); self._text(BLUE)
            self.cell(40, 4, f"PAGE {self.page_no()} OF {{nb}}", ln=False, align="R")
            self.set_xy(12, y+6)
            self.set_font("Helvetica", "", 6.5); self._text(G500)
            self.cell(0, 3, "Confidential - for internal Ember stakeholders only.", ln=False)

        # ---- section heading block (eyebrow + title + underline)
        def _section_block(self, eyebrow, title):
            x = self.get_x(); y = self.get_y()
            self.set_font("Helvetica", "B", 7); self._text(ORANGE)
            try:
                self.set_char_spacing(1.0)
                self.cell(200, 4, eyebrow.upper(), ln=True)
                self.set_char_spacing(0)
            except Exception:
                self.cell(200, 4, eyebrow.upper(), ln=True)
            self.set_x(x)
            self.set_font("Helvetica", "B", 15); self._text(BLUE)
            self.cell(200, 8, title, ln=True)
            # Accent rule
            self._fill(ORANGE); self.rect(x, self.get_y()+0.5, 14, 0.7, style="F")
            self.set_y(self.get_y() + 4)
            self._text(BLUE)

        # ---- sub-section heading (within a page): clean blue title
        #      with a short orange mini-rule. No redundant eyebrow — the
        #      main page heading already establishes the section.
        def sub_section(self, title):
            # Give breathing room and avoid landing on the footer
            if self.get_y() > 180:
                self.add_page()
            self.set_y(self.get_y() + 1)
            self.set_x(12)
            self.set_font("Helvetica", "B", 11); self._text(BLUE)
            self.cell(0, 6, safe(title), ln=True)
            # Mini orange rule under the title
            y_rule = self.get_y() + 0.2
            self._fill(ORANGE); self.rect(12, y_rule, 10, 0.5, style="F")
            self.set_y(y_rule + 2.5)

        # ---- vertical-space guard: force a new page if there isn't
        #      at least `mm` of vertical space left before the footer.
        #      Used to keep sub-section titles glued to their tables.
        def ensure_space(self, mm):
            if self.get_y() + mm > self.h - 20:
                self.add_page()

        # ---- branded table
        def branded_table(self, headers, rows, col_widths=None,
                          first_align="L", header_align=None):
            if not headers:
                return
            n = len(headers)
            usable = 297 - 24  # 12mm margins both sides
            if col_widths is None:
                first = min(70, usable * 0.35)
                rest = (usable - first) / max(n - 1, 1)
                col_widths = [first] + [rest] * (n - 1)
            total_w = sum(col_widths)

            hdr_h = 6.2
            row_h = 5.0

            def draw_header_row():
                # Top orange hairline
                y0 = self.get_y()
                self._fill(ORANGE); self.rect(12, y0, total_w, 0.5, style="F")
                self.set_xy(12, y0 + 0.6)
                self.set_font("Helvetica", "B", 7); self._text(WHITE)
                for i, h in enumerate(headers):
                    if header_align and i < len(header_align):
                        align = header_align[i]
                    else:
                        align = first_align if i == 0 else "R"
                    self._fill(BLUE)
                    self.cell(col_widths[i], hdr_h, safe(str(h))[:40],
                              border=0, fill=True, align=align)
                self.ln(hdr_h)

            draw_header_row()

            self.set_font("Helvetica", "", 7.2)
            for ri, row in enumerate(rows):
                # manual page break — leave room for footer (20mm margin)
                if self.get_y() + row_h > self.h - 20:
                    self.add_page()
                    draw_header_row()
                    self.set_font("Helvetica", "", 7.2)
                fill_rgb = ROW_ALT if ri % 2 == 0 else PAPER
                self._fill(fill_rgb); self._text(G700)
                self.set_x(12)
                for i, val in enumerate(row[:n]):
                    align = first_align if i == 0 else "R"
                    self.cell(col_widths[i], row_h, fmt_cell(val),
                              border=0, fill=True, align=align)
                self.ln(row_h)

            # Bottom hairline
            self._fill(ORANGE); self.rect(12, self.get_y()+0.5, total_w, 0.4, style="F")
            self.set_y(self.get_y() + 4)

        # ---- KPI card (for operations header)
        def kpi_card(self, x, y, w, h, label, value, caption="", accent=ORANGE):
            self._fill(WHITE); self._draw(G300); self.set_line_width(0.3)
            self.rect(x, y, w, h, style="DF")
            self._fill(accent); self.rect(x, y, 1.8, h, style="F")
            self.set_xy(x+6, y+4.5)
            self.set_font("Helvetica", "B", 6.5); self._text(accent)
            try:
                self.set_char_spacing(0.9)
                self.cell(w-8, 3.5, label.upper(), ln=False)
                self.set_char_spacing(0)
            except Exception:
                self.cell(w-8, 3.5, label.upper(), ln=False)
            self.set_xy(x+6, y+9)
            self.set_font("Helvetica", "B", 16); self._text(BLUE)
            self.cell(w-8, 9, str(value), ln=False)
            if caption:
                self.set_xy(x+6, y+h-5)
                self.set_font("Helvetica", "", 6.8); self._text(G500)
                self.cell(w-8, 3.5, caption, ln=False)

    pdf = PDF()
    pdf.add_page()

    # Keep legacy helper names so the existing per-report logic below still works
    def draw_section(title): pdf.sub_section(title)
    def draw_table(headers, rows, col_widths=None):
        pdf.branded_table(headers, rows, col_widths=col_widths)

    if report_type == "returns":
        # Helper: find the first year index where this project actually has
        # capital contributions. Avoids leading columns of zeros when a
        # project starts later than the portfolio-wide year range.
        def _first_contrib_idx(metrics):
            pref = {"Total LP Contributions", "LP Contributions",
                    "Total Equity Contributions", "Equity Contributions"}
            for m in metrics:
                if m.get("label") in pref:
                    yearly = m.get("yearly") or []
                    for i, v in enumerate(yearly):
                        try:
                            if v not in (None, "") and abs(float(v)) > 0:
                                return i
                        except (TypeError, ValueError):
                            pass
            # Fallback: first column with any non-zero value across metrics
            max_len = max((len(m.get("yearly") or []) for m in metrics), default=0)
            for i in range(max_len):
                for m in metrics:
                    yearly = m.get("yearly") or []
                    if i < len(yearly):
                        v = yearly[i]
                        try:
                            if v not in (None, "") and abs(float(v)) > 0:
                                return i
                        except (TypeError, ValueError):
                            pass
            return 0

        summary_cols = ["Project", "LP IRR", "Equity Multiple", "Total LP Profit", "Promote"]
        sum_rows = []
        for proj in data.get("projects", []):
            m = {m["label"]: m for m in proj.get("metrics", [])}
            sum_rows.append([
                proj["name"],
                f"{m.get('LP IRR',{}).get('total',0):.1%}" if m.get('LP IRR',{}).get('total') else "",
                f"{m.get('LP Equity Multiple',{}).get('total',0):.2f}x" if m.get('LP Equity Multiple',{}).get('total') else "",
                m.get('Total LP Profit',{}).get('total', ""),
                m.get('Promote',{}).get('total', ""),
            ])
        draw_section("Portfolio Summary")
        draw_table(summary_cols, sum_rows, [70, 22, 28, 30, 25])
        years = data.get("years", [])

        # Target: two projects per page, never split a table from its title.
        # Usable vertical region per page (after page-1 header/eyebrow/title):
        #   ~ y=46 start -> y=179 safe (footer at 199, auto-break margin 20).
        # Two projects in ~130mm -> budget ~62mm per project block.
        for proj in data.get("projects", []):
            metrics = proj.get("metrics", [])
            start_idx = _first_contrib_idx(metrics)
            proj_years = years[start_idx:start_idx + 10]
            # Estimate block height so we can keep the title + table together.
            # sub_section ~= 10mm; table header ~= 7mm; each row ~= 5mm; pad 6mm.
            est_block = 10 + 7 + len(metrics) * 5 + 6
            pdf.ensure_space(est_block)
            draw_section(proj["name"])
            hdrs = ["Metric", "Total"] + [str(y) for y in proj_years]
            rows_data = []
            for m in metrics:
                yearly_slice = (m.get("yearly") or [])[start_idx:start_idx + 10]
                row_vals = [m["label"], m.get("total", "")] + yearly_slice
                rows_data.append(row_vals)
            draw_table(hdrs, rows_data)

    elif report_type == "loans":
        mpc = data.get("mpc_loans", {})
        if mpc.get("headers") and mpc.get("rows"):
            draw_section("MPC Loan Capacities")
            rows_data = [[r.get(h, "") for h in mpc["headers"]] for r in mpc["rows"]]
            draw_table(mpc["headers"], rows_data)
        vl = data.get("vertical_loans", {})
        if vl.get("headers") and vl.get("rows"):
            draw_section("Vertical Loan Capacities")
            rows_data = [[r.get(h, "") for h in vl["headers"]] for r in vl["rows"]]
            draw_table(vl["headers"], rows_data)
        for sched in data.get("debt_schedules", []):
            months = sched.get("months", [])
            if not months:
                continue
            draw_section(f"Debt Schedule — {sched.get('project','')}")
            hdrs = [""] + [str(m) for m in months[:12]]
            rows_data = [
                ["Scheduled Payments"] + sched.get("payments", [])[:12],
                ["Cumulative Payments"] + sched.get("cumulative_payments", [])[:12],
                ["Lot Revenues"] + sched.get("revenues", [])[:12],
                ["Cumulative Revenues"] + sched.get("cumulative_revenues", [])[:12],
            ]
            draw_table(hdrs, rows_data)

    elif report_type == "operations":
        kpis = data.get("kpis", []) or []
        if kpis:
            def _fmt_kpi(v):
                if isinstance(v, (int, float)):
                    av = abs(v)
                    if av >= 1_000_000: return f"${v/1_000_000:,.1f}M"
                    if av >= 1_000:     return f"${v/1000:,.0f}K"
                    return f"${v:,.0f}"
                return safe(str(v))
            # Up to 4 KPIs per row, card grid
            cards = kpis[:8]
            per_row = 4 if len(cards) >= 4 else len(cards)
            gap = 4
            usable = 297 - 24
            cw = (usable - gap*(per_row-1)) / per_row if per_row else usable
            ch = 26
            y0 = pdf.get_y()
            for i, kpi in enumerate(cards):
                col = i % per_row
                row = i // per_row
                x = 12 + col*(cw+gap)
                y = y0 + row*(ch+gap)
                accent = ORANGE if (i % 2 == 0) else BLUE
                pdf.kpi_card(x, y, cw, ch,
                             kpi.get("label", ""),
                             _fmt_kpi(kpi.get("value", "")),
                             "", accent)
            rows_used = (len(cards) + per_row - 1)//per_row
            pdf.set_y(y0 + rows_used*(ch+gap) + 2)

        yr = data.get("yearly_rollup", {})
        if yr.get("years"):
            draw_section("Annual Revenue Forecast")
            hdrs = ["Revenue Source"] + [str(y) for y in yr["years"]]
            rows_data = [[row["label"]] + row.get("values", []) for row in yr.get("rows", [])]
            draw_table(hdrs, rows_data)
        mo = data.get("monthly", {})
        if mo.get("dates"):
            draw_section("Monthly Fee Revenue")
            dates = mo["dates"][:12]
            hdrs = ["Project / Category"] + [f"{d[5:7]}/{d[2:4]}" for d in dates]
            rows_data = [[f"{r['project']} — {r['category']}"] + r.get("values", [])[:12]
                         for r in mo.get("rows", [])]
            draw_table(hdrs, rows_data)

    return pdf.output()


def _gen_new_pdf_report(rt, data, uploaded_at=None):
    """Render the redesigned executive PDF for a given report type.

    The same WeasyPrint pipeline the live download buttons use, callable
    from outside a request context (e.g. APScheduler / the monthly email
    job). Returns PDF bytes on success, or None on any failure — the
    caller should fall back to the legacy fpdf2 `_gen_pdf_report` so
    users still receive *something* when WeasyPrint can't load.

    Reports covered:
        returns         → templates/returns_report.html        (8-page)
        ember_capital   → templates/capital_report.html        (3-page)
        operations      → templates/operations_report.html     (2-page)
        loans           → templates/loans_report.html          (5-page)
    """
    try:
        from weasyprint import HTML  # noqa: F401  (probe import only)
    except (ImportError, OSError):
        app.logger.warning("WeasyPrint unavailable; falling back to legacy PDF for %s", rt)
        return None

    # Manufacture a request context so render_template / url_for /
    # request.host_url all resolve. Base URL doesn't matter much because
    # _weasyprint_local_fetcher serves /static/* from disk.
    base_url = os.environ.get("APP_BASE_URL") or "http://localhost"
    try:
        with app.test_request_context(base_url=base_url):
            if rt == "returns":
                pdf_bytes = _render_returns_report_pdf(data, uploaded_at)
                return bytes(pdf_bytes)

            if rt == "ember_capital":
                from weasyprint import HTML
                ctx = _capital_report_context()
                html = render_template("capital_report.html", capital=ctx)
                return HTML(string=html, base_url=base_url,
                            url_fetcher=_weasyprint_local_fetcher).write_pdf()

            if rt == "operations":
                from weasyprint import HTML
                view_ctx = _build_operations_view_context(data, uploaded_at)
                if view_ctx is None:
                    return None
                rpt_ctx = _build_operations_report_context(view_ctx, run_date=uploaded_at)
                html = render_template("operations_report.html", **rpt_ctx)
                return HTML(string=html, base_url=base_url,
                            url_fetcher=_weasyprint_local_fetcher).write_pdf()

            if rt == "loans":
                from weasyprint import HTML
                view_ctx = _build_loans_view_context(data, uploaded_at)
                if view_ctx is None:
                    return None
                rpt_ctx = _build_loans_report_context(view_ctx, run_date=uploaded_at)
                html = render_template("loans_report.html", **rpt_ctx)
                return HTML(string=html, base_url=base_url,
                            url_fetcher=_weasyprint_local_fetcher).write_pdf()

            # Any future rt without a redesigned report — caller hits
            # the legacy fpdf2 generator.
            return None
    except Exception as e:
        app.logger.warning("_gen_new_pdf_report(%s) failed: %s: %s",
                           rt, type(e).__name__, e)
        return None


def _send_monthly_emails(force=False, recipient_ids=None):
    """Send monthly reports. Returns count of emails sent.

    Modes:
      • Default (cron):       all opted-in users, period-claim enforced.
      • force=True:           all opted-in users, period-claim bypassed.
      • recipient_ids=[...]:  ad-hoc send to a specific subset, used by
                              the per-user Send button on /admin/reports.
                              Period claim is skipped (only meaningful
                              for the once-a-month bulk send), and the
                              subscription filter is dropped — admin is
                              explicitly requesting these users; the
                              per-user loop still skips users with no
                              email or no resolvable subscriptions and
                              records the reason in `diag`.
    """
    now = datetime.datetime.utcnow()
    period = now.strftime("%Y-%m")

    conn = get_db()
    cur = conn.cursor()

    if not force and not recipient_ids:
        # Atomically claim this period BEFORE sending. The UNIQUE
        # constraint on report_sends.period means only one process can
        # win — if Railway runs the cron in two worker instances at the
        # same time (which happened on 2026-05-01 → duplicate emails),
        # the second one's INSERT silently no-ops and we bail out here.
        # The previous SELECT-then-INSERT-after pattern raced.
        cur.execute(
            "INSERT INTO report_sends (period) VALUES (%s) ON CONFLICT DO NOTHING",
            (period,)
        )
        claimed = cur.rowcount
        conn.commit()
        if not claimed:
            cur.close(); conn.close()
            print(f"[Reports] Period {period} already claimed by another instance; skipping.", flush=True)
            return 0

    if recipient_ids:
        # Ad-hoc per-user send: filter by id only, no email/subscription
        # gate at the SQL level so the per-recipient loop can record a
        # specific skip reason ("no email on file", "no subscriptions")
        # in the diag for each user the admin clicked.
        cur.execute("""
            SELECT id, username, email, report_format, report_subscriptions,
                   report_opt_in, page_access, first_name, last_name
            FROM users
            WHERE id = ANY(%s)
        """, (list(recipient_ids),))
    else:
        # Cron / bulk send: every user with an email + at least one
        # subscription. Checks both:
        #   (a) the new `report_subscriptions` JSONB (preferred source,
        #       set via the admin Reports page).
        #   (b) the legacy `report_opt_in` flag — kept for users who
        #       were set up before the new UI shipped and never re-saved.
        cur.execute("""
            SELECT id, username, email, report_format, report_subscriptions,
                   report_opt_in, page_access, first_name, last_name
            FROM users
            WHERE COALESCE(email, '') <> ''
              AND ( report_opt_in = TRUE
                    OR (report_subscriptions IS NOT NULL
                        AND report_subscriptions <> '{}'::jsonb) )
        """)
    recipients = cur.fetchall()

    # Fetch latest report data for all three types
    report_data = {}
    for rt in ("returns", "loans", "operations"):
        cur.execute("SELECT data FROM reports WHERE report_type = %s ORDER BY uploaded_at DESC LIMIT 1", (rt,))
        row = cur.fetchone()
        report_data[rt] = row["data"] if row else None

    cur.close(); conn.close()

    # Ember Capital uses a composite payload (returns + settings + commitments)
    try:
        ec_payload = _build_ember_capital_payload()
        report_data["ember_capital"] = ec_payload if ec_payload.get("projects") else None
    except Exception as e:
        print(f"Error building ember_capital payload: {e}")
        report_data["ember_capital"] = None

    if not recipients:
        return 0

    sendgrid_key = os.environ.get("SENDGRID_API_KEY", "")
    from_addr = os.environ.get("SMTP_FROM", "")

    if not sendgrid_key:
        raise ValueError("SENDGRID_API_KEY environment variable must be set")
    if not from_addr:
        raise ValueError("SMTP_FROM environment variable must be set (used as sender address)")

    subject = now.strftime("%B %Y") + " Ember Reports"

    report_labels = {
        "returns": "Active Project Returns",
        "loans": "Loan Capacities & Debt Schedules",
        "operations": "Ember Operating Revenues",
        "ember_capital": "Ember Capital Executive Report",
    }

    sg = SendGridAPIClient(sendgrid_key)
    sent_count = 0
    # Per-recipient diagnostic so the admin "Send Reports Now" UI can show
    # who got skipped and why. Also goes to stdout for Railway logs.
    diag = []

    # Load logo once
    # Email footer wordmark — the design-handoff blue wordmark, attached
    # via CID so it embeds reliably in every client (CID is widely
    # supported and avoids data-URI volume limits).
    logo_b64 = ""
    logo_path = os.path.join(os.path.dirname(__file__),
                              "static", "img", "ember_logo_blue_wordmark.png")
    if os.path.exists(logo_path):
        with open(logo_path, "rb") as f:
            logo_b64 = base64.b64encode(f.read()).decode()

    print(f"[Reports] {len(recipients)} candidate recipient(s) matched the filter", flush=True)

    for user in recipients:
        email_addr  = (user.get("email") or "").strip()
        fn          = (user.get("first_name") or "").strip()
        ln          = (user.get("last_name") or "").strip()
        display_name = f"{fn} {ln}".strip() or user["username"]
        legacy_fmt  = user.get("report_format") or "pdf"
        subs        = user.get("report_subscriptions") or {}
        pa          = user.get("page_access") or {}

        # Recipient might land here without an email when an admin
        # explicitly requested an ad-hoc send (recipient_ids path skips
        # the SQL-level email filter so we can record the skip reason).
        if not email_addr:
            print(f"[Reports]  · skip {display_name}: no email on file", flush=True)
            diag.append({"user": display_name, "email": "",
                         "status": "skipped", "reason": "no email on file"})
            continue

        # Resolve which reports + format this user gets:
        #   • If `report_subscriptions` has entries, use them as-is —
        #     {report_key: 'pdf'|'excel'}. This is what the new admin
        #     Reports modal writes.
        #   • Otherwise fall back to the legacy flag: send every report
        #     they have page access to, in their global preferred format.
        if subs:
            user_subs = {rt: subs[rt] for rt in subs
                         if rt in report_labels and subs[rt] in ("pdf", "excel")}
        else:
            user_subs = {rt: legacy_fmt for rt in report_labels
                         if pa.get(rt, True)}

        # Filter out reports we don't actually have data for, and respect
        # page_access. (page_access still gates: an admin can disable a
        # whole report category for a user.)
        accessible = {rt: report_labels[rt] for rt in user_subs
                      if report_data.get(rt) and pa.get(rt, True)}
        if not accessible:
            reason = ("no subscriptions" if not user_subs
                      else "no accessible reports (page_access or no data)")
            print(f"[Reports]  · skip {display_name} <{email_addr}>: {reason}", flush=True)
            diag.append({"user": display_name, "email": email_addr,
                         "status": "skipped", "reason": reason})
            continue

        # First name only for the greeting per the design handoff. Fall
        # back to the first token of display_name (handles "Javier Mendez"
        # → "Javier"), then to display_name itself, then to "there".
        greet_name = (fn or (display_name.split()[0] if display_name else "")
                         or display_name or "there")

        # Period label for header pill + lede (e.g. "May 2026"). Issue
        # number intentionally dropped — we don't track issues; the
        # design-handoff README marks it optional.
        period_label = now.strftime("%B %Y")

        # Reports list rows — index + title + tinted file-type pill.
        report_rows = []
        for i, (rt, label) in enumerate(accessible.items(), start=1):
            fmt = user_subs[rt]
            if fmt == "excel":
                pill_bg, pill_fg, pill_text = "#E2F1EA", "#1F7A4D", "EXCEL"
            else:  # pdf (default)
                pill_bg, pill_fg, pill_text = "#FEEEE7", "#A8330D", "PDF"
            border = "" if i == len(accessible) else "border-bottom:1px solid rgba(8,35,59,0.08);"
            row_label = label.replace("&", "&amp;")
            report_rows.append(f"""
      <tr><td style="padding:16px 0;{border}">
        <table cellpadding="0" cellspacing="0" border="0" width="100%" role="presentation">
          <tr>
            <td width="28" style="font-family:'JetBrains Mono',ui-monospace,Menlo,monospace;font-size:11px;color:#939598;letter-spacing:0.04em;vertical-align:middle;">{i:02d}</td>
            <td style="font-family:'Plus Jakarta Sans',system-ui,sans-serif;font-weight:600;font-size:15px;color:#13344E;line-height:1.35;letter-spacing:-0.005em;vertical-align:middle;padding:0 8px;">{row_label}</td>
            <td align="right" style="vertical-align:middle;">
              <span style="display:inline-block;padding:4px 10px;border-radius:999px;background:{pill_bg};color:{pill_fg};font-family:'Plus Jakarta Sans',system-ui,sans-serif;font-weight:700;font-size:10px;letter-spacing:0.10em;text-transform:uppercase;">{pill_text}</span>
            </td>
          </tr>
        </table>
      </td></tr>""")
        report_rows_html = "".join(report_rows)

        n_reports = len(accessible)
        files_label = f"{n_reports} file" + ("" if n_reports == 1 else "s")

        # Manage Delivery link target — must be an ABSOLUTE URL since
        # /home as a relative href is meaningless inside an email
        # client. Try PUBLIC_URL (admin override) first, then Railway's
        # auto-set RAILWAY_PUBLIC_DOMAIN. If neither is set we still
        # ship something clickable; the link just won't resolve until
        # one of those env vars is configured.
        public_url = (os.environ.get("PUBLIC_URL")
                      or os.environ.get("RAILWAY_PUBLIC_DOMAIN")
                      or "").strip()
        if public_url and not public_url.startswith(("http://", "https://")):
            public_url = "https://" + public_url
        public_url = public_url.rstrip("/")
        manage_url = (public_url + "/home") if public_url else "/home"

        # Bars motif — 5 vertical bars echoing the Ember mark. Heights:
        # 72/44/60/36/56; first 4 orange, last semi-transparent white.
        # Rendered as inline-block <span>s inside a right-aligned table
        # cell — earlier `position:absolute` got stripped by Gmail and
        # the bars fell back to inline flow on the LEFT of the eyebrow.
        bars_motif = "".join(
            f'<span style="display:inline-block;width:6px;height:{h}px;'
            f'background:{c};border-radius:0 0 999px 999px;'
            f'vertical-align:top;{("margin-left:6px;" if i else "")}font-size:0;line-height:0;"></span>'
            for i, (h, c) in enumerate(zip(
                [72, 44, 60, 36, 56],
                ["#F25929", "#F25929", "#F25929", "#F25929", "rgba(255,255,255,0.35)"]))
        )

        logo_img = ('<img src="cid:ember_logo" alt="Ember" height="22" '
                    'style="display:block;margin:0 auto 10px;height:22px;border:0;">'
                    if logo_b64 else "")

        html_body = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Ember Monthly Reports — {period_label}</title>
<link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700;800&display=swap">
</head>
<body style="margin:0;padding:0;background:#F8F4EE;font-family:'Plus Jakarta Sans',system-ui,-apple-system,'Segoe UI',sans-serif;color:#13344E;">
<table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#F8F4EE;padding:32px 16px;">
  <tr><td align="center">
    <table role="presentation" width="640" cellpadding="0" cellspacing="0" border="0" style="max-width:640px;background:#FFFFFF;border-radius:16px;overflow:hidden;box-shadow:0 2px 4px rgba(8,35,59,.06),0 1px 2px rgba(8,35,59,.04);">
      <!-- HEADER -- 2-column inner table so the bars motif sits in the
           top-right cell (valign=top, no top padding) while the text
           content fills the left cell with normal header padding. The
           outer header td has padding:0 so the dark-blue background
           still extends edge-to-edge. -->
      <tr><td style="background:#08233B;padding:0;color:#FFFFFF;">
        <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0">
          <tr>
            <td valign="top" style="padding:36px 0 32px 40px;">
              <div style="font-family:'Plus Jakarta Sans',system-ui,sans-serif;font-weight:600;font-size:12px;letter-spacing:0.20em;color:#F25929;text-transform:uppercase;margin-bottom:14px;">EMBER FINANCE &amp; ANALYTICS</div>
              <h1 style="margin:0 0 6px;font-family:'Plus Jakarta Sans',system-ui,sans-serif;font-weight:800;font-size:34px;line-height:1.05;letter-spacing:-0.02em;color:#FFFFFF;">Ember Monthly Reports</h1>
              <div style="font-family:'Plus Jakarta Sans',system-ui,sans-serif;font-size:14px;line-height:1.5;color:rgba(255,255,255,0.72);max-width:44ch;">Delivered on the first of each month.</div>
              <div style="display:inline-block;background:#FFFFFF;padding:8px 14px;border-radius:999px;box-shadow:0 2px 4px rgba(8,35,59,.10);margin-top:22px;">
                <span style="display:inline-block;width:6px;height:6px;border-radius:50%;background:#F25929;vertical-align:middle;margin-right:8px;"></span>
                <span style="font-family:'Plus Jakarta Sans',system-ui,sans-serif;font-weight:700;font-size:12px;letter-spacing:0.12em;text-transform:uppercase;color:#13344E;vertical-align:middle;">{period_label}</span>
              </div>
            </td>
            <td valign="top" align="right" width="110" style="padding:0 32px 0 0;font-size:0;line-height:0;">{bars_motif}</td>
          </tr>
        </table>
      </td></tr>

      <!-- BODY -->
      <tr><td style="padding:40px 40px 8px;background:#FFFFFF;">
        <div style="font-family:'Plus Jakarta Sans',system-ui,sans-serif;font-weight:700;font-size:22px;line-height:1.25;letter-spacing:-0.01em;color:#13344E;margin-bottom:14px;">Hello {greet_name},</div>
        <div style="font-family:'Plus Jakarta Sans',system-ui,sans-serif;font-size:16px;line-height:1.65;color:#58595B;max-width:56ch;margin-bottom:32px;">Your <strong style="color:#13344E;font-weight:700;">{period_label}</strong> reports are attached below.</div>

        <!-- Reports header row -->
        <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" style="border-bottom:1px solid rgba(8,35,59,0.12);padding-bottom:12px;margin-bottom:8px;">
          <tr>
            <td style="font-family:'Plus Jakarta Sans',system-ui,sans-serif;font-weight:700;font-size:11px;letter-spacing:0.20em;text-transform:uppercase;color:#F25929;padding-bottom:12px;">Reports included</td>
            <td align="right" style="font-family:'Plus Jakarta Sans',system-ui,sans-serif;font-weight:500;font-size:12px;color:#939598;padding-bottom:12px;">{files_label}</td>
          </tr>
        </table>

        <!-- Reports list -->
        <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-bottom:32px;">{report_rows_html}
        </table>

        <!-- Note callout -->
        <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#F8F4EE;border-radius:12px;">
          <tr>
            <td width="30" style="padding:14px 0 14px 18px;vertical-align:top;">
              <svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="#F25929" stroke-width="1.6" stroke-linecap="round" stroke-linejoin="round" style="display:block;"><circle cx="12" cy="12" r="10"/><path d="M12 8v.01M11 12h1v4h1"/></svg>
            </td>
            <td style="padding:14px 18px 14px 12px;font-family:'Plus Jakarta Sans',system-ui,sans-serif;font-size:13px;line-height:1.5;color:#58595B;">Generated on the 1st of each month. Reply to this email if a number looks wrong — we'd rather hear about it early.</td>
          </tr>
        </table>

        <div style="height:32px;line-height:32px;font-size:0;">&nbsp;</div>
      </td></tr>

      <!-- FOOTER -->
      <tr><td style="background:#F8F4EE;border-top:1px solid rgba(8,35,59,0.08);padding:28px 40px 32px;text-align:center;">
        {logo_img}
        <div style="font-family:'Plus Jakarta Sans',system-ui,sans-serif;font-weight:600;font-size:13px;letter-spacing:-0.005em;color:#13344E;">Ember Finance &amp; Analytics Team</div>
        <div style="font-family:'Plus Jakarta Sans',system-ui,sans-serif;font-size:12px;color:#939598;margin-top:6px;">Sent from the EmberApps platform <span style="color:#D1D3D4;margin:0 6px;">·</span> <a href="{manage_url}" style="color:#A8330D;font-weight:600;text-decoration:none;">Manage delivery</a></div>
      </td></tr>
    </table>
  </td></tr>
</table>
</body>
</html>"""

        plain_lines = [f"Hello {greet_name},", "",
                       f"Your {period_label} reports are attached below.", "",
                       "REPORTS INCLUDED:"]
        for i, (rt, label) in enumerate(accessible.items(), start=1):
            plain_lines.append(f"  {i:02d}. {label} ({user_subs[rt].upper()})")
        plain_lines.extend([
            "",
            "Generated on the 1st of each month. Reply to this email if a number looks wrong — we'd rather hear about it early.",
            "",
            "Ember Finance & Analytics Team",
            f"Sent from the EmberApps platform · Manage delivery: {manage_url}",
        ])
        plain_body = "\n".join(plain_lines)

        message = Mail(
            from_email=from_addr,
            to_emails=email_addr,
            subject=subject,
        )
        message.content = [
            Content("text/plain", plain_body),
            Content("text/html", html_body),
        ]

        # Collect ALL attachments first, then assign once. SendGrid's
        # Mail.attachment setter REPLACES on each call in modern SDK
        # versions — assigning per-iteration would silently drop every
        # attachment except the last one. Building a list and assigning
        # at the end keeps every PDF / Excel + the inline logo.
        attachments = []

        if logo_b64:
            logo_att = Attachment(
                FileContent(logo_b64),
                FileName("ember_logo.png"),
                FileType("image/png"),
                Disposition("inline"),
            )
            logo_att.content_id = "ember_logo"
            attachments.append(logo_att)

        attached_rts = []
        for rt, label in accessible.items():
            data = report_data[rt]
            fmt  = user_subs[rt]   # 'pdf' or 'excel', per-report from subscriptions
            try:
                if fmt == "excel":
                    if rt == "returns":
                        file_bytes = _gen_excel_returns(data)
                        filename = f"{label.replace(' ','_')}.xlsx"
                    elif rt == "loans":
                        file_bytes = _gen_excel_loans(data)
                        filename = "Loan_Capacities.xlsx"
                    elif rt == "ember_capital":
                        file_bytes = _gen_excel_ember_capital(data)
                        filename = "Ember_Capital_Executive_Report.xlsx"
                    else:
                        file_bytes = _gen_excel_operations(data)
                        filename = "Ember_Operating_Revenues.xlsx"
                    mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                else:
                    # Prefer the redesigned WeasyPrint PDF (same one the
                    # live download buttons serve). Fall back to the
                    # legacy fpdf2 generator only if WeasyPrint can't
                    # load — guarantees the email always carries an
                    # attachment instead of silently dropping it.
                    new_bytes = _gen_new_pdf_report(rt, data, uploaded_at=now)
                    file_bytes = bytes(new_bytes) if new_bytes else bytes(_gen_pdf_report(rt, data))
                    filename = f"{label.replace(' ','_')}.pdf"
                    mime_type = "application/pdf"

                attachments.append(Attachment(
                    FileContent(base64.b64encode(file_bytes).decode()),
                    FileName(filename),
                    FileType(mime_type),
                    Disposition("attachment"),
                ))
                attached_rts.append(f"{rt}({fmt})")
            except Exception as e:
                print(f"[Reports] {display_name}: failed to build {rt} {fmt}: "
                      f"{type(e).__name__}: {e}", flush=True)

        if not attachments or all(a.disposition.get() == "inline" if hasattr(a.disposition, 'get') else False
                                   for a in attachments):
            # Only the inline logo, no actual reports — don't bother sending.
            print(f"[Reports]  · skip {display_name} <{email_addr}>: "
                  f"no report attachments built", flush=True)
            diag.append({"user": display_name, "email": email_addr,
                         "status": "skipped", "reason": "no attachments built"})
            continue

        message.attachment = attachments

        try:
            sg.send(message)
            sent_count += 1
            print(f"[Reports]  · sent {display_name} <{email_addr}>: "
                  f"{', '.join(attached_rts)}", flush=True)
            diag.append({"user": display_name, "email": email_addr,
                         "status": "sent", "reports": attached_rts})
        except Exception as e:
            print(f"[Reports] SendGrid error sending to <{email_addr}>: "
                  f"{type(e).__name__}: {e}", flush=True)
            # If SendGrid returns a structured error body, log it too —
            # this is the only way to see why an email was rejected
            # (over quota, bounced, suppressed, etc.).
            err_body = getattr(getattr(e, "body", None), "decode", lambda *_: None)("utf-8") \
                       if hasattr(getattr(e, "body", None), "decode") else getattr(e, "body", None)
            if err_body:
                print(f"[Reports]    response: {err_body}", flush=True)
            diag.append({"user": display_name, "email": email_addr,
                         "status": "send_failed",
                         "reason": f"{type(e).__name__}: {e}"})

    print(f"[Reports] Done — {sent_count} email(s) sent, "
          f"{len(diag) - sent_count} skipped/failed.", flush=True)

    # NOTE: the period claim was inserted at the START of this function
    # (atomic INSERT ... ON CONFLICT DO NOTHING). We no longer record
    # the send a second time at the end — that double-write is what
    # used to leave a window for duplicates when two workers raced.

    # Stash the per-recipient breakdown so the admin "Send Reports Now"
    # endpoint can surface it back to the UI.
    _send_monthly_emails.last_diag = diag
    _send_monthly_emails.last_count = sent_count

    return sent_count


def _gen_excel_returns(data):
    """Extract the returns Excel generation logic for reuse in email sending."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    LABEL_MAP = {"LP IRR": "Net Cashflow", "LP Equity Multiple": "Cumulative Net Cashflow"}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Project Returns"

    PROJ_FILL   = PatternFill("solid", fgColor="F2EFE8")
    SUMM_FILL   = PatternFill("solid", fgColor="E8F0EE")
    HEADER_FILL = PatternFill("solid", fgColor="F7F6F3")
    thin = Side(style="thin", color="CCCCCC")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    TEXT="1A1A1A"; HDR_TEXT="555555"; PROJ_TEXT="6B4E1E"; SUMM_TEXT="2D6B5A"; ACCENT="7A5C1E"

    def _f(bold=False, color=TEXT, size=9):
        return Font(name="Calibri", size=size, bold=bold, color=color)
    def _set_num(cell, val):
        if isinstance(val, (int, float)) and val != 0:
            cell.value = val; cell.number_format = "#,##0"
        else:
            cell.value = None

    years = data.get("years", [])
    all_idxs = list(range(len(years)))
    num_cols = 2 + len(years)
    r = 1
    ws.cell(row=r, column=1, value="Consolidated Ember Project Returns").font = Font(name="Calibri", bold=True, size=14, color=PROJ_TEXT)
    r += 1
    ws.cell(row=r, column=1, value="($ in 000s)").font = _f(color="888888")
    r += 2

    SUMMARY_HDR_FILL = PatternFill("solid", fgColor="EDE8DF")
    summary_cols = ["Project", "LP IRR", "Equity Multiple", "Total LP Profit", "Promote"]
    for ci, h in enumerate(summary_cols, 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = _f(bold=True, color=HDR_TEXT); c.fill = SUMMARY_HDR_FILL; c.border = cell_border
        c.alignment = Alignment(horizontal="left" if ci==1 else "center")
    r += 1
    for proj in data.get("projects", []):
        metrics_by_label = {m["label"]: m for m in proj.get("metrics", [])}
        irr_val = metrics_by_label.get("LP IRR", {}).get("total"); em_val = metrics_by_label.get("LP Equity Multiple", {}).get("total")
        pft_val = metrics_by_label.get("Total LP Profit", {}).get("total"); prom_val = metrics_by_label.get("Promote", {}).get("total")
        nc = ws.cell(row=r, column=1, value=proj["name"]); nc.font = _f(bold=True, color=PROJ_TEXT); nc.border = cell_border
        ic = ws.cell(row=r, column=2); ic.font = _f(bold=True, color=ACCENT); ic.alignment = Alignment(horizontal="right"); ic.border = cell_border
        if isinstance(irr_val, (int, float)) and irr_val: ic.value = irr_val; ic.number_format = "0.0%"
        ec = ws.cell(row=r, column=3); ec.font = _f(bold=True, color=ACCENT); ec.alignment = Alignment(horizontal="right"); ec.border = cell_border
        if isinstance(em_val, (int, float)) and em_val: ec.value = em_val; ec.number_format = '0.00"x"'
        pc = ws.cell(row=r, column=4); pc.font = _f(); pc.alignment = Alignment(horizontal="right"); pc.border = cell_border; _set_num(pc, pft_val)
        prc = ws.cell(row=r, column=5); prc.font = _f(); prc.alignment = Alignment(horizontal="right"); prc.border = cell_border; _set_num(prc, prom_val)
        r += 1
    r += 1

    def write_section_header(r, title, fill, color):
        c = ws.cell(row=r, column=1, value=title); c.font = Font(name="Calibri", bold=True, size=10, color=color); c.fill = fill; c.border = cell_border
        for ci in range(2, num_cols+1): cell=ws.cell(row=r, column=ci); cell.fill=fill; cell.border=cell_border
        return r+1
    def write_col_headers(r, col_labels):
        ws.cell(row=r, column=1, value="Metric").font = _f(bold=True, color=HDR_TEXT); ws.cell(row=r, column=1).fill=HEADER_FILL; ws.cell(row=r, column=1).border=cell_border
        ws.cell(row=r, column=2, value="Total").font = _f(bold=True, color=HDR_TEXT); ws.cell(row=r, column=2).fill=HEADER_FILL; ws.cell(row=r, column=2).alignment=Alignment(horizontal="center"); ws.cell(row=r, column=2).border=cell_border
        for ci, lbl in enumerate(col_labels, 3):
            c=ws.cell(row=r, column=ci, value=lbl); c.font=_f(bold=True, color=HDR_TEXT); c.fill=HEADER_FILL; c.alignment=Alignment(horizontal="center"); c.border=cell_border
        return r+1
    def write_project(r, proj):
        r=write_section_header(r, proj["name"], PROJ_FILL, PROJ_TEXT); r=write_col_headers(r, years)
        for m in proj.get("metrics", []):
            label=m["label"]; display=LABEL_MAP.get(label, label); is_accent=label in ("LP IRR","LP Equity Multiple"); txt_color=ACCENT if is_accent else TEXT
            total = sum(v for v in m.get("yearly",[]) if isinstance(v,(int,float))) if label=="LP IRR" else ([v for v in m.get("yearly",[]) if isinstance(v,(int,float)) and v!=0] or [0])[-1] if label=="LP Equity Multiple" else m.get("total",0)
            lc=ws.cell(row=r, column=1, value=display); lc.font=_f(bold=is_accent, color=txt_color); lc.border=cell_border
            tc=ws.cell(row=r, column=2); tc.font=_f(bold=is_accent, color=txt_color); tc.alignment=Alignment(horizontal="right"); tc.border=cell_border; _set_num(tc, total)
            for ci, i in enumerate(all_idxs, 3):
                yc=ws.cell(row=r, column=ci); val=m["yearly"][i] if i<len(m.get("yearly",[])) else 0; yc.font=_f(color=txt_color); yc.alignment=Alignment(horizontal="right"); yc.border=cell_border; _set_num(yc, val)
            r+=1
        return r+1

    for proj in data.get("projects", []):
        r = write_project(r, proj)
    summary = data.get("summary", [])
    if summary:
        r=write_section_header(r, "Portfolio Summary", SUMM_FILL, SUMM_TEXT); r=write_col_headers(r, years)
        for s in summary:
            lc=ws.cell(row=r, column=1, value=s["label"]); lc.font=_f(); lc.border=cell_border
            tc=ws.cell(row=r, column=2); tc.font=_f(); tc.alignment=Alignment(horizontal="right"); tc.border=cell_border; _set_num(tc, s.get("total",0))
            for ci, i in enumerate(all_idxs, 3):
                yc=ws.cell(row=r, column=ci); val=s["yearly"][i] if i<len(s.get("yearly",[])) else 0; yc.font=_f(); yc.alignment=Alignment(horizontal="right"); yc.border=cell_border; _set_num(yc, val)
            r+=1

    ws.column_dimensions["A"].width=32; ws.column_dimensions["B"].width=13
    ws.column_dimensions["C"].width=14; ws.column_dimensions["D"].width=14; ws.column_dimensions["E"].width=13
    for ci in range(6, 3+len(years)): ws.column_dimensions[get_column_letter(ci)].width=11

    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out.read()


def _gen_excel_operations(data):
    """Extract the operations Excel generation logic for reuse in email sending."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Operating Revenues"
    GOLD="C8A96E"; HEADER_FILL=PatternFill("solid", fgColor="1E2535"); TOTALS_FILL=PatternFill("solid", fgColor="161B24")
    thin=Side(style="thin", color="2E3750"); cell_border=Border(left=thin, right=thin, top=thin, bottom=thin)
    def _hdr_font(bold=False): return Font(name="Calibri", size=9, bold=bold, color="8B95A8")
    def _val_font(bold=False): return Font(name="Calibri", size=9, bold=bold)
    def write_section(r, title):
        c=ws.cell(row=r, column=1, value=title); c.font=Font(name="Calibri", size=11, bold=True, color=GOLD); return r+1
    def write_table(r, col_headers, data_rows, totals):
        for ci, h in enumerate(col_headers, 1):
            c=ws.cell(row=r, column=ci, value=h); c.font=_hdr_font(bold=True); c.fill=HEADER_FILL; c.border=cell_border; c.alignment=Alignment(horizontal="left" if ci==1 else "center")
        r+=1
        for ri, row_data in enumerate(data_rows):
            for ci, val in enumerate(row_data, 1):
                c=ws.cell(row=r, column=ci, value=val if val else None); c.font=_val_font(); c.border=cell_border; c.alignment=Alignment(horizontal="left" if ci==1 else "right")
                if ci>1 and isinstance(val, (int,float)): c.number_format="#,##0"
            r+=1
        ws.cell(row=r, column=1, value="Total").font=_val_font(bold=True); ws.cell(row=r, column=1).border=cell_border; ws.cell(row=r, column=1).fill=TOTALS_FILL; ws.cell(row=r, column=1).alignment=Alignment(horizontal="left")
        for ci, v in enumerate(totals, 2):
            cell=ws.cell(row=r, column=ci, value=v if v else None); cell.font=_val_font(bold=True); cell.fill=TOTALS_FILL; cell.border=cell_border; cell.alignment=Alignment(horizontal="right")
            if isinstance(v, (int,float)): cell.number_format="#,##0"
        return r+2

    r=1; ws.cell(row=r, column=1, value="Ember Operating Revenues").font=Font(name="Calibri", bold=True, size=14, color=GOLD); r+=2
    r=write_section(r, "KPI Summary")
    for kpi in data.get("kpis", []):
        ws.cell(row=r, column=1, value=kpi["label"]).font=_val_font()
        vc=ws.cell(row=r, column=2, value=kpi["value"]); vc.font=_val_font(bold=True); vc.number_format="#,##0"; vc.alignment=Alignment(horizontal="right"); r+=1
    r+=1
    yr=data.get("yearly_rollup",{})
    if yr.get("years"):
        r=write_section(r, "Annual Revenue Forecast (Next 5 Years)")
        r=write_table(r, ["Revenue Source"]+[str(y) for y in yr["years"]], [[row["label"]]+row["values"] for row in yr.get("rows",[])], yr.get("totals",[]))
    mo=data.get("monthly",{})
    if mo.get("dates"):
        r=write_section(r, "Monthly Fee Revenue")
        r=write_table(r, ["Project / Category"]+[f"{d[5:7]}/{d[2:4]}" for d in mo["dates"]], [[f"{row['project']} — {row['category']}"]+row["values"] for row in mo.get("rows",[])], mo.get("totals",[]))
    n12=data.get("next_12_months",{})
    if n12.get("dates"):
        r=write_section(r, "Next 12 Months")
        r=write_table(r, ["Revenue Source"]+[f"{d[5:7]}/{d[2:4]}" for d in n12["dates"]], [[row["label"]]+row["values"] for row in n12.get("rows",[])], n12.get("totals",[]))
    qr=data.get("quarterly_rollup",{})
    if qr.get("quarters"):
        r=write_section(r, "Next 12 Quarters")
        r=write_table(r, ["Revenue Source"]+qr["quarters"], [[row["label"]]+row["values"] for row in qr.get("rows",[])], qr.get("totals",[]))

    ws.column_dimensions["A"].width=36
    for ci in range(2, 50): ws.column_dimensions[get_column_letter(ci)].width=11
    out=io.BytesIO(); wb.save(out); out.seek(0)
    return out.read()


# ─── MACRO DASHBOARD ──────────────────────────────────────────────────────────

@app.route("/macro")
@login_required
def macro_dashboard():
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("macro", True):
        return redirect(url_for("home"))
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT data, uploaded_at FROM reports WHERE report_type = 'macro' ORDER BY uploaded_at DESC LIMIT 1")
    row = cur.fetchone(); cur.close(); conn.close()
    data = row["data"] if row else None
    # Unified date format ("10 May 2026") used by both the header
    # "Last updated:" line and the new data-dates footer partial.
    uploaded_at = _fmt_data_date(row["uploaded_at"]) if row else None
    pa = session.get("page_access") or {}
    if session.get("is_admin"):
        pa = {k: True for k in ["mpc_underwriting","returns","loans","operations","macro","portfolio"]}

    # Fetch FRED series in parallel
    fred_data = {}
    fred_key = os.environ.get("FRED_API_KEY")
    if fred_key:
        start_18m = (datetime.datetime.now() - datetime.timedelta(days=548)).strftime("%Y-%m-%d")
        start_3yr = (datetime.datetime.now() - datetime.timedelta(days=1095)).strftime("%Y-%m-%d")
        fred_series = [
            ("MORTGAGE30US", start_18m, {}),
            ("FEDFUNDS",     start_3yr, {}),
            ("DPRIME",       start_3yr, {"frequency": "m", "aggregation_method": "avg"}),
            ("HSN1F",        start_3yr, {}),
            ("HOUSTS",       start_3yr, {}),
            ("MSPNHSUS",     start_3yr, {}),
            ("MSACSR",       start_3yr, {}),
            ("UMCSENT",      start_3yr, {}),
            ("CPIAUCSL",     start_3yr, {}),
            ("CUSR0000SEHC", start_3yr, {}),
        ]

        def _fetch_fred(sid, start, extra):
            params = {"series_id": sid, "api_key": fred_key,
                      "observation_start": start, "file_type": "json"}
            params.update(extra)
            for attempt in range(2):  # one retry on failure
                try:
                    r = requests.get("https://api.stlouisfed.org/fred/series/observations",
                                     params=params, timeout=10)
                    if r.ok:
                        obs = [o for o in r.json().get("observations", []) if o["value"] != "."]
                        return sid, {"dates": [o["date"] for o in obs],
                                     "values": [float(o["value"]) for o in obs]}
                    print(f"FRED {sid} {r.status_code}: {r.text[:120]}", flush=True)
                    break  # don't retry on HTTP errors (bad key, bad series, etc.)
                except Exception as e:
                    print(f"FRED {sid} attempt {attempt+1} error: {e}", flush=True)
            return sid, None

        # max_workers=5 avoids hitting FRED rate limits with simultaneous requests
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as ex:
            for sid, result in ex.map(lambda t: _fetch_fred(*t), fred_series):
                if result:
                    fred_data[sid] = result
    else:
        print("FRED_API_KEY not set", flush=True)

    # Data dates footer for macro: the parser doesn't yet expose a
    # per-series "as of" date, so we mirror the upload/refresh timestamp
    # for both "Data From" and "Uploaded". Can swap "Data From" to the
    # latest data month per series in a follow-up.
    return render_template("macro.html", data=data, uploaded_at=uploaded_at,
                           data_from=uploaded_at,
                           is_admin=session.get("is_admin"), page_access=pa,
                           fred_data=fred_data)


@app.route("/api/fred-audit")
@login_required
def fred_audit():
    """Check latest observation date for every series used in the dashboard."""
    if not session.get("is_admin"):
        return jsonify({"error": "forbidden"}), 403
    fred_key = os.environ.get("FRED_API_KEY")
    if not fred_key:
        return jsonify({"error": "FRED_API_KEY not set"})
    series_list = [
        "MORTGAGE30US", "FEDFUNDS", "DPRIME",
        "HSN1F", "HOUSTS", "MSPNHSUS", "MSACSR",
        "UMCSENT", "CPIAUCSL", "CUSR0000SEHC"
    ]
    def _latest(sid):
        try:
            r = requests.get("https://api.stlouisfed.org/fred/series/observations",
                params={"series_id": sid, "api_key": fred_key,
                        "sort_order": "desc", "limit": 1, "file_type": "json"}, timeout=8)
            if r.ok:
                obs = [o for o in r.json().get("observations", []) if o["value"] != "."]
                if obs:
                    return {"latest_date": obs[0]["date"], "latest_value": obs[0]["value"], "ok": True}
            return {"ok": False, "status": r.status_code, "error": r.text[:120]}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    with concurrent.futures.ThreadPoolExecutor(max_workers=5) as ex:
        results = dict(zip(series_list, ex.map(_latest, series_list)))
    return jsonify(results)


@app.route("/api/fred-test")
@login_required
def fred_test():
    if not session.get("is_admin"):
        return jsonify({"error": "forbidden"}), 403
    fred_key = os.environ.get("FRED_API_KEY")
    if not fred_key:
        return jsonify({"error": "FRED_API_KEY not set"})
    series_id = request.args.get("series", "MORTGAGE30US")
    try:
        resp = requests.get(
            "https://api.stlouisfed.org/fred/series/observations",
            params={"series_id": series_id, "api_key": fred_key,
                    "sort_order": "desc", "limit": 3, "file_type": "json"},
            timeout=8
        )
        data = resp.json() if resp.ok else {}
        obs = data.get("observations", [])
        return jsonify({
            "series": series_id, "status": resp.status_code, "ok": resp.ok,
            "latest_3": [{"date": o["date"], "value": o["value"]} for o in obs if o["value"] != "."],
            "body_preview": resp.text[:300] if not resp.ok else None
        })
    except Exception as e:
        return jsonify({"error": str(e)})


@app.route("/api/macro-status")
@login_required
def macro_status():
    if not session.get("is_admin"):
        return jsonify({"error": "forbidden"}), 403
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT uploaded_at FROM reports WHERE report_type = 'macro' ORDER BY uploaded_at DESC LIMIT 1")
    row = cur.fetchone(); cur.close(); conn.close()
    # Return ISO 8601 with timezone offset so the home-page status JS can
    # format the timestamp in the viewer's local timezone instead of the
    # server's. uploaded_at is a naive UTC timestamp from Postgres → tag
    # it tz-aware before isoformat() so the offset is explicit.
    if row and row["uploaded_at"]:
        ts = row["uploaded_at"]
        if ts.tzinfo is None:
            ts = ts.replace(tzinfo=datetime.timezone.utc)
        last_data_iso = ts.isoformat()
    else:
        last_data_iso = None
    return jsonify({**_refresh_state, "last_data_stored_iso": last_data_iso})


@app.route("/api/refresh-macro", methods=["POST"])
@login_required
def refresh_macro():
    if not session.get("is_admin"):
        return jsonify({"error": "forbidden"}), 403
    if _refresh_state["running"]:
        return jsonify({"ok": False, "message": "Refresh already in progress"})
    t = threading.Thread(target=_do_macro_refresh, daemon=True)
    t.start()
    return jsonify({"ok": True, "message": "Refresh started"})


@app.route("/api/upload-macro", methods=["POST"])
@login_required
def upload_macro():
    if not session.get("is_admin"):
        return jsonify({"error": "Admin only"}), 403
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file"}), 400
    try:
        file_bytes = f.read()
        data = parse_macro(file_bytes)
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM reports WHERE report_type = 'macro'")
    cur.execute("INSERT INTO reports (report_type, data, uploaded_by) VALUES (%s, %s, %s)",
                ("macro", json.dumps(data), session["user_id"]))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True, "counties": data.get("counties", [])})


# ─── COMMUNITY SALES TRACKER ──────────────────────────────────────────────────
@app.route("/sales")
@login_required
def sales_dashboard():
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("sales", True):
        return redirect(url_for("home"))
    pa = session.get("page_access") or {}
    if session.get("is_admin"):
        pa = {k: True for k in ["mpc_underwriting","returns","loans","operations","macro","portfolio","sales"]}
    return render_template("sales.html",
        username=session.get("username"),
        is_admin=session.get("is_admin"),
        page_access=pa,
        pipsy_configured=bool(os.environ.get("PIPSY_API_TOKEN")))


@app.route("/api/sales-data")
@login_required
def sales_data():
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("sales", True):
        return jsonify({"error": "Access denied"}), 403
    force = request.args.get("refresh") == "1" and session.get("is_admin")
    try:
        data = get_sales_dashboard_data(force_refresh=force)
        return jsonify(data)
    except RuntimeError as e:
        # Missing API token or Pipsy error — surface a clear message
        return jsonify({"error": str(e)}), 503
    except Exception as e:
        return jsonify({"error": f"Unexpected error: {e}"}), 500


def _parse_pipsy_property() -> int:
    """Resolve the ?property= query param to a numeric Pipsy property id.
    Accepts the numeric ids directly (1=GPD, 2=WRG, 3=HLD) or the short
    codes 'gpd'/'wrg'/'hld'. Defaults to GPD when unset/invalid so the
    legacy single-property callers don't break."""
    raw = (request.args.get("property") or "1").strip().lower()
    aliases = {"gpd": 1, "wrg": 2, "hld": 3}
    if raw in aliases:
        return aliases[raw]
    try:
        n = int(raw)
        return n if n in (1, 2, 3) else 1
    except ValueError:
        return 1


@app.route("/api/pipsy/home-sales")
@login_required
def pipsy_home_sales():
    """Net home sales (gross − cancel) bucketed by section + YYYY-MM.

    Used by the UW Actuals overlay: returns
    `{ "property": N, "sales": { "Section 1": { "2025-03": 4, ... } } }`.
    Same gate as /api/sales-data — admins or any user whose page_access
    grants the sales tab."""
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("sales", True):
        return jsonify({"error": "Access denied"}), 403
    prop = _parse_pipsy_property()
    try:
        data = pipsy_home_sales_by_section(prop)
        return jsonify({"property": prop, "sales": data})
    except RuntimeError as e:
        # Missing API token / GraphQL error — same shape as /api/sales-data
        return jsonify({"error": str(e)}), 503
    except Exception as e:
        return jsonify({"error": f"Unexpected error: {e}"}), 500


@app.route("/api/pipsy/lot-takedowns")
@login_required
def pipsy_lot_takedowns():
    """Lot takedowns bucketed by section + YYYY-MM of takedown_date.

    Used by the UW Actuals overlay: returns
    `{ "property": N, "takedowns": { "Section 1": { "2025-03": 2 } } }`.
    Same access gate as /api/pipsy/home-sales."""
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("sales", True):
        return jsonify({"error": "Access denied"}), 403
    prop = _parse_pipsy_property()
    try:
        data = pipsy_lot_takedowns_by_section(prop)
        return jsonify({"property": prop, "takedowns": data})
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 503
    except Exception as e:
        return jsonify({"error": f"Unexpected error: {e}"}), 500


@app.route("/api/upload-bohlke", methods=["POST"])
@login_required
def upload_bohlke():
    """Admin-only — accepts a Bohlke competitive-set xlsx, parses it,
    and stores the JSON in the reports table (report_type='bohlke')."""
    if not session.get("is_admin"):
        return jsonify({"error": "Admin only"}), 403
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file"}), 400
    try:
        file_bytes = f.read()
        data = parse_bohlke(file_bytes)
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM reports WHERE report_type = 'bohlke'")
    cur.execute("INSERT INTO reports (report_type, data, uploaded_by) VALUES (%s, %s, %s)",
                ("bohlke", json.dumps(data), session["user_id"]))
    conn.commit(); cur.close(); conn.close()
    return jsonify({
        "ok": True,
        "yearly_rows": len(data.get("yearly", [])),
        "comps": len(data.get("comps", [])),
        "tgp_builders": len(data.get("tgp_builders", [])),
    })


@app.route("/api/bohlke-data")
@login_required
def bohlke_data():
    """Return the latest stored Bohlke report data, or an empty shell
    if nothing has been uploaded yet."""
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("sales", True):
        return jsonify({"error": "Access denied"}), 403
    conn = get_db(); cur = conn.cursor()
    cur.execute("""
        SELECT data, uploaded_at, uploaded_by
        FROM reports
        WHERE report_type = 'bohlke'
        ORDER BY uploaded_at DESC
        LIMIT 1
    """)
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row:
        return jsonify({"yearly": [], "comps": [], "tgp_builders": [], "uploaded_at": None})
    data = row["data"]
    uploaded_at = row["uploaded_at"]
    # psycopg2 may return JSONB as dict already; fall back to json.loads if str
    if isinstance(data, str):
        data = json.loads(data)
    data["uploaded_at"] = uploaded_at.isoformat() if uploaded_at else None
    return jsonify(data)


@app.route("/api/upload-waller-monthly", methods=["POST"])
@login_required
def upload_waller_monthly():
    """Admin-only — accepts the raw 'Waller ISD Comps by Month' xlsx and
    stores the parsed JSON in reports (report_type='waller_monthly')."""
    if not session.get("is_admin"):
        return jsonify({"error": "Admin only"}), 403
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file"}), 400
    try:
        data = parse_waller_monthly(f.read())
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM reports WHERE report_type = 'waller_monthly'")
    cur.execute("INSERT INTO reports (report_type, data, uploaded_by) VALUES (%s, %s, %s)",
                ("waller_monthly", json.dumps(data), session["user_id"]))
    conn.commit(); cur.close(); conn.close()
    return jsonify({
        "ok": True,
        "months": len(data.get("months", [])),
        "communities": len(data.get("communities", [])),
        "has_waller_total": data.get("waller_total") is not None,
    })


@app.route("/api/waller-monthly-data")
@login_required
def waller_monthly_data():
    """Return the latest stored Waller Monthly report, or an empty shell."""
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("sales", True):
        return jsonify({"error": "Access denied"}), 403
    conn = get_db(); cur = conn.cursor()
    cur.execute("""
        SELECT data, uploaded_at, uploaded_by
        FROM reports
        WHERE report_type = 'waller_monthly'
        ORDER BY uploaded_at DESC
        LIMIT 1
    """)
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row:
        return jsonify({"months": [], "waller_total": None, "communities": [],
                        "uploaded_at": None})
    data = row["data"]
    if isinstance(data, str):
        data = json.loads(data)
    data["uploaded_at"] = row["uploaded_at"].isoformat() if row["uploaded_at"] else None
    return jsonify(data)


@app.route("/api/upload-hpermits", methods=["POST"])
@login_required
def upload_hpermits():
    """Admin-only — accepts the raw 'HPermits' xlsx and stores the parsed
    JSON in reports (report_type='hpermits')."""
    if not session.get("is_admin"):
        return jsonify({"error": "Admin only"}), 403
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file"}), 400
    try:
        data = parse_hpermits(f.read())
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM reports WHERE report_type = 'hpermits'")
    cur.execute("INSERT INTO reports (report_type, data, uploaded_by) VALUES (%s, %s, %s)",
                ("hpermits", json.dumps(data), session["user_id"]))
    conn.commit(); cur.close(); conn.close()
    return jsonify({
        "ok": True,
        "markets": len(data.get("markets", [])),
        "builders": len(data.get("builders", [])),
        "top_projects": len(data.get("top_projects", [])),
    })


@app.route("/api/hpermits-data")
@login_required
def hpermits_data():
    """Return the latest stored HPermits report, or an empty shell."""
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("sales", True):
        return jsonify({"error": "Access denied"}), 403
    conn = get_db(); cur = conn.cursor()
    cur.execute("""
        SELECT data, uploaded_at, uploaded_by
        FROM reports
        WHERE report_type = 'hpermits'
        ORDER BY uploaded_at DESC
        LIMIT 1
    """)
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row:
        return jsonify({
            "months_prev": [], "months_curr": [],
            "ttm_prev_label": "", "ttm_curr_label": "",
            "grand_total": None, "markets": [], "submarkets": [],
            "builders": [], "companies": [], "top_projects": [],
            "uploaded_at": None,
        })
    data = row["data"]
    if isinstance(data, str):
        data = json.loads(data)
    data["uploaded_at"] = row["uploaded_at"].isoformat() if row["uploaded_at"] else None
    return jsonify(data)


def _UW_EMPTY_PROP():
    return {"home_sales": {}, "lot_takedowns": {}, "bem": {}, "section_lots": {}, "sheet_names": []}


def _normalize_uw_payload(raw) -> dict:
    """Normalize whatever's in the DB to the per-property shape
    `{ gpd: {...}, wrg: {...} }`. Legacy rows were stored flat (the GPD-only
    era) — we lift those into the gpd slot so old uploads keep rendering.
    """
    if isinstance(raw, str):
        raw = json.loads(raw)
    if not isinstance(raw, dict):
        return {"gpd": _UW_EMPTY_PROP(), "wrg": _UW_EMPTY_PROP()}
    # Detect new (per-property) shape: has a 'gpd' or 'wrg' key with a dict value.
    looks_nested = any(
        isinstance(raw.get(k), dict) and {"home_sales", "lot_takedowns", "bem"} & set(raw.get(k, {}).keys())
        for k in ("gpd", "wrg")
    )
    if looks_nested:
        return {
            "gpd": {**_UW_EMPTY_PROP(), **(raw.get("gpd") or {})},
            "wrg": {**_UW_EMPTY_PROP(), **(raw.get("wrg") or {})},
        }
    # Legacy flat -> gpd
    legacy = {**_UW_EMPTY_PROP(), **{k: raw.get(k) for k in _UW_EMPTY_PROP() if k in raw}}
    return {"gpd": legacy, "wrg": _UW_EMPTY_PROP()}


def _uw_property_from_request() -> str:
    """Resolve the ?property= form/query param to 'gpd' or 'wrg'. Default
    gpd so legacy uploaders that don't pass the param keep working."""
    raw = (request.values.get("property") or "gpd").strip().lower()
    return "wrg" if raw == "wrg" else "gpd"


@app.route("/api/upload-uw", methods=["POST"])
@login_required
def upload_uw():
    """Admin-only — accepts the raw 'UW Performance Export' xlsx for the
    property indicated by `?property=gpd|wrg` (default gpd). Parses the
    workbook, merges into the existing per-property payload (so uploading
    GPD doesn't wipe WRG and vice versa), and saves under
    report_type='uw_performance'."""
    if not session.get("is_admin"):
        return jsonify({"error": "Admin only"}), 403
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file"}), 400
    prop = _uw_property_from_request()
    try:
        data = parse_uw(f.read())
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    conn = get_db(); cur = conn.cursor()
    # Load the existing payload so we merge instead of clobbering the other property.
    cur.execute(
        "SELECT data FROM reports WHERE report_type = 'uw_performance' "
        "ORDER BY uploaded_at DESC LIMIT 1"
    )
    row = cur.fetchone()
    existing = _normalize_uw_payload(row["data"] if row else None)
    existing[prop] = data  # Full replacement for the uploaded property
    cur.execute("DELETE FROM reports WHERE report_type = 'uw_performance'")
    cur.execute("INSERT INTO reports (report_type, data, uploaded_by) VALUES (%s, %s, %s)",
                ("uw_performance", json.dumps(existing), session["user_id"]))
    conn.commit(); cur.close(); conn.close()
    return jsonify({
        "ok": True,
        "property": prop,
        "home_sales_sections":   len(data.get("home_sales", {})),
        "lot_takedowns_sections": len(data.get("lot_takedowns", {})),
        "bem_sections":          len(data.get("bem", {})),
    })


@app.route("/api/uw-data")
@login_required
def uw_data():
    """Return the latest stored UW Performance report, normalized to the
    per-property shape `{ gpd: {...}, wrg: {...}, uploaded_at }`. Always
    returns both slots — empty shells if a property hasn't been uploaded."""
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("sales", True):
        return jsonify({"error": "Access denied"}), 403
    conn = get_db(); cur = conn.cursor()
    cur.execute("""
        SELECT data, uploaded_at, uploaded_by
        FROM reports
        WHERE report_type = 'uw_performance'
        ORDER BY uploaded_at DESC
        LIMIT 1
    """)
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row:
        return jsonify({
            "gpd": _UW_EMPTY_PROP(), "wrg": _UW_EMPTY_PROP(), "uploaded_at": None,
        })
    payload = _normalize_uw_payload(row["data"])
    payload["uploaded_at"] = row["uploaded_at"].isoformat() if row["uploaded_at"] else None
    return jsonify(payload)


@app.route("/api/starts", methods=["GET", "POST"])
@login_required
def starts_data():
    """Editable starts matrix per community (GPD/WRG).

    GET returns the latest saved payload `{ gpd, wrg, uploaded_at }`,
    where each community is `{ months: [..], dates: [..], rows: [{lot,
    bld, sec, del, metric, vals: [..]}], cellEdits?: {...} }`. Empty
    object if nothing has been saved.

    POST (admin only) replaces the saved payload. Sent body:
    `{ data: { gpd: {...}, wrg: {...} } }`. The body envelope matches
    /api/underwriting on the coworker dashboard so the same hydrate
    pattern works.
    """
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("sales", True):
        return jsonify({"error": "Access denied"}), 403

    if request.method == "GET":
        conn = get_db(); cur = conn.cursor()
        cur.execute("""
            SELECT data, uploaded_at, uploaded_by
            FROM reports
            WHERE report_type = 'starts'
            ORDER BY uploaded_at DESC
            LIMIT 1
        """)
        row = cur.fetchone()
        cur.close(); conn.close()
        if not row:
            return jsonify({"hasData": False, "data": {"gpd": None, "wrg": None}, "uploaded_at": None})
        payload = row["data"]
        if isinstance(payload, str):
            payload = json.loads(payload)
        return jsonify({
            "hasData": True,
            "data": payload if isinstance(payload, dict) else {"gpd": None, "wrg": None},
            "uploaded_at": row["uploaded_at"].isoformat() if row["uploaded_at"] else None,
            "uploaded_by": row["uploaded_by"],
        })

    # POST — admin only. Saves the full payload (single-row, latest wins).
    if not session.get("is_admin"):
        return jsonify({"error": "Admin only"}), 403
    body = request.get_json(silent=True) or {}
    data = body.get("data")
    if not isinstance(data, dict):
        return jsonify({"error": "Missing data object"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM reports WHERE report_type = 'starts'")
    cur.execute(
        "INSERT INTO reports (report_type, data, uploaded_by) VALUES (%s, %s, %s)",
        ("starts", json.dumps(data), session["user_id"]),
    )
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})


@app.route("/api/uw-template")
@login_required
def uw_template():
    """Serve the blank UW Performance Export template xlsx so admins can
    download it, fill it in, and re-upload. Same file for GPD and WRG —
    only the section list differs, and the parser reads sheet labels."""
    pa = session.get("page_access") or {}
    if not session.get("is_admin") and not pa.get("sales", True):
        return jsonify({"error": "Access denied"}), 403
    prop = _uw_property_from_request()
    from flask import send_from_directory
    template_dir = os.path.join(app.root_path, "static", "uw_templates")
    filename = "UW Performance Export.xlsx"
    return send_from_directory(
        template_dir, filename,
        as_attachment=True,
        download_name=f"UW Performance Export - {prop.upper()}.xlsx",
    )


# ─── Health endpoint (sales-dashboard parity item #1) ────────────────────────
# Trivial liveness probe for uptime monitoring. Returns 200 OK with a
# tiny JSON body. Doesn't touch the DB or Sage — intentionally cheap.
@app.route("/api/health")
def api_health():
    return jsonify({"status": "ok"})


# ─── Nightly sales-data refresh (sales-dashboard parity item #2) ─────────────
# Warms the in-memory Pipsy cache so the first morning page-view of
# /sales doesn't pay the cold-start cost (~10–20 sec to pull six
# GraphQL pages across the three properties).
def _refresh_sales_cache():
    """Background job — force-pull fresh sales data from Pipsy and
    discard the result. The cache inside sales_parser stays warm
    for the next request. Logs+swallows exceptions so a Pipsy outage
    can't kill the scheduler thread."""
    try:
        get_sales_dashboard_data(force_refresh=True)
        print("[scheduler] sales cache refreshed from Pipsy")
    except Exception as e:
        print(f"[scheduler] sales cache refresh failed: {e}")


# ─── SCHEDULER ────────────────────────────────────────────────────────────────
def _start_scheduler():
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
        # Pin to America/Chicago so the schedule shifts with DST (CST/CDT).
        # zoneinfo is stdlib on Python 3.9+; falls back to pytz if not.
        try:
            from zoneinfo import ZoneInfo
            tz = ZoneInfo("America/Chicago")
        except ImportError:
            import pytz
            tz = pytz.timezone("America/Chicago")
        scheduler = BackgroundScheduler()
        # Run on the 1st of every month at 9:00 AM Central time.
        scheduler.add_job(
            _send_monthly_emails, "cron",
            day=1, hour=9, minute=0,
            timezone=tz,
            id="monthly_reports",
        )
        # Daily Pipsy sales-data refresh at 06:30 America/Chicago,
        # so the first morning page-view of /sales hits a warm cache.
        scheduler.add_job(
            _refresh_sales_cache, "cron",
            hour=6, minute=30,
            timezone=tz,
            id="sales_cache_daily",
        )
        scheduler.start()
        print("APScheduler started — monthly reports (1st @ 09:00 CT) + "
              "sales cache warm-up (daily @ 06:30 CT)")
    except Exception as e:
        print(f"Scheduler failed to start: {e}")

_start_scheduler()


if __name__ == "__main__":
    init_db()
    app.run(debug=True, port=5001)
