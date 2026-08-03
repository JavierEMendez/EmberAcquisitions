"""
unit_economics_parser.py — Parses the "Unit Economics" tab of a master-planned
community pro-forma model (one workbook = one entity) into structured JSON.

The tab holds one ~49-row block per section (Revenues / Costs line items across
To Date, Remaining, Total, $/FF, $/Lot, $/Acre, % of Costs, % of Rev), an
"Additional Info" panel per section (front feet, acreage, lots, land purchase),
per-phase stats tables, and an entity-wide "Community Rollup" block.

Everything is located by label, never by fixed row/column: sections get added,
removed, and re-phased between model versions. Only section-level data and the
entity rollup are trusted from the workbook — phase and cross-entity rollups
are recomputed by the app (verified to match the model's own rollups exactly).

Layout facts the parser relies on (stable across model versions):
  * Section title cell matches "Section N" and the row below it, same column,
    is "Revenues" with "To Date" one cell right ("Total" header distinguishes
    the dollar block from the parallel "(Ks)" block).
  * The section's phase tag ("Phase N") sits in the same row, a few columns
    right of the title.
  * Sub-line-items (e.g. Lot Premiums under "Premiums, Escalations, Fence
    Fees") are indented one level; summary rows (Total, Gross Costs, Net
    Margin, ...) are bold — styles carry the hierarchy, so the parser loads
    the workbook with styles (not read_only).
"""

import io
import re
from datetime import date, datetime
from typing import Any

import openpyxl


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_SECTION_RE = re.compile(r"^Section\s+(\d+)$")
_PHASE_RE = re.compile(r"^Phase\s+(\d+)$")

# Sentinel that terminates a unit-economics block.
_LAST_ROW_LABEL = "net margin"

# Additional Info labels → JSON keys.
_INFO_KEYS = {
    "total front feet": "total_front_feet",
    "total acreage": "total_acreage",
    "total lots": "total_lots",
    "phase lots": "phase_lots",
    "blended price per acre": "blended_price_per_acre",
    "land purchase": "land_purchase",
    "price per acre": "price_per_acre",
    "phase acreage": "phase_acreage",
    "life of project acreage": "life_of_project_acreage",
    "life of project front feet": "life_of_project_front_feet",
    "phase front feet": "phase_front_feet",
}


def _num(val: Any) -> float | int | None:
    """Cell value → number, preserving None (blank) as None."""
    if val is None:
        return None
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return None
        try:
            val = float(s.replace(",", "").replace("$", ""))
        except ValueError:
            return None
    try:
        n = float(val)
    except (TypeError, ValueError):
        return None
    if n == int(n) and abs(n) < 1e15:
        return int(n)
    return round(n, 6)


def _str(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _date_iso(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    return str(val).strip()


def _find_sheet(wb, needle: str = "unit economics"):
    for name in wb.sheetnames:
        if needle in name.strip().lower():
            return wb[name]
    return None


# ---------------------------------------------------------------------------
# Block reader
# ---------------------------------------------------------------------------

def _read_block_rows(ws, header_row: int, label_col: int, max_row: int) -> list[dict]:
    """Read a unit-economics block starting at its "Revenues" header row.

    Returns the rows in sheet order. Each row carries a `group`:
      revenue / revenue_total / cost / summary
    plus indent + bold flags so the UI can mirror the Excel presentation.
    Stops at "Net Margin" (inclusive) or the first gap after the block.
    """
    rows: list[dict] = []
    group = "revenue"
    r = header_row + 1
    blanks = 0
    while r <= max_row:
        cell = ws.cell(row=r, column=label_col)
        label = _str(cell.value)
        if not label:
            blanks += 1
            if blanks >= 3:
                break
            r += 1
            continue
        blanks = 0
        low = label.lower()
        if low == "costs":
            group = "cost"
            r += 1
            continue
        vals = [ws.cell(row=r, column=label_col + i).value for i in range(1, 9)]
        row = {
            "label": label,
            "group": group,
            "indent": 1 if (cell.alignment.indent or 0) >= 1 else 0,
            "bold": bool(cell.font.bold),
            "to_date": _num(vals[0]),
            "remaining": _num(vals[1]),
            "total": _num(vals[2]),
            "per_ff": _num(vals[3]),
            "per_lot": _num(vals[4]),
            "per_acre": _num(vals[5]),
            "pct_costs": _num(vals[6]),
            "pct_rev": _num(vals[7]),
        }
        if group == "revenue" and low == "total":
            row["group"] = "revenue_total"
            group = "cost"          # "Costs" header follows; skip handled above
        elif low in ("gross costs", "gross margin", "operations & overhead",
                     "financing", "net costs", "net margin"):
            row["group"] = "summary"
            group = "summary"
        rows.append(row)
        if low == _LAST_ROW_LABEL:
            break
        r += 1
    return rows


def _read_additional_info(ws, header_row: int, info_col: int, max_row: int) -> dict:
    """Read the label/value pairs under a "Section N | Additional Info" cell."""
    info: dict = {}
    for r in range(header_row + 1, min(header_row + 20, max_row + 1)):
        label = _str(ws.cell(row=r, column=info_col).value).lower()
        if not label:
            continue
        key = _INFO_KEYS.get(label)
        if not key:
            continue
        raw = ws.cell(row=r, column=info_col + 1).value
        info[key] = _str(raw) if key == "land_purchase" else _num(raw)
    return info


def _find_phase_tag(ws, row: int, start_col: int, span: int = 12) -> str:
    """Phase tag ("Phase N") sits a few columns right of the section title."""
    for c in range(start_col + 1, start_col + span + 1):
        v = _str(ws.cell(row=row, column=c).value)
        if _PHASE_RE.match(v):
            return v
    return ""


# ---------------------------------------------------------------------------
# Main entry
# ---------------------------------------------------------------------------

def parse_unit_economics(file_bytes: bytes) -> dict:
    """Parse the Unit Economics tab. Raises ValueError on a missing tab or
    if no section blocks are found."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = _find_sheet(wb)
    if ws is None:
        raise ValueError('Workbook has no "Unit Economics" tab')
    max_row = ws.max_row
    max_col = min(ws.max_column, 100)

    # Actuals date: labelled cell near the top ("Actuals Date:").
    actuals_date = ""
    for r in range(1, min(6, max_row) + 1):
        for c in range(1, 12):
            if _str(ws.cell(row=r, column=c).value).lower().startswith("actuals date"):
                actuals_date = _date_iso(ws.cell(row=r, column=c + 1).value)
                break
        if actuals_date:
            break

    # ── Section blocks ──────────────────────────────────────────────────
    # A section anchor is a "Section N" title whose next row (same column)
    # is "Revenues" followed by "To Date", and whose Total header is plain
    # "Total" — that separates the dollar block from the "(Ks)" mirror.
    sections = []
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            m = _SECTION_RE.match(_str(ws.cell(row=r, column=c).value))
            if not m:
                continue
            if _str(ws.cell(row=r + 1, column=c).value).lower() != "revenues":
                continue
            if _str(ws.cell(row=r + 1, column=c + 1).value).lower() != "to date":
                continue
            if _str(ws.cell(row=r + 1, column=c + 3).value).lower() != "total":
                continue
            phase = _find_phase_tag(ws, r, c)
            rows = _read_block_rows(ws, r + 1, c, max_row)
            if not rows:
                continue
            # Additional Info panel: same row, first matching cell right of
            # the block ("Section N | Additional Info").
            info = {}
            for ic in range(c + 9, min(c + 16, max_col + 1)):
                v = _str(ws.cell(row=r, column=ic).value)
                if v.lower().endswith("additional info"):
                    info = _read_additional_info(ws, r, ic, max_row)
                    break
            sections.append({
                "key": f"Section {m.group(1)}",
                "number": int(m.group(1)),
                "phase": phase,
                "phase_num": int(_PHASE_RE.match(phase).group(1)) if _PHASE_RE.match(phase) else 0,
                "info": info,
                "rows": rows,
            })
            break  # one section per row; skip the "(Ks)" mirror block
    if not sections:
        raise ValueError('No section blocks found on the "Unit Economics" tab')
    sections.sort(key=lambda s: s["number"])

    # ── Entity rollup ("Community Rollup" block in the model) ───────────
    # The model's rollup includes to-date history from closed-out sections
    # that no longer appear as blocks, so it is parsed rather than summed.
    entity_rollup = None
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            if _str(ws.cell(row=r, column=c).value).lower() != "community rollup":
                continue
            if _str(ws.cell(row=r + 1, column=c).value).lower() == "revenues":
                entity_rollup = _read_block_rows(ws, r + 1, c, max_row)
            break
        if entity_rollup:
            break

    # ── Per-phase stats tables (lots / front feet / acreage) ────────────
    # Parsed for entity-level denominators and validation; phase-level
    # denominators are recomputed from section info at render time.
    phase_stats: dict[str, dict] = {}
    stat_titles = {
        "total lots per phase": "lots",
        "total ff per phase": "front_feet",
        "total acreage per phase": "acreage",
    }
    entity_stats: dict[str, float] = {}
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            key = stat_titles.get(_str(ws.cell(row=r, column=c).value).lower())
            if not key:
                continue
            for rr in range(r + 1, min(r + 16, max_row + 1)):
                label = _str(ws.cell(row=rr, column=c).value)
                val = _num(ws.cell(row=rr, column=c + 1).value)
                if _PHASE_RE.match(label):
                    phase_stats.setdefault(label, {})[key] = val or 0
                elif label.lower().startswith("total"):
                    entity_stats[key] = val or 0
                    break
                elif not label:
                    break

    # Entity-level unit denominators. Life-of-project figures cover sold-out
    # sections too, matching what the model's own rollup divides by.
    info0 = sections[0]["info"] if sections else {}
    entity_units = {
        "front_feet": info0.get("life_of_project_front_feet") or entity_stats.get("front_feet") or 0,
        "acreage": info0.get("life_of_project_acreage") or entity_stats.get("acreage") or 0,
        "lots": entity_stats.get("lots") or sum((s["info"].get("total_lots") or 0) for s in sections),
    }

    return {
        "actuals_date": actuals_date,
        "sections": sections,
        "entity_rollup": entity_rollup,
        "entity_units": entity_units,
        "phase_stats": phase_stats,
    }


# ---------------------------------------------------------------------------
# Aggregation — blends N section blocks (or N entity rollups) into one block
# ---------------------------------------------------------------------------

def blend_blocks(row_sets: list[list[dict]], units: dict) -> list[dict]:
    """Sum dollar columns across blocks and recompute per-unit and percentage
    columns against the blended denominators. Row identity is (group, label);
    ordering follows the first block, with unseen rows appended in place.

    Verified against the model: its own phase rollups equal this blend of
    their member sections to the dollar.
    """
    order: list[tuple] = []
    merged: dict[tuple, dict] = {}
    for rows in row_sets:
        for row in rows:
            key = (row["group"], row["label"].lower())
            if key not in merged:
                merged[key] = {
                    "label": row["label"], "group": row["group"],
                    "indent": row["indent"], "bold": row["bold"],
                    "to_date": None, "remaining": None, "total": None,
                }
                order.append(key)
            tgt = merged[key]
            for f in ("to_date", "remaining", "total"):
                if row[f] is not None:
                    tgt[f] = (tgt[f] or 0) + row[f]

    ff = units.get("front_feet") or 0
    lots = units.get("lots") or 0
    acres = units.get("acreage") or 0
    out = [merged[k] for k in order]

    rev_total = next((r["total"] for r in out if r["group"] == "revenue_total"), None) or 0
    gross_costs = next((r["total"] for r in out
                        if r["group"] == "summary" and r["label"].lower() == "gross costs"), None) or 0

    for row in out:
        total = row["total"] or 0
        row["per_ff"] = round(total / ff, 2) if ff else None
        row["per_lot"] = round(total / lots, 2) if lots else None
        row["per_acre"] = round(total / acres, 2) if acres else None
        # Same convention as the model's section blocks: revenue rows are a
        # share of total revenue in both % columns; cost/summary rows are a
        # share of gross costs and of total revenue.
        if row["group"] in ("revenue", "revenue_total"):
            row["pct_costs"] = round(total / rev_total, 4) if rev_total else None
            row["pct_rev"] = round(total / rev_total, 4) if rev_total else None
        else:
            row["pct_costs"] = round(total / gross_costs, 4) if gross_costs else None
            row["pct_rev"] = round(total / rev_total, 4) if rev_total else None
    return out


def sum_units(infos: list[dict]) -> dict:
    """Per-unit denominators for a set of sections (their info panels)."""
    return {
        "front_feet": sum((i.get("total_front_feet") or 0) for i in infos),
        "acreage": round(sum((i.get("total_acreage") or 0) for i in infos), 4),
        "lots": sum((i.get("total_lots") or 0) for i in infos),
    }
